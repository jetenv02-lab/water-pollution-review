# -*- coding: utf-8 -*-
"""Step 3e: 規則庫驅動的檢查器。

核心想法:
之前的 step3b/3d 寫死 5-6 種檢查, 只能涵蓋規則庫的 10% 左右,
所以結果都集中在「快混槽展現重金屬去除」這一類。

本檢查器改成「規則庫驅動」:
- 讀規則庫.xlsx 各槽體分頁的 299 筆規則
- 對申請文件每個單元, 用該標準槽體類型對應的規則, 逐條嘗試比對
- 用「關鍵字模糊比對」啟發式判斷規則是否觸發
- 觸發 → 列出該條規則 + 標記為「需審查」

對照項目關鍵字 (例如):
  pH        → 看單元 measure_params 是否有 pH 範圍, 範圍是否 < 規則建議
  停留時間  → 看 design_params 是否有停留時間數值
  液位計    → 看 equipment 是否含液位計
  排泥      → 看 equipment 是否含排泥
  攪拌機    → 看 equipment 是否含攪拌
  污泥迴流  → 看是否有相關設施
  質量平衡  → 看進出流質量差異
  有效位數  → 看數值的小數位
  ...

依賴: openpyxl
使用:
    from step3e_rule_driven_check import run_rule_driven_check
    findings = run_rule_driven_check(app_data)
"""
import csv
import os
import re
from collections import defaultdict

BASE = os.path.dirname(os.path.abspath(__file__))
RULES_CSV = os.path.join(BASE, "rules_extracted.csv")


# ──────────────────────────────────────────────────
# 規則庫載入 (直接讀 rules_extracted.csv, 避免依賴本機才有的 xlsx)
# ──────────────────────────────────────────────────

def load_rules_by_tank():
    """讀 rules_extracted.csv → { 標準槽體名稱: [規則 dict, ...] }。

    跳過「狀態 = ?」(待討論) 的規則。
    同義字: 「對照項目」會多存一個 normalized 版本到 `對照項目_標準` 欄。
    """
    if not os.path.exists(RULES_CSV):
        return {}
    # 嘗試載入同義字 (失敗不致命, 退化成 identity)
    try:
        import step3f_synonyms
        synonym_normalize = step3f_synonyms.normalize
    except Exception:
        synonym_normalize = lambda x: x

    rules_by_tank = defaultdict(list)
    skipped = 0
    with open(RULES_CSV, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # 狀態欄: 空白 / V → 跑, ? → 跳過
            status = (row.get("狀態") or "").strip()
            if status == "?":
                skipped += 1
                continue
            tank = (row.get("標準槽體名稱") or "未分類").strip()
            raw_compare = row.get("對照項目", "")
            # 把 CSV 欄位名稱對應到原本 xlsx 的欄位名稱
            rule = {
                "缺失ID": row.get("缺失ID", ""),
                "來源": row.get("來源", ""),
                "原文缺失": row.get("原文缺失", ""),
                "檢查類型": row.get("檢查類型", ""),
                "對照項目": raw_compare,
                "對照項目_標準": synonym_normalize(raw_compare),  # 同義字展開後的標準詞
                "規則": row.get("規則", ""),
                "比對位置": row.get("比對位置", ""),
                "判定邏輯": row.get("判定邏輯", ""),
                "技師姓名": row.get("技師姓名", ""),
                "序號": row.get("序號", ""),
                "原始槽體代號": row.get("原始槽體代號", ""),
                "狀態": status,
            }
            rules_by_tank[tank].append(rule)
    # 把跳過數記到模組級全域 (Streamlit UI 可讀)
    globals()["LAST_SKIPPED_COUNT"] = skipped
    return dict(rules_by_tank)


# 上次載入規則時, 因「狀態 = ?」被跳過的筆數
LAST_SKIPPED_COUNT = 0


def get_last_skipped_count():
    """回傳上次載入規則時被「狀態 = ?」跳過的筆數。"""
    return globals().get("LAST_SKIPPED_COUNT", 0)


# ──────────────────────────────────────────────────
# 啟發式檢查器 (對照項目 → 檢查函式)
# ──────────────────────────────────────────────────

PH_RANGE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*[~～\-]\s*(\d+(?:\.\d+)?)")
NUMBER_RE = re.compile(r"(\d+(?:\.\d+)?)")


def get_param_range(params_dict, key_substring):
    """從 design_params 或 measure_params 找名稱含 key_substring 的參數,回傳 (min, max, pname)。

    修訂: 只要 min 或 max 其中一個有值就回傳, 由呼叫端判斷是否需標「僅登載單邊」。
    這樣才不會把「有登載但只填下限」的槽誤判成完全未登載。
    """
    for pname, pval in params_dict.items():
        if key_substring in pname:
            if isinstance(pval, dict):
                lo = pval.get("min")
                hi = pval.get("max")
                if lo is not None or hi is not None:
                    try:
                        lo_f = float(lo) if lo is not None else None
                    except (TypeError, ValueError):
                        lo_f = None
                    try:
                        hi_f = float(hi) if hi is not None else None
                    except (TypeError, ValueError):
                        hi_f = None
                    return lo_f, hi_f, pname
    return None, None, None


def has_equipment(equipment_list, name_keyword):
    """檢查機具清單是否含某關鍵字。"""
    for eq in equipment_list:
        eq_name = str(eq.get("name", ""))
        if name_keyword in eq_name:
            return True
    return False


def get_concentration_diff(unit, item_name):
    """取得某水質項目的進出流濃度差 (進 - 出)。回 (c_in, c_out) 或 (None, None)。"""
    c_in = c_out = None
    for code, items in unit.get("influent", {}).items():
        if item_name in items:
            v = items[item_name]
            if isinstance(v, dict):
                try:
                    c_in = float(v.get("濃度", 0))
                except (TypeError, ValueError):
                    pass
                break
    for code, items in unit.get("effluent", {}).items():
        if item_name in items:
            v = items[item_name]
            if isinstance(v, dict):
                try:
                    c_out = float(v.get("濃度", 0))
                except (TypeError, ValueError):
                    pass
                break
    return c_in, c_out


# ──────────────────────────────────────────────────
# 學理去除率範圍查詢 (讀 _槽體學理 Y/Z 欄)
# ──────────────────────────────────────────────────

_REMOVAL_CACHE = None


def _load_removal_ranges():
    """從 規則庫.xlsx _槽體學理 分頁讀 Y (主要削減項目) + Z (削減率範圍%) 欄.

    回傳: dict[標準槽體 → dict[水質項目 → (min%, max%)]]
    例: {"沉澱池": {"SS": (80, 95), "銅": (60, 90), ...}, ...}
    """
    global _REMOVAL_CACHE
    if _REMOVAL_CACHE is not None:
        return _REMOVAL_CACHE

    xlsx_path = os.path.join(BASE, "規則庫.xlsx")
    result = {}
    if not os.path.exists(xlsx_path):
        _REMOVAL_CACHE = result
        return result
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        if "_槽體學理" not in wb.sheetnames:
            _REMOVAL_CACHE = result
            return result
        ws = wb["_槽體學理"]
        # Y = col 25, Z = col 26
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 26:
                continue
            tank = row[0]
            status = row[7]  # H 狀態
            items_str = row[24] if len(row) > 24 else None  # Y
            rates_str = row[25] if len(row) > 25 else None  # Z
            if not tank or str(status or "").strip() != "V":
                continue
            if not items_str or not rates_str:
                continue
            items = [s.strip() for s in str(items_str).split(";") if s.strip()]
            rates = [s.strip() for s in str(rates_str).split(";") if s.strip()]
            tank_map = {}
            for it, rt in zip(items, rates):
                # 解析 "80~95" 或 "80~95 (備註)"
                rt_clean = rt.split("(")[0].strip()
                if "~" in rt_clean:
                    lo_s, hi_s = rt_clean.split("~", 1)
                    try:
                        lo = float(lo_s.strip())
                        hi = float(hi_s.strip())
                        tank_map[it] = (lo, hi)
                    except ValueError:
                        continue
            if tank_map:
                result[str(tank).strip()] = tank_map
        wb.close()
    except Exception as e:
        print(f"[step3e _load_removal_ranges 失敗] {e}")
    _REMOVAL_CACHE = result
    return result


# 水質項目同義詞 (規則庫用中文, PDF 抽出來可能有 mg/L 後綴)
_ITEM_SYNONYM_MAP = {
    "SS": ["SS", "懸浮固體", "懸浮固體（mg/L）", "懸浮固體(mg/L)"],
    "懸浮固體": ["懸浮固體", "SS", "懸浮固體（mg/L）", "懸浮固體(mg/L)"],
    "COD": ["COD", "化學需氧量", "化學需氧量（mg/L）", "化學需氧量(mg/L)"],
    "BOD": ["BOD", "生化需氧量", "生化需氧量（mg/L）", "生化需氧量(mg/L)"],
    "氨氮": ["氨氮", "氨氮（mg/L）", "氨氮(mg/L)"],
    "油脂": ["油脂", "油脂（mg/L）", "油脂(mg/L)"],
    "真色色度": ["真色色度", "色度"],
    "色度": ["色度", "真色色度"],
    "銅": ["銅"], "鎳": ["鎳"], "鋅": ["鋅"], "鉛": ["鉛"],
    "鎘": ["鎘"], "總鉻": ["總鉻", "鉻"], "六價鉻": ["六價鉻"],
    "氰化物": ["氰化物"], "含水率": ["含水率", "含水率(%)", "含水率（%）"],
    "濁度": ["濁度"], "游離氯": ["游離氯"], "異味": ["異味"],
}


def _get_removal_range(std_tank, item):
    """查該槽體該項目的學理去除率範圍.

    Returns:
        (lo, hi) 或 (None, None) 若未登記
    """
    ranges = _load_removal_ranges()
    tank_map = ranges.get(std_tank)
    if not tank_map:
        # 別名 fallback
        for t, m in ranges.items():
            if t in std_tank or std_tank in t:
                tank_map = m
                break
    if not tank_map:
        return (None, None)

    # 直接匹配
    if item in tank_map:
        return tank_map[item]
    # 同義詞匹配
    candidates = _ITEM_SYNONYM_MAP.get(item, [item])
    for cand in candidates:
        if cand in tank_map:
            return tank_map[cand]
    # 反向: 規則庫的 key 可能是 item 的同義
    for key in tank_map:
        if key == item or key in item or item in key:
            return tank_map[key]
    return (None, None)


# ──────────────────────────────────────────────────
# 規則啟發式檢查
# ──────────────────────────────────────────────────

def check_rule_against_unit(rule, unit):
    """嘗試把一條規則套到一個單元, 啟發式判斷是否「應重點審查」。

    回傳 dict 或 None:
        {
            "嚴重度": "待確認",
            "對照項目": "...",
            "描述": "...",
            "規則來源": rule.get("缺失ID") + 來源,
        }
    """
    target_item = (rule.get("對照項目") or "").strip()
    rule_text = (rule.get("規則") or "").strip()
    check_type = (rule.get("檢查類型") or "").strip()
    judgment = (rule.get("判定邏輯") or "").strip()
    deficiency_id = rule.get("缺失ID", "")
    source = rule.get("來源", "")

    if not target_item:
        return None

    code = unit.get("raw_code", "")
    std_tank = unit.get("std_tank", "")

    design_params = unit.get("design_params", {})
    measure_params = unit.get("measure_params", {})
    equipment_list = unit.get("equipment", [])

    # ── 1. pH 規則 ──
    if "pH" in target_item or "pH" in rule_text:
        # 取 unit 的 pH 範圍
        for pname, pval in {**measure_params, **design_params}.items():
            if "pH" in pname:
                lo = pval.get("min") if isinstance(pval, dict) else None
                hi = pval.get("max") if isinstance(pval, dict) else None
                if lo is None:
                    continue
                # 規則裡的 pH 門檻
                threshold_match = re.search(
                    r"pH[\s\S]{0,20}?(?:下限|>|>=|超過|大於)\s*([0-9.]+)",
                    rule_text + judgment
                )
                if threshold_match:
                    try:
                        thr = float(threshold_match.group(1))
                        if lo < thr:
                            return _make_finding(
                                "待確認", check_type, code, std_tank, target_item,
                                f"{pname} 下限 {lo} 規則建議 > {thr} ({rule_text[:60]})",
                                deficiency_id, source
                            )
                    except ValueError:
                        pass
                else:
                    # 未抽到具體門檻 → 至少列出讓使用者人工檢查
                    return _make_finding(
                        "待確認", check_type, code, std_tank, target_item,
                        f"申請文件 {pname} 範圍 {lo}~{hi}, 規則: {rule_text[:80]}",
                        deficiency_id, source
                    )

    # ── 2. 停留時間 / 有效容量 / 有效水深 等設計參數 ──
    for design_key in ["停留時間", "有效容量", "有效水深", "表面溢流率", "溢流率",
                       "污泥迴流率", "迴流率", "MLSS", "DO", "溶氧", "F/M", "食微比",
                       "濾速", "上升流速", "體積負荷", "有機負荷", "攪拌"]:
        if design_key in target_item or design_key in rule_text:
            # Nick 2026-07-07: 跨槽對比類規則 skip
            # 例 D041a「快混與慢混槽之水力停留時間應有區隔」→ 只有廠內有配對槽才適用
            # 此處只有單 unit context, 無法跨槽比較 → 保守 skip
            # (真要抓需要專門的跨單元檢查, 不該在單槽 dispatch 這裡處理)
            if "快混" in rule_text and "慢混" in rule_text:
                continue
            if "沉澱" in rule_text and "濃縮" in rule_text and design_key == "污泥迴流率":
                continue

            lo, hi, pname = get_param_range({**design_params, **measure_params}, design_key)
            if pname:
                # 「轉速應為固定值而非範圍值」規則 (陳映嘉 D090b):
                # 若廠商已填固定值 (min == max), 就是合規, 不要誤觸發
                is_rpm_rule = (
                    ("轉速" in rule_text or "固定值" in rule_text
                     or "區間值" in rule_text or "攪拌機" in target_item)
                    and design_key == "攪拌"
                )
                if is_rpm_rule and lo is not None and hi is not None:
                    try:
                        if float(lo) == float(hi):
                            continue  # 固定值 → 合規, 不出 finding
                    except (TypeError, ValueError):
                        pass
                # 顯示: 兩邊都有 → "0.2~1.5"; 只有一邊 → "≥ 0.2" 或 "≤ 1.5"
                if lo is not None and hi is not None:
                    val_str = f"{lo}~{hi}"
                elif lo is not None:
                    val_str = f"≥ {lo} (未填上限)"
                else:
                    val_str = f"≤ {hi} (未填下限)"
                return _make_finding(
                    "待確認", check_type, code, std_tank, target_item,
                    f"申請文件 {pname} = {val_str}, 規則: {rule_text[:80]}",
                    deficiency_id, source
                )
            # 規則指該項目但單元沒登載 → 也標記
            elif design_key in target_item and design_key in ["停留時間", "有效容量"]:
                return _make_finding(
                    "待確認", check_type, code, std_tank, target_item,
                    f"申請文件未登載『{design_key}』, 規則: {rule_text[:80]}",
                    deficiency_id, source
                )

    # ── 3. 機具設施 檢查已由 check_required_equipment.py 接管 ──
    # (Nick 2026-07-06 定調: 用 _槽體學理.必備機具 過濾, 只該套的才套)
    # 這裡不再對每條 rule 都跳「機具未列 X」, 避免砂濾塔被套「攪拌機未列」等誤觸發

    # ── 4. 加藥量 ──
    if "加藥" in target_item or "加藥量" in rule_text or "藥品" in rule_text:
        # 找量測參數中含「加藥量」者
        for pname, pval in measure_params.items():
            if "加藥" in pname:
                if isinstance(pval, dict):
                    raw = pval.get("raw", "")
                    return _make_finding(
                        "待確認", check_type, code, std_tank, target_item,
                        f"{pname} = {raw[:50]}, 規則: {rule_text[:80]}",
                        deficiency_id, source
                    )

    # ── 5. 質量平衡 ──
    if "質量平衡" in check_type or "質量平衡" in rule_text or "質量" in target_item:
        # 規則牽涉質量平衡, 多半需人工檢視
        # 只在 unit 有進出流時才列, 避免假陽性
        if unit.get("influent") and unit.get("effluent"):
            return _make_finding(
                "待確認", "質量平衡", code, std_tank, target_item,
                f"質量平衡需檢驗, 規則: {rule_text[:80]}",
                deficiency_id, source
            )

    # ── 6. 去除率 / 重金屬 ──
    # 修訂 (2026-07-01 v2, Nick 補「去除率是要抓的」):
    #   分兩類:
    #   類 1 「無分離功能槽」(快混/慢混/pH/中和/氧化):
    #        → D054 學理, 顯著去除 (>= 30%) 就 finding「無分離不應去除」
    #   類 2 「有分離功能槽」(沉澱/浮除/砂濾/離子交換/活性碳/MBR):
    #        → 對照 _get_removal_range 讀 _槽體學理 Y/Z 欄的學理範圍
    #        → 太低: 待確認 (實測偏低, 需檢查加藥/pH/HRT)
    #        → 太高: 提醒 (實測偏高請確認數據合理性)
    #   共同過濾: (a) 排除放流池 (b) 槽體適用性
    heavy_metals = ["銅", "鎳", "鋅", "鉛", "鎘", "鉻", "總鉻", "六價鉻",
                    "錫", "汞", "總汞", "砷", "鉬"]
    SEPARATION_KEYWORDS = ["沉澱", "沉降", "浮除", "砂濾", "過濾器", "活性碳",
                           "離子交換", "膜", "MBR", "油脂分離"]
    rule_tank = (rule.get("標準槽體名稱") or "").strip()
    # 「金屬去除率」/「去除率」這種通稱也視為觸發對每個金屬跑
    is_generic_removal = any(g in target_item for g in
                              ["金屬去除率", "去除率", "重金屬去除"])
    for metal in heavy_metals:
        # 明確含金屬名 OR 是通稱去除率 (對每個金屬都跑)
        if metal in target_item or is_generic_removal:
            # (a) 排除放流池
            if "放流" in std_tank:
                continue
            # (b) 槽體適用性
            if rule_tank and rule_tank not in ("(文件類)", "(現場設備類)"):
                if rule_tank not in std_tank and std_tank not in rule_tank:
                    continue

            c_in, c_out = get_concentration_diff(unit, metal)
            if c_in is None or c_out is None or c_in <= 0:
                continue
            removal = (c_in - c_out) / c_in * 100

            is_separation_tank = any(kw in std_tank for kw in SEPARATION_KEYWORDS)
            expected_lo, expected_hi = _get_removal_range(std_tank, metal)

            if is_separation_tank:
                # 該降卻升: 學理登記為「該降」但實測反而上升 (削減率為負值)
                # → 廠商可能填錯、pH>11 兩性反溶、或漏列進流支線
                if expected_lo is not None and c_out > c_in and removal <= -10:
                    return _make_finding(
                        "待確認", "去除率", code, std_tank, target_item,
                        f"{metal} 進{c_in:.2f}→出{c_out:.2f} (削減率 {removal:.1f}%, "
                        f"反而上升). 依學理 {std_tank} {metal} 應為去除 "
                        f"{expected_lo:.0f}~{expected_hi:.0f}%, 實測反而增加, 可能原因: "
                        f"(1) 廠商水質表填錯 (進出欄反了); "
                        f"(2) 漏列進流支線 (回流/反洗液); "
                        f"(3) pH 控制不當 (Zn/Al 兩性氫氧化物於 pH>11 反溶). "
                        f"規則: {rule_text[:60]}",
                        deficiency_id, source
                    )
                # 類 2: 對照學理範圍
                if expected_lo is not None and expected_hi is not None:
                    if 0 <= removal < expected_lo:
                        return _make_finding(
                            "待確認", "去除率", code, std_tank, target_item,
                            f"{metal} 進{c_in:.2f}→出{c_out:.2f} (去除 {removal:.1f}%). "
                            f"依學理 {std_tank} {metal} 去除率應為 {expected_lo:.0f}~{expected_hi:.0f}%, "
                            f"實測偏低, 請重新審視加藥量/pH/HRT/助凝劑等操作參數. "
                            f"規則: {rule_text[:60]}",
                            deficiency_id, source
                        )
                    if removal > expected_hi:
                        return _make_finding(
                            "提醒", "去除率", code, std_tank, target_item,
                            f"{metal} 進{c_in:.2f}→出{c_out:.2f} (去除 {removal:.1f}%). "
                            f"依學理 {std_tank} {metal} 去除率通常為 {expected_lo:.0f}~{expected_hi:.0f}%, "
                            f"實測偏高請確認數據合理性 (是否進流高估或出流低於檢量極限).",
                            deficiency_id, source
                        )
                else:
                    # 學理未登記, fallback
                    if removal >= 30:
                        return _make_finding(
                            "提醒", "去除率", code, std_tank, target_item,
                            f"{metal} 進{c_in:.2f}→出{c_out:.2f} (去除 {removal:.1f}%). "
                            f"請確認是否符合學理範圍 (規則庫尚未登記 {std_tank} {metal} 學理值). "
                            f"規則: {rule_text[:60]}",
                            deficiency_id, source
                        )
            else:
                # 類 1: 無分離功能, D054 學理
                if removal >= 30:
                    return _make_finding(
                        "待確認", "去除率", code, std_tank, target_item,
                        f"{metal} 進{c_in:.2f}→出{c_out:.2f} (去除 {removal:.1f}%). "
                        f"該單元 ({std_tank}) 無明顯分離機制, 不應表現重金屬去除. "
                        f"規則: {rule_text[:60]}",
                        deficiency_id, source
                    )

    # ── 7. SS / 懸浮固體 / BOD / COD ──
    # 同樣邏輯: 分離槽對照學理範圍, 非分離槽用顯著變動
    for water_item in ["SS", "懸浮固體", "BOD", "COD", "氨氮", "總氮", "總磷",
                       "硝酸鹽氮", "硼", "氟鹽", "氰化物", "油脂", "真色色度"]:
        if water_item in target_item:
            if "放流" in std_tank:
                continue
            if rule_tank and rule_tank not in ("(文件類)", "(現場設備類)"):
                if rule_tank not in std_tank and std_tank not in rule_tank:
                    continue
            c_in, c_out = get_concentration_diff(unit, water_item)
            if c_in is None or c_out is None:
                for variant in [water_item + "（mg/L）", water_item + "(mg/L)"]:
                    c_in, c_out = get_concentration_diff(unit, variant)
                    if c_in is not None:
                        break
            if c_in is None or c_out is None or c_in <= 0:
                continue
            change_pct = abs(c_in - c_out) / c_in * 100

            is_separation_tank = any(kw in std_tank for kw in SEPARATION_KEYWORDS)
            # 統一項目名 (規則庫用 SS/COD/BOD 簡化)
            std_item = "SS" if water_item == "懸浮固體" else water_item
            expected_lo, expected_hi = _get_removal_range(std_tank, std_item)

            if is_separation_tank and expected_lo is not None:
                # 該降卻升: 學理登記為「該降」但實測反而上升
                if c_out > c_in:
                    inc_pct = (c_out - c_in) / c_in * 100
                    if inc_pct >= 10:
                        return _make_finding(
                            "待確認", "去除率", code, std_tank, target_item,
                            f"{water_item} 進{c_in:.2f}→出{c_out:.2f} (反而上升 {inc_pct:.1f}%). "
                            f"依學理 {std_tank} {water_item} 應為去除 {expected_lo:.0f}~{expected_hi:.0f}%, "
                            f"實測反而增加, 可能原因: "
                            f"(1) 廠商水質表填錯 (進出欄反了); "
                            f"(2) 該槽有加藥引入 (例如快混加 PAC 使 SS↑, 但沉澱池不應如此); "
                            f"(3) 有其他高濃度支線未列入. "
                            f"規則: {rule_text[:60]}",
                            deficiency_id, source
                        )
                else:
                    removal = (c_in - c_out) / c_in * 100
                    if removal < expected_lo:
                        return _make_finding(
                            "待確認", "去除率", code, std_tank, target_item,
                            f"{water_item} 進{c_in:.2f}→出{c_out:.2f} (去除 {removal:.1f}%). "
                            f"依學理 {std_tank} {water_item} 去除率應為 {expected_lo:.0f}~{expected_hi:.0f}%, "
                            f"實測偏低, 請重新審視操作參數.",
                            deficiency_id, source
                        )
                    if removal > expected_hi:
                        return _make_finding(
                            "提醒", "去除率", code, std_tank, target_item,
                            f"{water_item} 進{c_in:.2f}→出{c_out:.2f} (去除 {removal:.1f}%). "
                            f"依學理 {std_tank} {water_item} 去除率通常為 {expected_lo:.0f}~{expected_hi:.0f}%, "
                            f"實測偏高請確認數據合理性.",
                            deficiency_id, source
                        )
            elif change_pct >= 30:
                # 無學理範圍或非分離槽, 顯著變動 fallback
                direction = "下降" if c_out < c_in else "上升"
                return _make_finding(
                    "待確認", check_type or "水質標準", code, std_tank, target_item,
                    f"{water_item} 進{c_in:.2f}→出{c_out:.2f} ({direction} {change_pct:.1f}%), "
                    f"規則: {rule_text[:80]}",
                    deficiency_id, source
                )

    # ── 8. 操作條件 (反洗頻率、更換頻率) ──
    for op_kw in ["反洗", "更換頻率", "操作參數", "操作時間"]:
        if op_kw in target_item:
            # 看 measure_params 或 design_params 有沒有相關記錄
            for pname in {**measure_params, **design_params}.keys():
                if op_kw in pname:
                    return _make_finding(
                        "待確認", check_type or "操作條件", code, std_tank, target_item,
                        f"申請文件有相關參數 ({pname}), 規則: {rule_text[:80]}",
                        deficiency_id, source
                    )

    # 沒命中任何啟發式 → 不產生 finding
    return None


def _make_finding(severity, check_type, code, std_tank, target_item, desc, deficiency_id, source):
    return {
        "嚴重度": severity,
        "類型": check_type or "其他",
        "單元": code,
        "標準槽體": std_tank,
        "對照項目": target_item,
        "描述": desc,
        "依據": f"規則庫 {deficiency_id} ({source})",
    }


# ──────────────────────────────────────────────────
# 主入口
# ──────────────────────────────────────────────────

def run_rule_driven_check(app_data, max_findings_per_unit=20):
    """規則庫驅動的全面審查。

    對申請文件每個單元, 拿該標準槽體類型的所有規則跑啟發式檢查。

    Args:
        app_data: step2_extract_v2 輸出的 JSON
        max_findings_per_unit: 每單元最多 finding 數 (避免轟炸)

    Returns:
        list of finding dict
    """
    rules_by_tank = load_rules_by_tank()
    if not rules_by_tank:
        return []

    findings = []
    for code, unit in app_data.get("units", {}).items():
        std_tank = unit.get("std_tank", "")
        # 取該槽體類型的規則 (含通用的「文件類」「現場設備類」)
        applicable_rules = []
        applicable_rules.extend(rules_by_tank.get(std_tank, []))
        # 不額外加文件類/現場設備類, 避免重複過多

        unit_findings = []
        seen_keys = set()  # 避免同一單元同一「檢查情境」重複
        for rule in applicable_rules:
            try:
                f = check_rule_against_unit(rule, unit)
                if f is None:
                    continue
                # dedup key = (類型, 描述前 40 字)
                # 不同規則但檢查同 metric (例: 沉澱池銅 49.8%) 時描述前綴會相同.
                # 對照項目不放進 key, 因為多條規則的對照項目字面不同
                # ("金屬去除率" / "去除率估算合理性" / "去除率與表面溢流率")
                # 但實際檢查的都是同一件事.
                desc_head = (f.get("描述") or "")[:40]
                key = (f.get("類型"), desc_head)
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                unit_findings.append(f)
            except Exception as e:
                # 不讓單一規則失敗影響整體
                pass

        # B (2026-07-01): 若該單元有「具體」finding, 移除同單元的「通用」finding
        # 通用 finding 特徵: 描述含 "需檢驗" / "需人工檢視" / "多半需人工" 等模糊語彙
        # 或 對照項目 = "水質濃度及質量" 這種抽象名稱
        # 具體 finding 特徵: 描述含具體數字 (進XX→出XX / XX%)
        unit_findings = _dedup_generic_when_specific_exists(unit_findings)

        # 限制每單元 finding 數量
        unit_findings = unit_findings[:max_findings_per_unit]
        findings.extend(unit_findings)

    return findings


# ──────────────────────────────────────────────────
# 通用 finding vs 具體 finding 分類
# ──────────────────────────────────────────────────
_GENERIC_PHRASES = [
    "需檢驗", "需人工檢視", "多半需人工", "需檢核",
    "水質濃度及質量", "各項水質濃度",
    "應保持不變", "應予檢核",
]


def _is_generic_finding(f):
    """判斷是否為通用/技師本來就懂的 finding.
    特徵:
    - 描述含模糊語彙 (需檢驗/需人工/應予檢核)
    - 對照項目為抽象名稱 (水質濃度及質量)
    - 且描述**沒有**具體數字 (進XX→出XX / XX%)
    """
    if not isinstance(f, dict):
        return False
    desc = str(f.get("描述") or "")
    target = str(f.get("對照項目") or "")

    # 有具體數字 → 不是通用
    import re as _re
    if _re.search(r"進\s*-?\d|\d+\s*→\s*\d|\d+\.\d+\s*%|\d+%|去除\s*-?\d", desc):
        return False

    # 匹配模糊語彙
    for phrase in _GENERIC_PHRASES:
        if phrase in desc or phrase in target:
            return True
    return False


def _dedup_generic_when_specific_exists(unit_findings):
    """若該單元有具體 finding, 移除通用 finding.
    但若通用 finding 是**唯一**的 finding, 保留 (無其他資訊時仍要提示).
    """
    if not unit_findings:
        return unit_findings
    specific = [f for f in unit_findings if not _is_generic_finding(f)]
    generic = [f for f in unit_findings if _is_generic_finding(f)]
    if specific:
        return specific  # 有具體 → 只保留具體
    return generic  # 沒具體 → 保留通用 (別漏抓)


if __name__ == "__main__":
    import io
    import json
    import sys

    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    except Exception:
        pass

    jsons = sorted([f for f in os.listdir(BASE) if f.startswith("application_") and f.endswith(".json")])
    if not jsons:
        print("找不到 application_*.json")
        sys.exit(0)
    with open(os.path.join(BASE, jsons[-1]), "r", encoding="utf-8") as f:
        app_data = json.load(f)

    print(f"=== 規則庫驅動檢查: {app_data.get('source_pdf', '?')} ===")
    print(f"單元數: {app_data.get('total_units', 0)}")
    findings = run_rule_driven_check(app_data)
    print(f"\n總 finding 數: {len(findings)}")

    # 統計
    from collections import Counter
    types = Counter()
    for f in findings:
        types[(f["嚴重度"], f["類型"])] += 1
    print("\n=== 類型分布 ===")
    for k, v in sorted(types.items()):
        print(f"  {k[0]:6s} | {k[1]:20s} | {v}")

    # 展示前 10 筆
    print("\n=== 前 10 筆 finding ===")
    for i, f in enumerate(findings[:10], 1):
        print(f"\n{i}. [{f['嚴重度']}|{f['類型']}] {f['單元']} ({f['標準槽體']}) - {f['對照項目']}")
        print(f"   {f['描述'][:120]}")
        print(f"   依據: {f['依據']}")
