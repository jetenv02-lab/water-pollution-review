# -*- coding: utf-8 -*-
"""水措審查系統 — Streamlit 線上版 (v2 邏輯,經使用者明確授權覆蓋)。

本檔內容已切換為 streamlit_app_v2.py 的邏輯,使 Streamlit Cloud
預設 Main file (streamlit_app.py) 就能跑 v2 版抽取器。

v2 重大改進:
- 改用成熟版抽取器 step2_extract_v2 (覆蓋率 100%、單元類型歸類正確)
- 顯示更豐富的單元資訊(設計參數/量測參數/機具設施/進出流水質)
- 加入「水量平衡缺口提示」(尚未做 OCR,但會說明限制)
"""
import csv
import io
import json
import os
import tempfile
from collections import defaultdict
from datetime import datetime

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import 內部模組 - 用 try/except 包起避免單一模組失敗讓整個 App 掛掉
# (Python 3.14 + Streamlit Cloud 偶發 KeyError 在 module loading)
_import_errors = []

try:
    from step2_extract_v2 import extract_application
except Exception as e:
    _import_errors.append(("step2_extract_v2", str(e)))
    extract_application = None

try:
    from step2b_locate_sections import locate_sections, compress_ranges
except Exception as e:
    _import_errors.append(("step2b_locate_sections", str(e)))
    locate_sections = None
    def compress_ranges(p): return str(p)

try:
    from step2c_ocr_diagram import ocr_diagram_pages
except Exception as e:
    _import_errors.append(("step2c_ocr_diagram", str(e)))
    def ocr_diagram_pages(*a, **k): return {"error": f"OCR 模組載入失敗: {_import_errors[-1][1]}"}

try:
    from step3b_balance_check import run_balance_checks
except Exception as e:
    _import_errors.append(("step3b_balance_check", str(e)))
    def run_balance_checks(*a, **k): return []

try:
    from step3c_unit_db import BUSINESS_TYPES
except Exception as e:
    _import_errors.append(("step3c_unit_db", str(e)))
    BUSINESS_TYPES = {}

try:
    from step3d_principle_check import run_advanced_checks
except Exception as e:
    _import_errors.append(("step3d_principle_check", str(e)))
    def run_advanced_checks(*a, **k): return []

try:
    from step3e_rule_driven_check import run_rule_driven_check
except Exception as e:
    _import_errors.append(("step3e_rule_driven_check", str(e)))
    def run_rule_driven_check(*a, **k): return []

try:
    from step4_flow_graph import build_flow_graph, get_unit_neighbors
except Exception as e:
    _import_errors.append(("step4_flow_graph", str(e)))
    def build_flow_graph(*a, **k): return {"nodes": [], "edges": [], "wm_sources": [], "discharges": []}
    def get_unit_neighbors(*a, **k): return {"upstream": [], "downstream": []}

# ───────────────────────────── 頁面設定 ─────────────────────────────

st.set_page_config(
    page_title="水措審查系統 v3",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #2F5496 0%, #1a365d 100%);
        color: white;
        padding: 20px 30px;
        border-radius: 8px;
        margin-bottom: 20px;
    }
    .main-header h1 { margin: 0; font-size: 28px; }
    .main-header p { margin: 4px 0 0 0; opacity: 0.9; }
    .version-badge {
        background: #38a169;
        color: white;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 13px;
        margin-left: 8px;
    }
    .ocr-warning {
        background: #fef3c7;
        border-left: 4px solid #f59e0b;
        padding: 12px 16px;
        border-radius: 4px;
        margin: 12px 0;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>水措審查系統 <span class="version-badge">v3</span></h1>
    <p>自動化「水污染防治措施」申請文件審查 · 比對環工技師查核缺失資料庫</p>
    <p style="font-size: 13px; opacity: 0.85; margin-top: 6px;">
        v3 新增: 章節動態定位 · OCR 流向圖解析 · 智能審查 (學理檢查)
    </p>
</div>
""", unsafe_allow_html=True)

CSV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rules_extracted.csv")


@st.cache_data(show_spinner=False)
def load_rules():
    if not os.path.exists(CSV_PATH):
        return [], {}
    rules = []
    with open(CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rules.append(row)
    by_tank = defaultdict(list)
    for r in rules:
        tank = (r.get("標準槽體名稱") or "未分類").strip()
        by_tank[tank].append(r)
    return rules, dict(by_tank)


# ───────────────────────────── Excel 輸出 ─────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SEVERITY_FILL = {
    "不合理": PatternFill("solid", fgColor="FCE4D6"),
    "待確認": PatternFill("solid", fgColor="FFF2CC"),
    "提醒": PatternFill("solid", fgColor="E2F0D9"),
    "合理": PatternFill("solid", fgColor="E2EFDA"),
}
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_unit_excel(app_data):
    """產出『各單元資料表』Excel (含設計/量測/機具/進出水質)。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 總表分頁
    ws = wb.create_sheet("_單元總表")
    headers = ["單元代號", "原始名稱", "標準槽體類型", "代碼", "出現頁數",
               "設計參數數", "量測參數數", "機具項目數", "進流數", "出流數"]
    widths = [12, 30, 16, 8, 16, 12, 12, 12, 10, 10]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    ws.freeze_panes = "A2"

    for code, info in sorted(app_data["units"].items()):
        ws.append([
            code, info.get("name_in_doc", ""), info.get("std_tank", ""),
            info.get("code_id", ""),
            ", ".join(map(str, info.get("pages_found", []))),
            len(info.get("design_params", {})),
            len(info.get("measure_params", {})),
            len(info.get("equipment", [])),
            len(info.get("influent", {})),
            len(info.get("effluent", {})),
        ])

    # 每個單元一張分頁
    for code, info in sorted(app_data["units"].items()):
        sname = code[:31]
        ws = wb.create_sheet(sname)
        row = 1

        # 標題
        ws.cell(row=row, column=1, value=f"{code} {info.get('name_in_doc', '')}").font = Font(size=14, bold=True)
        row += 1
        ws.cell(row=row, column=1, value=f"標準類型: {info.get('std_tank', '')} | 代碼: {info.get('code_id', '')} | 頁: {info.get('pages_found', [])}")
        row += 2

        # 單元尺寸
        size_info = info.get("size") or {}
        if size_info:
            ws.cell(row=row, column=1, value="單元尺寸").font = Font(bold=True, color="2F5496")
            row += 1
            _size_keys = ["材質", "長/直徑", "寬", "高", "有效水深", "有效容量", "數量", "其他"]
            _units_lookup = {
                "長/直徑": "公尺", "寬": "公尺", "高": "公尺",
                "有效水深": "公尺", "有效容量": "m³", "數量": "座",
            }
            ws.cell(row=row, column=1, value="項目")
            ws.cell(row=row, column=2, value="值")
            ws.cell(row=row, column=3, value="單位")
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for k in _size_keys:
                if k not in size_info:
                    continue
                ws.cell(row=row, column=1, value=k)
                ws.cell(row=row, column=2, value=size_info[k])
                ws.cell(row=row, column=3, value=_units_lookup.get(k, ""))
                row += 1
            row += 1

        # 設計參數
        if info.get("design_params"):
            ws.cell(row=row, column=1, value="設計操作參數").font = Font(bold=True, color="2F5496")
            row += 1
            ws.cell(row=row, column=1, value="參數")
            ws.cell(row=row, column=2, value="最小值")
            ws.cell(row=row, column=3, value="最大值")
            ws.cell(row=row, column=4, value="原始")
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for pname, pval in info["design_params"].items():
                ws.cell(row=row, column=1, value=pname)
                ws.cell(row=row, column=2, value=pval.get("min", ""))
                ws.cell(row=row, column=3, value=pval.get("max", ""))
                ws.cell(row=row, column=4, value=pval.get("raw", ""))
                row += 1
            row += 1

        # 量測參數
        if info.get("measure_params"):
            ws.cell(row=row, column=1, value="量測操作參數").font = Font(bold=True, color="2F5496")
            row += 1
            ws.cell(row=row, column=1, value="參數")
            ws.cell(row=row, column=2, value="最小值")
            ws.cell(row=row, column=3, value="最大值")
            ws.cell(row=row, column=4, value="原始")
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for pname, pval in info["measure_params"].items():
                ws.cell(row=row, column=1, value=pname)
                ws.cell(row=row, column=2, value=pval.get("min", ""))
                ws.cell(row=row, column=3, value=pval.get("max", ""))
                ws.cell(row=row, column=4, value=pval.get("raw", ""))
                row += 1
            row += 1

        # 機具
        if info.get("equipment"):
            ws.cell(row=row, column=1, value="相關機具設施").font = Font(bold=True, color="2F5496")
            row += 1
            ws.cell(row=row, column=1, value="名稱")
            ws.cell(row=row, column=2, value="位置")
            ws.cell(row=row, column=3, value="數量")
            ws.cell(row=row, column=4, value="馬力(kW)")
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for eq in info["equipment"]:
                ws.cell(row=row, column=1, value=eq.get("name", ""))
                ws.cell(row=row, column=2, value=eq.get("位置", ""))
                ws.cell(row=row, column=3, value=eq.get("數量", ""))
                ws.cell(row=row, column=4, value=eq.get("馬力_kW", ""))
                row += 1
            row += 1

        # 進流水質
        for infl_code, q_data in info.get("influent", {}).items():
            ws.cell(row=row, column=1, value=f"進流水質 — {infl_code}").font = Font(bold=True, color="2F5496")
            row += 1
            ws.cell(row=row, column=1, value="水質項目")
            ws.cell(row=row, column=2, value="濃度")
            ws.cell(row=row, column=3, value="質量")
            ws.cell(row=row, column=4, value="範圍(pH/水溫)")
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for item, val in q_data.items():
                ws.cell(row=row, column=1, value=item)
                ws.cell(row=row, column=2, value=val.get("濃度", ""))
                ws.cell(row=row, column=3, value=val.get("質量", ""))
                ws.cell(row=row, column=4, value=val.get("範圍", ""))
                row += 1
            row += 1

        # 出流水質
        for effl_code, q_data in info.get("effluent", {}).items():
            ws.cell(row=row, column=1, value=f"出流水質 — {effl_code}").font = Font(bold=True, color="38a169")
            row += 1
            ws.cell(row=row, column=1, value="水質項目")
            ws.cell(row=row, column=2, value="濃度")
            ws.cell(row=row, column=3, value="質量")
            ws.cell(row=row, column=4, value="範圍(pH/水溫)")
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for item, val in q_data.items():
                ws.cell(row=row, column=1, value=item)
                ws.cell(row=row, column=2, value=val.get("濃度", ""))
                ws.cell(row=row, column=3, value=val.get("質量", ""))
                ws.cell(row=row, column=4, value=val.get("範圍", ""))
                row += 1
            row += 1

        # 欄寬
        for c in range(1, 6):
            ws.column_dimensions[get_column_letter(c)].width = [22, 18, 18, 36][min(c - 1, 3)]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ───────────────────────────── 介面 ─────────────────────────────

rules, rules_by_tank = load_rules()

with st.sidebar:
    st.header("系統狀態")
    st.metric("規則總數", len(rules))
    st.metric("槽體分類數", len(rules_by_tank))
    # 顯示 CSV 檔案修改時間 + 大小 (台北時區)
    if os.path.exists(CSV_PATH):
        import datetime as _dt
        try:
            from zoneinfo import ZoneInfo as _ZI
            _tz_tpe = _ZI("Asia/Taipei")
        except Exception:
            _tz_tpe = _dt.timezone(_dt.timedelta(hours=8))
        mtime = _dt.datetime.fromtimestamp(os.path.getmtime(CSV_PATH), _tz_tpe)
        st.caption(f"CSV: {os.path.getsize(CSV_PATH)//1024} KB · {mtime.strftime('%m-%d %H:%M')}")
    if st.button("🔄 清快取重載", help="若顯示舊資料,點此清快取"):
        st.cache_data.clear()
        st.rerun()
    if st.button("🔄 重新載入規則庫", help="清本機快取 + 若 規則庫.xlsx 有變動則 commit/push 到 GitHub (Streamlit Cloud 會自動 redeploy)"):
        # 1. 清本機快取
        try:
            import tank_chemistry as _tc_reload
            _tc_reload.clear_cache()
            st.cache_data.clear()
            _local_ok = True
        except Exception as _e:
            st.error(f"本機快取清除失敗: {_e}")
            _local_ok = False
        # 2. 若 規則庫.xlsx 有變動 → 自動 commit + push
        if _local_ok:
            try:
                import subprocess as _sp
                _r = _sp.run(['git', 'status', '--porcelain', '規則庫.xlsx'], capture_output=True, text=True, encoding='utf-8')
                if not _r.stdout.strip():
                    st.success("✅ 本機規則庫快取已清除 (規則庫.xlsx 無變動, 不需推送線上版)")
                else:
                    _sp.run(['git', 'add', '規則庫.xlsx'], check=True)
                    import datetime as _dt
                    _msg = f"規則庫更新: {_dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    _sp.run(['git', 'commit', '-m', _msg], check=True, capture_output=True, text=True, encoding='utf-8')
                    _pr = _sp.run(['git', 'push'], capture_output=True, text=True, encoding='utf-8')
                    if _pr.returncode == 0:
                        st.success(f"✅ 本機快取已清除 + 已推送 GitHub: {_msg}。Streamlit Cloud 將於 1~3 分鐘內 redeploy。")
                    else:
                        st.warning(f"本機快取已清除, 但推送 GitHub 失敗: {_pr.stderr}")
            except Exception as _e:
                st.warning(f"本機快取已清除, 但 git 操作失敗: {_e}")
    st.divider()

    # ───────── 本次瀏覽審查歷史 (重整頁面後即消失) ─────────
    history = st.session_state.get("_review_history", [])
    if history:
        st.markdown("**📋 本次審查歷史**")
        st.caption("(本次瀏覽紀錄 · 完整歷史請看「📊 規則庫瀏覽 → 📋 歷次審查紀錄」)")
        # 雲端寫入狀態 — 對使用者隱藏細節, 只顯示成功訊息
        log_status = st.session_state.get("_last_sheet_log_status")
        if log_status and log_status.startswith("✅"):
            st.caption(log_status)
        for i, h in enumerate(history[:5]):
            with st.container(border=True):
                fname_short = h["filename"]
                if len(fname_short) > 22:
                    fname_short = fname_short[:20] + "..."
                st.caption(f"⏱ {h['time']} · {h['elapsed_sec']}s")
                st.markdown(f"**{fname_short}**")
                st.caption(
                    f"{h['units']} 單元 · 🔴 {h['unreasonable']} · 🟡 {h['manual']}"
                )
        if len(history) > 5:
            st.caption(f"還有 {len(history)-5} 筆紀錄...")
        if st.button("🗑 清除歷史", key="_clear_history"):
            st.session_state.pop("_review_history", None)
            st.rerun()
        st.divider()

    st.markdown("**v3 版本能力**")
    st.text("- 38 單元完整抽取")
    st.text("- 設計/量測/機具/水質")
    st.text("- 章節動態定位")
    st.text("- OCR 流向圖解析")
    st.text("- 智能審查 (學理檢查)")
    st.text("- 水流串接圖")
    st.divider()
    st.markdown("**專案連結**")
    st.markdown("[GitHub 專案頁面](https://github.com/jetenv02-lab/water-pollution-review)")

    # 顯示 import 失敗模組 (debug 用)
    if _import_errors:
        with st.expander(f"⚠ 模組載入問題 ({len(_import_errors)})"):
            for mod, err in _import_errors:
                st.caption(f"**{mod}**: {err[:150]}")

# ───────── 全頁面 busy overlay + 停止偵測 ─────────
# Streamlit 的執行模型: 每次 widget 互動都會 rerun 整個 script;
# 若上一次審查中按了其他 widget (如 file_uploader 換檔), 那次審查的 thread
# 會被中斷, try/finally 沒走完 → _busy 卡在 True。
# 對策: 用 _busy_run_id 標記「這次 rerun 是審查本身的 rerun 嗎?」
#   - 進審查時, 把當下 timestamp 寫到 _busy_run_id
#   - 每次 rerun 進來, 若 _busy=True 但 _busy_run_id 是 N 秒前的舊值, 視為中斷
import time as _time_busy
_now_ts = _time_busy.time()

# 上次 rerun 不是審查本身, 但 _busy=True → 中斷殘留, 自動解鎖
if st.session_state.get("_busy", False):
    _bid = st.session_state.get("_busy_run_id", 0)
    _bkind = st.session_state.get("_busy_kind", "")
    _pdf_name_check = st.session_state.get("_pdf_filename", "")

    # zombie 偵測 (立刻解鎖):
    # - 主審查標 busy 但 _pdf_filename 是空的 → 一定是掛了 (主審查最先存檔名)
    # - _busy_run_id=0 (從來沒寫過) 但 _busy=True
    is_zombie = (
        (_bkind == "main" and not _pdf_name_check)
        or (_bid == 0)
    )

    # 超時 (180 秒, 主審查正常 ~2-3 分鐘, 跑示意圖 +30 秒)
    is_timeout = (_now_ts - _bid) > 180

    if is_zombie or is_timeout:
        st.session_state["_busy"] = False
        st.session_state.pop("_busy_run_id", None)
        st.session_state.pop("_busy_kind", None)
        reason = "(偵測到無檔名的殘留狀態)" if is_zombie else "(超過 180 秒未完成)"
        st.warning(f"⚠️ 偵測到上次審查被中途中斷, 已自動解鎖 {reason}。請重新點「開始完整審查」。")
    else:
        # 還在合理時間內, 提供「手動解鎖」按鈕讓使用者可主動恢復
        _elapsed = int(_now_ts - _bid)
        cu1, cu2 = st.columns([4, 1])
        with cu1:
            st.info(f"⏳ 審查進行中 (已 {_elapsed} 秒)。若實際已停止, 可點右邊按鈕強制解鎖。")
        with cu2:
            if st.button("🔓 強制解鎖", key="_force_unlock_top"):
                st.session_state["_busy"] = False
                st.session_state.pop("_busy_run_id", None)
                st.session_state.pop("_busy_kind", None)
                st.session_state.pop("_cancel_requested", None)
                st.rerun()

# 若使用者點過停止鈕 → 顯示已停止
if st.session_state.pop("_cancel_just_done", False):
    st.warning("🛑 審查已停止。已完成的步驟結果有保留, 你可重新點「開始完整審查」繼續。")

# 審查成功完成 → 顯示成功訊息
_rjd = st.session_state.pop("_review_just_done", None)
if _rjd:
    st.success(
        f"✅ **基本審查已完成**! 共 {_rjd['units']} 單元 · "
        f"找出 {_rjd['unreasonable']} 項不合理 / {_rjd['manual']} 項待確認複核 · "
        f"耗時 {_rjd['elapsed']} 秒"
    )
    st.info(
        "💡 **下一步**: 你可以繼續往下捲動,展開「📊 水量平衡示意圖解析」 "
        "讓 Gemini Vision 讀流向圖, 得到跨單元的完整流向結構 + 質量平衡檢核"
    )

tab1, tab2, tab_sync, tab_import, tab3 = st.tabs(["🚀 開始審查", "📊 規則庫瀏覽", "🔄 規則庫管理", "📥 匯入新規則", "📖 使用說明"])

with tab1:
    st.subheader("上傳申請文件 PDF")
    st.caption("v3 抽取器會解析: 處理設施資料表 + 進出水質資料表。流向示意圖與水量平衡示意圖可透過 OCR 自動讀取。")

    st.markdown("""
    <div class="ocr-warning">
    <strong>v3 已支援的審查能力:</strong><br>
    - 自動定位本份文件的水量平衡圖與流向示意圖頁碼(不寫死)<br>
    - OCR 讀取圖片頁面: 單元代號、流量 Q、加藥量、含水率<br>
    - 智能審查: 質量平衡守恆、快混槽不應展現重金屬去除、沉澱池溢流率等學理檢查
    </div>
    """, unsafe_allow_html=True)

    # busy 時鎖 file_uploader — 避免使用者中途換檔觸發 rerun 把審查 thread 打斷
    # (Streamlit 每次 widget 互動都會從頭重跑 script, file_uploader 重新選檔
    #  會 rerun, 把正在跑的審查中斷, _busy 因為 try/finally 沒走完而卡在 True)
    _busy_now = st.session_state.get("_busy", False)
    _busy_kind = st.session_state.get("_busy_kind", "")  # "main" 主審查 / "flow" 示意圖解析

    if _busy_now and _busy_kind == "main":
        # 只在「主審查」進行中顯示這個鎖定提示
        # (示意圖解析的 overlay 已經蓋滿全頁, 不需要再顯示這個小提示)
        _prev_name = st.session_state.get("_pdf_filename", "(未知)")
        st.info(
            f"🔒 **審查進行中** — 已鎖定上傳功能,避免中斷處理\n\n"
            f"目前正在審查的檔案: **{_prev_name}**"
        )
        uploaded = None  # 走「顯示區」路線, 不會觸發底下的 button 邏輯
    elif _busy_now:
        # 示意圖解析 (或其他非主審查的 busy) — 不顯示提示也不顯示 uploader
        # overlay 蓋全頁了, 這裡顯示什麼都看不到, 留空避免「框露半邊」的詭異畫面
        uploaded = None
    else:
        uploaded = st.file_uploader("選擇 PDF", type=["pdf"])

    # 判斷: 這個 uploaded 是不是「剛剛跑完審查的同一個檔」?
    # 跑完後 file_uploader 因為 widget 記憶, 重 rerun 仍然會回同一個檔,
    # 進入「if uploaded is not None」會重新顯示整個表單, 讓使用者誤以為
    # 「審查又自動開始 / 還沒結束」。改成: 已跑過審查 → 顯示「✅ 已完成」+
    # 「重新審查」按鈕, 不再顯示 checkbox / selectbox / 開始審查鈕。
    _already_reviewed = (
        uploaded is not None
        and st.session_state.get("_check_findings") is not None
        and st.session_state.get("_pdf_filename") == uploaded.name
    )

    if _already_reviewed:
        st.success(
            f"✅ 已審查: **{uploaded.name}** ({uploaded.size // 1024} KB) — "
            f"結果在下方顯示"
        )
        cc1, cc2 = st.columns([1, 4])
        with cc1:
            if st.button("🔄 重新審查", help="清掉本次結果, 重新跑審查",
                         disabled=_busy_now):
                # 清掉上次審查的結果, 但保留 PDF (使用者沒換檔, 不用重上傳)
                for k in ("_check_findings", "_app_data", "_sections",
                          "_ocr_result", "_flow_extract_result", "_flow_graph"):
                    st.session_state.pop(k, None)
                # 清示意圖頁快取
                for k in list(st.session_state.keys()):
                    if k.startswith("_balance_pages_cache::"):
                        st.session_state.pop(k, None)
                st.rerun()
        with cc2:
            st.caption("👇 結果可在下方「📊 各單元詳細頁」「智能審查結果」展開")

    # 一鍵跑完整流程: 抽取 + OCR + 智能審查 (只在「上傳了但還沒跑審查」時顯示)
    if uploaded is not None and not _already_reviewed:
        st.success(f"已上傳: **{uploaded.name}** ({uploaded.size // 1024} KB)")

        # 章節擷取 toggle — 只處理「參、水污染防治措施資料」段
        col_opt1, col_opt2 = st.columns([2, 3])
        with col_opt1:
            only_chapter_iii = st.checkbox(
                "✂️ 只處理「參、水污染防治措施資料」章節",
                value=True,
                key="_only_ch3",
                disabled=_busy_now,
                help="自動偵測該章節, 只抽該段內容 (省 token + 加速)。"
                     "找不到章節會自動 fallback 用全文。",
            )
        with col_opt2:
            if only_chapter_iii:
                st.caption(
                    "✅ 系統會先掃 PDF 找「參、」章節, 切出該段給 OCR/Gemini, 通常加速 5-10 倍"
                )
            else:
                st.caption("⚠️ 處理整本 PDF (慢、貴, 但完整)")

        # 事業類別選擇 (供智能審查使用) - 提前到按鈕之前讓使用者一次設定
        business_type = st.selectbox(
            "事業類別 (用於檢查申報項目是否完整)",
            ["(不檢查)"] + list(BUSINESS_TYPES.keys()),
            key="_business_type",
            disabled=_busy_now,
            help="選了之後智能審查會檢查該事業類別應申報的項目是否漏項",
        )

        # 是否同時跑「示意圖解析」(Step 4) — 用 Gemini Vision 抽流向圖
        # 預先檢查 Gemini key, 沒設就 disable + 改說明
        try:
            import gemini_extractor as _gex_check
            _gex_ok = _gex_check.check_gemini_status().get("ok", False)
        except Exception:
            _gex_ok = False

        also_flow = st.checkbox(
            "📊 同時跑「水量平衡示意圖解析」(Step 4)",
            value=_gex_ok,  # Gemini 沒設就預設不勾
            key="_also_flow",
            disabled=_busy_now or not _gex_ok,
            help=(
                "用 AI 視覺辨識讀流向示意圖, 拿到精確跨單元流向 + Q 值, "
                "並跟反推 Q 互相驗算。約 +30 秒, 約 $0.005-0.01 / 份。\n\n"
                "若不勾, 反推 Q 還是會算 (來自水質表), 只是不會做圖面比對。"
                if _gex_ok else "AI 影像辨識尚未啟用, 此選項不可用"
            ),
        )

        # 自動儲存快照: 跑完主審查時, 把內部覆核 Excel + JSON 存到 review_runs/
        # 預設開啟 (使用者要求: 「每次審查都記」)
        auto_snapshot = st.checkbox(
            "💾 自動儲存內部覆核快照 (存 review_runs/ + Sheets 記錄 run_id)",
            value=st.session_state.get("_auto_snapshot", True),
            key="_auto_snapshot",
            disabled=_busy_now,
            help="跑完主審查時自動產 internal Excel + JSON, 存到 review_runs/ 資料夾。30 天後自動壓縮, 90 天後只留 JSON。Sheets _審查紀錄 那一列會記下 run_id / 檔案路徑。",
        )

        # 🔒 若正在跑審查, button 禁用
        _is_busy = st.session_state.get("_busy", False)
        if _is_busy:
            st.warning(
                "⏳ **審查進行中,請勿關閉頁面或操作其他功能** — 此操作通常需要 30秒~3分鐘"
            )
        if st.button("🚀 開始完整審查", type="primary",
                     disabled=_is_busy,
                     help="一次跑完: 抽取單元 → 章節定位 → OCR 流向圖 → 智能審查"):
            # 標記正在跑「主審查」(鎖其他主要按鈕 + 觸發鎖定提示)
            st.session_state["_busy"] = True
            st.session_state["_busy_kind"] = "main"
            import time as _t_lock
            st.session_state["_busy_run_id"] = _t_lock.time()
            # 暫存 PDF 到 disk 供多步驟使用
            pdf_bytes = uploaded.read()
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tf:
                tf.write(pdf_bytes)
                tmp_path = tf.name

            # 若啟用章節擷取, 先找參章位置 (給後續 step 參考)
            chapter_info = None
            if only_chapter_iii:
                try:
                    import pdf_chapter_extractor
                    chapter_info = pdf_chapter_extractor.find_chapter_iii_pages(pdf_bytes)
                    if chapter_info.get("ok") and chapter_info.get("found"):
                        st.info(
                            f"✂️ 找到「參、水污染防治措施資料」: "
                            f"頁 {chapter_info['start_page']}~{chapter_info['end_page']} "
                            f"(全本 {chapter_info['total_pages']} 頁, "
                            f"節省 {100 - (chapter_info['end_page'] - chapter_info['start_page'] + 1) * 100 // chapter_info['total_pages']}%)"
                        )
                    elif chapter_info.get("ok"):
                        st.warning("⚠️ 找不到「參、水污染防治措施資料」章節, 改用全文")
                        chapter_info = None
                except Exception as _e:
                    st.caption(f"章節定位失敗 ({_e}), 改用全文")
                    chapter_info = None
            # 把章節範圍存到 session_state, 給其他模組讀
            st.session_state["_chapter_iii_info"] = chapter_info

            # 進度條 + 剩餘時間估算
            import time as _time
            t_start = _time.time()

            # 基準預估時間 (秒, 來自實測秋棠案例)
            # Step1: 60s, Step2 OCR: 90s, Step3: 5s, Step4 示意圖: 25s
            # 3-step: 155s, 4-step: 180s
            _BASELINE_TOTAL = 180 if st.session_state.get("_also_flow") else 155

            def _eta(percent_done):
                """估算剩餘秒數。

                早期 (已耗時 < 3 秒) 用基準預估值,
                晚期改用「已耗時 / 已完成%」實測推估 (更準)。
                """
                if percent_done <= 0:
                    return "預估中..."
                elapsed = _time.time() - t_start
                # 早期用基準, 避免「跑 0.5 秒 就以為總共 5 秒」
                if elapsed < 3:
                    remaining = _BASELINE_TOTAL * (100 - percent_done) / 100
                else:
                    total_est = elapsed / (percent_done / 100)
                    remaining = max(0, total_est - elapsed)
                if remaining < 1:
                    return "即將完成"
                if remaining < 60:
                    return f"約 {remaining:.0f} 秒"
                return f"約 {remaining/60:.1f} 分鐘"

            # ─── 全頁 overlay (CSS) 蓋住整個畫面, 中央顯示審查中 + 停止鈕 ───
            # 用 fixed position + 高 z-index 蓋掉 main + sidebar + tab 列
            st.markdown("""
            <style>
            .reviewing-overlay {
                position: fixed; top: 0; left: 0; right: 0; bottom: 0;
                background: rgba(15, 23, 42, 0.85);
                z-index: 9999;
                display: flex; align-items: center; justify-content: center;
                backdrop-filter: blur(4px);
            }
            .reviewing-card {
                background: white;
                border-radius: 16px;
                padding: 32px 40px;
                max-width: 560px; width: 90%;
                box-shadow: 0 20px 60px rgba(0,0,0,0.4);
                text-align: center;
            }
            .reviewing-card h2 { margin: 0 0 8px 0; font-size: 22px; color: #1e293b; }
            .reviewing-card .sub { color: #64748b; font-size: 14px; margin-bottom: 18px; }
            .reviewing-card .stage {
                background: #eff6ff; color: #1d4ed8;
                padding: 12px 16px; border-radius: 8px;
                font-weight: 600; margin: 12px 0;
                border-left: 4px solid #2563eb;
                text-align: left;
            }
            .reviewing-card .pct {
                font-size: 36px; font-weight: 700; color: #2563eb;
                margin: 8px 0 4px 0;
            }
            .reviewing-card .bar-bg {
                background: #e2e8f0; border-radius: 999px; height: 10px;
                overflow: hidden; margin: 8px 0 16px 0;
            }
            .reviewing-card .bar-fill {
                background: linear-gradient(90deg, #3b82f6, #2563eb);
                height: 100%; border-radius: 999px;
                transition: width 0.3s ease;
            }
            .reviewing-card .eta {
                color: #64748b; font-size: 13px;
                display: flex; justify-content: space-between; margin-top: 4px;
            }
            </style>
            """, unsafe_allow_html=True)

            # 這個 placeholder 會被各 step 不斷更新 (inject HTML)
            overlay_slot = st.empty()
            # 停止鈕 — 用獨立 form, 點下去會 rerun → 中斷正在跑的 Python 主執行緒
            cancel_slot = st.empty()

            def _render_overlay(percent, stage_text, eta_text):
                """更新 overlay 內容。"""
                overlay_slot.markdown(f"""
                <div class="reviewing-overlay">
                  <div class="reviewing-card">
                    <h2>🔍 審查進行中</h2>
                    <div class="sub">請勿關閉頁面 · 切勿點頁面其他按鈕</div>
                    <div class="stage">{stage_text}</div>
                    <div class="pct">{percent}%</div>
                    <div class="bar-bg"><div class="bar-fill" style="width: {percent}%;"></div></div>
                    <div class="eta"><span>已耗時 {int(_time.time() - t_start)} 秒</span><span>剩餘 {eta_text}</span></div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

            # 停止鈕 — 用 form 確保只有 submit 才 rerun, 而非每次互動
            with cancel_slot.container():
                # 用 columns 把按鈕推到右下角
                _cc1, _cc2, _cc3 = st.columns([3, 2, 1])
                with _cc3:
                    if st.button("🛑 停止審查",
                                 key=f"_stop_btn_{int(t_start)}",
                                 type="secondary",
                                 help="會中斷剩下的步驟, 已完成的結果會保留"):
                        # 設 cancel flag, 但因為這個 button click 本身就會觸發 rerun
                        # 中斷正在跑的審查 thread, 所以直接 set flag + rerun
                        st.session_state["_cancel_requested"] = True
                        st.session_state["_cancel_just_done"] = True
                        st.session_state["_busy"] = False
                        st.session_state.pop("_busy_run_id", None)
                        st.session_state.pop("_busy_kind", None)
                        st.rerun()

            try:
                # 動態 step 數: also_flow 勾就 4 step, 不勾就 3 step
                _total_steps = 4 if also_flow else 3
                _S = lambda n: f"Step {n}/{_total_steps}"

                # 進度條斷點
                if also_flow:
                    p1_end, p2_end, p3_end, p4_end = 30, 55, 75, 92
                else:
                    p1_end, p2_end, p3_end, p4_end = 40, 75, 96, 96

                # ─── Step 1: 章節定位 + 單元抽取 ───
                _render_overlay(10, f"{_S(1)} · 解析 PDF 章節與處理單元", _eta(10))
                sections_local = locate_sections(tmp_path, verbose=False)
                _render_overlay(p1_end - 10, f"{_S(1)} · 抽取處理單元結構化資料", _eta(p1_end - 10))
                if st.session_state.get("_cancel_requested"):
                    raise RuntimeError("使用者已要求停止")
                app_data_local = extract_application(tmp_path, verbose=False)
                # 修正: extract_application 拿 tmp_path 當 source_pdf, 蓋回真正的上傳檔名
                if app_data_local and getattr(uploaded, 'name', None):
                    app_data_local['source_pdf'] = uploaded.name
                _render_overlay(p1_end, f"{_S(1)} · 完成 (共 {app_data_local['total_units']} 個處理單元)", _eta(p1_end))

                # 存到 session
                ocr_target_pages = sorted(set(
                    sections_local.get("flow_diagram", []) +
                    sections_local.get("balance_diagram", [])
                ))
                st.session_state["_sections"] = sections_local
                st.session_state["_app_data"] = app_data_local
                st.session_state["_pdf_bytes"] = pdf_bytes
                st.session_state["_pdf_filename"] = uploaded.name
                st.session_state["_ocr_target_pages"] = ocr_target_pages
                # 建立水流串接圖
                st.session_state["_flow_graph"] = build_flow_graph(app_data_local)

                # ─── Step 2: OCR (若有流向圖頁) ───
                if st.session_state.get("_cancel_requested"):
                    raise RuntimeError("使用者已要求停止")
                _p2_mid = (p1_end + p2_end) // 2
                if ocr_target_pages:
                    n_ocr_pages = len(ocr_target_pages)
                    _render_overlay(_p2_mid, f"{_S(2)} · OCR 解析 {n_ocr_pages} 頁流向圖 / 水量平衡圖", _eta(_p2_mid))
                    ocr_result = ocr_diagram_pages(tmp_path, ocr_target_pages, verbose=False)
                    st.session_state["_ocr_result"] = ocr_result
                    if "error" not in ocr_result:
                        summary = ocr_result["summary"]
                        _render_overlay(p2_end, f"{_S(2)} · 完成 (識別 {summary['total_units']} 單元 / {summary['total_flows']} 流量)", _eta(p2_end))
                    else:
                        _render_overlay(p2_end, f"{_S(2)} · OCR 略過", _eta(p2_end))
                else:
                    st.session_state.pop("_ocr_result", None)
                    _render_overlay(p2_end, f"{_S(2)} · 無流向圖頁面, 跳過 OCR", _eta(p2_end))

                # ─── Step 3: 智能審查 (3 層: 質量平衡 + 學理 + 規則庫驅動) ───
                if st.session_state.get("_cancel_requested"):
                    raise RuntimeError("使用者已要求停止")
                _p3_a = p2_end + (p3_end - p2_end) // 3
                _p3_b = p2_end + 2 * (p3_end - p2_end) // 3
                _render_overlay(_p3_a, f"{_S(3)} · 質量平衡檢查", _eta(_p3_a))
                findings_basic = run_balance_checks(app_data_local)
                _render_overlay(_p3_b, f"{_S(3)} · 學理檢查 (環工設計準則)", _eta(_p3_b))
                bt = None if business_type == "(不檢查)" else business_type
                findings_adv = run_advanced_checks(app_data_local, business_type=bt)
                _render_overlay(p3_end, f"{_S(3)} · 規則庫驅動檢查 (規則庫 299 筆環工技師缺失)", _eta(p3_end))
                findings_rule = run_rule_driven_check(app_data_local)

                # ─── 新增: RPM 攪拌轉速檢查 + 放流水標準檢查 (2026-06-30, be0ddc1) ───
                findings_rpm = []
                findings_discharge = []
                try:
                    from check_rpm import check_all_units_rpm
                    findings_rpm = check_all_units_rpm(app_data_local.get("units") or {})
                except Exception as _e_rpm:
                    print(f"[check_rpm 失敗] {_e_rpm}")

                try:
                    from discharge_standard_loader import check_all_discharge_units
                    from extract_production_scale import extract_section_5
                    _industry = ""
                    try:
                        _scale = extract_section_5(tmp_path)
                        _industry_raw = _scale.get("業別") or ""
                        for _kw in ["電鍍", "PCB", "印刷電路板", "化工", "化學",
                                    "食品", "紙板", "造紙", "金屬基本", "金屬表面"]:
                            if _kw in _industry_raw:
                                _industry = _kw
                                break
                    except Exception:
                        pass
                    if _industry:
                        findings_discharge = check_all_discharge_units(
                            app_data_local.get("units") or {}, _industry
                        )
                except Exception as _e_dis:
                    print(f"[check_discharge 失敗] {_e_dis}")

                # ─── 新增: 進=出完全相同 偵測 (廠商偷懶填表, 2026-06-30) ───
                findings_identical = []
                try:
                    from check_identical_inout import check_all_units_identical_inout
                    findings_identical = check_all_units_identical_inout(
                        app_data_local.get("units") or {}
                    )
                except Exception as _e_id:
                    print(f"[check_identical 失敗] {_e_id}")

                # B+C (2026-07-01): 跨層 dedup
                # B: 若同單元同時有具體 finding + 通用 finding → 移除通用
                #    通用特徵: 描述含 "需檢驗/需人工/應保持不變" 等模糊語彙, 且沒具體數字
                # C: 若同單元有其他 finding + 進=出偷懶 → 進=出併入描述末尾
                _all_findings = (
                    findings_basic + findings_adv + findings_rule
                    + findings_rpm + findings_discharge
                )

                _GENERIC_PHRASES = ["需檢驗", "需人工檢視", "多半需人工", "需檢核",
                                    "水質濃度及質量", "各項水質濃度",
                                    "應保持不變", "應予檢核"]

                def _is_generic_finding(_f):
                    import re as _re
                    _d = str(_f.get("描述") or "")
                    _t = str(_f.get("對照項目") or "")
                    # 有具體數字 → 非通用
                    if _re.search(r"進\s*-?\d|\d+\s*→\s*\d|\d+\.\d+\s*%|\d+%|去除\s*-?\d", _d):
                        return False
                    for _p in _GENERIC_PHRASES:
                        if _p in _d or _p in _t:
                            return True
                    return False

                # 找出「有具體 finding」的單元
                _units_specific = set()
                for _f in _all_findings:
                    if not _is_generic_finding(_f):
                        _u = _f.get("單元")
                        if _u:
                            _units_specific.add(_u)

                # B: 移除同單元的通用 finding (但保留該單元「僅有」通用 finding 的情況)
                _kept_findings = []
                for _f in _all_findings:
                    if _is_generic_finding(_f) and _f.get("單元") in _units_specific:
                        continue
                    _kept_findings.append(_f)

                # C: 進=出偷懶 — 同單元有其他 finding 就併入描述, 沒有就獨立列
                _identical_standalone = []
                _identical_annotations = {}
                _units_with_kept = set(
                    _f.get("單元") for _f in _kept_findings if _f.get("單元")
                )
                for _f in findings_identical:
                    _code = _f.get("單元")
                    if _code and _code in _units_with_kept:
                        _desc = _f.get("描述", "")
                        _brief = _desc.split("。")[0][:80] if _desc else "進出水質高度雷同"
                        _identical_annotations[_code] = f" [附註: 該單元{_brief}]"
                    else:
                        _identical_standalone.append(_f)

                # 加註 (每單元只加 1 次)
                for _f in _kept_findings:
                    _code = _f.get("單元")
                    if _code in _identical_annotations:
                        _f["描述"] = (_f.get("描述") or "") + _identical_annotations[_code]
                        del _identical_annotations[_code]

                st.session_state["_check_findings"] = (
                    _kept_findings + _identical_standalone
                )

                # ─── Step 4: 水量平衡示意圖解析 (AI 視覺辨識, 可選) ───
                if also_flow:
                    if st.session_state.get("_cancel_requested"):
                        raise RuntimeError("使用者已要求停止")
                    try:
                        import flow_diagram_extractor as _fde_main
                        # 先找示意圖頁
                        _render_overlay(p3_end + 1, f"{_S(4)} · 定位示意圖頁面", _eta(p3_end + 1))
                        _loc = _fde_main.find_balance_diagram_pages(pdf_bytes)
                        _img_pages_count = len(_loc.get("image_pages", [])) if _loc.get("ok") else 0
                        if _img_pages_count > 0:
                            # 預設處理上限 5 張避免太貴 (示意圖通常 1-3 張)
                            _max_pages = min(_img_pages_count, 5)
                            _render_overlay(
                                p3_end + 2,
                                f"{_S(4)} · 找到 {_img_pages_count} 張示意圖, 將處理 {_max_pages} 張 (AI 視覺辨識)",
                                _eta(p3_end + 2),
                            )

                            def _flow_cb(cur, tot, msg):
                                # 把示意圖內部進度 (0~100%) 映射到 p3_end+2 ~ p4_end-1
                                _range = p4_end - p3_end - 2
                                inner = (cur / tot) if tot else 0
                                _pct = p3_end + 2 + int(inner * _range)
                                _render_overlay(
                                    min(_pct, p4_end - 1),
                                    f"{_S(4)} · 處理第 {cur}/{tot} 張示意圖 · {msg}",
                                    _eta(_pct),
                                )

                            _fr = _fde_main.extract_all_balance_diagrams(
                                pdf_bytes,
                                max_pages=_max_pages,
                                progress_callback=_flow_cb,
                            )
                            st.session_state["_flow_extract_result"] = _fr
                            if _fr.get("ok"):
                                _render_overlay(
                                    p4_end,
                                    f"{_S(4)} · 完成 (處理 {_fr.get('pages_processed', 0)} 張, "
                                    f"抽出 {len(_fr.get('all_units', []))} 單元 / "
                                    f"{len(_fr.get('all_flows', []))} 流向)",
                                    _eta(p4_end),
                                )
                            else:
                                _render_overlay(p4_end, f"{_S(4)} · AI 解析失敗 (已略過, 可稍後手動重試)", _eta(p4_end))
                        else:
                            _render_overlay(p4_end, f"{_S(4)} · 找不到示意圖頁, 已跳過", _eta(p4_end))
                    except Exception as _flow_err:
                        # 示意圖解析失敗不致命, 繼續主流程
                        _render_overlay(p4_end, f"{_S(4)} · 解析略過 ({str(_flow_err)[:40]})", _eta(p4_end))

                # 不要立刻顯示 100% — 還有後續工作 (統計 / 寫 Sheet 歷史紀錄)
                _render_overlay(97, "整理結果並存檔...", "幾秒")

                total_elapsed = _time.time() - t_start
                stats = {"不合理": 0, "待確認": 0}
                for f in st.session_state["_check_findings"]:
                    sev = f.get("嚴重度")
                    if sev in stats:
                        stats[sev] += 1
                # 完成後標記要顯示成功訊息 (在 finally 清完 overlay 後顯示)
                st.session_state["_review_just_done"] = {
                    "units": app_data_local["total_units"],
                    "unreasonable": stats["不合理"],
                    "manual": stats["待確認"],
                    "elapsed": int(total_elapsed),
                }

                # 記錄到 session 歷史 (重整就消失, 但本次瀏覽可看)
                # 強制台北時區 (Streamlit Cloud 預設可能是 UTC)
                from datetime import datetime as _dt
                try:
                    from zoneinfo import ZoneInfo as _ZI
                    _tpe_tz = _ZI("Asia/Taipei")
                except Exception:
                    from datetime import timezone as _tz, timedelta as _td
                    _tpe_tz = _tz(_td(hours=8))
                if "_review_history" not in st.session_state:
                    st.session_state["_review_history"] = []
                review_record = {
                    "time": _dt.now(_tpe_tz).strftime("%H:%M:%S"),
                    "filename": uploaded.name,
                    "units": app_data_local["total_units"],
                    "unreasonable": stats["不合理"],
                    "manual": stats["待確認"],
                    "elapsed_sec": int(total_elapsed),
                }
                st.session_state["_review_history"].insert(0, review_record)
                # 限制歷史只保留最近 10 筆
                st.session_state["_review_history"] = st.session_state["_review_history"][:10]

                # ─── 自動儲存內部覆核快照 (依 checkbox 控制) ───
                if st.session_state.get("_auto_snapshot", True):
                    try:
                        _render_overlay(98, "儲存內部覆核快照...", "5 秒")
                        import export_report as _xrp_snap
                        import review_snapshot as _rs_snap
                        _findings_all = st.session_state.get("_check_findings") or []
                        # 業別
                        _bt_snap = st.session_state.get("_business_type", "") or ""
                        if _bt_snap == "(不檢查)":
                            _bt_snap = ""
                        _snap_opts = {"business_type": _bt_snap}
                        # 用 internal 預設值
                        for k, v in _xrp_snap.get_default_options("internal").items():
                            _snap_opts[k] = v
                        _base = os.path.splitext(uploaded.name)[0]
                        _excel_bytes, _, _ = _xrp_snap.build_export(
                            "internal", app_data_local, _findings_all, _snap_opts, base_name=_base
                        )
                        _json_bytes, _, _ = _xrp_snap.build_export(
                            "json", app_data_local, _findings_all, _snap_opts, base_name=_base
                        )
                        _snap_fc = {"不合理": stats["不合理"], "待確認": stats["待確認"], "錯誤": stats.get("錯誤", 0)}
                        _snap_result = _rs_snap.save_snapshot(
                            excel_bytes=_excel_bytes,
                            json_bytes=_json_bytes,
                            base_name=_base,
                            findings_count=_snap_fc,
                        )
                        if _snap_result.get("ok"):
                            # 把 run_id / paths 塞進 review_record, 讓 Sheets 一起寫進去
                            review_record["run_id"] = _snap_result["run_id"]
                            review_record["snapshot_report_path"] = _snap_result["report_path"]
                            review_record["snapshot_json_path"] = _snap_result["json_path"]
                            st.session_state["_last_snapshot_info"] = _snap_result
                    except Exception as _snap_err:
                        # 快照失敗不致命, 主流程繼續
                        st.session_state["_last_snapshot_error"] = str(_snap_err)

                # 寫入歷史紀錄 (網路請求可能要 5-10 秒)
                _render_overlay(99, "寫入審查紀錄 (雲端存檔)...", "幾秒")
                try:
                    import review_history
                    sheet_result = review_history.append_review_record(review_record)
                    if sheet_result.get("ok"):
                        st.session_state["_last_sheet_log_status"] = (
                            f"✅ 紀錄已存檔 (本文件第 {sheet_result['review_times']} 次審查)"
                        )
                    else:
                        # 失敗不顯示給使用者 — 改成內部記錄即可, 不影響主流程
                        st.session_state["_last_sheet_log_status"] = ""
                except Exception:
                    st.session_state["_last_sheet_log_status"] = ""

                # 真正完成 — 短暫顯示 100% 後 finally 會清掉 overlay 並 rerun
                total_elapsed = _time.time() - t_start
                _render_overlay(100, f"✅ 全部完成! 共耗時 {total_elapsed:.0f} 秒", "0 秒")
            except RuntimeError as _re:
                # 使用者按了停止鈕 (在 step 之間檢查到 _cancel_requested)
                if "停止" in str(_re):
                    st.session_state["_cancel_just_done"] = True
                else:
                    raise
            finally:
                try:
                    os.unlink(tmp_path)
                except:
                    pass
                # 清掉 overlay 跟停止鈕
                try:
                    overlay_slot.empty()
                    cancel_slot.empty()
                except Exception:
                    pass
                # 解鎖 UI
                st.session_state["_busy"] = False
                st.session_state.pop("_busy_run_id", None)
                st.session_state.pop("_busy_kind", None)
                st.session_state.pop("_cancel_requested", None)
                # 重新 rerun, 讓主畫面正常顯示結果
                st.rerun()

    # ───────── 顯示區 (永遠基於 session_state, 不被 button rerun 影響) ─────────
    # ── 水量平衡示意圖解析 (Gemini Vision 抽結構化流向) ──
    pdf_bytes_for_flow = st.session_state.get("_pdf_bytes")
    if pdf_bytes_for_flow:
        with st.expander("📊 水量平衡示意圖解析 (跨單元流向 + 質平檢核)", expanded=False):
            st.caption(
                "從申請文件的「參、水污染防治措施資料 / 水質水量平衡示意圖」抽出流向結構,"
                " 可得知: 每個單元有幾條進流/出流、來源/去處、流量 Q, 並做跨單元質量平衡檢核。"
            )

            try:
                import flow_diagram_extractor as _fde
                import gemini_extractor as _ge_chk
                fde_ok = True
            except Exception as _fe:
                st.error(f"無法載入 flow_diagram_extractor: {_fe}")
                fde_ok = False

            if fde_ok:
                _fstat = _ge_chk.check_gemini_status()
                if not _fstat["ok"]:
                    st.warning(f"⚠️ {_fstat['message']} — 此功能需要 Gemini Vision")
                else:
                    # 偵測圖頁 — 用 session 快取避免每次 rerun 都掃整本 PDF
                    # (修 bug: 沒快取的話, 使用者勾 checkbox 或任何 widget 互動都會觸發
                    #  rerun, 此函式會跑 5-10 秒掃 400+ 頁 PDF, 看起來像「系統還在執行」)
                    _pdf_filename_now = st.session_state.get("_pdf_filename", "")
                    _cache_key = f"_balance_pages_cache::{_pdf_filename_now}"
                    if _cache_key not in st.session_state:
                        with st.spinner("偵測示意圖頁..."):
                            st.session_state[_cache_key] = _fde.find_balance_diagram_pages(
                                pdf_bytes_for_flow
                            )
                    loc = st.session_state[_cache_key]
                    if loc.get("ok") and loc.get("image_pages"):
                        img_pages = loc["image_pages"]
                        st.info(
                            f"📍 找到 **{len(img_pages)}** 張水量平衡示意圖頁: "
                            f"p{', p'.join(map(str, img_pages))}"
                        )
                        # 修: 只有 1 張時不能用 slider (min==max), 直接固定處理那 1 張
                        if len(img_pages) <= 1:
                            max_pages = len(img_pages)
                            st.caption(f"將處理全部 **{max_pages}** 張示意圖 (約 ${max_pages * 0.005:.3f})")
                        else:
                            max_pages = st.slider(
                                "處理頁數 (省 token, 先測一張看效果)",
                                min_value=1, max_value=len(img_pages),
                                value=min(2, len(img_pages)),
                                key="_flow_max_pages",
                                help=f"每張圖約 $0.002~0.005, 共 {len(img_pages)} 張全跑約 ${len(img_pages) * 0.005:.3f}"
                            )

                        if "_flow_extract_result" not in st.session_state:
                            st.session_state["_flow_extract_result"] = None

                        # 若已有結果 → 顯示「✅ 已解析」, 按鈕變成「🔄 重抽」
                        _existing_fr = st.session_state.get("_flow_extract_result")
                        if _existing_fr and _existing_fr.get("ok"):
                            _btn_label = "🔄 重新解析 (重抽 Gemini)"
                            _summary_n_flows = len(_existing_fr.get("all_flows", []))
                            _summary_n_units = len(_existing_fr.get("all_units", []))
                            st.success(
                                f"✅ 已解析: 處理 {_existing_fr.get('pages_processed', 0)} 張 / "
                                f"{_summary_n_units} 單元 / {_summary_n_flows} 流向 — "
                                f"結果已套用到上方各單元的「Q (示意圖解析)」"
                            )
                        else:
                            _btn_label = "🤖 開始解析"

                        if st.button(_btn_label, type="primary", key="_btn_flow_extract",
                                     width="stretch",
                                     disabled=st.session_state.get("_busy", False)):
                            # 跟主審查一樣套全頁 overlay, 讓使用者明確知道「系統正在跑」
                            import time as _t_flow
                            st.session_state["_busy"] = True
                            st.session_state["_busy_kind"] = "flow"  # 示意圖解析, 不顯示主鎖定提示
                            st.session_state["_busy_run_id"] = _t_flow.time()

                            # CSS overlay 樣式 (跟主審查同一套)
                            flow_overlay = st.empty()
                            _t_flow_start = _t_flow.time()

                            def _render_flow_overlay(percent, stage_text):
                                elapsed = int(_t_flow.time() - _t_flow_start)
                                flow_overlay.markdown(f"""
                                <style>
                                .reviewing-overlay {{
                                    position: fixed; top: 0; left: 0; right: 0; bottom: 0;
                                    background: rgba(15, 23, 42, 0.85);
                                    z-index: 9999;
                                    display: flex; align-items: center; justify-content: center;
                                    backdrop-filter: blur(4px);
                                }}
                                .reviewing-card {{
                                    background: white; border-radius: 16px;
                                    padding: 32px 40px; max-width: 560px; width: 90%;
                                    box-shadow: 0 20px 60px rgba(0,0,0,0.4);
                                    text-align: center;
                                }}
                                .reviewing-card h2 {{ margin: 0 0 8px 0; font-size: 22px; color: #1e293b; }}
                                .reviewing-card .sub {{ color: #64748b; font-size: 14px; margin-bottom: 18px; }}
                                .reviewing-card .stage {{
                                    background: #eff6ff; color: #1d4ed8;
                                    padding: 12px 16px; border-radius: 8px;
                                    font-weight: 600; margin: 12px 0;
                                    border-left: 4px solid #2563eb;
                                    text-align: left;
                                }}
                                .reviewing-card .pct {{ font-size: 36px; font-weight: 700; color: #2563eb; margin: 8px 0 4px 0; }}
                                .reviewing-card .bar-bg {{
                                    background: #e2e8f0; border-radius: 999px; height: 10px;
                                    overflow: hidden; margin: 8px 0 16px 0;
                                }}
                                .reviewing-card .bar-fill {{
                                    background: linear-gradient(90deg, #3b82f6, #2563eb);
                                    height: 100%; border-radius: 999px;
                                    transition: width 0.3s ease;
                                }}
                                .reviewing-card .eta {{ color: #64748b; font-size: 13px; text-align: right; margin-top: 4px; }}
                                </style>
                                <div class="reviewing-overlay">
                                  <div class="reviewing-card">
                                    <h2>📊 水量平衡示意圖解析中</h2>
                                    <div class="sub">Gemini Vision 處理中 · 請勿關閉頁面 · 切勿點頁面其他按鈕</div>
                                    <div class="stage">{stage_text}</div>
                                    <div class="pct">{percent}%</div>
                                    <div class="bar-bg"><div class="bar-fill" style="width: {percent}%;"></div></div>
                                    <div class="eta">已耗時 {elapsed} 秒</div>
                                  </div>
                                </div>
                                """, unsafe_allow_html=True)

                            def _cb(cur, tot, msg):
                                pct = int((cur / tot) * 100) if tot else 0
                                _render_flow_overlay(min(pct, 99), f"處理第 {cur} / {tot} 張示意圖 · {msg}")

                            try:
                                _render_flow_overlay(5, f"準備解析 {max_pages} 張示意圖...")
                                fr = _fde.extract_all_balance_diagrams(
                                    pdf_bytes_for_flow,
                                    max_pages=max_pages,
                                    progress_callback=_cb,
                                )
                                _render_flow_overlay(100, "✅ 解析完成!")
                            finally:
                                # 清掉 overlay
                                try:
                                    flow_overlay.empty()
                                except Exception:
                                    pass
                                st.session_state["_flow_extract_result"] = fr
                                # 解鎖
                                st.session_state["_busy"] = False
                                st.session_state.pop("_busy_run_id", None)
                                st.session_state.pop("_busy_kind", None)
                                st.rerun()

                        fr = st.session_state.get("_flow_extract_result")
                        if fr:
                            if not fr.get("ok"):
                                st.error(f"❌ 失敗 ({fr.get('stage')}): {fr.get('error')}")
                            else:
                                usage = fr.get("gemini_usage", {})
                                st.success(
                                    f"✅ **水量平衡示意圖解析完成** · "
                                    f"處理 {fr['pages_processed']} 頁 / "
                                    f"抽出 **{len(fr['all_units'])}** 單元 / "
                                    f"**{len(fr['all_flows'])}** 流向 / "
                                    f"**{len(fr['all_external_inputs'])}** 外部輸入 / "
                                    f"**{len(fr['all_discharge_points'])}** 放流口"
                                )
                                st.caption(
                                    f"Token: in={usage.get('input_tokens', '?')} "
                                    f"out={usage.get('output_tokens', '?')}"
                                )

                                if fr.get("errors"):
                                    with st.expander(f"⚠️ {len(fr['errors'])} 頁解析失敗"):
                                        for e in fr["errors"]:
                                            st.caption(f"p{e['page']}: {e['error'][:100]}")

                                tabs = st.tabs([
                                    "🌊 流向總表", "📦 單元清單",
                                    "⚖️ 質平檢核", "🔗 編號一致性",
                                    "🔄 跟系統比對", "📄 各頁原始 JSON",
                                ])

                                with tabs[0]:
                                    import pandas as _pd
                                    if fr["all_flows"]:
                                        df_f = _pd.DataFrame(fr["all_flows"]).astype(str)
                                        st.dataframe(df_f, width="stretch", hide_index=True)
                                    else:
                                        st.info("沒有抽到流向")

                                    if fr["all_external_inputs"]:
                                        st.markdown("**📥 外部原廢水 (WM)**")
                                        st.dataframe(
                                            _pd.DataFrame(fr["all_external_inputs"]).astype(str),
                                            width="stretch", hide_index=True
                                        )

                                    if fr["all_discharge_points"]:
                                        st.markdown("**📤 放流口 (D)**")
                                        st.dataframe(
                                            _pd.DataFrame(fr["all_discharge_points"]).astype(str),
                                            width="stretch", hide_index=True
                                        )

                                with tabs[1]:
                                    if fr["all_units"]:
                                        import pandas as _pd
                                        df_u = _pd.DataFrame(fr["all_units"]).astype(str)
                                        st.dataframe(df_u, width="stretch", hide_index=True)
                                        st.caption(f"圖中共 {len(fr['all_units'])} 個處理單元")
                                    else:
                                        st.info("沒有抽到單元")

                                with tabs[2]:
                                    bal = _fde.check_water_balance(fr)
                                    s = bal["summary"]
                                    cb1, cb2, cb3, cb4 = st.columns(4)
                                    cb1.metric("總單元", s["total_units"])
                                    cb2.metric("✅ 平衡 (<1%)", s["balanced_count"])
                                    cb3.metric("⚠️ 偏差 (1~5%)", s["warning_count"])
                                    cb4.metric("❌ 異常 (>5%)", s["error_count"])

                                    rows_b = []
                                    for code, info in bal["by_unit"].items():
                                        rows_b.append({
                                            "單元": code,
                                            "Σ進流 (CMD)": info["in_total_cmd"],
                                            "Σ出流 (CMD)": info["out_total_cmd"],
                                            "差異 (%)": info["diff_pct"],
                                            "警告": info["warning"] or "",
                                        })
                                    if rows_b:
                                        import pandas as _pd
                                        st.dataframe(
                                            _pd.DataFrame(rows_b).astype(str),
                                            width="stretch", hide_index=True
                                        )

                                with tabs[3]:
                                    # 編號一致性檢核
                                    consistency = _fde.check_stream_consistency(fr)
                                    s = consistency["summary"]
                                    sc1, sc2, sc3, sc4 = st.columns(4)
                                    sc1.metric("總編號數", s["total_streams"])
                                    sc2.metric("⚠️ 警告", s["warnings_count"])
                                    sc3.metric("🔍 上游有編號\n但下游漏抽", s["unmatched_from_count"])
                                    sc4.metric("🔍 下游有編號\n但上游漏抽", s["unmatched_to_count"])

                                    if consistency["warnings"]:
                                        st.markdown("**⚠️ 一致性警告**")
                                        warn_rows = []
                                        for w in consistency["warnings"]:
                                            warn_rows.append({
                                                "類型": w["type"],
                                                "編號": w.get("stream", ""),
                                                "訊息": w["message"],
                                            })
                                        import pandas as _pd
                                        st.dataframe(
                                            _pd.DataFrame(warn_rows).astype(str),
                                            width="stretch", hide_index=True
                                        )
                                    else:
                                        st.success("✅ 沒有編號一致性問題")

                                    if consistency["unmatched_from"]:
                                        with st.expander(
                                            f"🔍 上游 WTA 編號 {len(consistency['unmatched_from'])} 條, 下游沒抽到對應 WTB"
                                        ):
                                            import pandas as _pd
                                            st.dataframe(
                                                _pd.DataFrame(consistency["unmatched_from"]).astype(str),
                                                width="stretch", hide_index=True
                                            )
                                            st.caption("可能原因: 圖中沒標下游 WTB 編號 / Gemini 漏抽 / 該箭頭實際進入外部")

                                    if consistency["unmatched_to"]:
                                        with st.expander(
                                            f"🔍 下游 WTB 編號 {len(consistency['unmatched_to'])} 條, 上游沒抽到對應 WTA"
                                        ):
                                            import pandas as _pd
                                            st.dataframe(
                                                _pd.DataFrame(consistency["unmatched_to"]).astype(str),
                                                width="stretch", hide_index=True
                                            )
                                            st.caption("可能原因: 來自外部原廢水 (WM) / Gemini 漏抽上游編號")

                                    with st.expander("📋 全部編號詳細位置 (debug)"):
                                        cg = consistency["code_groups"]
                                        cg_rows = []
                                        for code, info in cg.items():
                                            cg_rows.append({
                                                "編號": code,
                                                "as_from 次數": len(info["as_from"]),
                                                "as_to 次數": len(info["as_to"]),
                                                "Q 值集合": str(info["Q_set"]),
                                            })
                                        if cg_rows:
                                            import pandas as _pd
                                            st.dataframe(
                                                _pd.DataFrame(cg_rows).astype(str),
                                                width="stretch", hide_index=True
                                            )

                                with tabs[4]:
                                    # 跟現有 app_data 比對
                                    app_data_cmp = st.session_state.get("_app_data") or {}
                                    if not app_data_cmp:
                                        st.info("尚未做完整審查, 先按上方「開始完整審查」")
                                    else:
                                        sys_units = set(app_data_cmp.get("units", {}).keys())
                                        img_units = {u["code"] for u in fr["all_units"]}
                                        only_in_sys = sys_units - img_units
                                        only_in_img = img_units - sys_units
                                        both = sys_units & img_units

                                        cmp1, cmp2, cmp3 = st.columns(3)
                                        cmp1.metric("共同有", len(both))
                                        cmp2.metric("僅系統有", len(only_in_sys))
                                        cmp3.metric("僅圖中有", len(only_in_img))

                                        if only_in_img:
                                            st.warning(
                                                f"圖上有但系統沒抽到的單元: {', '.join(sorted(only_in_img))}"
                                            )
                                        if only_in_sys:
                                            st.caption(
                                                f"系統有但圖上沒看到 (可能是污泥/反洗等支線): "
                                                f"{', '.join(sorted(only_in_sys))}"
                                            )

                                        # 比對進流數
                                        st.markdown("**進流數比對 (圖中 vs 系統抽取)**")
                                        cmp_rows = []
                                        in_cnt_img = {}
                                        for flow in fr["all_flows"]:
                                            to_u = flow.get("to_unit")
                                            if to_u:
                                                in_cnt_img[to_u] = in_cnt_img.get(to_u, 0) + 1
                                        for ext in fr["all_external_inputs"]:
                                            to_u = ext.get("to_unit")
                                            if to_u:
                                                in_cnt_img[to_u] = in_cnt_img.get(to_u, 0) + 1
                                        for code in sorted(both):
                                            sys_in = len(app_data_cmp["units"][code].get("influent", {}))
                                            img_in = in_cnt_img.get(code, 0)
                                            cmp_rows.append({
                                                "單元": code,
                                                "圖中進流數": img_in,
                                                "系統抽取進流數": sys_in,
                                                "差異": img_in - sys_in,
                                                "備註": "✅" if img_in == sys_in else "⚠️ 不一致",
                                            })
                                        if cmp_rows:
                                            import pandas as _pd
                                            st.dataframe(
                                                _pd.DataFrame(cmp_rows).astype(str),
                                                width="stretch", hide_index=True
                                            )

                                with tabs[5]:
                                    for pr in fr.get("per_page_results", []):
                                        st.markdown(f"**p{pr['page']} - {pr.get('system_title', '')}**")
                                        st.json(pr["data"])
                    else:
                        st.caption("此 PDF 沒找到水量平衡示意圖純圖頁")

    # ── 補充: 圖片局部判讀 (Gemini Vision) ──
    with st.expander("📷 補充: 上傳圖片做局部判讀 (Gemini Vision)", expanded=False):
        st.caption(
            "用途: 上傳申請文件的截圖 (流向圖 / 水量平衡 / 數據表 / 設計尺寸) 讓 Gemini 判讀。"
            " 適用於 PDF 抽取不夠完整、想針對特定區塊深入分析的時候。"
        )
        try:
            import gemini_vision as _gv
            _gv_ok = True
        except Exception as _e:
            st.error(f"無法載入 gemini_vision: {_e}")
            _gv_ok = False

        if _gv_ok:
            # 認證狀態
            try:
                import gemini_extractor as _ge
                _gs = _ge.check_gemini_status()
            except Exception:
                _gs = {"ok": False, "message": "?"}

            if not _gs.get("ok"):
                st.warning(f"⚠️ {_gs.get('message')} (請見『匯入新規則』分頁設定)")
            else:
                imgs = st.file_uploader(
                    "拖入 1~N 張圖片 (PNG / JPG / WEBP)",
                    type=["png", "jpg", "jpeg", "webp"],
                    accept_multiple_files=True,
                    key="_review_img_uploader",
                )
                focus_hint = st.text_input(
                    "提示 (可選, 告訴 Gemini 該注意什麼)",
                    placeholder="例: 請特別注意流向是否標示反洗水來源",
                    key="_review_img_hint",
                )

                if imgs:
                    cols_thumb = st.columns(min(len(imgs), 4))
                    for i, img in enumerate(imgs):
                        with cols_thumb[i % 4]:
                            st.image(img, caption=img.name, width=150)

                    if st.button("🔍 開始判讀", key="_btn_interpret_images",
                                 type="primary"):
                        with st.spinner(f"Gemini Vision 判讀 {len(imgs)} 張圖片…"):
                            img_bytes_list = [img.getvalue() for img in imgs]
                            _gr = _gv.process_images(
                                img_bytes_list,
                                mode="interpret_diagram",
                                focus_hint=focus_hint,
                            )
                        if _gr.get("ok"):
                            result = _gr.get("result", {})
                            st.success(
                                f"✅ 判讀完成 ({_gr['image_count']} 張) · "
                                f"信心度 {result.get('confidence', '?')} · "
                                f"tokens in {_gr['gemini_usage'].get('input_tokens')} "
                                f"out {_gr['gemini_usage'].get('output_tokens')}"
                            )
                            st.markdown(f"**圖表類型**: {result.get('diagram_type', '(未判斷)')}")

                            if result.get("observations"):
                                st.markdown("**🔍 關鍵發現**")
                                for obs in result["observations"]:
                                    st.markdown(f"- {obs}")

                            if result.get("concerns"):
                                st.markdown("**⚠️ 可能不合理之處**")
                                for c in result["concerns"]:
                                    st.markdown(f"- {c}")

                            if result.get("units"):
                                st.markdown(f"**📦 槽體資料 ({len(result['units'])} 個)**")
                                import pandas as _pd
                                df_u = _pd.DataFrame(result["units"]).astype(str)
                                st.dataframe(df_u, width="stretch", hide_index=True)

                            if result.get("flows"):
                                st.markdown(f"**🌊 流向資料 ({len(result['flows'])} 條)**")
                                import pandas as _pd
                                df_f = _pd.DataFrame(result["flows"]).astype(str)
                                st.dataframe(df_f, width="stretch", hide_index=True)

                            with st.expander("Raw text (圖片裡看得到的所有文字)"):
                                st.text(result.get("raw_text", ""))
                        else:
                            st.error(f"❌ 失敗 ({_gr.get('stage')}): {_gr.get('error')}")

    if st.session_state.get("_app_data"):
        app_data = st.session_state["_app_data"]
        sections = st.session_state["_sections"]
        pdf_filename = st.session_state.get("_pdf_filename", "")

        st.success(f"✅ 抽取完成! 共 **{app_data['total_units']}** 個處理單元 · 來源: {pdf_filename}")

        # ───────── 防呆檢查 (2026-07-01): 偵測「PDF 結構不支援」case ─────────
        # 避免同事上傳客製化格式 PDF (例如 P-01 而非 T01) 時看到「總審查項 4 / 不合理 0」誤以為完美
        _n_units = app_data.get("total_units", 0)
        _pdf_pages = app_data.get("total_pages", 0)
        _facility_pages = len(sections.get("facility_table", []))
        _quality_pages = len(sections.get("quality_data", []))

        if _n_units == 0:
            st.error(
                "🚨 **系統未偵測到任何處理單元**\n\n"
                "可能原因:\n"
                "1. 該 PDF 使用非 T01-XX 格式的單元代號 (例如 P-01 / HU-01, 系統目前只認 T01/T02 系列)\n"
                "2. 該 PDF 處理設施表 / 水質表為純圖片 (需 OCR)\n"
                "3. 上傳的可能不是水措計畫本文 (而是附件/結論報告/流程圖冊)\n\n"
                "**建議**: 確認上傳檔案是水措計畫申請本文, 或用下方「補充: 上傳圖片做局部判讀 (Gemini Vision)」補抓流程圖。"
            )
        elif _n_units < 5 and _pdf_pages >= 30:
            st.warning(
                f"⚠️ **抽取數異常**: {_pdf_pages} 頁 PDF 只抽到 **{_n_units} 個單元**, 可能漏抓\n\n"
                "常見原因:\n"
                "- 該 PDF 用「一頁一單元橫向格式」(需 step2 額外處理)\n"
                "- 該 PDF 單元代號非 T01-XX 格式\n"
                "- 處理設施表為圖片\n\n"
                "**建議**: 先看下方「本文件章節定位」判斷是否有章節「未找到」, 再決定要不要繼續審查。"
            )
        elif _facility_pages == 0 and _quality_pages == 0 and _n_units > 0:
            st.warning(
                "⚠️ **核心章節缺失**: 系統抽到單元, 但「處理設施資料表」與「進出水質資料表」都未定位到\n\n"
                "後續質平/HRT/加藥/放流檢查可能不準, 建議人工核對關鍵單元後再送審。"
            )

        # 技師註解抽到後僅供「AI vs 人工比對」用, 不顯示總覽
        # (註解仍在 app_data['reviewer_notes'], 供 AI/人工 recall 評估或單元詳細頁配對)

        # ───────── 章節動態定位區塊 ─────────
        st.subheader("📍 本文件章節定位")
        st.caption("不同 PDF 頁碼會浮動。系統用『章節標題』找到各區段所在頁,而非寫死頁碼。")

        sec_labels = {
            "flow_diagram": ("廢(污)水流向示意圖", "🌊", "純圖片,需 OCR"),
            "balance_diagram": ("水質水量平衡示意圖", "📊", "純圖片,需 OCR"),
            "quality_data": ("進出水質資料表", "💧", "文字,已抽取"),
            "facility_table": ("處理設施資料表", "🔧", "文字,已抽取"),
            "raw_water": ("原廢水水質", "🧪", "文字"),
            "discharge": ("放流口資料", "🚰", "文字"),
            "emergency": ("緊急應變方法", "⚠", "文字"),
            "sludge": ("污泥處理", "♻", "文字"),
        }
        section_rows = []
        for sec_type, (label, emoji, note) in sec_labels.items():
            pages = sections.get(sec_type, [])
            if pages:
                section_rows.append({
                    "區段": f"{emoji} {label}",
                    "頁碼範圍": compress_ranges(pages),
                    "頁數": len(pages),
                    "狀態": note,
                })
            else:
                section_rows.append({
                    "區段": f"{emoji} {label}",
                    "頁碼範圍": "(未找到)",
                    "頁數": 0,
                    "狀態": "未在文件中偵測到",
                })
        st.dataframe(section_rows, width="stretch", hide_index=True)

        # 統計卡片
        total_design = sum(len(u["design_params"]) for u in app_data["units"].values())
        total_measure = sum(len(u["measure_params"]) for u in app_data["units"].values())
        total_eq = sum(len(u["equipment"]) for u in app_data["units"].values())
        total_in = sum(len(u["influent"]) for u in app_data["units"].values())
        total_out = sum(len(u["effluent"]) for u in app_data["units"].values())

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("處理單元", app_data["total_units"])
        c2.metric("設計參數", total_design)
        c3.metric("量測參數", total_measure)
        c4.metric("機具", total_eq)
        c5.metric("水質流向", f"進{total_in}/出{total_out}")

        # 單元清單 (精簡 + 加削減率)
        st.subheader("📋 處理單元清單")
        st.caption(
            "**進/出** = 進出流水流股數　·　"
            "**削減率** = 質量平衡: Σ(進流質量) → Σ(出流質量), 跨各水質項目平均 "
            "(排除 pH/水溫)。⚠️ = 有項目異常 (出 > 進 或 > 99.5%)"
        )

        # 先算所有單元的削減率
        try:
            import unit_removal_rate as _urr
            removal_results = _urr.compute_all_units_removal(app_data)
        except Exception as _re:
            st.caption(f"削減率計算失敗: {_re}")
            removal_results = {}
            _urr = None

        unit_rows = []
        for code, info in sorted(app_data["units"].items()):
            rr = removal_results.get(code, {})
            removal_text = (_urr.format_removal_short(rr.get("summary", {}))
                            if _urr and removal_results else "-")
            unit_rows.append({
                "代號": code,
                "原始名稱": info["name_in_doc"],
                "標準類型": info["std_tank"],
                "頁數": ", ".join(map(str, info["pages_found"][:3])),
                "📥進": len(info["influent"]),
                "📤出": len(info["effluent"]),
                "📉削減率": removal_text,
            })
        st.dataframe(unit_rows, width="stretch", hide_index=True)

        # 削減率異常清單
        if _urr and removal_results:
            warned_units = [
                (code, rr["summary"]) for code, rr in removal_results.items()
                if rr.get("summary", {}).get("warnings")
            ]
            if warned_units:
                with st.expander(f"⚠️ {len(warned_units)} 個單元有削減率異常項目"):
                    for code, summ in warned_units[:30]:
                        unit_info = app_data["units"].get(code, {})
                        unit_name = unit_info.get("name_in_doc", "")
                        std_name = unit_info.get("std_tank", "")
                        st.markdown(
                            f"**{code} {unit_name}** ({std_name}) — "
                            f"平均削減率 {summ.get('avg_removal_pct', '?')}%"
                        )
                        for w in summ["warnings"][:5]:
                            st.caption(f"　• {w}")
                        if len(summ["warnings"]) > 5:
                            st.caption(f"　… 還有 {len(summ['warnings']) - 5} 項異常")

        # 篩選器 — 看單元詳情 (現在用 session_state 保留選擇)
        st.subheader("🔎 單元詳情")
        unit_codes = sorted(app_data["units"].keys())
        selected = st.selectbox(
            "選擇單元查看詳情",
            unit_codes,
            key="_selected_unit",  # key 讓 Streamlit 自動把選擇存進 session
        )
        if selected and selected in app_data["units"]:
            unit = app_data["units"][selected]

            # 顯示水流上下游 (如果有 flow_graph)
            graph = st.session_state.get("_flow_graph")
            if graph:
                nb = get_unit_neighbors(graph, selected)
                upstream = list(nb.get("upstream", []))
                downstream = list(nb.get("downstream", []))

                # 補上 WMxx (原廢水) → 本單元 / 本單元 → Dxx (放流口) 連結
                # 來源優先順序:
                #   1. PDF「參、第六項 原廢水水量水質資料」(step2_raw_water)
                #      用「水質指紋」比對 → 配對「外部進入」的 WTBxx
                #   2. 示意圖解析 (Gemini Vision) 補方向

                # 來源 1: PDF WM 資料配對 — 取代「(外部進入)」這個含糊標籤
                raw_water_dict = app_data.get("raw_water", {}) or {}
                if raw_water_dict:
                    # 找出 upstream 中 from_unit=="(外部進入)" 的 WTB 編號
                    # 對每個用水質指紋配對 WM
                    new_upstream = []
                    for u in upstream:
                        # 判斷「未配對」: 新版 step4 用「? 來源未配對」, 舊版用「(外部進入)」
                        from_unit = u.get("from_unit", "")
                        if from_unit and not (from_unit.startswith("?") or from_unit == "(外部進入)"):
                            new_upstream.append(u)
                            continue
                        wtb_code = u.get("to_stream", "")
                        # 取這條 WTB 的水質
                        infl_q = (unit.get("influent") or {}).get(wtb_code, {})
                        # 跟每個 WM 比對 (簡單指紋: 3+ 項濃度匹配)
                        matched_wm = None
                        for wm_code, wm_data in raw_water_dict.items():
                            wm_quality = wm_data.get("quality") or {}
                            if not wm_quality or not isinstance(infl_q, dict):
                                continue
                            common = match = 0
                            for item_name, wm_val in wm_quality.items():
                                wm_key = item_name.replace(" ", "")
                                for infl_item, infl_val_dict in infl_q.items():
                                    if not isinstance(infl_val_dict, dict):
                                        continue
                                    if wm_key not in infl_item.replace(" ", ""):
                                        continue
                                    common += 1
                                    try:
                                        v1 = float(str(wm_val).replace("~", "-").split("-")[0].strip())
                                        v2 = float(infl_val_dict.get("濃度", 0) or 0)
                                        if v1 > 0 and abs(v1 - v2) / max(v1, v2) < 0.05:
                                            match += 1
                                    except (ValueError, TypeError):
                                        pass
                                    break
                            if common >= 3 and match >= max(2, common * 0.6):
                                matched_wm = wm_code
                                break

                        if matched_wm:
                            wm_data = raw_water_dict[matched_wm]
                            new_upstream.append({
                                "from_unit": f"🌊 {matched_wm}",
                                "from_stream": matched_wm,
                                "to_stream": wtb_code,
                                "confidence": "高",
                                "method": "PDF 原廢水水質指紋",
                                "_source": "pdf_wm",
                                "_q_cmd": wm_data.get("q_cmd"),
                                "_sources": wm_data.get("sources", []),
                            })
                        else:
                            new_upstream.append(u)  # 保留原本的「? 來源未配對」
                    upstream = new_upstream

                # 來源 2: 示意圖解析 (補拓樸, 不覆蓋 PDF)
                # 原則: PDF (step2) 為主, Vision 只補「拓樸資訊」, 不覆蓋 Q 值
                # 不一致時加警告供使用者交叉比對
                _fr = st.session_state.get("_flow_extract_result")
                if _fr and _fr.get("ok"):
                    # 用 to_stream 判斷是否重複 (而非 from_stream, 因為 PDF 配對失敗時 from_stream 是 "WM?")
                    existing_to_streams = {str(u.get("to_stream", "")) for u in upstream}
                    existing_to_units = {str(d.get("to_unit", "")) for d in downstream}

                    for ei in _fr.get("all_external_inputs", []) or []:
                        if ei.get("to_unit") != selected:
                            continue
                        vision_to_stream = ei.get("to_stream") or ""
                        vision_q = ei.get("Q_cmd")

                        # 找該 to_stream 是否已存在 (PDF 已提供)
                        match_idx = None
                        for i, u in enumerate(upstream):
                            if str(u.get("to_stream", "")) == vision_to_stream:
                                match_idx = i
                                break

                        if match_idx is not None:
                            # PDF 已有此 stream, 用 Vision 補資訊 (例: WM 名稱), 不覆蓋 Q
                            existing = upstream[match_idx]
                            pdf_q = existing.get("_q_cmd")
                            # 補 WM code 名稱 (若 PDF 配對失敗, from_stream="?")
                            vision_code = ei.get("code")
                            if vision_code and existing.get("from_stream") in ("WM?", "?", None, ""):
                                existing["from_stream"] = vision_code
                                existing["from_unit"] = f"🌊 {vision_code} (PDF 配對失敗, Vision 補)"
                                existing["_vision_supplement"] = True
                            # Q 不一致時加警告
                            if pdf_q is not None and vision_q is not None:
                                try:
                                    if pdf_q > 0 and abs(float(vision_q) - float(pdf_q)) / float(pdf_q) > 0.2:
                                        existing["_q_warning"] = (
                                            f"⚠️ PDF Q={pdf_q:g} vs 示意圖 Q={vision_q:g} "
                                            f"(差 {abs(float(vision_q)-float(pdf_q))/float(pdf_q)*100:.0f}%)。以 PDF 為準。"
                                        )
                                except (ValueError, TypeError, ZeroDivisionError):
                                    pass
                        else:
                            # PDF 沒此 stream, 完全新增 (Vision 補拓樸)
                            upstream.append({
                                "from_unit": f"🌊 {ei.get('code', 'WM?')}",
                                "from_stream": ei.get("code", "WM?"),
                                "to_stream": vision_to_stream or "(原廢水)",
                                "confidence": "中",
                                "method": "示意圖解析 (Gemini Vision, 拓樸補充)",
                                "_source": "diagram",
                                "_extra_name": ei.get("name", ""),
                                "_q_cmd": vision_q,
                            })
                    for dp in _fr.get("all_discharge_points", []) or []:
                        if dp.get("from_unit") != selected:
                            continue
                        vision_dp_code = dp.get("code", "D?")
                        # 找下游是否已有此放流口
                        match_idx = None
                        for i, d in enumerate(downstream):
                            if str(d.get("to_unit", "")).endswith(vision_dp_code):
                                match_idx = i
                                break

                        if match_idx is not None:
                            existing_d_item = downstream[match_idx]
                            pdf_q = existing_d_item.get("_q_cmd")
                            vision_q = dp.get("Q_cmd")
                            if pdf_q is not None and vision_q is not None:
                                try:
                                    if pdf_q > 0 and abs(float(vision_q) - float(pdf_q)) / float(pdf_q) > 0.2:
                                        existing_d_item["_q_warning"] = (
                                            f"⚠️ PDF Q={pdf_q:g} vs 示意圖 Q={vision_q:g} "
                                            f"(差 {abs(float(vision_q)-float(pdf_q))/float(pdf_q)*100:.0f}%)。以 PDF 為準。"
                                        )
                                except (ValueError, TypeError, ZeroDivisionError):
                                    pass
                        else:
                            downstream.append({
                                "from_stream": dp.get("from_stream") or "(放流)",
                                "to_unit": f"🏁 {vision_dp_code}",
                                "to_stream": vision_dp_code,
                                "confidence": "中",
                                "method": "示意圖解析 (Gemini Vision, 拓樸補充)",
                                "_source": "diagram",
                                "_extra_name": dp.get("name", ""),
                                "_q_cmd": dp.get("Q_cmd"),
                            })

                if upstream or downstream:
                    st.markdown("##### 🔗 水流串接")
                    cu, cd = st.columns(2)
                    with cu:
                        st.markdown(f"**上游 ({len(upstream)} 條進入)**")
                        if upstream:
                            up_rows = []
                            warnings_up = []
                            for u in upstream:
                                row = {
                                    "來源單元": u["from_unit"],
                                    "出流編號": u["from_stream"],
                                    "→ 進流編號": u["to_stream"],
                                }
                                if u.get("_q_cmd") is not None:
                                    row["Q (CMD)"] = f"{u['_q_cmd']:g}"
                                up_rows.append(row)
                                if u.get("_q_warning"):
                                    warnings_up.append(u["_q_warning"])
                            st.dataframe(up_rows, width="stretch", hide_index=True)
                            if any(u.get("_source") == "diagram" for u in upstream):
                                st.caption("🌊 = 原廢水 (來自示意圖解析)")
                            if any(u.get("_vision_supplement") for u in upstream):
                                st.caption("💡 「PDF 配對失敗, Vision 補」 = step2 找不到配對, 但示意圖補了 WM 編號")
                            for _w in warnings_up:
                                st.warning(_w)
                        else:
                            st.caption("(無偵測到上游, 可能是原廢水進入點或未串接)")
                    with cd:
                        st.markdown(f"**下游 ({len(downstream)} 條流出)**")
                        if downstream:
                            dn_rows = []
                            warnings_dn = []
                            for d in downstream:
                                row = {
                                    "出流編號": d["from_stream"],
                                    "→ 目標單元": d["to_unit"],
                                    "目標進流編號": d["to_stream"],
                                }
                                if d.get("_q_cmd") is not None:
                                    row["Q (CMD)"] = f"{d['_q_cmd']:g}"
                                dn_rows.append(row)
                                if d.get("_q_warning"):
                                    warnings_dn.append(d["_q_warning"])
                            st.dataframe(dn_rows, width="stretch", hide_index=True)
                            if any(d.get("_source") == "diagram" for d in downstream):
                                st.caption("🏁 = 放流口 (來自示意圖解析)")
                            for _w in warnings_dn:
                                st.warning(_w)
                        else:
                            st.caption("(無偵測到下游, 可能是放流口或未串接)")

                    # 提示: 沒跑示意圖解析 + PDF 也沒抽到 WM 資料 → 才顯示提示
                    # (現在主流程會抽 PDF「參、第六項 原廢水」, 通常 raw_water 都會有)
                    _has_wm_or_d = (
                        bool(app_data.get("raw_water"))
                        or bool(app_data.get("discharge"))
                        or (_fr and _fr.get("ok"))
                    )
                    if not _has_wm_or_d:
                        st.info(
                            "💡 想看「原廢水 WMxx → 本單元」「本單元 → 放流口 Dxx」的連結? "
                            "請開啟「📊 水量平衡示意圖解析」(Step 4) 跑 Gemini Vision 解析。"
                        )
                    st.divider()

            # 流向備註 (水量分流, 非異常)
            _topo_notes = unit.get("topology_notes") or []
            if _topo_notes:
                with st.expander(f"ℹ️ 流向備註 ({len(_topo_notes)} 則) — 水量分流/匯流, 非異常", expanded=False):
                    for _n in _topo_notes:
                        st.markdown(f"- {_n}")

            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"**單元代號**: {selected}")
                st.markdown(f"**原始名稱**: {unit['name_in_doc']}")
                st.markdown(f"**標準類型**: {unit['std_tank']}")
                st.markdown(f"**內部代碼**: {unit.get('code_id', '')}")
                st.markdown(f"**出現頁數**: {unit['pages_found']}")

                # 單元尺寸 (材質 / 長/直徑 / 寬 / 高 / 有效水深 / 有效容量 / 數量)
                size_info = unit.get("size") or {}
                if size_info:
                    # 用 markdown 列點顯示, 有效容量 + 有效水深 用粗體 (學理計算最常用)
                    st.markdown("**單元尺寸**:")
                    _size_order = ["材質", "長/直徑", "寬", "高", "有效水深", "有效容量", "數量", "其他"]
                    _units_map = {
                        "長/直徑": "公尺", "寬": "公尺", "高": "公尺",
                        "有效水深": "公尺", "有效容量": "m³", "數量": "座",
                    }
                    for k in _size_order:
                        if k not in size_info:
                            continue
                        v = size_info[k]
                        unit_label = _units_map.get(k, "")
                        if k in ("有效容量", "有效水深"):
                            st.markdown(f"- **{k}**: `{v}` {unit_label}")
                        else:
                            st.markdown(f"- {k}: `{v}` {unit_label}")
            with c2:
                if unit.get("design_params"):
                    st.markdown("**設計操作參數**:")
                    for k, v in unit["design_params"].items():
                        st.markdown(f"- {k}: `{v.get('raw', '')}`")
                if unit.get("measure_params"):
                    st.markdown("**量測操作參數**:")
                    for k, v in unit["measure_params"].items():
                        st.markdown(f"- {k}: `{v.get('raw', '')}`")

            if unit.get("equipment"):
                st.markdown("**相關機具設施**:")
                eq_rows = [{"名稱": str(e["name"]), "位置": str(e.get("位置", "")),
                            "數量": str(e.get("數量", "")), "馬力(kW)": str(e.get("馬力_kW", ""))}
                           for e in unit["equipment"]]
                st.dataframe(eq_rows, width="stretch", hide_index=True)

            # ── 設計參數體檢卡片 (HRT / SOR / G 值) ──
            try:
                import step3h_design_metrics as _dm
                import tank_chemistry as _tc
                _metrics = _dm.compute_all_metrics(unit)
                _rule = _tc.get_rule_for_unit(unit)
            except Exception:
                _metrics = None
                _rule = None

            if _metrics and _rule:
                st.divider()
                st.markdown("##### ⚙️ 設計參數體檢")

                def _verdict(value, vmin, vmax, unit_str):
                    """回 (display_str, delta_str, delta_color) — 給 st.metric 用。"""
                    if value is None:
                        return ("—", "(無法計算)", "off")
                    val_str = f"{value:.3f}" if value < 1 else f"{value:.1f}"
                    if vmin is None and vmax is None:
                        return (f"{val_str} {unit_str}", "(不檢查)", "off")
                    v_min = vmin if vmin is not None else 0
                    v_max = vmax if vmax is not None else float("inf")
                    range_str = (
                        f"{vmin or '-'} ~ {vmax or '∞'}" if vmin or vmax
                        else "(不檢查)"
                    )
                    if v_min <= value <= v_max:
                        return (f"{val_str} {unit_str}", f"✅ 範圍 {range_str}", "normal")
                    elif value < v_min:
                        return (f"{val_str} {unit_str}", f"⚠️ 過低 (應 ≥ {vmin})", "inverse")
                    else:
                        return (f"{val_str} {unit_str}", f"⚠️ 過高 (應 ≤ {vmax})", "inverse")

                mc1, mc2, mc3 = st.columns(3)

                # HRT
                hrt = _metrics["hrt_hr"]
                hrt_disp, hrt_note, hrt_color = _verdict(
                    hrt, _rule.get("HRT_min"), _rule.get("HRT_max"), "hr"
                )
                with mc1:
                    st.metric("⏱ 水力停留時間 (HRT)", hrt_disp, hrt_note, delta_color=hrt_color)
                    if hrt is not None:
                        st.caption(f"= {hrt*60:.1f} 分鐘 (V={_metrics['volume_m3']:.2f} m³ ÷ Q={_metrics['main_q_cmd']:.1f} CMD × 24)")

                # SOR
                sor = _metrics["sor_m3_m2_d"]
                sor_max = _rule.get("SOR_max")
                if sor_max and _metrics["is_lamella"]:
                    sor_max_disp = sor_max * 2.4
                else:
                    sor_max_disp = sor_max
                sor_disp, sor_note, sor_color = _verdict(
                    sor, None, sor_max_disp, "m³/m²·d"
                )
                with mc2:
                    st.metric("🌊 表面溢流率 (SOR)", sor_disp, sor_note, delta_color=sor_color)
                    if sor is not None:
                        lam_tag = " · 斜板" if _metrics["is_lamella"] else ""
                        st.caption(
                            f"= Q={_metrics['main_q_cmd']:.1f} ÷ A={_metrics['surface_area_m2']:.2f} m²"
                            f"{lam_tag}"
                        )

                # G 值
                g = _metrics["g_value_s_inv"]
                g_disp, g_note, g_color = _verdict(
                    g, _rule.get("G_min"), _rule.get("G_max"), "1/s"
                )
                with mc3:
                    st.metric("🔀 G 值 (速度梯度)", g_disp, g_note, delta_color=g_color)
                    if g is not None and _metrics["motor_power_w"]:
                        st.caption(
                            f"= √(P={_metrics['motor_power_w']:.0f}W / "
                            f"(μ×V={_metrics['volume_m3']:.2f}))"
                        )

            # ── 技師審查註解 (PDF 上的便利貼) ──
            # 從整份 app_data 抽出本單元相關頁的註解
            _all_reviewer_notes = app_data.get("reviewer_notes") or []
            if _all_reviewer_notes:
                unit_pages = unit.get("pages_found", [])
                # 找這幾頁 (含 ±1 鄰近頁) 的註解
                _ref_pages = set()
                for p in unit_pages:
                    _ref_pages.update([p - 1, p, p + 1])
                _unit_notes = [n for n in _all_reviewer_notes if n.get("page") in _ref_pages]
                if _unit_notes:
                    st.divider()
                    st.markdown(f"##### 📌 PDF 技師審查註解 ({len(_unit_notes)} 筆)")
                    st.caption(f"這是 PDF 上原本就有的審查技師便利貼註解 (出現在頁 {sorted(set(n.get('page') for n in _unit_notes))})")
                    for n in _unit_notes:
                        author = n.get("author", "user")
                        author_tag = f" by {author}" if author and author != "user" else ""
                        with st.container(border=True):
                            st.markdown(f"📌 **頁 {n.get('page')}**{author_tag}")
                            st.markdown(f"> {n.get('contents', '')}")

            # ── 本單元審查發現摘要 ──
            # 從全廠 findings 篩出這個 selected 單元的, 上方紅黃綠 + 詳列
            _all_findings = st.session_state.get("_check_findings") or []
            unit_findings = [f for f in _all_findings if f.get("單元") == selected]
            st.divider()
            if unit_findings:
                from collections import Counter as _Counter
                _sev = _Counter(f.get("嚴重度", "?") for f in unit_findings)
                _type = _Counter(f.get("類型", "?") for f in unit_findings)

                # 上方三張卡片
                _cf1, _cf2, _cf3 = st.columns(3)
                _cf1.metric("📋 本單元發現", len(unit_findings))
                _cf2.metric("🔴 不合理", _sev.get("不合理", 0))
                _cf3.metric("🟡 待確認", _sev.get("待確認", 0))

                # 詳細列表 (依嚴重度排序, 不合理在前)
                _sev_priority = {"不合理": 0, "待確認": 1, "提醒": 2, "錯誤": 3}
                unit_findings_sorted = sorted(
                    unit_findings,
                    key=lambda f: (_sev_priority.get(f.get("嚴重度"), 9), f.get("類型", "")),
                )
                with st.expander(
                    f"🔍 **本單元審查發現** ({len(unit_findings)} 筆) — "
                    + " · ".join(f"{k} {v}" for k, v in _type.most_common()),
                    expanded=True,
                ):
                    for f in unit_findings_sorted:
                        sev = f.get("嚴重度", "?")
                        emoji = {"不合理": "🔴", "待確認": "🟡", "提醒": "💡", "錯誤": "⚪"}.get(sev, "⚪")
                        with st.container(border=True):
                            st.markdown(
                                f"{emoji} **{f.get('類型', '?')}** · {f.get('對照項目', '?')}"
                            )
                            st.caption(str(f.get("描述", "")))
                            if f.get("依據"):
                                st.caption(f"📌 依據: {f['依據']}")
            else:
                # 沒發現問題, 給綠色提示
                if _all_findings:  # 有跑過審查
                    st.success(f"✅ **本單元無審查發現** — 此單元沒有被任何學理規則 / 規則庫條目觸發")
                # 如果整份還沒跑審查, 不顯示 (避免誤導)

            # ── 反推 Q 對照表 (主要來源) + 示意圖解析 Q (驗算用) ──
            # 主來源: step2 抽取階段已從「質量÷濃度×1000」反推, 覆蓋率 100%
            # 驗算: 示意圖解析 (Gemini Vision) 跑完後可比對, 不一致就標 ⚠️
            stream_q_reverse = unit.get("stream_q", {}) or {}

            stream_q_diagram = {}
            _fr = st.session_state.get("_flow_extract_result")
            if _fr:
                for f in _fr.get("all_flows", []) or []:
                    q = f.get("Q_cmd")
                    if q is None: continue
                    for k in ("from_stream", "to_stream"):
                        c = f.get(k)
                        if c: stream_q_diagram.setdefault(str(c), q)
                for ei in _fr.get("external_inputs", []) or []:
                    q = ei.get("Q_cmd")
                    if q is None: continue
                    for k in ("code", "to_stream"):
                        c = ei.get(k)
                        if c: stream_q_diagram.setdefault(str(c), q)
                for dp in _fr.get("discharge_points", []) or []:
                    q = dp.get("Q_cmd")
                    if q is None: continue
                    for k in ("code", "from_stream"):
                        c = dp.get(k)
                        if c: stream_q_diagram.setdefault(str(c), q)

            def _fmt_q_label(stream_code):
                """編號旁邊的 Q 標籤。

                優先用反推 Q (來自水質表, 100% 覆蓋), 加標「反推」備註。
                若示意圖解析也有, 比對兩者: 一致用 ✅, 不一致 ⚠️。
                """
                rev = stream_q_reverse.get(str(stream_code))
                diag = stream_q_diagram.get(str(stream_code))

                if rev and rev.get("ok"):
                    q = rev["q_cmd"]
                    consistent = rev.get("consistent", True)
                    spread = rev.get("spread_pct", 0)
                    items_n = rev.get("items_count", 0)
                    if not consistent:
                        # 水質表填寫不一致 (19 項算出來差 > 5%)
                        return f" · ⚠️ Q ≈ {q:g} CMD (反推, {items_n} 項差 {spread:.1f}%)"
                    label = f" · Q = {q:g} CMD (反推自水質表)"
                    # 若示意圖也有, 驗算
                    if diag is not None and q > 0:
                        try:
                            diff = abs(float(diag) - q) / q * 100
                            if diff > 10:
                                label += f" ⚠️ 示意圖 = {float(diag):g} (差 {diff:.0f}%)"
                            else:
                                label += " ✅"
                        except (TypeError, ValueError):
                            pass
                    return label

                # 沒反推到 → 退用示意圖解析的 Q
                if diag is not None:
                    try:
                        return f" · Q = {float(diag):g} CMD (示意圖解析)"
                    except (TypeError, ValueError):
                        return f" · Q = {diag} CMD (示意圖解析)"
                return ""

            # 預先算: 每個水質項目的整體削減率 (給出流表用)
            # 跟每個項目「Σ所有進流質量」(給進流表算佔比用)
            _to_float = lambda v: (float(v) if v not in (None, "", "-") else None)

            def _mass_of(stream_dict, item_name):
                v = stream_dict.get(item_name) if isinstance(stream_dict, dict) else None
                if isinstance(v, dict):
                    try:
                        return float(v.get("質量")) if v.get("質量") not in (None, "") else None
                    except (TypeError, ValueError):
                        return None
                return None

            # 蒐集所有水質項目 (兩邊聯集)
            _all_items_in_unit = set()
            for _s in list((unit.get("influent") or {}).values()) + list((unit.get("effluent") or {}).values()):
                if isinstance(_s, dict):
                    _all_items_in_unit.update(_s.keys())

            # Σ進流質量 / Σ出流質量 → 削減率
            item_total_in = {}
            item_total_out = {}
            for item in _all_items_in_unit:
                tin = 0.0
                tout = 0.0
                has_in = False
                has_out = False
                for s in (unit.get("influent") or {}).values():
                    m = _mass_of(s, item)
                    if m is not None:
                        tin += m
                        has_in = True
                for s in (unit.get("effluent") or {}).values():
                    m = _mass_of(s, item)
                    if m is not None:
                        tout += m
                        has_out = True
                if has_in:
                    item_total_in[item] = tin
                if has_out:
                    item_total_out[item] = tout

            def _removal_pct_of(item):
                """該項目的整體削減率 % (Σ進 → Σ出), 沒得算回 None。"""
                tin = item_total_in.get(item)
                tout = item_total_out.get(item)
                if tin is None or tout is None or tin <= 0:
                    return None
                return (tin - tout) / tin * 100

            def _share_pct(stream_item_mass, item):
                """這條流的某項目質量佔該項目總進流質量的 %。"""
                tin = item_total_in.get(item)
                if tin is None or tin <= 0 or stream_item_mass is None:
                    return None
                try:
                    m = float(stream_item_mass) if stream_item_mass not in (None, "") else None
                except (TypeError, ValueError):
                    return None
                if m is None:
                    return None
                return m / tin * 100

            def _fmt_removal(pct):
                """格式化削減率, 標記異常。"""
                if pct is None:
                    return "—"
                if pct < -10:
                    return f"⚠️ {pct:+.1f}%"  # 出流 > 進流
                if pct > 99.5:
                    return f"⚠️ {pct:.1f}%"   # 異常高
                if pct < 0:
                    return f"{pct:+.1f}%"
                return f"{pct:.1f}%"

            if unit.get("influent"):
                _hint = " — Q 自水質表反推 (質量÷濃度×1000)"
                st.markdown(f"**進流水質** ({len(unit['influent'])} 流向){_hint}")
                infl_items = list(unit["influent"].items())
                for i, (infl_code, qdata) in enumerate(infl_items):
                    q_label = _fmt_q_label(infl_code)
                    # 第一條預設展開, 其他收起
                    expanded = (i == 0)
                    with st.expander(f"📥 **{infl_code}**{q_label}", expanded=expanded):
                        # 若是反推不一致, 額外提示
                        rev = stream_q_reverse.get(str(infl_code), {})
                        if rev.get("ok") and not rev.get("consistent", True):
                            st.warning(
                                f"⚠️ 水質表填寫可能有誤: 19 個水質項目分別反推 Q, "
                                f"算出 {rev['q_min']:.2f} ~ {rev['q_max']:.2f} CMD "
                                f"(中位 {rev['q_cmd']:.2f}, 差異 {rev['spread_pct']:.1f}%)"
                            )
                        # 進流欄位: 多一欄「佔進流總質量 %」+ 「整體削減率」
                        # 佔比: 這條流的這項目質量 ÷ 全部進流加總
                        # 削減率: 該項目整體削減率 (Σ進→Σ出)
                        q_rows = []
                        for k, v in qdata.items():
                            if not isinstance(v, dict):
                                row = {"水質項目": str(k), "濃度": "", "質量": "",
                                       "佔進流%": "—", "整體削減率": "—"}
                            else:
                                mass_str = str(v.get("質量", ""))
                                share = _share_pct(v.get("質量"), k)
                                share_txt = f"{share:.1f}%" if share is not None else "—"
                                removal = _removal_pct_of(k)
                                row = {
                                    "水質項目": str(k),
                                    "濃度": str(v.get("濃度", v.get("範圍", ""))),
                                    "質量": mass_str,
                                    "佔進流%": share_txt,
                                    "整體削減率": _fmt_removal(removal),
                                }
                            q_rows.append(row)
                        st.dataframe(q_rows, width="stretch", hide_index=True)

            if unit.get("effluent"):
                st.markdown(f"**出流水質** ({len(unit['effluent'])} 流向)")
                effl_items = list(unit["effluent"].items())
                for i, (effl_code, qdata) in enumerate(effl_items):
                    q_label = _fmt_q_label(effl_code)
                    expanded = (i == 0)
                    with st.expander(f"📤 **{effl_code}**{q_label}", expanded=expanded):
                        rev = stream_q_reverse.get(str(effl_code), {})
                        if rev.get("ok") and not rev.get("consistent", True):
                            st.warning(
                                f"⚠️ 水質表填寫可能有誤: 19 個水質項目分別反推 Q, "
                                f"算出 {rev['q_min']:.2f} ~ {rev['q_max']:.2f} CMD "
                                f"(中位 {rev['q_cmd']:.2f}, 差異 {rev['spread_pct']:.1f}%)"
                            )
                        # 出流欄位: 多一欄「整體削減率 %」
                        q_rows = []
                        for k, v in qdata.items():
                            if not isinstance(v, dict):
                                row = {"水質項目": str(k), "濃度": "", "質量": "",
                                       "整體削減率": "—"}
                            else:
                                removal = _removal_pct_of(k)
                                row = {
                                    "水質項目": str(k),
                                    "濃度": str(v.get("濃度", v.get("範圍", ""))),
                                    "質量": str(v.get("質量", "")),
                                    "整體削減率": _fmt_removal(removal),
                                }
                            q_rows.append(row)
                        st.dataframe(q_rows, width="stretch", hide_index=True)

            # 削減率分項表 (本單元的所有水質項目)
            if _urr:
                unit_rr = removal_results.get(selected, {})
                items = unit_rr.get("items", {})
                summ = unit_rr.get("summary", {})
                if items:
                    st.markdown("**📉 削減率分項 (Σ進流質量 → Σ出流質量)**")
                    rr_rows = []
                    for item, r in sorted(items.items()):
                        pct = r["removal_pct"]
                        flag = ""
                        if pct < -10:
                            flag = "⚠️ 出流>進流"
                        elif pct > 99.5 and not any(k in item for k in ("懸浮", "SS")):
                            flag = "⚠️ 過高"
                        rr_rows.append({
                            "水質項目": item,
                            "進流總質量 (g/d)": r["in_mass"],
                            "出流總質量 (g/d)": r["out_mass"],
                            "削減率 (%)": pct,
                            "備註": flag,
                        })
                    st.dataframe(rr_rows, width="stretch", hide_index=True)
                    cc1, cc2, cc3, cc4 = st.columns(4)
                    cc1.metric("涵蓋項目", summ["items_count"])
                    cc2.metric("平均削減率", f"{summ['avg_removal_pct']}%")
                    cc3.metric("最低", f"{summ['min_removal_pct']}%")
                    cc4.metric("最高", f"{summ['max_removal_pct']}%")
                else:
                    st.caption("(此單元無進+出兩邊都有資料的水質項目, 無法計算削減率)")

        # 下載
        st.divider()
        st.subheader("📥 下載抽取結果")
        base_name = os.path.splitext(pdf_filename)[0] if pdf_filename else "result"

        # ───── 統一匯出: 選對象 → checkbox 自動套對應預設 ─────
        # 強制重 import (避免 Streamlit Cloud 用舊版 cache)
        import importlib
        import export_report as _xrp
        try:
            _xrp = importlib.reload(_xrp)
        except Exception:
            pass
        # 防呆: 若舊版 cache 還沒被擠掉, 提示 reboot
        if not hasattr(_xrp, "get_default_options") or not hasattr(_xrp, "OPTION_LABELS"):
            st.error(
                "⚠️ 偵測到 export_report 模組為舊版本 (Streamlit Cloud 模組 cache)。"
                "請點右下角「Manage app」→「Reboot app」強制重啟, 即可生效。"
            )
            st.stop()

        _findings = st.session_state.get("_check_findings") or []
        _bt = st.session_state.get("_business_type", "") or ""
        if _bt == "(不檢查)": _bt = ""

        with st.expander("📤 匯出審查結果", expanded=False):
            _target_labels = {
                "internal": "🗂 內部覆核 (Excel, 多分頁含細節)",
                "vendor":   "📋 廠商通知 (Word, 改善建議書)",
                "json":     "💾 AI 再分析 (JSON, 整合資料)",
            }
            # 初始化: 第一次套 internal 預設
            if "_export_target" not in st.session_state:
                st.session_state["_export_target"] = "internal"
                for k, v in _xrp.get_default_options("internal").items():
                    st.session_state[f"_export_opt_{k}"] = v

            _prev_target = st.session_state.get("_export_prev_target", st.session_state["_export_target"])
            _target = st.radio(
                "匯出對象",
                options=list(_target_labels.keys()),
                format_func=lambda k: _target_labels[k],
                key="_export_target",
                horizontal=False,
            )
            # 偵測對象變更 → 自動套對應預設 (12 個 checkbox 全部刷新)
            if _target != _prev_target:
                for k, v in _xrp.get_default_options(_target).items():
                    st.session_state[f"_export_opt_{k}"] = v
                st.session_state["_export_prev_target"] = _target
                st.rerun()
            st.session_state["_export_prev_target"] = _target

            # 全選 / 全取消 按鈕
            st.markdown("**📦 包含內容** (一律含: 摘要 + 🔴不合理 + 🟡待確認)")
            bcol1, bcol2, _bcol3 = st.columns([1, 1, 4])
            if bcol1.button("☑ 全選", key="_export_select_all"):
                for k in _xrp.OPTION_LABELS:
                    st.session_state[f"_export_opt_{k}"] = True
                st.rerun()
            if bcol2.button("☐ 全取消", key="_export_select_none"):
                for k in _xrp.OPTION_LABELS:
                    st.session_state[f"_export_opt_{k}"] = False
                st.rerun()

            # 依分組顯示 12 個 checkbox
            from collections import defaultdict
            _grouped = defaultdict(list)
            for opt_key, (group, label) in _xrp.OPTION_LABELS.items():
                _grouped[group].append((opt_key, label))
            for group_name, opts_in_group in _grouped.items():
                st.caption(group_name)
                _cols = st.columns(3)
                for i, (opt_key, label) in enumerate(opts_in_group):
                    _cols[i % 3].checkbox(label, key=f"_export_opt_{opt_key}")

            if not _findings:
                st.caption("💡 還沒跑「開始完整審查」, 匯出檔只會有抽取資料, 沒有 findings。")

            # 收集 checkbox → options
            _opts = {"business_type": _bt}
            for k in _xrp.OPTION_LABELS:
                _opts[k] = bool(st.session_state.get(f"_export_opt_{k}", False))

            try:
                _data, _fname, _mime = _xrp.build_export(
                    _target, app_data, _findings, _opts, base_name=base_name
                )
                st.download_button(
                    f"📥 下載 ({len(_data)/1024:.0f} KB)",
                    data=_data,
                    file_name=_fname,
                    mime=_mime,
                    width="stretch",
                    type="primary",
                )


            except RuntimeError as _e:
                st.error(f"產生失敗: {_e}")
                if "python-docx" in str(_e):
                    st.code("pip install python-docx", language="bash")
            except Exception as _e:
                st.error(f"匯出錯誤: {_e}")

    # ───────── 智能審查結果顯示 (基於 session, 由「開始完整審查」按鈕產生) ─────────
    if st.session_state.get("_check_findings") is not None:
        st.divider()
        st.subheader("智能審查結果")
        st.caption(
            "根據環工技師 299 筆查核缺失歸納的學理規則。"
            " 結果依「審查類型」分組,涵蓋質量平衡/機具設施/設計參數/去除率等多面向。"
        )

        findings = st.session_state["_check_findings"]

        # 統計 (合併「不合理」和「待確認」, 改用「審查類型」分組)
        from collections import Counter
        type_counter = Counter(f.get("類型", "其他") for f in findings)
        sev_counter = Counter(f.get("嚴重度", "?") for f in findings)

        # 上方四張卡片: 總覽 + 三種嚴重度
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("📋 總審查項", len(findings))
        c2.metric("🔴 明顯不合理", sev_counter.get("不合理", 0),
                  help="系統能 100% 自動判定為違反學理 (如快混槽展現重金屬去除)")
        c3.metric("🟡 應人工複核", sev_counter.get("待確認", 0),
                  help="系統找出規則涉及的具體單元數值,需技師人工判讀是否合理")
        c4.metric("📊 涵蓋類型", len(type_counter))

        st.divider()

        # 篩選器
        col1, col2, col3 = st.columns([2, 2, 3])
        with col1:
            filter_type = st.selectbox(
                "依類型篩選",
                ["(全部)"] + sorted(type_counter.keys()),
                key="_filter_type",
            )
        with col2:
            unique_units = sorted({f["單元"] for f in findings})
            filter_unit = st.selectbox(
                "依單元篩選",
                ["(全部)"] + unique_units,
                key="_filter_unit",
            )
        with col3:
            filter_kw = st.text_input("關鍵字搜尋 (描述/規則)", key="_filter_kw")

        filtered = findings
        if filter_type != "(全部)":
            filtered = [f for f in filtered if f.get("類型") == filter_type]
        if filter_unit != "(全部)":
            filtered = [f for f in filtered if f.get("單元") == filter_unit]
        if filter_kw:
            kw = filter_kw.lower()
            filtered = [
                f for f in filtered
                if kw in str(f.get("描述", "")).lower()
                or kw in str(f.get("依據", "")).lower()
            ]

        st.caption(f"顯示 {len(filtered)} / {len(findings)} 筆")

        # 依類型分組顯示
        by_type = defaultdict(list)
        for f in filtered:
            by_type[f.get("類型", "其他")].append(f)

        # 類型顯示順序 (重要的放前面)
        type_priority = [
            "質量平衡", "去除率", "設計參數", "機具設施", "水質標準",
            "操作條件", "流向/示意圖", "文件一致性", "單位/有效位數", "其他"
        ]
        ordered_types = (
            [t for t in type_priority if t in by_type]
            + [t for t in sorted(by_type.keys()) if t not in type_priority]
        )

        for type_name in ordered_types:
            items = by_type[type_name]
            if not items:
                continue
            # 嚴重度色碼 emoji
            sev_in_group = Counter(f["嚴重度"] for f in items)
            sev_summary_parts = []
            if sev_in_group.get("不合理"):
                sev_summary_parts.append(f"🔴 {sev_in_group['不合理']} 不合理")
            if sev_in_group.get("待確認"):
                sev_summary_parts.append(f"🟡 {sev_in_group['待確認']} 待確認")
            sev_summary = " · ".join(sev_summary_parts)

            with st.expander(f"**{type_name}** ({len(items)} 筆) — {sev_summary}", expanded=True):
                # 每個類型用表格顯示
                rows = []
                for f in items:
                    sev_emoji = {"不合理": "🔴", "待確認": "🟡", "提醒": "💡"}.get(f["嚴重度"], "⚪")
                    rows.append({
                        "嚴重": sev_emoji,
                        "單元": str(f["單元"]),
                        "標準槽體": str(f["標準槽體"]),
                        "對照項目": str(f["對照項目"]),
                        "描述": str(f["描述"])[:150],
                        "依據": str(f.get("依據", ""))[:80],
                    })
                st.dataframe(rows, width="stretch", hide_index=True)

        if not findings:
            st.success("本份文件未偵測到明顯不合理之處 (基於目前內建的學理規則)")

        # 把 Counter 轉為 dict 給 JSON 序列化
        findings_json = json.dumps({
            "source": st.session_state.get("_pdf_filename", "?"),
            "total_findings": len(findings),
            "severity_stats": dict(sev_counter),
            "type_stats": dict(type_counter),
            "findings": findings,
        }, ensure_ascii=False, indent=2)
        st.download_button(
            "下載審查結果 JSON",
            data=findings_json.encode("utf-8"),
            file_name="智能審查結果.json",
            mime="application/json",
        )

    # ───────── OCR 流向圖結果顯示 (由「開始完整審查」按鈕產生) ─────────
    if st.session_state.get("_ocr_result"):
        st.divider()
        st.subheader("OCR 解析流向圖 / 水量平衡圖")
        ocr_pages = st.session_state.get("_ocr_target_pages", [])
        if ocr_pages:
            st.caption(f"OCR 解析了 {len(ocr_pages)} 頁 (頁 {compress_ranges(ocr_pages)})")

        ocr_result = st.session_state["_ocr_result"]
        if "error" in ocr_result:
            st.error(f"OCR 失敗: {ocr_result['error']}")
        else:
            summary = ocr_result["summary"]

            with st.expander(
                f"📷 **OCR 識別結果** — "
                f"{summary['total_units']} 單元 / "
                f"{summary['total_flows']} 流量 / "
                f"{summary['total_doses']} 加藥 / "
                f"{summary['total_moistures']} 含水率  (點此展開)",
                expanded=False,
            ):
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("單元", summary["total_units"])
                c2.metric("流量(Q)", summary["total_flows"])
                c3.metric("加藥", summary["total_doses"])
                c4.metric("含水率", summary["total_moistures"])

                # 用座標找最近的單元 — 給每個 OCR 結果關聯一個對應單元
                ocr_units = ocr_result.get("all_units", [])
                def _nearest_unit(item):
                    """找跟 (x, y) 最近的單元代號 (Manhattan 距離)。"""
                    if not ocr_units:
                        return ""
                    ix, iy = item.get("x", 0), item.get("y", 0)
                    best = None
                    best_d = float("inf")
                    for u in ocr_units:
                        dx = abs(u.get("x", 0) - ix)
                        dy = abs(u.get("y", 0) - iy)
                        d = dx + dy
                        if d < best_d:
                            best_d = d
                            best = u.get("code", "")
                    # 距離太遠 (>500 px) 就不認為有對應
                    return best if best_d < 500 else ""

                if ocr_result.get("all_flows"):
                    with st.expander(
                        f"💧 識別到的流量 Q (CMD) — {len(ocr_result['all_flows'])} 筆",
                        expanded=True,
                    ):
                        flow_rows = [
                            {
                                "對應單元": _nearest_unit(f) or "(未對應)",
                                "Q (CMD)": str(f["q"]),
                                "OCR 原文": str(f["text"]),
                            }
                            for f in ocr_result["all_flows"]
                        ]
                        st.dataframe(flow_rows, width="stretch", hide_index=True)

                if ocr_result.get("all_doses"):
                    with st.expander(
                        f"💊 識別到的加藥量 — {len(ocr_result['all_doses'])} 筆",
                        expanded=False,
                    ):
                        dose_rows = [
                            {
                                "對應單元": _nearest_unit(d) or "(未對應)",
                                "化學品": str(d["chemical"]),
                                "用量": str(d["amount"]),
                                "單位": str(d.get("unit", "")),
                                "OCR 原文": str(d["text"]),
                            }
                            for d in ocr_result["all_doses"]
                        ]
                        st.dataframe(dose_rows, width="stretch", hide_index=True)

                if ocr_result.get("all_moistures"):
                    with st.expander(
                        f"💦 識別到的含水率 — {len(ocr_result['all_moistures'])} 筆",
                        expanded=False,
                    ):
                        mois_rows = [
                            {
                                "對應單元": _nearest_unit(m) or "(未對應)",
                                "含水率 (%)": str(m["value_pct"]),
                                "OCR 原文": str(m["text"]),
                            }
                            for m in ocr_result["all_moistures"]
                        ]
                        st.dataframe(mois_rows, width="stretch", hide_index=True)

                if ocr_result.get("all_units"):
                    with st.expander(
                        f"📦 OCR 識別到的單元代號 — {len(ocr_result['all_units'])} 筆",
                        expanded=False,
                    ):
                        unit_rows = [
                            {"代號": str(u["code"]), "OCR 原文": str(u["text"])}
                            for u in ocr_result["all_units"]
                        ]
                        st.dataframe(unit_rows, width="stretch", hide_index=True)

            pdf_name = st.session_state.get("_pdf_filename", "ocr_result")
            base_ocr = os.path.splitext(pdf_name)[0]
            ocr_json = json.dumps(ocr_result, ensure_ascii=False, indent=2)
            st.download_button(
                "下載 OCR 結果 JSON",
                data=ocr_json.encode("utf-8"),
                file_name=f"{base_ocr}_ocr.json",
                mime="application/json",
            )

with tab2:
    st.subheader("規則庫內容")
    st.caption(f"目前載入 {len(rules)} 筆規則，分布於 {len(rules_by_tank)} 個槽體分類。")

    if rules:
        col1, col2 = st.columns([1, 3])
        with col1:
            selected_tank = st.selectbox("選擇槽體", ["(全部)"] + sorted(rules_by_tank.keys()))
        with col2:
            keyword = st.text_input("關鍵字搜尋")

        display_rules = rules
        if selected_tank != "(全部)":
            display_rules = rules_by_tank.get(selected_tank, [])
        if keyword:
            kw = keyword.lower()
            display_rules = [r for r in display_rules if kw in str(r).lower()]

        st.caption(f"顯示 {len(display_rules)} / {len(rules)} 筆")
        st.dataframe(
            [{
                "ID": r.get("缺失ID"),
                "槽體": r.get("標準槽體名稱"),
                "技師": r.get("技師姓名"),
                "檢查類型": r.get("檢查類型"),
                "對照項目": r.get("對照項目"),
                "規則": r.get("規則"),
                "原文": (r.get("原文缺失") or "")[:100],
            } for r in display_rules[:500]],
            width="stretch", hide_index=True,
        )

    # ── 歷次審查紀錄 ──
    st.divider()
    st.subheader("📋 歷次審查紀錄")
    st.caption("每次「開始完整審查」完成後會自動記錄一筆。")

    col_h1, col_h2 = st.columns([1, 4])
    with col_h1:
        refresh_history = st.button("🔄 重新載入", key="_refresh_history_btn")
    if refresh_history or "_review_history_sheet" not in st.session_state:
        try:
            import review_history as _rh
            with st.spinner("從 Sheet 載入歷史…"):
                result = _rh.load_review_history(limit=200)
            st.session_state["_review_history_sheet"] = result
        except Exception as e:
            st.session_state["_review_history_sheet"] = {"ok": False, "error": str(e)}

    sheet_history = st.session_state.get("_review_history_sheet", {})
    if sheet_history.get("ok"):
        rows = sheet_history.get("rows", [])
        if rows:
            with col_h2:
                st.caption(f"共 {len(rows)} 筆 (最新在最上面)")
            try:
                import pandas as _pd
                df_h = _pd.DataFrame(rows)
                # 把所有欄位轉成字串 (避免 PyArrow 錯誤)
                df_h = df_h.astype(str)
                st.dataframe(df_h, width="stretch", hide_index=True)
            except Exception as _e:
                st.error(f"無法顯示表格: {_e}")
                st.json(rows[:10])
        else:
            st.info("尚無歷史紀錄。執行一次「🚀 開始審查」後會自動寫進來。")
    else:
        st.warning("⚠️ 暫時無法載入歷史紀錄,請稍後重試。")

# ─────────────────────────────────────────────
# 規則庫管理 (Google Sheets 同步) — v2: xlsx 為主, csv 自動衍生
# ─────────────────────────────────────────────
with tab_sync:
    st.subheader("🔄 規則庫 ↔ 線上協作表 同步")
    st.markdown(
        "**規則庫.xlsx** 是審查系統的主檔。\n"
        "**線上協作表** 是供審查人員集體編輯的雲端版本 (結構與 xlsx 對應, 每個槽體一個分頁)。\n"
        "**rules_extracted.csv** 為自動衍生檔, 由 xlsx 產出供審查程式讀取。"
    )

    # 載入同步模組
    try:
        import sheets_sync
        sync_ok = True
    except Exception as e:
        st.error(f"無法載入同步模組: {e}")
        sync_ok = False

    if sync_ok:
        # ── 連線狀態 (不洩露帳號細節) ──
        status = sheets_sync.check_auth_status()
        if status["ok"]:
            st.success("✅ 已連線到線上協作表")
        else:
            st.warning("⚠️ 尚未連線到線上協作表 (相關功能無法使用)")

        st.divider()

        # ── 協作表連結 (用按鈕導向, 不直接秀 ID) ──
        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheets_sync.DEFAULT_SHEET_ID}/edit"
        col_link_a, col_link_b = st.columns([3, 2])
        with col_link_a:
            st.markdown(f"**📄 線上協作表**: [打開協作表]({sheet_url})")
            st.caption("審查人員可在此瀏覽 / 編輯規則內容")
        with col_link_b:
            xlsx_exists = os.path.exists(sheets_sync.RULES_XLSX)
            csv_exists = os.path.exists(sheets_sync.RULES_CSV)
            st.markdown("**📁 本機檔案狀態**")
            st.caption(f"規則庫.xlsx: {'✅ 存在' if xlsx_exists else '❌ 不存在'}")
            st.caption(f"rules_extracted.csv: {'✅ 存在' if csv_exists else '❌ 不存在'}")

        st.divider()

        # ── 四個動作按鈕 ──
        col_a, col_b, col_c, col_d = st.columns(4)

        with col_a:
            st.markdown("#### ⬆️ 上傳")
            st.caption("xlsx → Sheet (清空後整批寫入,自動加狀態下拉選單)")
            if st.button("上傳 xlsx → Sheet", type="primary",
                         disabled=not status["ok"] or st.session_state.get("_busy", False),
                         width="stretch", key="upload_btn"):
                st.session_state["_busy"] = True
                with st.spinner("上傳中… (29 個分頁)"):
                    r = sheets_sync.upload_xlsx_to_sheets()
                st.session_state["_busy"] = False
                if r.get("ok"):
                    st.success(
                        f"✅ 上傳成功: {len(r['sheets_written'])} 個分頁 / "
                        f"{r['total_data_rows']} 筆資料"
                    )
                    if r.get("sheets_removed"):
                        st.caption(f"清掉舊分頁: {', '.join(r['sheets_removed'])}")
                    st.caption(f"狀態欄下拉選單: 已設 {r.get('validations_added', 0)} 個分頁")
                    st.caption(f"時間: {r['timestamp']}")
                else:
                    st.error(f"❌ 失敗: {r.get('error', '?')}")

        with col_b:
            st.markdown("#### 🔍 預覽差異")
            st.caption("比對 Sheet vs xlsx 哪些規則有差異 (依槽體+缺失ID)")
            if st.button("預覽差異",
                         disabled=not status["ok"] or st.session_state.get("_busy", False),
                         width="stretch", key="preview_btn"):
                st.session_state["_busy"] = True
                with st.spinner("比對中…"):
                    r = sheets_sync.preview_diff()
                st.session_state["_busy"] = False
                if r.get("ok"):
                    st.success(
                        f"Sheet {r['sheet_total']} 筆 vs xlsx {r['xlsx_total']} 筆 / "
                        f"新增 {r['total_added']} / 刪除 {r['total_removed']} / "
                        f"異動 {r['total_changed']}"
                    )
                    if r.get("by_tank"):
                        with st.expander(f"差異明細 ({len(r['by_tank'])} 個槽體有變化)"):
                            for tank, d in r["by_tank"].items():
                                line = f"**{tank}**: "
                                parts = []
                                if d["added"]:
                                    parts.append(f"➕ {len(d['added'])} 筆 ({', '.join(d['added'][:3])}…)")
                                if d["removed"]:
                                    parts.append(f"➖ {len(d['removed'])} 筆 ({', '.join(d['removed'][:3])}…)")
                                if d["changed"]:
                                    parts.append(f"✏️ {len(d['changed'])} 筆 ({', '.join(d['changed'][:3])}…)")
                                st.markdown(line + " / ".join(parts))
                    else:
                        st.info("Sheet 跟 xlsx 完全一致 ✅")
                else:
                    st.error(f"❌ 失敗: {r.get('error', '?')}")

        with col_c:
            st.markdown("#### ⬇️ 下載")
            st.caption("Sheet → xlsx (自動備份舊版 + 自動產 csv)")
            confirm_dl = st.checkbox("我確定要覆寫 xlsx", key="confirm_download")
            if st.button("下載 Sheet → xlsx", type="secondary",
                         disabled=not (status["ok"] and confirm_dl) or st.session_state.get("_busy", False),
                         width="stretch", key="download_btn"):
                st.session_state["_busy"] = True
                with st.spinner("下載 + 備份中…"):
                    r = sheets_sync.download_sheets_to_xlsx()
                st.session_state["_busy"] = False
                if r.get("ok"):
                    st.success(
                        f"✅ 下載成功: {r['sheets_read']} 個分頁 / "
                        f"{r['total_data_rows']} 筆資料"
                    )
                    if r.get("backup"):
                        st.caption(f"已備份: `{os.path.basename(r['backup'].get('xlsx',''))}`")
                    if r.get("csv_export", {}).get("ok"):
                        st.caption(f"已產出 csv: {r['csv_export']['rows_written']} 筆")

                    # 顯示自動 git push 結果 (修同步架構)
                    gp = r.get("git_push") or {}
                    gp_status = gp.get("status")
                    if gp_status == "pushed":
                        st.success("📤 " + str(gp.get('message', '已推送 GitHub')) + "。Cloud 將於 1~3 分鐘自動 redeploy")
                    elif gp_status == "no_change":
                        st.caption("📋 規則庫.xlsx 跟 GitHub 一致, 無需推送")
                    elif gp_status == "no_git":
                        st.warning(
                            "⚠️ 此環境 (Streamlit Cloud 容器) 無法自動推送 GitHub。"
                            "規則僅在本次 session 生效, reboot 後會失效! "
                            "解法: 在本機跑 streamlit, 點此按鈕即可自動推送; "
                            "或請系統管理員從 GitHub repo 手動更新 規則庫.xlsx"
                        )
                    elif gp_status == "no_credentials":
                        st.warning(
                            "⚠️ Cloud 環境無 GitHub 推送權限。本次規則 reboot 後會失效。 細節: " + str(gp.get('message', ''))
                        )
                    elif gp_status == "push_failed":
                        st.warning("⚠️ 自動推送失敗: " + str(gp.get('message', '')) + "。請用「📤 推送規則庫到 GitHub」按鈕重試")
                else:
                    st.error(f"❌ 失敗: {r.get('error', '?')}")

        with col_d:
            st.markdown("#### 🔄 重產 csv")
            st.caption("從 xlsx 重新產出 csv (不動 Sheet)")
            if st.button("重產 csv", width="stretch", key="export_btn"):
                with st.spinner("產出中…"):
                    r = sheets_sync.export_xlsx_to_csv()
                if r.get("ok"):
                    st.success(f"✅ 完成: {r['rows_written']} 筆規則寫入 csv")
                else:
                    st.error(f"❌ 失敗: {r.get('error', '?')}")

        st.divider()

        # ── 狀態欄統計 ──
        try:
            import step3e_rule_driven_check as _s3e
            rules = _s3e.load_rules_by_tank()
            # 排除 _ 開頭的 key
            tank_rules = {k: v for k, v in rules.items() if not k.startswith("_")}
            total = sum(len(v) for v in tank_rules.values())
            skipped = _s3e.get_last_skipped_count()
            st.markdown("#### 📊 規則狀態統計")
            cols = st.columns(3)
            cols[0].metric("可用規則 (V/空白)", total)
            cols[1].metric("跳過 (狀態=?)", skipped)
            cols[2].metric("規則庫總數", total + skipped)
        except Exception as e:
            st.caption(f"無法載入規則統計: {e}")

        st.divider()

        # ── 備份清單 ──
        st.markdown("#### 📦 本機備份歷史")
        st.caption("每次「下載」前自動備份 xlsx + csv (timestamped)")
        backups = sheets_sync.list_backups()
        if backups:
            import pandas as _pd
            df_bk = _pd.DataFrame(backups)
            df_bk = df_bk[["mtime", "name", "size_kb"]]
            df_bk.columns = ["時間", "檔名", "大小 KB"]
            st.dataframe(df_bk, width="stretch", hide_index=True)
            st.caption(f"備份目錄: `{sheets_sync.BACKUP_DIR}` (本機保留)")
        else:
            st.info("尚無備份。第一次「下載 Sheet → CSV」後會自動產生。")

# ─────────────────────────────────────────────
# 📥 匯入新規則 (半自動,從 NotebookLM 結果貼上)
# ─────────────────────────────────────────────
with tab_import:
    st.subheader("📥 匯入新審查意見規則")
    st.markdown(
        "**半自動流程**: 上傳審查意見 PDF / 圖片 / CSV / xlsx, 或直接貼上文字 → "
        "系統解析 + 預覽 + 衝突檢查 → 確認後自動寫入主檔。"
    )

    try:
        import rule_importer
        importer_ok = True
    except Exception as e:
        st.error(f"無法載入 rule_importer: {e}")
        importer_ok = False

    if importer_ok:
        from datetime import date as _date, datetime as _dt2
        try:
            from zoneinfo import ZoneInfo as _ZI2
            _tz_tpe2 = _ZI2("Asia/Taipei")
        except Exception:
            from datetime import timezone as _tz2, timedelta as _td2
            _tz_tpe2 = _tz2(_td2(hours=8))

        def _today_tpe():
            """今天的日期 (台北時區)。"""
            return _dt2.now(_tz_tpe2).date()

        # 預先讀現有 _來源清單 算下個 S 編號
        try:
            _state = rule_importer._get_existing_state()
            next_s = f"S{_state['max_s_num'] + 1:02d}"
            next_d = f"D{_state['max_d_num'] + 1:03d}"
        except Exception:
            next_s = "S?"
            next_d = "D?"

        st.info(f"📌 系統會自動分配: 來源代號 = `{next_s}` / 第一筆缺失ID = `{next_d}` 起跳")

        # ── Step 1: 提供資料 (兩種方式) ──
        st.markdown("### Step 1 — 提供規則資料")

        input_mode = st.radio(
            "選擇輸入方式",
            [
                "🤖 上傳 PDF (Gemini 自動抽)",
                "📷 上傳圖片 (Gemini Vision 抽規則)",
                "📎 上傳檔案 (CSV / xlsx)",
                "📝 貼上文字",
            ],
            key="imp_input_mode",
            horizontal=True,
        )

        parse_result = None
        auto_filename = ""
        auto_technicians = ""
        is_gemini_result = False  # 標記是否來自 Gemini (給後面複核 UI 用)

        if input_mode.startswith("🤖"):
            # Gemini PDF 模式
            try:
                import gemini_extractor
                gemini_ok = True
            except Exception as e:
                st.error(f"無法載入 gemini_extractor: {e}")
                gemini_ok = False

            if gemini_ok:
                g_status = gemini_extractor.check_gemini_status()
                if g_status["ok"]:
                    st.success("✅ AI 抽取服務已就緒")
                else:
                    st.warning("⚠️ AI 抽取服務尚未啟用 (此功能暫不可用)")

                uploaded_pdf = st.file_uploader(
                    "拖 PDF 到這裡 (審查意見書)",
                    type=["pdf"],
                    key="imp_pdf_uploader",
                    disabled=not g_status["ok"],
                    help="Gemini 會自動讀整份 PDF, 抽出每筆缺失成結構化規則"
                )

                if "gemini_extract_result" not in st.session_state:
                    st.session_state["gemini_extract_result"] = None
                    st.session_state["gemini_pdf_key"] = ""

                if uploaded_pdf:
                    pdf_key = f"{uploaded_pdf.name}_{uploaded_pdf.size}"
                    last_key = st.session_state.get("gemini_pdf_key", "")

                    col_btn_e, col_st = st.columns([1, 3])
                    with col_btn_e:
                        re_extract = st.button(
                            "🔄 重抽" if last_key == pdf_key else "🤖 開始抽取",
                            type="primary",
                            disabled=not g_status["ok"] or st.session_state.get("_busy", False),
                            width="stretch",
                            key="imp_gemini_btn",
                        )
                    with col_st:
                        if last_key == pdf_key and st.session_state["gemini_extract_result"]:
                            st.caption(
                                f"✅ 已抽過: {uploaded_pdf.name} (按重抽會再呼叫 Gemini, 會花費 token)"
                            )
                        else:
                            st.caption(f"準備抽: {uploaded_pdf.name} ({uploaded_pdf.size/1024:.0f} KB)")

                    if re_extract:
                        st.session_state["_busy"] = True
                        with st.spinner("📄 讀 PDF → 🤖 呼叫 Gemini → 解析… (10-30 秒)"):
                            ex_result = gemini_extractor.extract_rules_from_pdf(
                                uploaded_pdf.getvalue(), uploaded_pdf.name
                            )
                        st.session_state["gemini_extract_result"] = ex_result
                        st.session_state["gemini_pdf_key"] = pdf_key
                        st.session_state["_busy"] = False

                    ex_result = st.session_state.get("gemini_extract_result")
                    if ex_result and st.session_state.get("gemini_pdf_key") == pdf_key:
                        if not ex_result.get("ok"):
                            st.error(
                                f"❌ 失敗 ({ex_result.get('stage')}): {ex_result.get('error')}"
                            )
                            if ex_result.get("raw_response"):
                                with st.expander("Gemini raw response (debug)"):
                                    st.code(ex_result["raw_response"])
                        else:
                            cd = ex_result.get("confidence_dist", {})
                            usage = ex_result.get("gemini_usage", {})
                            st.success(
                                f"✅ Gemini 抽出 **{ex_result['row_count']}** 筆 "
                                f"(高信心 {cd.get('high', 0)} / "
                                f"中 {cd.get('medium', 0)} / "
                                f"低 {cd.get('low', 0)})"
                            )
                            st.caption(
                                f"PDF: {ex_result['pdf_pages']} 頁 {ex_result['pdf_chars']} 字 / "
                                f"Gemini tokens: in {usage.get('input_tokens', '?')} "
                                f"out {usage.get('output_tokens', '?')}"
                            )

                            parse_result = {
                                "ok": True,
                                "rows": ex_result["rows"],
                                "format": "gemini-pdf",
                                "row_count": ex_result["row_count"],
                                "filename": uploaded_pdf.name,
                            }
                            auto_filename = uploaded_pdf.name.rsplit(".", 1)[0]
                            is_gemini_result = True
        elif input_mode.startswith("📷"):
            # Gemini Vision 圖片模式: 把審查意見書的截圖 → 結構化規則
            try:
                import gemini_vision
                import gemini_extractor
                vision_ok = True
            except Exception as _ve:
                st.error(f"無法載入 gemini_vision: {_ve}")
                vision_ok = False

            if vision_ok:
                _vstat = gemini_extractor.check_gemini_status()
                if _vstat["ok"]:
                    st.success("✅ AI 影像辨識已就緒")
                else:
                    st.warning("⚠️ AI 影像辨識尚未啟用 (此功能暫不可用)")

                uploaded_imgs = st.file_uploader(
                    "拖入 1~N 張審查意見書截圖 (PNG / JPG / WEBP)",
                    type=["png", "jpg", "jpeg", "webp"],
                    accept_multiple_files=True,
                    key="imp_img_uploader",
                    disabled=not _vstat["ok"],
                    help="可一次上傳多張, Gemini 會綜合判斷"
                )

                if "vision_extract_result" not in st.session_state:
                    st.session_state["vision_extract_result"] = None
                    st.session_state["vision_img_key"] = ""

                if uploaded_imgs:
                    # session key 用「檔名 + size 列表」(換不同圖片組就重抽)
                    img_key = "|".join(f"{i.name}_{i.size}" for i in uploaded_imgs)
                    last_key = st.session_state.get("vision_img_key", "")

                    # 縮圖
                    cols_t = st.columns(min(len(uploaded_imgs), 4))
                    for i, img in enumerate(uploaded_imgs):
                        with cols_t[i % 4]:
                            st.image(img, caption=img.name, width=130)

                    btn_label = "🔄 重抽" if last_key == img_key else f"📷 開始抽取 ({len(uploaded_imgs)} 張)"
                    if st.button(btn_label, type="primary",
                                 disabled=not _vstat["ok"], width="stretch",
                                 key="imp_vision_btn"):
                        with st.spinner(f"Gemini Vision 處理 {len(uploaded_imgs)} 張…"):
                            img_bytes_list = [img.getvalue() for img in uploaded_imgs]
                            _vr = gemini_vision.process_images(
                                img_bytes_list, mode="extract_rules"
                            )
                        st.session_state["vision_extract_result"] = _vr
                        st.session_state["vision_img_key"] = img_key

                    ex_result = st.session_state.get("vision_extract_result")
                    if ex_result and st.session_state.get("vision_img_key") == img_key:
                        if not ex_result.get("ok"):
                            st.error(
                                f"❌ 失敗 ({ex_result.get('stage')}): {ex_result.get('error')}"
                            )
                            if ex_result.get("raw_response"):
                                with st.expander("Gemini raw response (debug)"):
                                    st.code(ex_result["raw_response"])
                        else:
                            cd = ex_result.get("confidence_dist", {})
                            usage = ex_result.get("gemini_usage", {})
                            st.success(
                                f"✅ Vision 抽出 **{ex_result['row_count']}** 筆 "
                                f"(高 {cd.get('high', 0)} / 中 {cd.get('medium', 0)} / 低 {cd.get('low', 0)})"
                            )
                            st.caption(
                                f"圖片: {ex_result['image_count']} 張 / "
                                f"tokens in {usage.get('input_tokens', '?')} "
                                f"out {usage.get('output_tokens', '?')}"
                            )

                            parse_result = {
                                "ok": True,
                                "rows": ex_result["rows"],
                                "format": "gemini-vision",
                                "row_count": ex_result["row_count"],
                            }
                            # 用第一張圖的名稱當預設檔名
                            first_img_name = uploaded_imgs[0].name
                            auto_filename = f"圖片_{first_img_name.rsplit('.', 1)[0]}"
                            is_gemini_result = True
        elif input_mode.startswith("📎"):
            uploaded = st.file_uploader(
                "拖檔案到這裡, 或點選檔案",
                type=["csv", "xlsx", "xls"],
                key="imp_uploader",
                help="CSV: 單一扁平表 / xlsx: 可單一分頁也可多分頁 (每分頁名稱 = 槽體名稱)"
            )
            if uploaded:
                with st.spinner(f"解析 {uploaded.name}…"):
                    parse_result = rule_importer.parse_uploaded_file(uploaded)
                auto_filename = uploaded.name.rsplit(".", 1)[0]

                if parse_result.get("ok") and parse_result.get("sheets"):
                    sheets_info = parse_result["sheets"]
                    if len(sheets_info) > 1:
                        st.caption(
                            f"📑 偵測到 {len(sheets_info)} 個分頁: "
                            + ", ".join(f"{s['name']}({s['rows']})" for s in sheets_info[:8])
                            + ("..." if len(sheets_info) > 8 else "")
                        )
        else:
            st.caption(
                "**支援**: TSV (Excel 複製貼上預設) / CSV / Markdown 表格。"
                " 必要欄位: `原文缺失` / `檢查類型` / `對照項目` / `規則` / `標準槽體名稱`。"
            )

            with st.expander("📋 範本格式 (點開複製)"):
                st.code("""缺失ID\t原文缺失\t檢查類型\t對照項目\t規則\t比對位置\t判定邏輯\t技師姓名\t序號\t標準槽體名稱\t原始槽體代號
\t出水池排放口缺裝液位計\t機具設施\t液位計\t放流池排放口應設液位計監測水位\t廢污水處理設施操作條件\t若 放流池 且 無液位計 → 標記:缺機具\t範例技師\t序1 範例技師 (1)\t放流池\tT01-15
""", language="text")

            pasted_text = st.text_area(
                "貼上你的規則表 (TSV / CSV / Markdown)",
                key="imp_pasted",
                height=200,
                placeholder="從 NotebookLM 或 Excel/Sheet 複製整個表格 (含表頭) 直接貼這裡",
            )
            if pasted_text.strip():
                parse_result = rule_importer.parse_input_text(pasted_text)

        # ── 後續步驟 (只有 parse_result OK 才往下) ──
        if parse_result:
            if not parse_result.get("ok"):
                st.error(f"❌ 解析失敗: {parse_result.get('error')}")
            else:
                rows = parse_result["rows"]

                # 自動推導: 從上傳資料抽出技師姓名 (去重)
                techs_in_data = sorted({
                    (r.get("技師姓名") or "").strip()
                    for r in rows
                    if (r.get("技師姓名") or "").strip()
                })
                if techs_in_data:
                    auto_technicians = " / ".join(techs_in_data[:5])
                    if len(techs_in_data) > 5:
                        auto_technicians += f" 等 {len(techs_in_data)} 位"

                st.success(
                    f"✅ 解析成功: {parse_result['row_count']} 筆 "
                    f"({parse_result['format'].upper()})"
                )

                # ── Gemini 結果 → 顯示複核 UI (data_editor) ──
                if is_gemini_result:
                    st.markdown("### Step 1.5 — 人工複核 (預設全勾選)")
                    st.caption(
                        "Gemini 抽出的結果可能有誤, 請逐筆確認:\n"
                        "- ☑️ **選取** 欄: 取消勾選不要匯入的\n"
                        "- 其他欄位都可直接點擊修改\n"
                        "- **信心度** = Gemini 自評, low 建議仔細看"
                    )

                    import pandas as _pd
                    # 把 rows 轉成 DataFrame, 加「選取」欄預設 True, 加「信心度」欄
                    df_review = _pd.DataFrame(rows).fillna("")
                    df_review.insert(0, "選取", True)
                    if "_confidence" in df_review.columns:
                        df_review.rename(columns={"_confidence": "信心度"}, inplace=True)
                    else:
                        df_review["信心度"] = "unknown"

                    # 欄位順序
                    ordered_cols = [
                        "選取", "信心度", "標準槽體名稱", "原始槽體代號",
                        "原文缺失", "檢查類型", "對照項目", "規則",
                        "比對位置", "判定邏輯", "技師姓名", "序號", "缺失ID", "狀態",
                    ]
                    ordered_cols = [c for c in ordered_cols if c in df_review.columns]
                    df_review = df_review[ordered_cols]

                    # 排序: 低信心度排前面 (讓使用者先注意)
                    conf_order = {"low": 0, "medium": 1, "high": 2, "unknown": 3}
                    df_review["_sort"] = df_review["信心度"].map(lambda x: conf_order.get(x, 4))
                    df_review = df_review.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

                    # 槽體下拉選單選項
                    try:
                        import gemini_extractor as _ge
                        tank_options = _ge.STANDARD_TANKS
                        check_options = _ge.CHECK_TYPES
                    except Exception:
                        tank_options = []
                        check_options = []

                    # data_editor 設定
                    column_config = {
                        "選取": st.column_config.CheckboxColumn(
                            "✓", help="勾選 = 匯入", default=True, width="small"
                        ),
                        "信心度": st.column_config.SelectboxColumn(
                            "信心",
                            options=["high", "medium", "low", "unknown"],
                            width="small",
                        ),
                        "標準槽體名稱": st.column_config.SelectboxColumn(
                            "標準槽體",
                            options=tank_options,
                            required=False,
                        ) if tank_options else st.column_config.TextColumn("標準槽體"),
                        "檢查類型": st.column_config.SelectboxColumn(
                            "檢查類型",
                            options=check_options,
                        ) if check_options else st.column_config.TextColumn("檢查類型"),
                        "原文缺失": st.column_config.TextColumn(
                            "原文缺失", width="large"
                        ),
                        "規則": st.column_config.TextColumn(
                            "規則", width="medium"
                        ),
                        "原始槽體代號": st.column_config.TextColumn("原代號", width="small"),
                        "對照項目": st.column_config.TextColumn("對照項目", width="small"),
                        "比對位置": st.column_config.TextColumn("比對位置"),
                        "判定邏輯": st.column_config.TextColumn("判定邏輯"),
                        "技師姓名": st.column_config.TextColumn("技師", width="small"),
                        "序號": st.column_config.TextColumn("序號", width="small"),
                        "缺失ID": st.column_config.TextColumn("缺失ID", width="small"),
                        "狀態": st.column_config.SelectboxColumn(
                            "狀態", options=["", "V", "?"], width="small",
                        ),
                    }

                    edited_df = st.data_editor(
                        df_review,
                        column_config=column_config,
                        hide_index=True,
                        width="stretch",
                        num_rows="fixed",  # 不讓使用者亂加列
                        key="imp_gemini_editor",
                    )

                    # 統計勾選
                    selected_count = int(edited_df["選取"].sum())
                    total_count = len(edited_df)
                    col_sel1, col_sel2, col_sel3 = st.columns(3)
                    col_sel1.metric("勾選筆數", selected_count)
                    col_sel2.metric("總筆數", total_count)
                    col_sel3.metric(
                        "取消勾選", total_count - selected_count,
                        delta=f"-{total_count - selected_count}" if selected_count < total_count else None,
                    )

                    # 把編輯+勾選後的結果回灌成 rows (給後續流程用)
                    selected_rows = edited_df[edited_df["選取"] == True].drop(columns=["選取"])
                    if "信心度" in selected_rows.columns:
                        selected_rows = selected_rows.drop(columns=["信心度"])
                    rows = selected_rows.to_dict(orient="records")

                    if not rows:
                        st.warning("⚠️ 沒有勾選任何規則, 請至少勾一筆")

                else:
                    # 非 Gemini 路徑: 顯示傳統前 5 筆預覽
                    with st.expander(f"📋 前 5 筆預覽 (共 {len(rows)} 筆)"):
                        import pandas as _pd
                        df_preview = _pd.DataFrame(rows).fillna("")
                        st.dataframe(df_preview.head(5), width="stretch", hide_index=True)

                # ── Step 2: 預覽匯入結果 ──
                if rows:
                    st.markdown("### Step 2 — 預覽匯入結果")
                    preview = rule_importer.preview_import(rows)

                    col_p1, col_p2, col_p3, col_p4 = st.columns(4)
                    col_p1.metric("總筆數", preview["total"])
                    col_p2.metric("可匯入", preview["ok_to_import"])
                    col_p3.metric("涵蓋槽體", len(preview["tanks_in_import"]))
                    col_p4.metric("新槽體", len(preview["new_tanks"]))

                    if preview["new_tanks"]:
                        st.warning(
                            f"⚠️ 將新建 {len(preview['new_tanks'])} 個槽體分頁: "
                            f"{', '.join(preview['new_tanks'])}\n\n"
                            f"請確認名稱跟 RULE_AUTHORING.md 第 1 章的標準槽體一致。"
                        )

                    if preview["id_conflicts"]:
                        st.info(
                            f"ℹ️ {len(preview['id_conflicts'])} 筆缺失ID 跟現有重複, "
                            f"系統會自動改成 `{next_d}` 起的新編號: "
                            f"{', '.join(preview['id_conflicts'][:5])}"
                            f"{'...' if len(preview['id_conflicts']) > 5 else ''}"
                        )

                    if preview["missing_required"]:
                        with st.expander(f"⚠️ {len(preview['missing_required'])} 筆缺必填欄 (會被跳過)"):
                            for m in preview["missing_required"][:20]:
                                st.write(f"- 第 {m['row_idx']} 列: 缺 {', '.join(m['missing'])} (預覽: {m['row_preview']})")
                            if len(preview["missing_required"]) > 20:
                                st.caption(f"… 還有 {len(preview['missing_required']) - 20} 筆")

                    if preview["tanks_in_import"]:
                        with st.expander(f"📂 涵蓋槽體清單 ({len(preview['tanks_in_import'])} 個)"):
                            st.write(", ".join(preview["tanks_in_import"]))

                    # ── Step 3: 基本資料 (自動帶入, 可選修改) ──
                    with st.expander("📝 基本資料 (系統已自動帶入, 可選擇修改)", expanded=False):
                        col_m1, col_m2 = st.columns(2)
                        today_str = _today_tpe().strftime("%Y-%m-%d")
                        with col_m1:
                            src_filename = st.text_input(
                                "審查意見檔名 (自動)", value=auto_filename,
                                key="imp_filename",
                                help="預設使用上傳檔案的檔名 (扣副檔名)",
                            )
                            src_technician = st.text_input(
                                "技師姓名 (自動)", value=auto_technicians,
                                key="imp_technician",
                                help="從上傳資料的「技師姓名」欄自動統計",
                            )
                            src_cert = st.text_input("技師證書字號 (可選)", value="",
                                                      key="imp_cert", placeholder="留空或 (見原文)")
                        with col_m2:
                            src_date = st.text_input("查核日期 (預設今天)", value=today_str,
                                                      key="imp_date")
                            src_company = st.text_input("簽證事業名稱 (可選)", value="",
                                                         key="imp_company", placeholder="例: (多家) 或 公司名")
                            src_note = st.text_input("備註 (可選)", value="",
                                                      key="imp_note", placeholder="留空使用預設")

                    # 沒展開的話 src_xxx 變數不會被設, 在這邊補預設
                    if "imp_filename" not in st.session_state:
                        src_filename = auto_filename or "未命名審查意見"
                        src_technician = auto_technicians
                        src_cert = ""
                        src_date = _today_tpe().strftime("%Y-%m-%d")
                        src_company = ""
                        src_note = ""
                    else:
                        src_filename = st.session_state.get("imp_filename", auto_filename) or auto_filename or "未命名審查意見"
                        src_technician = st.session_state.get("imp_technician", auto_technicians)
                        src_cert = st.session_state.get("imp_cert", "")
                        src_date = st.session_state.get("imp_date", _today_tpe().strftime("%Y-%m-%d"))
                        src_company = st.session_state.get("imp_company", "")
                        src_note = st.session_state.get("imp_note", "")

                    # ── Step 4: 確認匯入 ──
                    st.markdown("### Step 3 — 確認匯入")
                    st.info(
                        f"即將匯入 **{preview['ok_to_import']}** 筆 → "
                        f"來源代號 **{preview['next_source_code']}** ({src_filename}) / "
                        f"從 **{next_d}** 開始分配 ID"
                    )

                    confirm_import = st.checkbox(
                        "我確定要寫入 規則庫.xlsx (會自動備份舊版)",
                        key="imp_confirm",
                    )

                    if st.button("📥 執行匯入", type="primary",
                                 disabled=not confirm_import or st.session_state.get("_busy", False),
                                 width="stretch"):
                        st.session_state["_busy"] = True
                        metadata = {
                            "檔名": src_filename,
                            "技師姓名": src_technician,
                            "技師證書字號": src_cert,
                            "查核日期": src_date,
                            "簽證事業名稱": src_company,
                            "備註": src_note or "Streamlit 半自動匯入",
                        }
                        with st.spinner("寫入中…"):
                            result = rule_importer.commit_import(rows, metadata, skip_missing=True)
                        st.session_state["_busy"] = False

                        if result.get("ok"):
                            st.success(
                                f"✅ 匯入成功! 來源 **{result['source_code']}** / "
                                f"匯入 **{result['imported_count']}** 筆 / "
                                f"跳過 **{result['skipped_count']}** 筆"
                            )
                            new_tanks = result.get("new_tanks_created") or []
                            if new_tanks:
                                st.info(f"🆕 新建分頁: {', '.join(new_tanks)}")
                                # ★ 順手提醒: 新槽體還沒有學理規則, 建議補
                                try:
                                    import tank_chemistry
                                    chem_rules = tank_chemistry.load_rules()
                                    missing_chem = [t for t in new_tanks if t not in chem_rules
                                                    and not t.startswith("(")]
                                    if missing_chem:
                                        st.warning(
                                            f"⚠️ **新槽體缺學理規則** — "
                                            f"`{', '.join(missing_chem)}` 還沒有對應的「應變動/不應變動」設定。\n\n"
                                            f"建議到雲端協作表 **`_槽體學理`** 分頁補上, "
                                            f"否則送審申請文件時, 系統無法判斷這些槽體的水質是否合理變動。"
                                        )
                                except Exception:
                                    pass
                            if result.get("backup"):
                                st.caption(f"備份: `{os.path.basename(result['backup'])}`")
                            st.info(
                                "💡 後續流程:\n"
                                "1. 切到「🔄 規則庫管理」分頁\n"
                                "2. 按「⬆️ 上傳 xlsx → 協作表」讓新規則同步到雲端\n"
                                "3. 若有新槽體, 請到 `_槽體學理` 分頁補上學理規則\n"
                                "4. 通知系統管理員將新規則部署到線上版"
                            )
                        else:
                            st.error(f"❌ 失敗: {result.get('error', '?')}")

with tab3:
    st.subheader("📖 水措審查系統 — 使用說明")

    st.markdown("""
## 🎯 系統用途

協助環工技師審查「水污染防治措施」申請文件 (PDF), 自動找出**不符合學理**或**需人工複核**的項目, 並比對審查意見規則庫。

---

## 🚀 快速開始 (一鍵完整審查)

### Step 1 · 上傳申請文件
在「**🚀 開始審查**」分頁, 把申請文件 PDF 拖到上傳區。

### Step 2 · 選擇選項
- **✂️ 只處理「參、水污染防治措施資料」章節** (預設打勾, 加速 5–10 倍)
- **事業類別**: 選了之後系統會檢查該類別應申報的項目是否漏項; 不選就跳過此檢查。

### Step 3 · 按「🚀 開始完整審查」
系統會自動執行三大步驟, 過程中**整個畫面會被遮罩**, 中央顯示:
- 目前進行到哪個 step
- 完成百分比 + 進度條
- 已耗時 / 預估剩餘時間

如需提前結束, 按右下角的「🛑 停止審查」(會中斷剩下的步驟, 已完成的會保留)。

### Step 4 · 看結果
完成後會出現「**智能審查結果**」區域, 依**類型分組** (質量平衡 / 去除率 / 設計參數 / 機具設施 / 水質標準 …), 每筆都有:
- 🔴 / 🟡 嚴重度
- 涉及單元
- 數值佐證 (例: 表面溢流率 67 m³/m²·d)
- 學理依據

可用上方的「依類型 / 依單元 / 關鍵字」三個篩選器縮小範圍。

---

## 🔍 額外功能

### 📊 水量平衡示意圖解析 (跨單元流向)
PDF 抽取完成後, 在主畫面往下捲動會看到「**📊 水量平衡示意圖解析**」摺疊區, 展開後按「🤖 開始解析」。系統會用 AI 視覺辨識讀取流向示意圖, 得到:
- 每條箭頭的 `from_unit` / `to_unit` / `Q (CMD)` / 對應的 WTA/WTB 編號
- 外部進入 (原廢水 WMxx) 與放流口 (Dxx)
- **跨單元質量平衡檢核** (Σ進 ≈ Σ出)
- **編號一致性檢核** (同一條水兩端編號流量是否一致)

解析完之後, 單元詳細頁的「進/出流水質」編號旁邊會直接帶出該條流的 `Q = XXX CMD`。

### 🖼 OCR 流向圖 (備用)
如果 AI 視覺辨識不可用, 系統會用本地 OCR 把流向圖上的:
- 處理單元代號 (T01-01, T02-03 …)
- 流量數值 (Q = 47.5 CMD)
- 加藥量、含水率

讀取出來, 並依座標自動配對到最近的處理單元。

### 🔗 水流串接圖
每個單元的詳細頁有「**🔗 水流串接**」, 顯示這個單元的上游 (誰流進來) 與下游 (流到哪去), 系統會用水質指紋自動配對 WTA → WTB。

### 📉 削減率分項
每個單元若進+出流兩邊都有水質質量資料, 系統會自動算:
- 每個水質項目的「進流總質量 → 出流總質量 → 削減率 %」
- 異常標記 (削減率 < -10% 或 > 99.5% 視單元類型)

### 📥 下載結果
- **Excel** (各單元設計參數 / 量測參數 / 機具 / 進出流水質 一頁一單元)
- **JSON** (完整結構化資料, 可供其他程式使用)

---

## 📊 規則庫瀏覽

切到「**📊 規則庫瀏覽**」分頁:
- 看目前規則庫總筆數、依槽體分組、依嚴重度分布
- 「📋 歷次審查紀錄」會列出所有跑過的審查 (檔名 / 單元數 / 不合理 / 待確認 / 耗時)

## 🔄 規則庫管理

切到「**🔄 規則庫管理**」分頁:
- **⬆️ 上傳 xlsx → 協作表**: 把本機規則庫推到雲端線上協作表
- **⬇️ 下載協作表 → xlsx**: 把雲端最新規則拉回本機 (會自動備份舊版)
- **🔄 重產 csv**: 從 xlsx 重新產出 csv (供審查引擎讀取)

## 📥 匯入新規則 (半自動)

當有新的「審查意見書」想加入規則庫, 切到「**📥 匯入新規則**」分頁, 提供 4 種輸入方式:

| 方式 | 適用 |
|------|------|
| 🤖 上傳 PDF | 整份審查意見書 PDF, AI 自動結構化抽取 |
| 📷 上傳圖片 | 截圖/掃描的審查意見書, AI 視覺辨識 |
| 📎 上傳檔案 | 已整理好的 CSV / xlsx |
| 📝 貼上文字 | 從其他工具複製貼上 |

系統會解析 → 預覽 → 衝突檢查 → 確認後寫入主檔。匯入完成後請切到「🔄 規則庫管理」上傳到協作表。

---

## ⚙️ 智能審查涵蓋規則

### 質量平衡
- 溶解性物質 (硝酸鹽 / 硼 / Cl⁻) 在無濃縮機制單元不應自行濃縮
- 跨單元: Σ 所有進流 Q ≈ Σ 所有出流 Q

### 去除位置學理
- 快混槽 / pH 調整槽**無固液分離**, 不應展現重金屬去除
- pH 調整槽除 pH 外, 其他水質應不變

### 設計參數
- 沉澱池表面溢流率應 < 50 m³/m²·d
- 各槽體的停留時間 / 有效容積 / 加藥量範圍

### 機具設施
- 各槽體應有的液位計、pH 計、排泥裝置、攪拌機等

### 規則庫驅動
- 依環工技師過往查核缺失歸納的學理規則

---

## ❓ 常見問題

**Q: 為什麼按了「開始完整審查」之後其他按鈕都不能點?**
A: 為了避免中途中斷處理, 整個畫面會被遮罩。若真的要中止, 用右下角「🛑 停止審查」按鈕。

**Q: 流向示意圖辨識失敗怎麼辦?**
A: 系統會自動 fallback 到本地 OCR; 可在主畫面下方「💧 識別到的流量 Q」摺疊區看 OCR 直接讀到的資料。

**Q: 進流水質的 Q 值哪裡來?**
A: 來自「📊 水量平衡示意圖解析」的結果 — 圖上實際標的 `Q = XXX CMD`, 不是計算的。沒跑過示意圖解析的話, 編號旁邊不會顯示 Q。

**Q: 大型 PDF (>100MB) 跑很慢?**
A: 建議勾選「✂️ 只處理「參、水污染防治措施資料」章節」, 可省下 5–10 倍時間。
""")

st.divider()
st.caption("水措審查系統 · 章節定位 + OCR + AI 視覺辨識 + 智能審查")
