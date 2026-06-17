# -*- coding: utf-8 -*-
"""加藥機制學理檢查 — 鎳系/輕系/各系分流系統。

學理:
    不同廢水分流系統對應不同加藥組合 (環工設計準則):

    | 分流 | 應加 | 不該加 | 學理目的 |
    |---|---|---|---|
    | 鎳系 | NaOH (大量, pH 10.5~11) | PAC/Polymer (沒沉澱) | Ni(OH)₂ 沉澱 |
    | 輕系 | NaOH (少量) | PAC/Polymer (沒必要) | 中和 |
    | 各系 | PAC + Polymer | 大量 NaOH (除非中和) | 混凝絮凝 |

判定:
    1. 從單元名稱 (name_in_doc) 偵測屬於哪個分流系統
        - 含「鎳系」「鎳水洗」「化銅」「化學鎳」「電鍍」 → 鎳系
        - 含「輕系」「輕細」「輕研磨」「酸鹼」 → 輕系
        - 含「各系」「綜合」「混合」「總匯」 → 各系
    2. 從 measure_params 看實際加什麼藥
    3. 比對學理應該加什麼, 不一致就標 ⚠️

同義字標準化:
    「聶系」→「鎳系」 (常見筆誤)
"""

# 分流系統 → (該加藥劑關鍵字, 不該加藥劑關鍵字)
DOSING_RULES = {
    "鎳系": {
        "should": ["NaOH", "氫氧化鈉", "燒鹼", "液鹼", "Ca(OH)", "氫氧化鈣", "石灰"],
        "should_not": ["PAC", "Polymer", "聚氯化鋁"],
        "ph_min": 10.5,  # pH 至少要拉到 10.5 才能讓 Ni(OH)₂ 完全沉澱
        "說明": "鎳系廢水含高濃度 Ni²⁺, 學理上應加大量 NaOH 將 pH 拉到 10.5~11, "
                "讓 Ni²⁺ → Ni(OH)₂ 沉澱 (綠色污泥)。若加 PAC/Polymer 沒鹼, 鎳無法去除",
    },
    "輕系": {
        "should": ["NaOH", "氫氧化鈉", "燒鹼", "液鹼", "硫酸", "H2SO4", "HCl"],
        "should_not": ["PAC", "Polymer", "聚氯化鋁"],
        "ph_min": None,  # 中和到 6~8 即可
        "說明": "輕系廢水污染輕、水量大, 學理上只需酸鹼中和 (pH 6~8)。"
                "加 PAC/Polymer 是浪費 (污染物濃度不夠形成絮羽)",
    },
    "各系": {
        "should": ["PAC", "Polymer", "聚氯化鋁", "硫酸鋁", "FeCl3", "高分子", "助凝劑"],
        "should_not": [],  # NaOH 可同時加 (中和+混凝)
        "ph_min": None,
        "說明": "各系是綜合廢水, 主要污染是膠體 + 細小 SS。學理上應加 PAC 混凝 + Polymer "
                "助凝形成大絮羽, 讓後端沉澱池能去除。只加 NaOH 沒混凝就沉不下來",
    },
}

def load_dosing_rules_from_xlsx(xlsx_path=None):
    """從 規則庫.xlsx 的 _加藥規則 分頁讀規則 (14 欄 schema), fallback 到 DOSING_RULES.

    新 schema (14 欄):
        A 編號 / B 分流系統 / C 槽體類別 / D 藥劑 / E 加藥角色 (應加/禁加/可加)
        F~G 配比 min/max (kg/ton) / H~I 商品濃度 min/max (%)
        J~K pH min/max / L 學理對應水質 / M 污泥代碼 / N 別名/說明

    返回:
        dict[分流系統] = {
            'should': [(藥劑, 配比min, 配比max, 商品濃度min, 商品濃度max, ph_min, ph_max), ...],
            'should_not': [(藥劑,), ...],
            'aliases': [別名 list],
            'ph_min': float (最嚴格的 pH_min),
            'ph_max': float,
            'sludge_code': str,
            '說明': str,
        }
    """
    import os
    if xlsx_path is None:
        xlsx_path = os.path.join(os.path.dirname(__file__), '規則庫.xlsx')
    if not os.path.exists(xlsx_path):
        return DOSING_RULES
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if '_加藥規則' not in wb.sheetnames:
            wb.close()
            return DOSING_RULES
        ws = wb['_加藥規則']
        header = [c.value for c in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception:
        return DOSING_RULES

    def _split(s):
        if not s: return []
        return [x.strip() for x in str(s).replace('，', ',').split(',') if x.strip()]

    def _to_float(v):
        try: return float(v) if v not in (None, '') else None
        except: return None

    # 解析 14 欄 schema
    rules = {}
    for r in rows:
        if not r or not r[0]: continue
        r = list(r) + [None] * (17 - len(r))
        rid, sys_name, tank_type, drug, role, ratio_min, ratio_max, conc_min, conc_max, ph_min, ph_max, water_quality, sludge_code, alias_desc, stage, orp_min, orp_max = r[:17]
        if not sys_name or not drug:
            continue
        sys_name = str(sys_name).strip()
        drug = str(drug).strip()
        role = (role or '').strip()

        # 別名: 從說明欄解析「別名: xxx」 (因為合併到「別名/說明」欄)
        aliases = []
        desc = str(alias_desc or '').strip()
        if desc.startswith('別名:') or desc.startswith('別名：'):
            # 取第一句「別名: X,Y,Z. 後續說明...」
            m_text = desc[3:].lstrip(':：').strip()
            first_dot = m_text.find('.')
            if first_dot > 0:
                aliases = _split(m_text[:first_dot])
            else:
                aliases = _split(m_text)

        # 初始化系統
        if sys_name not in rules:
            rules[sys_name] = {
                'should': [],
                'should_not': [],
                'aliases': [],
                'ph_min': None,
                'ph_max': None,
                'sludge_code': str(sludge_code or '').strip(),
                '說明': '',
            }

        # 加別名 (累積, 去重)
        for a in aliases:
            if a and a not in rules[sys_name]['aliases']:
                rules[sys_name]['aliases'].append(a)

        # 依角色加入
        if role == '應加':
            rules[sys_name]['should'].append({
                'drug': drug,
                'ratio_min': _to_float(ratio_min),
                'ratio_max': _to_float(ratio_max),
                'conc_min': _to_float(conc_min),
                'conc_max': _to_float(conc_max),
                'ph_min': _to_float(ph_min),
                'ph_max': _to_float(ph_max),
                'tank_type': str(tank_type or '').strip(),
                'water_quality': str(water_quality or '').strip(),
                'desc': desc,
                # 新增 3 個欄位 (17 欄 schema)
                'stage': str(stage or '').strip(),
                'orp_min': _to_float(orp_min),
                'orp_max': _to_float(orp_max),
            })
            # 更新系統級 pH 限制 (取最嚴格的)
            p_min = _to_float(ph_min)
            p_max = _to_float(ph_max)
            if p_min is not None:
                rules[sys_name]['ph_min'] = p_min if rules[sys_name]['ph_min'] is None else max(rules[sys_name]['ph_min'], p_min)
            if p_max is not None:
                rules[sys_name]['ph_max'] = p_max if rules[sys_name]['ph_max'] is None else min(rules[sys_name]['ph_max'], p_max)
        elif role == '禁加':
            rules[sys_name]['should_not'].append({
                'drug': drug,
                'tank_type': str(tank_type or '').strip(),
                'desc': desc,
            })

        # 累積說明
        if desc and desc not in rules[sys_name]['說明']:
            rules[sys_name]['說明'] = (rules[sys_name]['說明'] + ' ' + desc).strip()

    if not rules:
        return DOSING_RULES

    # 為了兼容舊呼叫端 (check_dosing_chemistry), 加 'should' / 'should_not' 簡化欄
    # 舊版要 list[str], 新版是 list[dict] — 加適配
    for sys_name, r in rules.items():
        r['should_drugs'] = [item['drug'] for item in r['should']]
        r['should_not_drugs'] = [item['drug'] for item in r['should_not']]
    return rules


# 模組載入時讀一次規則 (cache)
_CACHED_RULES = None

def get_rules():
    global _CACHED_RULES
    if _CACHED_RULES is None:
        _CACHED_RULES = load_dosing_rules_from_xlsx()
    return _CACHED_RULES

def clear_rules_cache():
    global _CACHED_RULES
    _CACHED_RULES = None


# 鎳系/輕系/各系 偵測關鍵字
SYSTEM_DETECT = [
    # (關鍵字 list, 系統名)
    (["鎳系", "鎳水洗", "化銅", "化學鎳", "鍍鎳"], "鎳系"),
    (["聶系"], "鎳系"),  # 常見筆誤 (聶 = 鎳)
    (["輕系", "輕細", "輕研磨", "輕污染"], "輕系"),
    (["各系", "綜合", "混合廢水", "總匯", "總混"], "各系"),
]


def detect_system(unit):
    """從單元名稱偵測屬於哪個分流系統。"""
    name = unit.get("name_in_doc", "") or ""
    for keywords, system in SYSTEM_DETECT:
        for kw in keywords:
            if kw in name:
                return system, kw
    return None, None


def has_chemical(measure_params, keywords):
    """檢查 measure_params 中是否含某類藥劑。"""
    if not measure_params:
        return False
    for pname in measure_params.keys():
        for kw in keywords:
            if kw in str(pname):
                return True
    return False


def get_ph_max(measure_params):
    """從 measure_params 找 pH 值的最大值。"""
    if not measure_params:
        return None
    for pname, pval in measure_params.items():
        if "pH" in str(pname):
            if isinstance(pval, dict):
                m = pval.get("max")
                try:
                    return float(m) if m is not None else None
                except (TypeError, ValueError):
                    return None
    return None


def check_dosing_chemistry(unit):
    """檢查單元的加藥機制是否符合分流系統學理。

    限縮: 只查「反應」型槽體 (pH調整池/快混/慢混/中和/氧化還原)
    收集池/調整池/沉澱池/儲槽 = 非反應型, 本來就不該加藥, 跳過

    Returns: list of findings
    """
    # ── 類型過濾: 只查反應型槽體 ──
    # 用 _槽體學理 的「類型」欄判斷
    try:
        import tank_chemistry as _tc
        rule = _tc.get_rule_for_unit(unit)
        if rule:
            unit_type = rule.get("類型", "")
            # 非反應型直接跳過 (收集/儲存/分離/生物/污泥 本來就不該加藥反應)
            if unit_type and unit_type != "反應":
                return []
    except Exception:
        # 規則庫沒載入時, 用單元名稱簡單篩選 (fallback)
        pass

    # ── 名稱 fallback 篩選 ──
    # 沒類型欄時, 用單元名稱判斷
    name = (unit.get("name_in_doc") or "").lower()
    # 一定不查的槽體類型 (從名稱)
    skip_keywords = ["收集池", "調整池", "儲槽", "貯槽", "貯留", "暫存",
                     "沉澱池", "沉降池", "濃縮槽", "污泥", "脫水",
                     "中間池", "緩衝", "放流池", "曝氣槽", "接觸氧化",
                     "活性碳", "離子交換", "砂濾"]
    for kw in skip_keywords:
        if kw in name:
            return []
    findings = []
    system, kw = detect_system(unit)
    if not system:
        return findings  # 無法判斷分流系統

    code = unit.get("raw_code") or unit.get("code") or "?"
    std_tank = unit.get("std_tank", "")
    rule = get_rules().get(system)
    if not rule:
        return findings

    measure = unit.get("measure_params") or {}

    # 1. 名稱含「聶系」筆誤 → 提示
    if kw == "聶系":
        findings.append({
            "嚴重度": "待人工",
            "類型": "文件一致性",
            "單元": code,
            "標準槽體": std_tank,
            "對照項目": "分流系統命名",
            "描述": (
                f"單元名稱「{unit.get('name_in_doc')}」含「聶系」, 應為「鎳系」筆誤 "
                f"(聶 vs 鎳, 同音字常見錯誤)。建議統一為「鎳系」。"
            ),
            "依據": "業界慣例: 鎳系 (Nickel) 廢水分流",
        })

    # 2. 鎳系應加 NaOH, 但 measure 沒有任何鹼劑 → 違反
    # 兼容兩種 schema: 舊 hardcoded DOSING_RULES (list[str]) 跟 新 _加藥規則 (list[dict])
    should_drugs = rule.get("should_drugs") or rule.get("should") or []
    if should_drugs and isinstance(should_drugs[0], dict):
        should_drugs = [d['drug'] for d in should_drugs]
    if should_drugs:
        has_should = has_chemical(measure, should_drugs)
        if not has_should:
            findings.append({
                "嚴重度": "不合理",
                "類型": "加藥機制",
                "單元": code,
                "標準槽體": std_tank,
                "對照項目": f"{system}應加藥劑",
                "描述": (
                    f"此單元屬於「{system}」, 學理上應加 {', '.join(should_drugs[:3])} 等。"
                    f"但 PDF 量測參數沒看到任何對應加藥。{rule['說明']}"
                ),
                "依據": f"學理: {system}分流加藥規則",
            })

    # 3. 鎳系不該加 PAC/Polymer, 但 measure 有 → 不合理 (學理硬性禁加)
    should_not_drugs = rule.get("should_not_drugs") or rule.get("should_not") or []
    if should_not_drugs and isinstance(should_not_drugs[0], dict):
        should_not_drugs = [d['drug'] for d in should_not_drugs]
    if should_not_drugs:
        has_not = has_chemical(measure, should_not_drugs)
        if has_not:
            findings.append({
                "嚴重度": "不合理",
                "類型": "加藥機制",
                "單元": code,
                "標準槽體": std_tank,
                "對照項目": f"{system}不該加的藥劑",
                "描述": (
                    f"此單元屬於「{system}」, 學理上不該加 {', '.join(should_not_drugs[:3])}。"
                    f"但 PDF 看到對應加藥, 請確認是否誤加或單元歸類錯誤。{rule['說明']}"
                ),
                "依據": f"學理: {system}分流加藥規則",
            })

    # 4. 鎳系 pH 應拉到 10.5+ 才能沉澱
    if rule.get("ph_min"):
        ph_max = get_ph_max(measure)
        if ph_max is not None and ph_max < rule["ph_min"]:
            findings.append({
                "嚴重度": "不合理",
                "類型": "加藥機制",
                "單元": code,
                "標準槽體": std_tank,
                "對照項目": "鎳系 pH 操作值",
                "描述": (
                    f"此單元屬於「{system}」, pH 量測最高 {ph_max} < 學理需求 {rule['ph_min']}。"
                    f"Ni²⁺ 沉澱需 pH ≥ 10.5, 否則放流會超標。"
                ),
                "依據": f"學理: Ni(OH)₂ 完全沉澱需 pH ≥ {rule['ph_min']}",
            })

    return findings


def check_sludge_classification(app_data):
    """檢查 B: 鎳系污泥沒區別於各系污泥 (廢棄物分類問題)。

    學理:
        - 鎳系污泥 = Ni(OH)₂ 沉澱物, 屬「有害事業廢棄物」(代碼 D-1101)
        - 各系污泥 = 混凝沉澱物, 屬「一般事業廢棄物」(代碼 D-2299)
        - 兩者法規上必須分開收集、貯存、清運

    檢查:
        - 廠內若有鎳系廢水處理 → 應該有獨立的「鎳系污泥儲槽」
        - 鎳系污泥 + 各系污泥混合進同一槽 → ❌ 廢棄物分類錯誤
    """
    findings = []
    units = app_data.get("units", {})

    # 是否有鎳系廢水處理?
    has_ni_system = False
    for unit in units.values():
        n = unit.get("name_in_doc", "") or ""
        if any(k in n for k in ["鎳系", "聶系", "鎳水洗", "化學鎳", "鍍鎳", "化銅"]):
            has_ni_system = True
            break

    if not has_ni_system:
        return findings

    # 有鎳系 → 找污泥相關單元
    sludge_units = []
    ni_sludge_units = []
    general_sludge_units = []
    for code, unit in units.items():
        n = unit.get("name_in_doc", "") or ""
        st = unit.get("std_tank", "") or ""
        # 是污泥單元嗎
        is_sludge = ("污泥" in n or "污泥" in st or
                     "脫水" in n or "脫水" in st or
                     "濃縮" in n or "濃縮" in st)
        if not is_sludge:
            continue
        sludge_units.append((code, n))
        # 鎳系污泥 vs 各系污泥
        if any(k in n for k in ["鎳系", "聶系", "鎳", "重金屬"]):
            ni_sludge_units.append((code, n))
        elif any(k in n for k in ["各系", "綜合", "混合", "一般"]):
            general_sludge_units.append((code, n))

    if not sludge_units:
        return findings  # 沒污泥單元, 不查

    # 異常 1: 沒看到「鎳系污泥」獨立槽
    if has_ni_system and not ni_sludge_units:
        findings.append({
            "嚴重度": "不合理",
            "類型": "加藥機制",
            "單元": "(全廠)",
            "標準槽體": "污泥儲槽",
            "對照項目": "鎳系污泥分類",
            "描述": (
                f"廠內有鎳系廢水處理 (檢出 {sum(1 for u in units.values() if any(k in (u.get('name_in_doc','') or '') for k in ['鎳系','聶系']))} 個鎳系單元), "
                f"但沒看到獨立的「鎳系污泥」儲槽/脫水機。"
                f"鎳系污泥屬有害事業廢棄物 (代碼 D-1101), 法規上應跟一般污泥分開收集、貯存、清運; "
                f"若混入各系污泥槽會造成廢棄物分類錯誤。"
            ),
            "依據": "事業廢棄物清理法 + 環境部公告事業廢棄物代碼 (D-1101: 含重金屬污泥)",
        })

    # 異常 2: 鎳系污泥 + 各系污泥同一槽
    # (這需要更精細的單元對應, 目前先以「都是污泥單元」但沒鎳系標記 = 可疑判斷)
    if has_ni_system and not ni_sludge_units and general_sludge_units:
        for code, n in general_sludge_units:
            findings.append({
                "嚴重度": "不合理",
                "類型": "加藥機制",
                "單元": code,
                "標準槽體": "污泥儲槽",
                "對照項目": "鎳系污泥分類",
                "描述": (
                    f"此單元 ({n}) 是非鎳系污泥儲槽, 但廠內有鎳系廢水處理。"
                    f"請確認鎳系污泥是否混入此槽; 若是, 違反廢棄物分類規定。"
                ),
                "依據": "事業廢棄物清理法 + 環境部公告事業廢棄物代碼",
            })

    return findings


def run_dosing_checks(app_data):
    """對所有單元跑加藥機制檢查 (含污泥分類)。"""
    # B: 全廠級別 — 鎳系污泥分類
    findings = list(check_sludge_classification(app_data))
    # 原本逐單元檢查 (A: 已限縮到反應型)
    for code, unit in app_data.get("units", {}).items():
        try:
            findings.extend(check_dosing_chemistry(unit))
        except Exception as e:
            findings.append({
                "嚴重度": "錯誤",
                "類型": "系統",
                "單元": code,
                "標準槽體": unit.get("std_tank", ""),
                "對照項目": "check_dosing_chemistry",
                "描述": f"檢查器錯誤: {e}",
                "依據": "(內部)",
            })
    return findings
