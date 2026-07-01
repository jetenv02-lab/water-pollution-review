# -*- coding: utf-8 -*-
"""攪拌機轉速 (RPM) 合理性檢查 — 純新增模組。

用途:
    - 讀 PDF 抽出的「攪拌機轉速」(design_params 或 measure_params)
    - 對照 _槽體學理 分頁的 RPM_min / RPM_max
    - 超出範圍 → 出 finding (提醒級)

規則庫欄位 (24 欄的 V/W/X):
    V. RPM_min (轉/分)
    W. RPM_max (轉/分)
    X. RPM 備註

對外 API:
    load_rpm_ranges() → dict[標準槽體 → (min, max, note)]
    check_unit_rpm(unit, rpm_ranges=None) → list[finding]
    check_all_units_rpm(units, rpm_ranges=None) → list[finding]

整合到主 pipeline (你要親手貼的 2 行):
    在 streamlit_app.py 或跑審查的地方:
        from check_rpm import check_all_units_rpm
        rpm_findings = check_all_units_rpm(app_data['units'])
        findings.extend(rpm_findings)
"""
import os
import re
from functools import lru_cache

import openpyxl

DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "規則庫.xlsx")
SHEET_NAME = "_槽體學理"


# 廠商 PDF 上「攪拌機轉速」的多種寫法
RPM_KEYWORDS = [
    "攪拌機轉速", "攪拌轉速", "轉速", "攪拌速度",
    "RPM", "rpm", "馬達轉速", "槳葉轉速",
]

# 槽體別名 (跟 dosing_rules_loader 對應)
TANK_ALIAS = {
    "pH 調整槽": "pH調整槽",
    "pH調整池": "pH調整槽",  # 若規則庫「pH調整池」跟「pH調整槽」都存在, 這個 alias 不生效, load 時會分別讀
    "中和槽": "中和池",
    "快混槽": "快混池",
    "慢混槽": "慢混池",
    "沉降池": "沉澱池",
    "混凝膠凝池": "混凝膠凝池",
    "沉降槽": "沉澱池",
}


@lru_cache(maxsize=1)
def load_rpm_ranges(xlsx_path=None):
    """從 _槽體學理 讀 RPM 範圍。

    回傳: dict[標準槽體 → {min: float, max: float, note: str}]
    """
    path = xlsx_path or DEFAULT_XLSX
    if not os.path.exists(path):
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[check_rpm] 讀 {path} 失敗: {e}")
        return {}
    if SHEET_NAME not in wb.sheetnames:
        return {}
    ws = wb[SHEET_NAME]

    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 24:
            continue
        tank = row[0]     # A
        status = row[7]   # H 狀態
        rmin = row[21]    # V
        rmax = row[22]    # W
        note = row[23]    # X

        if not tank or str(status or "").strip() != "V":
            continue
        # min/max 至少一個要有值
        rmin_val = _to_float(rmin)
        rmax_val = _to_float(rmax)
        if rmin_val is None and rmax_val is None:
            continue
        out[str(tank).strip()] = {
            "min": rmin_val,
            "max": rmax_val,
            "note": str(note or "").strip(),
        }
    return out


def _to_float(v):
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def _find_rpm_in_params(params):
    """從 design_params / measure_params 找攪拌轉速條目。

    回傳: (min, max, raw_str) 或 None
    """
    if not isinstance(params, dict):
        return None
    for key, val in params.items():
        key_str = str(key)
        # 找關鍵字
        for kw in RPM_KEYWORDS:
            if kw in key_str:
                # 排除「進料轉速」「泵浦轉速」等不相關的
                if any(bad in key_str for bad in ["泵", "馬達馬力", "功率"]):
                    continue
                if isinstance(val, dict):
                    return (
                        _to_float(val.get("min")),
                        _to_float(val.get("max")),
                        str(val.get("raw") or ""),
                    )
                else:
                    # 直接是數值
                    num = _to_float(val)
                    if num is not None:
                        return (num, num, str(val))
    return None


def check_unit_rpm(unit, rpm_ranges=None):
    """檢查單一單元的 RPM 是否合理。

    Args:
        unit: 處理單元 dict (有 std_tank + design_params / measure_params)
        rpm_ranges: dict from load_rpm_ranges(), None 自動載入

    Returns:
        list[finding]
    """
    if rpm_ranges is None:
        rpm_ranges = load_rpm_ranges()

    std_tank = unit.get("std_tank")
    if not std_tank:
        return []

    # 查規則
    key = std_tank
    if key not in rpm_ranges:
        # 試別名
        alias = TANK_ALIAS.get(std_tank)
        if alias and alias in rpm_ranges:
            key = alias
        else:
            return []

    rule = rpm_ranges[key]
    r_min = rule["min"]
    r_max = rule["max"]
    note = rule["note"]

    # 若規則兩邊都空 → 該槽不需檢查
    if r_min is None and r_max is None:
        return []

    # 找廠商申報的 RPM
    found = _find_rpm_in_params(unit.get("design_params") or {})
    if not found:
        found = _find_rpm_in_params(unit.get("measure_params") or {})
    if not found:
        return []  # PDF 沒寫轉速, 跳過

    doc_min, doc_max, raw = found
    code = unit.get("raw_code") or unit.get("code_id") or "?"

    findings = []

    # 判斷: 廠商 max 超過規則 max, 或廠商 min 低於規則 min
    problems = []
    if r_max is not None and doc_max is not None and doc_max > r_max:
        problems.append(f"上限 {doc_max:.0f} 超過學理範圍 {r_max:.0f}")
    if r_min is not None and doc_min is not None and doc_min < r_min:
        problems.append(f"下限 {doc_min:.0f} 低於學理範圍 {r_min:.0f}")

    if problems:
        range_str = ""
        if r_min is not None and r_max is not None:
            range_str = f"{r_min:.0f}~{r_max:.0f} RPM"
        elif r_min is not None:
            range_str = f"≥ {r_min:.0f} RPM"
        elif r_max is not None:
            range_str = f"≤ {r_max:.0f} RPM"

        # 判斷嚴重度: 偏差 > 50% 為待確認, 否則為提醒
        severity = "提醒"
        if r_max is not None and doc_max is not None and doc_max > r_max * 1.5:
            severity = "待確認"
        if r_min is not None and doc_min is not None and doc_min < r_min * 0.5:
            severity = "待確認"

        findings.append({
            "嚴重度": severity,
            "類型": "設計參數",
            "單元": code,
            "標準槽體": std_tank,
            "對照項目": "攪拌機轉速",
            "描述": (
                f"廠商申報攪拌機轉速 {doc_min:.0f}~{doc_max:.0f} RPM, "
                f"學理建議 {range_str}。"
                f" {' / '.join(problems)}。"
                f"{('備註: ' + note) if note else ''}"
            ),
            "依據": f"_槽體學理 RPM 範圍 ({std_tank})",
        })

    return findings


def check_all_units_rpm(units, rpm_ranges=None):
    """批次檢查所有單元的 RPM。

    Args:
        units: dict[code → unit] 或 list[unit]
        rpm_ranges: dict from load_rpm_ranges(), None 自動載入

    Returns:
        list[finding]
    """
    if rpm_ranges is None:
        rpm_ranges = load_rpm_ranges()

    findings = []
    if isinstance(units, dict):
        unit_iter = units.values()
    else:
        unit_iter = units

    fast_mix_units = []   # (code, doc_max)
    slow_mix_units = []   # (code, doc_max)

    for unit in unit_iter:
        if not isinstance(unit, dict):
            continue
        # 補 code_id 若沒有
        unit_copy = dict(unit)
        unit_copy.setdefault("code_id", unit.get("raw_code"))
        findings.extend(check_unit_rpm(unit_copy, rpm_ranges))

        # 收集快混/慢混 RPM 準備做比值檢查
        std = unit.get("std_tank") or ""
        rpm = _find_rpm_in_params(unit.get("design_params") or {}) or \
              _find_rpm_in_params(unit.get("measure_params") or {})
        if rpm:
            doc_min, doc_max, _ = rpm
            rpm_val = doc_max or doc_min
            code = unit.get("raw_code") or unit.get("code_id") or "?"
            if "快混" in std:
                fast_mix_units.append((code, rpm_val))
            elif "慢混" in std:
                slow_mix_units.append((code, rpm_val))

    # 快混/慢混 比值檢查: 快混應比慢混快 3~10 倍
    # 若同一廠內 快混 RPM ≤ 慢混 RPM 或 差 < 2 倍 → 提醒
    if fast_mix_units and slow_mix_units:
        for f_code, f_rpm in fast_mix_units:
            for s_code, s_rpm in slow_mix_units:
                if not f_rpm or not s_rpm:
                    continue
                if f_rpm <= s_rpm * 2:
                    findings.append({
                        "嚴重度": "待確認",
                        "類型": "設計參數",
                        "單元": f"{f_code} vs {s_code}",
                        "標準槽體": "快混/慢混",
                        "對照項目": "快混/慢混 RPM 比值",
                        "描述": (
                            f"{f_code} 快混 {f_rpm:.0f} RPM 與 "
                            f"{s_code} 慢混 {s_rpm:.0f} RPM 差距 < 2 倍 "
                            f"(學理: 快混應比慢混快 3~10 倍, 快混劇烈打散凝集劑, "
                            f"慢混低速讓絮體長大而不打散)。"
                            f"請確認轉速設定是否符合槽體功能。"
                        ),
                        "依據": "混凝原理: G 值 (速度梯度) 快混需 500~1000 s⁻¹, 慢混 30~80 s⁻¹",
                    })

    return findings


def _self_test():
    """自我測試 — 拿秋棠 T01-01 (90~220 RPM, 中和池) 當範例。"""
    ranges = load_rpm_ranges()
    print(f"[1] 讀到 {len(ranges)} 個槽體 RPM 範圍")
    for tank in ["中和池", "快混池", "慢混池", "沉澱池", "pH調整槽"]:
        r = ranges.get(tank)
        if r:
            print(f"  {tank}: {r['min']}~{r['max']} RPM ({r['note']})")

    print()

    # Test 1: 中和池 90~220, 學理 60~200, 上限超 → 應出提醒
    fake_unit = {
        "std_tank": "中和池",
        "raw_code": "T01-01",
        "design_params": {
            "攪拌機轉速": {"min": 90, "max": 220, "raw": "90~220 轉/分"},
        },
    }
    fs = check_unit_rpm(fake_unit, ranges)
    print(f"Test 1 (中和池 90~220, 學理 60~200): {len(fs)} finding")
    for f in fs:
        print(f"  {f['嚴重度']} - {f['對照項目']}: {f['描述']}")

    # Test 2: 快混槽 100~150, 學理 150~300, 下限低 → 應出提醒
    fake_unit2 = {
        "std_tank": "快混池",
        "raw_code": "T01-02",
        "design_params": {"攪拌轉速": {"min": 100, "max": 150}},
    }
    fs2 = check_unit_rpm(fake_unit2, ranges)
    print(f"\nTest 2 (快混池 100~150, 學理 150~300): {len(fs2)} finding")
    for f in fs2:
        print(f"  {f['嚴重度']} - {f['描述']}")

    # Test 3: 慢混池 200 RPM, 學理 20~60 → 應出待確認 (超 3 倍)
    fake_unit3 = {
        "std_tank": "慢混池",
        "raw_code": "T01-03",
        "design_params": {"攪拌轉速": {"min": 200, "max": 200}},
    }
    fs3 = check_unit_rpm(fake_unit3, ranges)
    print(f"\nTest 3 (慢混池 200 RPM, 學理 20~60): {len(fs3)} finding")
    for f in fs3:
        print(f"  {f['嚴重度']} - {f['描述']}")

    # Test 4: 中和池 100~150 都在範圍內 → 0 finding
    fake_unit4 = {
        "std_tank": "中和池",
        "raw_code": "T01-04",
        "design_params": {"攪拌機轉速": {"min": 100, "max": 150}},
    }
    fs4 = check_unit_rpm(fake_unit4, ranges)
    print(f"\nTest 4 (中和池 100~150, 全在範圍內): {len(fs4)} finding (預期 0)")
    assert len(fs4) == 0

    print("\n[OK] 自我測試通過")


if __name__ == "__main__":
    _self_test()
