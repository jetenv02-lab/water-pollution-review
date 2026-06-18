# -*- coding: utf-8 -*-
"""建立水措審查「規則庫」Excel 範本(.xlsx),可直接匯入 Google Sheets 供三人協作。

結構(已與使用者確認):
  _來源清單    (固定) 每份審查意見PDF的來源、技師、日期
  _槽體對照表  (固定) 槽體別名歸一:原始名稱/序號/代碼 → 標準槽體名稱 + 待確認旗標
  <槽體類型分頁> 每種槽體一個分頁(pH調整槽、廢水調整池、慢混池、沉澱池...)

每個槽體分頁欄位:
  缺失ID | 來源 | 原文缺失 | 檢查類型 | 對照項目 | 規則 | 比對位置 | 判定邏輯
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"/mnt/user-data/outputs/規則庫_範本.xlsx"

# ---- 樣式 ----
HEADER_FILL = PatternFill("solid", fgColor="2F5496")     # 深藍
FIXED_FILL = PatternFill("solid", fgColor="548235")      # 綠(固定表)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols, fill=HEADER_FILL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fill_rows(ws, rows, ncols):
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ncols):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER


# ===== 各槽體分頁的欄位定義 =====
TANK_HEADERS = ["缺失ID", "來源", "原文缺失", "檢查類型",
                "對照項目", "規則(萃取/可比對判斷式)", "比對位置(依標題,非頁碼)", "判定邏輯(條件→結論)"]
TANK_WIDTHS = [10, 14, 46, 14, 16, 40, 30, 40]

# 預先建立的代表性槽體分頁(之後比對引擎會自動新增更多)
SEED_TANKS = ["pH調整槽", "廢水調整池", "慢混池", "沉澱池", "中和池", "放流池"]

# 非特定槽體的缺失 → 分兩類分頁(使用者決定)
GENERIC_TABS = ["現場設備類", "文件類"]

# 三個使用者範例 → 寫入對應槽體分頁作為示範資料
SAMPLE_DATA = {
    "pH調整槽": [
        ["D001", "序4 張維欽/洪瑞敏技師",
         "18/42頁 T01-05 pH調整兼快混槽設施資料表,本化混程序擬去除重金屬,控制之pH為6~9,其低限值過低,非適宜去除重金屬之pH範圍。",
         "設計參數", "pH",
         "化混去除重金屬 pH 下限 須 > 6(6~9 之低限過低)",
         "參、水污染防治措施資料/廢(污)水(前)處理設施資料表/三、處理單元名稱及操作參數/(二)設計操作參數",
         "若 用途含『去除重金屬/化混』且 pH下限<=6 → 標記:pH下限過低,非適宜範圍"],
        ["D002", "序6 劉暐廷技師",
         "6) P7_質平計算,於各單元進出流水質濃度及質量有效位數應一致,如 T01-4 pH調整槽進出流水質除pH外,餘各項水質濃度及質量不應有變化。",
         "質量平衡", "各項水質濃度/質量",
         "pH調整槽:除 pH 外,進流=出流(濃度與質量皆不應變化);有效位數一致",
         "參、水污染防治措施資料/水質水量平衡示意圖/二、處理單元之進出水質資料",
         "比對該單元進/出流;除pH外任一項濃度或質量有變化 → 標記:質平不應變化"],
        ["D004", "序3 施紹揚技師",
         "T01 化混程序去除銅,控制之pH為6~9,惟銅之鹼性沉降最佳pH約8.5~9.5,所列下限偏低。",
         "設計參數", "pH",
         "去除銅(鹼性沉降)pH 宜 8.5~9.5;6~9 之下限偏低",
         "參、水污染防治措施資料/廢(污)水(前)處理設施資料表/三、處理單元名稱及操作參數/(二)設計操作參數",
         "若 用途含『去除銅』且 pH下限<8.5 → 標記:pH下限偏低,非銅鹼性沉降最佳範圍"],
        ["D005", "序12 徐振利技師",
         "T01-03 pH中和槽控制pH為5~9,其低限值偏低,不利後續生物處理(微生物適宜pH約6~8)。",
         "設計參數", "pH",
         "中和槽後接生物處理 pH 宜 6~8;5~9 之下限偏低",
         "參、水污染防治措施資料/廢(污)水(前)處理設施資料表/三、處理單元名稱及操作參數/(二)設計操作參數",
         "若 後接生物處理 且 pH下限<6 → 標記:pH下限偏低,不利微生物"],
    ],
    "廢水調整池": [
        ["D003", "序8 李俊坤技師",
         "頁次19/65 T01-02 廢水調整槽相關機具設施未填液位計。",
         "機具設施", "液位計",
         "廢水調整池(調整槽)之(四)相關機具設施 應含『液位計』",
         "參、水污染防治措施資料/廢(污)水(前)處理設施資料表/三、處理單元名稱及操作參數/(四)相關機具設施",
         "若 該單元(四)相關機具設施 未列『液位計』 → 標記:未填液位計"],
    ],
    "快混槽": [
        ["D006", "序9 汪正倫技師",
         "T01-08 快混槽控制之pH為8~12,其下限值8偏低,非適宜去除鎳(鎳鹼性最佳沉降約pH 10~10.5)之範圍。",
         "設計參數", "pH",
         "去除鎳(鹼性沉降)pH 宜 10~10.5;8~12 之下限8偏低",
         "參、水污染防治措施資料/廢(污)水(前)處理設施資料表/三、處理單元名稱及操作參數/(二)設計操作參數",
         "若 用途含『去除鎳』且 pH下限<10 → 標記:pH下限偏低,非鎳鹼性沉降最佳範圍"],
    ],
    "曝氣槽": [
        ["D007", "序10 劉家豪技師",
         "曝氣槽溶氧(DO)控制範圍1~10 mg/L,範圍過大且上限過高,一般生物處理曝氣槽DO宜維持約1~2 mg/L。",
         "設計參數", "溶氧DO",
         "生物處理曝氣槽 DO 宜約 1~2 mg/L;1~10 範圍過大、上限過高",
         "參、水污染防治措施資料/廢(污)水(前)處理設施資料表/三、處理單元名稱及操作參數/(二)設計操作參數",
         "若 DO上限>2(或範圍過大) → 標記:DO範圍不合理"],
    ],
}

# 非特定槽體的缺失 → 寫入 現場設備類 / 文件類
GENERIC_DATA = {
    "現場設備類": [
        ["G001", "序8 李俊坤技師",
         "T01-04 暫存槽1、T01-28 貯留槽 相關機具設施未填液位計。",
         "機具設施", "液位計",
         "暫存/貯留類槽體(四)相關機具設施 應含『液位計』",
         "參、水污染防治措施資料/廢(污)水(前)處理設施資料表/三、處理單元名稱及操作參數/(四)相關機具設施",
         "若 暫存/貯留槽(四)相關機具設施 未列『液位計』 → 標記:未填液位計"],
    ],
    "文件類": [
        ["G002", "序7 陳映嘉技師",
         "質量平衡表中有機物濃度於各單元間自行濃縮(出流>進流且無外加負荷),不符質量守恆與處理學理。",
         "質量平衡", "有機物濃度",
         "無外加污染源之單元 出流濃度 不得無故 > 進流濃度",
         "參、水污染防治措施資料/水質水量平衡示意圖/二、處理單元之進出水質資料",
         "若 該單元出流濃度>進流濃度 且 無外加負荷 → 標記:濃度不應自行濃縮,違反質量守恆"],
    ],
}


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- _來源清單 ----
    ws = wb.create_sheet("_來源清單")
    hdr = ["來源代號", "審查意見檔名", "技師姓名", "技師證書字號",
           "查核日期", "簽證事業名稱", "匯入日期", "備註"]
    ws.append(hdr)
    set_widths(ws, [10, 30, 12, 18, 14, 24, 14, 24])
    style_header(ws, len(hdr), FIXED_FILL)
    fill_rows(ws, [
        ["S01", "水污染業務查核缺失.pdf", "(多位)", "(見原文)", "111-114", "(多家)", "", "環工技師查核缺失彙整,序1-14"],
    ], len(hdr))

    # ---- _槽體對照表 ----
    ws = wb.create_sheet("_槽體對照表")
    hdr = ["原始名稱/序號/代碼", "出現來源", "猜測標準槽體名稱", "待確認", "確認人", "備註"]
    ws.append(hdr)
    set_widths(ws, [28, 16, 18, 10, 12, 28])
    style_header(ws, len(hdr), FIXED_FILL)
    fill_rows(ws, [
        ["T01-05 pH調整兼快混槽", "洪瑞敏技師", "pH調整槽", "否", "", "範例:含pH→歸pH調整槽"],
        ["T01-4 pH調整槽", "(範例)", "pH調整槽", "否", "", ""],
        ["T01-02 廢水調整槽", "(範例)", "廢水調整池", "是", "", "別名:廢水調整槽/鉻系廢水調整池→廢水調整池?"],
        ["鉻系廢水調整池", "秋棠", "廢水調整池", "是", "", "待確認確認"],
        ["中和池 T01-01 代碼120", "秋棠", "中和池", "否", "", "秋棠 p72"],
        ["放流池 T01-02 代碼417", "秋棠", "放流池", "否", "", "秋棠 p73"],
        ["螯合鎳系廢水收集池 T02-01 代碼414", "秋棠", "(待新增分頁)", "是", "", "秋棠 p81;新槽體類型"],
    ], len(hdr))

    # ---- _審查紀錄(每份審查文件的索引) ----
    ws = wb.create_sheet("_審查紀錄")
    hdr = ["審查文件", "不合理數量", "審查結果", "審查時間", "審查次數", "比對結果檔案位置"]
    ws.append(hdr)
    set_widths(ws, [34, 12, 16, 18, 10, 40])
    style_header(ws, len(hdr), FIXED_FILL)
    fill_rows(ws, [
        ["05 申請資料(秋棠)(1150119)final.pdf", "(待比對)", "(待比對)", "", "", "(比對引擎產出 Word 後填入)"],
    ], len(hdr))

    # ---- 各槽體分頁 ----
    for tank in SEED_TANKS:
        ws = wb.create_sheet(tank)
        ws.append(TANK_HEADERS)
        set_widths(ws, TANK_WIDTHS)
        style_header(ws, len(TANK_HEADERS))
        fill_rows(ws, SAMPLE_DATA.get(tank, []), len(TANK_HEADERS))

    # ---- 通用缺失分頁:現場設備類 / 文件類(非特定槽體) ----
    for tab in GENERIC_TABS:
        ws = wb.create_sheet(tab)
        ws.append(TANK_HEADERS)
        set_widths(ws, TANK_WIDTHS)
        style_header(ws, len(TANK_HEADERS))
        fill_rows(ws, GENERIC_DATA.get(tab, []), len(TANK_HEADERS))

    # ---- 說明分頁(放最前) ----
    ws = wb.create_sheet("_說明", 0)
    ws.append(["水措審查 規則庫 — 使用說明"])
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")
    notes = [
        "",
        "【分頁規則】",
        "  • 以『槽體類型名稱』命名分頁(不看序號,因每份文件序號不同)。",
        "  • 審查意見提到某槽體 → 整條記到該槽體分頁。新槽體類型自動新增分頁。",
        "  • _ 開頭為固定系統表:_來源清單、_槽體對照表。",
        "",
        "【_槽體對照表】",
        "  • 程式自動依關鍵字猜標準槽體名稱(含pH→pH調整槽),不確定者『待確認=是』,供三人修正。",
        "",
        "【槽體分頁欄位】",
        "  缺失ID｜來源｜原文缺失｜檢查類型(設計參數/水質標準/質量平衡/機具設施)｜對照項目｜規則｜比對位置(依標題)｜判定邏輯",
        "",
        "【比對位置 = 依標題定位(非頁碼)】每份文件頁碼不同,引擎以章節標題尋找:",
        "  • 進出水質 → 參/水質水量平衡示意圖/二、處理單元之進出水質資料",
        "  • 設計參數/量測參數/相關機具設施 → 參/廢(污)水(前)處理設施資料表/三、處理單元名稱及操作參數",
        "",
        "【協作】此檔可『檔案→匯入』到 Google Sheets,分享給3人共同編輯。",
    ]
    for n in notes:
        ws.append([n])
    ws.column_dimensions["A"].width = 90
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1):
        row[0].alignment = WRAP

    wb.save(OUT)
    print("已建立規則庫範本:", OUT)
    print("分頁:", [s.title for s in wb.worksheets])


if __name__ == "__main__":
    main()
