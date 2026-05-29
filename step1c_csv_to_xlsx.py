# -*- coding: utf-8 -*-
"""Step 1c: 把 rules_extracted.csv 寫入 規則庫.xlsx 的各槽體分頁。

冪等：可重複執行。每次會清空各槽體分頁(保留標題列)，再依 CSV 重新寫入。
不會動 _說明 / _來源清單 / _槽體對照表 / _審查紀錄 等系統表。

依賴: openpyxl
用法: python step1c_csv_to_xlsx.py
"""
import csv
import os
import shutil
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r"C:\Users\jeten\Desktop\AI\水措審查"
CSV_IN = os.path.join(BASE, "rules_extracted.csv")
XLSX_TEMPLATE = os.path.join(BASE, "規則庫_範本.xlsx")
XLSX_OUT = os.path.join(BASE, "規則庫.xlsx")

# ---- 樣式 ----
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TANK_HEADERS = ["缺失ID", "來源", "原文缺失", "檢查類型",
                "對照項目", "規則(萃取/可比對判斷式)",
                "比對位置(依標題,非頁碼)", "判定邏輯(條件→結論)",
                "技師姓名", "序號", "原始槽體代號"]
TANK_WIDTHS = [10, 14, 46, 12, 14, 36, 30, 36, 12, 18, 18]

GENERIC_TANKS = {"(文件類)", "(現場設備類)"}


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def normalize_tank_name(name):
    """把 (文件類)/(現場設備類) 轉成合法分頁名稱(Excel 分頁名不能含某些字元)。"""
    name = (name or "").strip()
    if name in ("", None):
        return "未分類"
    # Excel 分頁名不能含 \ / ? * [ ] :
    for ch in r"\/?*[]:":
        name = name.replace(ch, "")
    return name[:31]  # Excel 分頁名最長 31 字元


def clear_tank_sheet(ws):
    """清空槽體分頁的資料列(保留標題列)。"""
    if ws.max_row <= 1:
        return
    ws.delete_rows(2, ws.max_row - 1)


def ensure_tank_sheet(wb, tank_name):
    """確保分頁存在;不存在就建立並寫入標題列。"""
    sname = normalize_tank_name(tank_name)
    if sname in wb.sheetnames:
        ws = wb[sname]
    else:
        ws = wb.create_sheet(sname)
        ws.append(TANK_HEADERS)
        set_widths(ws, TANK_WIDTHS)
        style_header(ws, len(TANK_HEADERS))
    return ws


def append_row(ws, row_values):
    ws.append(row_values)
    last = ws.max_row
    for cell in ws[last]:
        cell.alignment = WRAP
        cell.border = BORDER


def update_source_list(wb, rules):
    """把 CSV 出現的『來源』寫入 _來源清單(若 _來源清單 已存在)。"""
    if "_來源清單" not in wb.sheetnames:
        return
    ws = wb["_來源清單"]
    existing = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            existing.add(str(v).strip())
    sources = sorted({r["來源"].split(" ", 1)[0] for r in rules if r.get("來源")})
    for sid in sources:
        if sid in existing:
            continue
        ws.append([sid, "水污染業務查核缺失.pdf", "(多位)", "(見原文)",
                   "111-114", "(多家)", "", "由 step1c 自動匯入"])


def update_tank_alias_table(wb, rules):
    """把 CSV 出現的『原始槽體代號』寫入 _槽體對照表(若已存在)。"""
    if "_槽體對照表" not in wb.sheetnames:
        return
    ws = wb["_槽體對照表"]
    existing = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            existing.add(str(v).strip())
    seen = {}  # alias -> (sources_set, std_tank)
    for r in rules:
        std = r.get("標準槽體名稱", "").strip()
        raw = (r.get("原始槽體代號") or "").strip()
        if not raw:
            continue
        for alias in [a.strip() for a in raw.split(";")]:
            if not alias:
                continue
            if alias in seen:
                seen[alias][0].add(r.get("技師姓名", ""))
            else:
                seen[alias] = ({r.get("技師姓名", "")}, std)
    for alias, (srcs, std) in sorted(seen.items()):
        if alias in existing:
            continue
        ws.append([alias, ", ".join(sorted(s for s in srcs if s)),
                   std, "是" if std in GENERIC_TANKS else "否",
                   "", "step1c 自動匯入"])


def main():
    if not os.path.exists(CSV_IN):
        print(f"找不到 CSV: {CSV_IN}")
        return

    # 從範本複製一份當工作檔(若工作檔不存在)
    if not os.path.exists(XLSX_OUT):
        if os.path.exists(XLSX_TEMPLATE):
            shutil.copy(XLSX_TEMPLATE, XLSX_OUT)
            print(f"從範本複製 → {XLSX_OUT}")
        else:
            wb_new = openpyxl.Workbook()
            wb_new.remove(wb_new.active)
            wb_new.save(XLSX_OUT)
            print(f"新建空白 xlsx → {XLSX_OUT}")

    # 讀 CSV
    rules = []
    with open(CSV_IN, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rules.append(row)
    print(f"讀入 {len(rules)} 筆規則")

    # 開 xlsx
    wb = openpyxl.load_workbook(XLSX_OUT)

    # 統計 → 知道哪些槽體分頁要重寫
    tank_groups = defaultdict(list)
    for r in rules:
        tank = r.get("標準槽體名稱", "").strip() or "未分類"
        tank_groups[tank].append(r)

    # 1) 清空現有槽體分頁(以底線開頭的系統表略過)
    for sname in list(wb.sheetnames):
        if sname.startswith("_"):
            continue
        if sname not in tank_groups:
            # 既有的分頁,但 CSV 沒資料 → 清空但保留
            clear_tank_sheet(wb[sname])

    # 2) 對 CSV 出現的每個槽體 → 確保分頁存在 → 清空 → 重新寫入
    for tank, rows in tank_groups.items():
        ws = ensure_tank_sheet(wb, tank)
        clear_tank_sheet(ws)
        # 若分頁標題與新 schema 不一致 → 重寫標題列
        if [c.value for c in ws[1]][:len(TANK_HEADERS)] != TANK_HEADERS:
            for c in range(1, len(TANK_HEADERS) + 1):
                ws.cell(row=1, column=c, value=TANK_HEADERS[c - 1])
            set_widths(ws, TANK_WIDTHS)
            style_header(ws, len(TANK_HEADERS))
        for r in rows:
            append_row(ws, [
                r.get("缺失ID", ""),
                r.get("來源", ""),
                r.get("原文缺失", ""),
                r.get("檢查類型", ""),
                r.get("對照項目", ""),
                r.get("規則", ""),
                r.get("比對位置", ""),
                r.get("判定邏輯", ""),
                r.get("技師姓名", ""),
                r.get("序號", ""),
                r.get("原始槽體代號", ""),
            ])

    # 3) 更新 _來源清單 / _槽體對照表
    update_source_list(wb, rules)
    update_tank_alias_table(wb, rules)

    wb.save(XLSX_OUT)
    print(f"已寫入 {XLSX_OUT}")
    print("分頁清單:")
    for s in wb.sheetnames:
        ws = wb[s]
        print(f"  {s}: {ws.max_row - 1 if not s.startswith('_') else ws.max_row} 列")


if __name__ == "__main__":
    main()
