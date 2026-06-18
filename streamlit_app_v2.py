# -*- coding: utf-8 -*-
"""水措審查系統 — Streamlit 線上版 v2。

v2 重大改進(相對 v1 streamlit_app.py):
- 改用成熟版抽取器 step2_extract_v2 (覆蓋率 100%、單元類型歸類正確)
- 顯示更豐富的單元資訊(設計參數/量測參數/機具設施/進出流水質)
- 加入「水量平衡缺口提示」(尚未做 OCR,但會說明限制)

部署:
  Streamlit Cloud → Main file: streamlit_app_v2.py
"""
import csv
import io
import json
import os
import tempfile
from collections import defaultdict
from datetime import datetime

import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 直接 import 我們的 v2 抽取器
from step2_extract_v2 import extract_application

# ───────────────────────────── 頁面設定 ─────────────────────────────

st.set_page_config(
    page_title="水措審查系統 v2",
    page_icon="💧",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #2F5496 0%, #1a365d 100%);
        color: white;
        padding: 20px 30px;
        border-radius: 8px;
        margin-bottom: 20px;
    }
    .main-header h1 { margin: 0; font-size: 28px; }
    .main-header p { margin: 4px 0 0 0; opacity: 0.9; }
    .v2-badge {
        background: #38a169;
        color: white;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 13px;
        margin-left: 8px;
    }
    .ocr-warning {
        background: #fef3c7;
        border-left: 4px solid #f59e0b;
        padding: 12px 16px;
        border-radius: 4px;
        margin: 12px 0;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>💧 水措審查系統 <span class="v2-badge">v2</span></h1>
    <p>自動化「水污染防治措施」申請文件審查 · 比對環工技師查核缺失資料庫</p>
</div>
""", unsafe_allow_html=True)

CSV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rules_extracted.csv")


@st.cache_data(show_spinner=False)
def load_rules():
    if not os.path.exists(CSV_PATH):
        return [], {}
    rules = []
    with open(CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rules.append(row)
    by_tank = defaultdict(list)
    for r in rules:
        tank = (r.get("標準槽體名稱") or "未分類").strip()
        by_tank[tank].append(r)
    return rules, dict(by_tank)


# ───────────────────────────── Excel 輸出 ─────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SEVERITY_FILL = {
    "不合理": PatternFill("solid", fgColor="FCE4D6"),
    "待確認": PatternFill("solid", fgColor="FFF2CC"),
    "合理": PatternFill("solid", fgColor="E2EFDA"),
}
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_unit_excel(app_data):
    """產出『各單元資料表』Excel (含設計/量測/機具/進出水質)。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 總表分頁
    ws = wb.create_sheet("_單元總表")
    headers = ["單元代號", "原始名稱", "標準槽體類型", "代碼", "出現頁數",
               "設計參數數", "量測參數數", "機具項目數", "進流數", "出流數"]
    widths = [12, 30, 16, 8, 16, 12, 12, 12, 10, 10]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    ws.freeze_panes = "A2"

    for code, info in sorted(app_data["units"].items()):
        ws.append([
            code, info.get("name_in_doc", ""), info.get("std_tank", ""),
            info.get("code_id", ""),
            ", ".join(map(str, info.get("pages_found", []))),
            len(info.get("design_params", {})),
            len(info.get("measure_params", {})),
            len(info.get("equipment", [])),
            len(info.get("influent", {})),
            len(info.get("effluent", {})),
        ])

    # 每個單元一張分頁
    for code, info in sorted(app_data["units"].items()):
        sname = code[:31]
        ws = wb.create_sheet(sname)
        row = 1

        # 標題
        ws.cell(row=row, column=1, value=f"{code} {info.get('name_in_doc', '')}").font = Font(size=14, bold=True)
        row += 1
        ws.cell(row=row, column=1, value=f"標準類型: {info.get('std_tank', '')} | 代碼: {info.get('code_id', '')} | 頁: {info.get('pages_found', [])}")
        row += 2

        # 設計參數
        if info.get("design_params"):
            ws.cell(row=row, column=1, value="設計操作參數").font = Font(bold=True, color="2F5496")
            row += 1
            ws.cell(row=row, column=1, value="參數")
            ws.cell(row=row, column=2, value="最小值")
            ws.cell(row=row, column=3, value="最大值")
            ws.cell(row=row, column=4, value="原始")
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for pname, pval in info["design_params"].items():
                ws.cell(row=row, column=1, value=pname)
                ws.cell(row=row, column=2, value=pval.get("min", ""))
                ws.cell(row=row, column=3, value=pval.get("max", ""))
                ws.cell(row=row, column=4, value=pval.get("raw", ""))
                row += 1
            row += 1

        # 量測參數
        if info.get("measure_params"):
            ws.cell(row=row, column=1, value="量測操作參數").font = Font(bold=True, color="2F5496")
            row += 1
            ws.cell(row=row, column=1, value="參數")
            ws.cell(row=row, column=2, value="最小值")
            ws.cell(row=row, column=3, value="最大值")
            ws.cell(row=row, column=4, value="原始")
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for pname, pval in info["measure_params"].items():
                ws.cell(row=row, column=1, value=pname)
                ws.cell(row=row, column=2, value=pval.get("min", ""))
                ws.cell(row=row, column=3, value=pval.get("max", ""))
                ws.cell(row=row, column=4, value=pval.get("raw", ""))
                row += 1
            row += 1

        # 機具
        if info.get("equipment"):
            ws.cell(row=row, column=1, value="相關機具設施").font = Font(bold=True, color="2F5496")
            row += 1
            ws.cell(row=row, column=1, value="名稱")
            ws.cell(row=row, column=2, value="位置")
            ws.cell(row=row, column=3, value="數量")
            ws.cell(row=row, column=4, value="馬力(kW)")
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for eq in info["equipment"]:
                ws.cell(row=row, column=1, value=eq.get("name", ""))
                ws.cell(row=row, column=2, value=eq.get("位置", ""))
                ws.cell(row=row, column=3, value=eq.get("數量", ""))
                ws.cell(row=row, column=4, value=eq.get("馬力_kW", ""))
                row += 1
            row += 1

        # 進流水質
        for infl_code, q_data in info.get("influent", {}).items():
            ws.cell(row=row, column=1, value=f"進流水質 — {infl_code}").font = Font(bold=True, color="2F5496")
            row += 1
            ws.cell(row=row, column=1, value="水質項目")
            ws.cell(row=row, column=2, value="濃度")
            ws.cell(row=row, column=3, value="質量")
            ws.cell(row=row, column=4, value="範圍(pH/水溫)")
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for item, val in q_data.items():
                ws.cell(row=row, column=1, value=item)
                ws.cell(row=row, column=2, value=val.get("濃度", ""))
                ws.cell(row=row, column=3, value=val.get("質量", ""))
                ws.cell(row=row, column=4, value=val.get("範圍", ""))
                row += 1
            row += 1

        # 出流水質
        for effl_code, q_data in info.get("effluent", {}).items():
            ws.cell(row=row, column=1, value=f"出流水質 — {effl_code}").font = Font(bold=True, color="38a169")
            row += 1
            ws.cell(row=row, column=1, value="水質項目")
            ws.cell(row=row, column=2, value="濃度")
            ws.cell(row=row, column=3, value="質量")
            ws.cell(row=row, column=4, value="範圍(pH/水溫)")
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = HEADER_FILL
                ws.cell(row=row, column=c).font = HEADER_FONT
            row += 1
            for item, val in q_data.items():
                ws.cell(row=row, column=1, value=item)
                ws.cell(row=row, column=2, value=val.get("濃度", ""))
                ws.cell(row=row, column=3, value=val.get("質量", ""))
                ws.cell(row=row, column=4, value=val.get("範圍", ""))
                row += 1
            row += 1

        # 欄寬
        for c in range(1, 6):
            ws.column_dimensions[get_column_letter(c)].width = [22, 18, 18, 36][min(c - 1, 3)]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ───────────────────────────── 介面 ─────────────────────────────

rules, rules_by_tank = load_rules()

with st.sidebar:
    st.header("📋 系統狀態")
    st.metric("規則總數", len(rules))
    st.metric("槽體分類數", len(rules_by_tank))
    st.divider()
    st.markdown("**版本資訊**")
    st.text("v2: 成熟版抽取器")
    st.text("✓ 38 單元覆蓋")
    st.text("✓ 設計/量測/機具/水質")
    st.text("✗ 流向圖 OCR (待做)")
    st.divider()
    st.markdown("**專案連結**")
    st.markdown("[GitHub](https://github.com/jetenv02-lab/water-pollution-review)")

tab1, tab2, tab3 = st.tabs(["🚀 開始審查", "📊 規則庫瀏覽", "📖 使用說明"])

with tab1:
    st.subheader("上傳申請文件 PDF")
    st.caption("v2 抽取器會解析: 處理設施資料表 + 進出水質資料表。流向示意圖(頁9-13)若為圖片,目前暫無法解析,需後續加入 OCR。")

    st.markdown("""
    <div class="ocr-warning">
    <strong>⚠ v2 限制告知:</strong><br>
    水量平衡示意圖(頁 9-13)在原 PDF 通常是圖片,本版尚未啟用 OCR。
    要做完整質量平衡審查需要該圖的流量/加藥/含水率資料,目前該部分需人工檢視。
    所有「處理設施資料表」「進出水質資料表」的文字內容皆已成功抽取。
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader("選擇 PDF", type=["pdf"])

    if uploaded is not None:
        st.success(f"已上傳: **{uploaded.name}** ({uploaded.size // 1024} KB)")
        if st.button("🚀 開始解析", type="primary"):
            with st.spinner("📖 解析 PDF 中... (大檔案可能需要 1-3 分鐘)"):
                pdf_bytes = uploaded.read()
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tf:
                    tf.write(pdf_bytes)
                    tmp_path = tf.name
                try:
                    app_data = extract_application(tmp_path, verbose=False)
                finally:
                    try:
                        os.unlink(tmp_path)
                    except:
                        pass

            st.success(f"✅ 抽取完成! 共 **{app_data['total_units']}** 個處理單元")

            # 統計卡片
            total_design = sum(len(u["design_params"]) for u in app_data["units"].values())
            total_measure = sum(len(u["measure_params"]) for u in app_data["units"].values())
            total_eq = sum(len(u["equipment"]) for u in app_data["units"].values())
            total_in = sum(len(u["influent"]) for u in app_data["units"].values())
            total_out = sum(len(u["effluent"]) for u in app_data["units"].values())

            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("處理單元", app_data["total_units"])
            c2.metric("設計參數", total_design)
            c3.metric("量測參數", total_measure)
            c4.metric("機具", total_eq)
            c5.metric("水質流向", f"進{total_in}/出{total_out}")

            # 單元清單
            st.subheader("📋 處理單元清單")
            unit_rows = []
            for code, info in sorted(app_data["units"].items()):
                unit_rows.append({
                    "代號": code,
                    "原始名稱": info["name_in_doc"],
                    "標準類型": info["std_tank"],
                    "代碼": info.get("code_id", ""),
                    "頁數": ", ".join(map(str, info["pages_found"][:3])),
                    "設計": len(info["design_params"]),
                    "量測": len(info["measure_params"]),
                    "機具": len(info["equipment"]),
                    "進": len(info["influent"]),
                    "出": len(info["effluent"]),
                })
            st.dataframe(unit_rows, use_container_width=True, hide_index=True)

            # 篩選器 — 看單元詳情
            st.subheader("🔎 單元詳情")
            unit_codes = sorted(app_data["units"].keys())
            selected = st.selectbox("選擇單元查看詳情", unit_codes)
            if selected:
                unit = app_data["units"][selected]
                c1, c2 = st.columns(2)
                with c1:
                    st.markdown(f"**單元代號**: {selected}")
                    st.markdown(f"**原始名稱**: {unit['name_in_doc']}")
                    st.markdown(f"**標準類型**: {unit['std_tank']}")
                    st.markdown(f"**內部代碼**: {unit.get('code_id', '')}")
                    st.markdown(f"**出現頁數**: {unit['pages_found']}")
                with c2:
                    if unit.get("design_params"):
                        st.markdown("**設計操作參數**:")
                        for k, v in unit["design_params"].items():
                            st.markdown(f"- {k}: `{v.get('raw', '')}`")
                    if unit.get("measure_params"):
                        st.markdown("**量測操作參數**:")
                        for k, v in unit["measure_params"].items():
                            st.markdown(f"- {k}: `{v.get('raw', '')}`")

                if unit.get("equipment"):
                    st.markdown("**相關機具設施**:")
                    eq_rows = [{"名稱": e["name"], "位置": e.get("位置", ""),
                                "數量": e.get("數量", ""), "馬力(kW)": e.get("馬力_kW", "")}
                               for e in unit["equipment"]]
                    st.dataframe(eq_rows, use_container_width=True, hide_index=True)

                if unit.get("influent"):
                    st.markdown(f"**進流水質** ({len(unit['influent'])} 流向)")
                    for infl_code, qdata in unit["influent"].items():
                        st.caption(f"📥 {infl_code}")
                        q_rows = [{"水質項目": k,
                                   "濃度": v.get("濃度", v.get("範圍", "")),
                                   "質量": v.get("質量", "")} for k, v in qdata.items()]
                        st.dataframe(q_rows, use_container_width=True, hide_index=True)

                if unit.get("effluent"):
                    st.markdown(f"**出流水質** ({len(unit['effluent'])} 流向)")
                    for effl_code, qdata in unit["effluent"].items():
                        st.caption(f"📤 {effl_code}")
                        q_rows = [{"水質項目": k,
                                   "濃度": v.get("濃度", v.get("範圍", "")),
                                   "質量": v.get("質量", "")} for k, v in qdata.items()]
                        st.dataframe(q_rows, use_container_width=True, hide_index=True)

            # 下載
            st.divider()
            st.subheader("📥 下載抽取結果")
            base_name = os.path.splitext(uploaded.name)[0]

            col1, col2 = st.columns(2)
            with col1:
                excel_buf = build_unit_excel(app_data)
                st.download_button(
                    "📊 下載各單元詳細 Excel",
                    data=excel_buf,
                    file_name=f"{base_name}_單元資料.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col2:
                json_str = json.dumps(app_data, ensure_ascii=False, indent=2)
                st.download_button(
                    "💾 下載 JSON 結構化資料",
                    data=json_str.encode("utf-8"),
                    file_name=f"{base_name}_抽取結果.json",
                    mime="application/json",
                    use_container_width=True,
                )

with tab2:
    st.subheader("規則庫內容")
    st.caption(f"目前載入 {len(rules)} 筆規則，分布於 {len(rules_by_tank)} 個槽體分類。")

    if rules:
        col1, col2 = st.columns([1, 3])
        with col1:
            selected_tank = st.selectbox("選擇槽體", ["(全部)"] + sorted(rules_by_tank.keys()))
        with col2:
            keyword = st.text_input("關鍵字搜尋")

        display_rules = rules
        if selected_tank != "(全部)":
            display_rules = rules_by_tank.get(selected_tank, [])
        if keyword:
            kw = keyword.lower()
            display_rules = [r for r in display_rules if kw in str(r).lower()]

        st.caption(f"顯示 {len(display_rules)} / {len(rules)} 筆")
        st.dataframe(
            [{
                "ID": r.get("缺失ID"),
                "槽體": r.get("標準槽體名稱"),
                "技師": r.get("技師姓名"),
                "檢查類型": r.get("檢查類型"),
                "對照項目": r.get("對照項目"),
                "規則": r.get("規則"),
                "原文": (r.get("原文缺失") or "")[:100],
            } for r in display_rules[:500]],
            use_container_width=True, hide_index=True,
        )

with tab3:
    st.subheader("水措審查系統 v2 — 使用說明")
    st.markdown("""
### v2 主要改進

| 項目 | v1 | v2 |
|------|----|----|
| 單元偵測覆蓋率 | 10/38 (26%) | **38/38 (100%)** |
| 單元類型歸類 | 多錯(放流池/廢水調整池) | **完全正確** |
| 設計參數抽取 | ❌ 無 | **✅ 17 筆** |
| 量測參數抽取 | ❌ 無 | **✅ 15 筆** |
| 機具設施抽取 | ❌ 無 | **✅ 83 筆** |
| 水質數據抽取 | ❌ 無 | **✅ 35 進流 / 35 出流** |
| 全形破折號處理 | ❌ 不支援 | **✅ 自動轉換** |

### 抽取邏輯

1. **掃描所有頁,找兩種關鍵區段**
   - 處理設施資料表（含「(一)處理單元名稱：xxx 序號：T01-01 代碼：120」）
   - 進出水質資料表（含「單元序號：T01-01」、「進流水流編號：WTB...」、「出流水流編號：WTA...」）
2. **解析設施資料表**: 抽 (二)設計參數、(三)量測參數、(四)機具設施
3. **解析水質表**: 抽各水質項目的濃度、質量、pH 範圍

### 已知限制

#### 🚫 水量平衡示意圖(頁 9-13)無法自動解析

原 PDF 該頁面是圖片,pdfplumber 抽不到文字。後續需引入:
- **OCR** (Tesseract / PaddleOCR) - 免費但中文準確度有限
- **Vision API** (Claude/GPT-4V) - 最準但需 API key

#### 🚫 比對引擎尚未針對新結構優化

目前比對引擎(step3)還在用舊邏輯,只機械化解析 pH 範圍。新版抽取已能提供:
- 完整設計參數(可比對停留時間、有效容量、加藥量等)
- 完整水質進出流(可做質量平衡檢查、檢測溶解性物質自行濃縮)
- 完整機具清單(可檢查液位計、流量計等是否齊備)

下一階段會擴充比對引擎,利用這些新增資訊做更深入的審查。
""")

st.divider()
st.caption("水措審查系統 v2 · 抽取器升級版 · [GitHub](https://github.com/jetenv02-lab/water-pollution-review)")
