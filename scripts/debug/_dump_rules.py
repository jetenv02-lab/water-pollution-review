import openpyxl
wb = openpyxl.load_workbook('規則庫.xlsx')
ws = wb['_槽體學理']
rows = []
for r in range(2, ws.max_row+1):
    tank = ws.cell(r, 1).value
    allow = ws.cell(r, 3).value or ''
    deny = ws.cell(r, 4).value or ''
    ttype = ws.cell(r, 14).value or ''
    status = ws.cell(r, 8).value or ''
    if not tank or status != 'V':
        continue
    rows.append(f'{tank} | {ttype} | 應變:{allow[:25]} | 不變:{str(deny)[:15]}')
open('_rules_dump.txt', 'w', encoding='utf-8').write('\n'.join(rows))
print('rows:', len(rows))
