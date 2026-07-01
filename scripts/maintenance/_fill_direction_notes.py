# -*- coding: utf-8 -*-
"""把「應變動方向備註」寫入 規則庫.xlsx 的 _槽體學理 分頁 U 欄。

設計:
  - U1 寫表頭「應變動方向備註」
  - U2~ 寫各槽體方向備註
  - 對齊 A 欄「標準槽體」名稱, 不依賴行號 (避免 row 順序變動)
  - 找不到對應的槽體 → 印警告但不中斷
"""
import openpyxl
from pathlib import Path

XLSX = Path(__file__).parent / "規則庫.xlsx"
SHEET = "_槽體學理"
NEW_COL_LETTER = "U"
NEW_COL_INDEX = 21  # U = 21st column
HEADER = "應變動方向備註"

# 方向備註 — 鍵: 標準槽體名稱; 值: 方向備註內容
DIRECTION_NOTES = {
    # 反應槽類
    "pH調整槽": "pH:↕(加酸鹼); 水溫:⇉",
    "pH調整池": "pH:↕; 水溫:⇉",
    "中和池": "pH:↕(中和至中性); 水溫:⇉",
    "pH調整暨快混池": "pH:↕; SS:↑(加PAC形成新固體); 水溫:⇉",
    "pH調整池暨快混池": "pH:↕; SS:↑; 水溫:⇉",
    "快混槽": "SS:↑(加PAC/絮凝劑); 水溫:⇉",
    "快混池": "SS:↑; 水溫:⇉",
    "慢混池": "SS:↑(加PAM絮凝, 顆粒變大但質量略增); 水溫:⇉",
    "慢混槽": "SS:↑; 水溫:⇉",
    "混凝膠凝池": "SS:↑(混凝形成絮體); 水溫:⇉",
    "氧化池": "COD:↓↓; BOD:↓↓; 真色色度:↓; pH:↕(加氧化劑改變)",
    "氰系氧化槽": "氰化物:↓↓(NaClO破氰); pH:↑(鹼性條件)",
    "鉻系還原槽": "六價鉻:↓↓(還原成3價); pH:↓(酸性條件)",
    "批次反應槽": "視批次操作而定 (酸/鹼/氧化/還原皆可能)",

    # 分離槽類
    "沉澱池": "SS:↓↓(主流出口); SS:↑(污泥支線); 重金屬:↓(隨絮體沉降); COD/BOD:↓(伴隨SS)",
    "沉降池": "SS:↓↓; SS污泥支線:↑; 重金屬:↓; COD/BOD:↓",
    "浮除槽": "SS:↓↓; 油脂:↓↓(主流出口); 浮渣支線:↑",
    "砂濾塔": "SS:↓↓(過濾); 濁度:↓↓; 反洗水支線:SS↑",
    "砂濾器": "SS:↓↓; 濁度:↓↓; 反洗水:SS↑",
    "活性碳吸附塔": "COD:↓; BOD:↓; 色度:↓↓; 異味:↓↓; 游離氯:↓↓; 微量有機物:↓↓",
    "活性碳吸附裝置": "COD:↓; BOD:↓; 色度:↓↓; 異味:↓↓; 游離氯:↓↓; 微量有機物:↓↓",
    "離子交換樹脂塔": "重金屬離子:↓↓(Cu/Ni/Zn/Cr/Cd); 硬度離子(Ca/Mg):↓; 反洗液支線:↑",
    "油脂分離槽": "油脂:↓↓(主流); 油脂渣支線:↑",
    "預處理池": "SS:↓(粗篩); 重金屬:↓(初步去除); 油脂:↓; 水溫:↕",

    # 生物處理類
    "曝氣槽": "BOD:↓↓; COD:↓; 氨氮:↓(硝化); DO:↑(鼓風); MLSS:↑(微生物增生)",
    "厭氧池": "BOD:↓↓; COD:↓↓; 沼氣:↑; pH:↕(可能酸化)",
    "接觸氧化池": "COD:↓; BOD:↓; 氨氮:↓",
    "MBR": "COD:↓↓; BOD:↓↓; 氨氮:↓↓; SS:↓↓(膜過濾, 出水極低)",

    # 污泥/儲存類
    "脫水機": "含水率:↓↓(脫水至70~85%); 濾液:含水率↑",
    "污泥離心式脫水機": "含水率:↓↓; 濾液:↑",
    "污泥帶濾式脫水機": "含水率:↓↓; 濾液:↑",
    "濃縮槽": "含水率:↓(濃縮至95~97%); 上澄液:↑",
    "污泥濃縮設施": "含水率:↓; 上澄液:↑",
    "污泥儲槽": "含水率:可能↓(沉降進一步濃縮); 一般⇉",
    "濾液池": "SS:↕(濾液SS可能高於主流); 含水率:↑",

    # 通用/中性類
    "暫存槽": "⇉ (應守恆, 僅暫存)",
    "貯留槽": "⇉ (應守恆)",
    "廢水調整池": "pH:↕(均勻化); 其他⇉",
    "廢水收集池": "⇉ (僅收集匯流)",
    "調勻池": "pH:↕; 水質波動:↓(均勻化); 質量⇉",
    "放流池": "⇉ (放流前緩衝)",
}


def main():
    if not XLSX.exists():
        print(f"[X] 找不到 {XLSX}")
        return

    wb = openpyxl.load_workbook(XLSX)
    if SHEET not in wb.sheetnames:
        print(f"[X] {XLSX} 沒有分頁 {SHEET}")
        return

    ws = wb[SHEET]

    # 檢查 U 欄是否已有表頭
    existing_header = ws.cell(1, NEW_COL_INDEX).value
    if existing_header and existing_header != HEADER:
        print(f"[!] U1 已有內容: '{existing_header}', 不覆蓋")
    else:
        ws.cell(1, NEW_COL_INDEX, HEADER)
        print(f"[OK] 寫入表頭 U1 = '{HEADER}'")

    # 對齊 A 欄, 寫入備註
    filled = 0
    missing = []
    for r in range(2, ws.max_row + 1):
        tank = ws.cell(r, 1).value
        status = ws.cell(r, 8).value or ""
        if not tank:
            continue
        if tank in DIRECTION_NOTES:
            existing = ws.cell(r, NEW_COL_INDEX).value
            if existing:
                print(f"[skip] R{r} {tank} 已有備註: {existing[:30]}")
                continue
            ws.cell(r, NEW_COL_INDEX, DIRECTION_NOTES[tank])
            filled += 1
        elif status == "V":
            missing.append((r, tank))

    wb.save(XLSX)
    print(f"\n[OK] 已寫入 {filled} 條方向備註到 {XLSX}")

    if missing:
        print(f"\n[!] 以下 active 槽體未在備註表中 (可能要補):")
        for r, tank in missing:
            print(f"    R{r}: {tank}")


if __name__ == "__main__":
    main()
