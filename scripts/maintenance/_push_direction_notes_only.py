# -*- coding: utf-8 -*-
"""只推「應變動方向備註」U 欄到 Google Sheets 的 _槽體學理 分頁。

不動其他欄、不動其他分頁、不重算統計。
"""
import os
import openpyxl

import sheets_sync as ss

XLSX = os.path.join(os.path.dirname(__file__), "規則庫.xlsx")
SHEET = "_槽體學理"
COL_LETTER = "U"
COL_INDEX = 21


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    # 蒐集 A 欄 + U 欄 (含表頭)
    rows_to_push = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, COL_INDEX).value
        rows_to_push.append([val if val is not None else ""])
    print(f"準備推 {len(rows_to_push)} 列到 {SHEET}!{COL_LETTER}1:{COL_LETTER}{len(rows_to_push)}")

    # 用 sheets_sync 內部 client
    client, source = ss._get_gspread_client()
    print(f"認證來源: {source}")

    sh = client.open_by_key(ss.DEFAULT_SHEET_ID)
    ws_sheet = sh.worksheet(SHEET)

    # 先確認/擴充欄數 (Sheet 預設可能少於 21)
    if ws_sheet.col_count < COL_INDEX:
        print(f"Sheet 目前 {ws_sheet.col_count} 欄, 擴充到 {COL_INDEX} 欄")
        ws_sheet.add_cols(COL_INDEX - ws_sheet.col_count)

    # gspread 用 A1 格式更新
    cell_range = f"{COL_LETTER}1:{COL_LETTER}{len(rows_to_push)}"
    ws_sheet.update(values=rows_to_push, range_name=cell_range)
    print(f"[OK] 已推 U 欄 ({len(rows_to_push)} 列) 到 Sheets")
    print(f"URL: https://docs.google.com/spreadsheets/d/{ss.DEFAULT_SHEET_ID}")


if __name__ == "__main__":
    main()
