# -*- coding: utf-8 -*-
"""把 3 個放流標準分頁推到 Sheets (不動其他分頁)。"""
import os
import openpyxl
import sheets_sync as ss

XLSX = os.path.join(os.path.dirname(__file__), "規則庫.xlsx")
SHEETS = ["_放流標準_附表五", "_放流標準_附表十六", "_放流標準_附表四"]


def main():
    wb = openpyxl.load_workbook(XLSX)
    client, source = ss._get_gspread_client()
    print(f"認證來源: {source}")
    sh = client.open_by_key(ss.DEFAULT_SHEET_ID)

    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        # 讀所有列
        data = []
        for r in range(1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            row = [v if v is not None else "" for v in row]
            data.append(row)

        # Sheets 若無此分頁, 新增
        existing_names = [w.title for w in sh.worksheets()]
        if sheet_name not in existing_names:
            ws_sheet = sh.add_worksheet(title=sheet_name, rows=len(data) + 20, cols=len(data[0]))
            print(f"[NEW] 建 Sheets 分頁 {sheet_name}")
        else:
            ws_sheet = sh.worksheet(sheet_name)
            # 確保尺寸夠
            need_rows = len(data)
            need_cols = len(data[0])
            if ws_sheet.row_count < need_rows:
                ws_sheet.add_rows(need_rows - ws_sheet.row_count)
            if ws_sheet.col_count < need_cols:
                ws_sheet.add_cols(need_cols - ws_sheet.col_count)
            # 清舊資料
            ws_sheet.clear()

        # 一次寫入
        cell_range = f"A1:{chr(64 + len(data[0]))}{len(data)}"
        ws_sheet.update(values=data, range_name=cell_range)
        print(f"[OK] 推 {sheet_name} ({len(data)} 列)")

    print(f"\nURL: https://docs.google.com/spreadsheets/d/{ss.DEFAULT_SHEET_ID}")


if __name__ == "__main__":
    main()
