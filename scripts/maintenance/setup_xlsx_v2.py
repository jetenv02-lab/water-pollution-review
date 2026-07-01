# -*- coding: utf-8 -*-
"""一次性腳本: 把 規則庫.xlsx 升級成 v2 結構。

執行前: **請先關閉 Excel**!

升級內容:
1. 每個槽體分頁加「狀態」欄 (預設空白)
2. 重寫 `_說明` 分頁, 給同事完整使用指引
3. 自動備份原 xlsx 到 backup/

用法:
    python setup_xlsx_v2.py
"""
import os
import shutil
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE, "規則庫.xlsx")
BACKUP_DIR = os.path.join(BASE, "backup")

# _說明 分頁完整內容
INSTRUCTION_LINES = [
    ["水措審查系統 — 規則庫 使用說明 (v2)"],
    [""],
    ["═══════════════════════════════════════════════════════════════"],
    ["這份規則庫是什麼?"],
    ["═══════════════════════════════════════════════════════════════"],
    ["• 收錄歷次水污染防治措施審查意見書中, 環工技師標註的缺失項目"],
    ["• 系統會用這些規則去自動審查新申請的水措 PDF"],
    ["• 每個處理單元 (槽體) 用一個分頁, 例如「快混槽」「沉澱池」"],
    ["• 4 個 meta 分頁 (開頭是底線 _ 的): _說明 / _來源清單 / _槽體對照表 / _審查紀錄"],
    [""],
    ["═══════════════════════════════════════════════════════════════"],
    ["欄位說明"],
    ["═══════════════════════════════════════════════════════════════"],
    ["欄位", "說明"],
    ["缺失ID", "唯一編號 D001 ~ D299 (新增時留空, 系統下載時自動補)"],
    ["來源", "出自哪份審查意見 (S01 / S02 …)"],
    ["原文缺失", "技師原文 (盡量保留, 不要改寫)"],
    ["檢查類型", "設計參數 / 機具設施 / 質量平衡 / 操作條件 / 流向示意圖 / 水質標準 / 文件一致性 / 去除率 / 其他"],
    ["對照項目", "這條規則針對什麼 (例如「反洗水來源」「pH」「攪拌轉速」)"],
    ["規則(萃取/可比對判斷式)", "規則白話描述 (例: 反洗水來源未標示於流向示意圖)"],
    ["比對位置(依標題,非頁碼)", "在申請文件的哪段去查 (例: 廢(污)水產生與水污染防治措施流向示意圖)"],
    ["判定邏輯(條件→結論)", "什麼條件下標記什麼 (例: 若 設備具反洗功能 且 未標示來源 → 標記:未標示來源)"],
    ["技師姓名", "原文出自哪位技師"],
    ["序號", "原文序號 (例: 序1 方天志技師 (1))"],
    ["原始槽體代號", "原文寫的槽體序號 (例: T01-05) — 可能跟標準名稱不一致"],
    ["狀態", "規則狀態: 空白=已核可 / ? = 待討論 (系統會跳過)"],
    [""],
    ["═══════════════════════════════════════════════════════════════"],
    ["狀態欄怎麼標 ?"],
    ["═══════════════════════════════════════════════════════════════"],
    ["狀態值", "意義", "系統行為", "什麼時候用"],
    ["(空白)", "預設 = V = 已核可", "✓ 跑這條規則", "規則沒問題, 不用動"],
    ["V", "已核可 (跟空白相同)", "✓ 跑這條規則", "你想顯式標記「我看過了沒問題」"],
    ["?", "有疑慮 / 待討論", "⊗ 跳過 (不報錯)", "不確定這條對不對, 先擱著"],
    [""],
    ["• 想刪除 (X)? → 直接整列刪掉 (Excel/Sheet 上右鍵刪除整列), 不用標 X"],
    ["• 狀態欄是 Sheet 上的下拉選單, 只能選 V / ? / 空白"],
    [""],
    ["═══════════════════════════════════════════════════════════════"],
    ["怎麼校對規則 ?"],
    ["═══════════════════════════════════════════════════════════════"],
    ["1. 規則被分到錯誤的槽體 ?"],
    ["   → 整列剪下 (含「缺失ID」) → 貼到正確的分頁"],
    ["   → 為了讓自己事後好認, 可以把該列文字標紅色字體 (純視覺, 程式不會讀)"],
    [""],
    ["2. 規則內容有疑慮 ?"],
    ["   → 在「狀態」欄選「?」"],
    ["   → 必要時在「原文缺失」末尾加說明 (例: 「[待討論] 是否應該包含 …」)"],
    [""],
    ["3. 規則是錯的, 該刪除 ?"],
    ["   → 整列刪除"],
    [""],
    ["4. 有新的審查意見, 想新增規則 ?"],
    ["   → 在對應槽體分頁的最下方加新列"],
    ["   → 缺失ID 留空 (系統下次同步時會自動補 D300, D301 ...)"],
    ["   → 必填欄位: 來源 / 原文缺失 / 檢查類型 / 對照項目 / 規則 / 判定邏輯 / 標準槽體名稱"],
    [""],
    ["═══════════════════════════════════════════════════════════════"],
    ["槽體分頁清單"],
    ["═══════════════════════════════════════════════════════════════"],
    ["這份規則庫目前涵蓋以下 25 種槽體:"],
    ["• 預處理: 廢水收集池 / 廢水調整池 / 調勻池 / 暫存槽 / 中和池"],
    ["• 化學處理: pH調整槽 / 快混槽 / 慢混池 / 沉澱池"],
    ["• 生物處理: 曝氣槽 / 氧化池"],
    ["• 高級處理: 活性碳吸附塔 / 活性碳吸附裝置 / 砂濾塔 / 離子交換樹脂塔"],
    ["• 污泥處理: 濃縮槽 / 污泥儲槽 / 脫水機"],
    ["• 排放: 放流池 / 貯留槽"],
    ["• 批次: 批次反應槽"],
    ["• 其他: (文件類) / (現場設備類) / 文件類 / 現場設備類"],
    [""],
    ["新增槽體分頁: 請通知系統管理員 (Nick), 因為要對應到「標準槽體名稱」"],
    [""],
    ["═══════════════════════════════════════════════════════════════"],
    ["重要 — 同步時注意"],
    ["═══════════════════════════════════════════════════════════════"],
    ["• 你們編輯的是 Google Sheets (不是 Excel)"],
    ["• Nick 會定期按「下載 Sheet → xlsx」, 你們的編輯就會落到主檔"],
    ["• 改完跟 Nick 說一聲, 他才知道何時可以拉回"],
    ["• 同時多人編輯沒問題, Sheet 自動合併"],
    [""],
    ["• 不要動 「缺失ID」欄 (除非是新增列, 才能留空)"],
    ["• 不要改分頁名稱 (= 標準槽體名稱, 程式靠這個分類)"],
    ["• 不要改第 1 列 (= 表頭)"],
    [""],
    ["═══════════════════════════════════════════════════════════════"],
    ["問題回報"],
    ["═══════════════════════════════════════════════════════════════"],
    ["• 系統 / 規則撰寫指引: https://github.com/jetenv02-lab/water-pollution-review"],
    ["• 線上版審查系統: https://water-review.streamlit.app/"],
    ["• 規則撰寫指引: RULE_AUTHORING.md (見 GitHub)"],
]


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"找不到 {XLSX_PATH}")
        return 1

    # 0) 確認檔案沒被開啟
    try:
        with open(XLSX_PATH, "r+b"):
            pass
    except PermissionError:
        print("❌ 規則庫.xlsx 被 Excel 鎖住了, 請先關閉 Excel 再執行")
        return 1

    # 1) 備份
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"規則庫_{ts}_pre_v2.xlsx")
    shutil.copy2(XLSX_PATH, backup_path)
    print(f"✓ 備份原檔 → {backup_path}")

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(XLSX_PATH)

    # 2) 處理每個槽體分頁: 加「狀態」欄
    tank_sheets_count = 0
    added_count = 0
    for sn in wb.sheetnames:
        if sn.startswith("_"):
            continue
        ws = wb[sn]
        if ws.max_row < 1:
            continue
        # 讀現有表頭
        headers = [c.value for c in ws[1]]
        if "狀態" in headers:
            print(f"  [{sn}] 已有「狀態」欄, 略過")
        else:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value="狀態")
            # 表頭設粗體
            ws.cell(row=1, column=new_col).font = Font(bold=True)
            added_count += 1
            print(f"  [{sn}] 加「狀態」欄 (第 {new_col} 欄)")
        tank_sheets_count += 1

    print(f"✓ 已處理 {tank_sheets_count} 個槽體分頁, 新增「狀態」欄 {added_count} 處")

    # 3) 重寫 _說明 分頁
    if "_說明" in wb.sheetnames:
        wb.remove(wb["_說明"])
    # 把 _說明 放第一個位置
    ws_intro = wb.create_sheet("_說明", 0)
    for row in INSTRUCTION_LINES:
        ws_intro.append(row)
    # 標題列加粗加底色
    title_fill = PatternFill("solid", fgColor="FFE4B5")
    ws_intro["A1"].font = Font(bold=True, size=14)
    ws_intro["A1"].fill = title_fill
    # 自動換行 + 適度欄寬
    ws_intro.column_dimensions["A"].width = 35
    ws_intro.column_dimensions["B"].width = 50
    ws_intro.column_dimensions["C"].width = 30
    ws_intro.column_dimensions["D"].width = 35
    for row in ws_intro.iter_rows(min_row=1, max_row=ws_intro.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    print(f"✓ 重寫 _說明 分頁 ({len(INSTRUCTION_LINES)} 行)")

    # 4) 存檔
    wb.save(XLSX_PATH)
    print(f"\n✅ 完成! 規則庫.xlsx 已升級為 v2 結構")
    print(f"   - 備份: {backup_path}")
    print(f"   - 新增 25 個槽體分頁的「狀態」欄")
    print(f"   - 重寫 _說明 分頁")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
