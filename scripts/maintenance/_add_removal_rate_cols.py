# -*- coding: utf-8 -*-
"""_槽體學理 加 3 欄 (Y/Z/AA) 削減率範圍.

新增欄位:
  Y  主要削減項目          — 該槽最重要的 1~3 項水質, 用「;」分隔
  Z  削減率範圍(%)         — 對照 Y 順序, 用「;」分隔的 min~max
  AA 削減率學理依據         — 一句話備註

不動既有 24 欄 (U 應變動方向備註 保留當前內容, 是「方向」的 sugar).

推 Sheets 用: scripts/maintenance/_push_removal_rate_cols.py
step3e 讀 Y/Z 判斷去除率是否在合理範圍.
"""
import os
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))

import openpyxl

XLSX = REPO_ROOT / "規則庫.xlsx"
SHEET = "_槽體學理"

NEW_HEADERS = {
    25: "主要削減項目",
    26: "削減率範圍(%)",
    27: "削減率學理依據",
}

# 資料: 由 Nick 學理討論確認的 40+ 條
# 格式: 標準槽體 → (主要削減項目, 削減率範圍, 學理依據)
# 空 tuple = 該槽本來就不做去除 (儲存/中性類)
RATES = {
    # ═══ 反應類 (無分離功能, D054 學理: 不應表現重金屬去除) ═══
    "pH調整槽":         ("", "", "純調 pH, 不去除任何項目"),
    "pH調整池":         ("", "", "純調 pH, 不去除任何項目"),
    "中和池":           ("", "", "純中和, 不去除任何項目"),
    "pH調整暨快混池":   ("SS", "0~5 (加PAC反而略升)", "調pH + 快混, 加PAC後 SS 略升, 未沉澱不去除重金屬"),
    "pH調整池暨快混池": ("SS", "0~5 (加PAC反而略升)", "同上"),
    "快混槽":           ("SS", "0~5 (加PAC反而略升)", "D054 學理: 快混無沉澱功能, 不應表現重金屬去除"),
    "快混池":           ("SS", "0~5 (加PAC反而略升)", "同快混槽"),
    "慢混池":           ("SS", "0~5 (加PAM 顆粒變大但未沉澱)", "絮凝但未分離, 質量守恆"),
    "慢混槽":           ("SS", "0~5", "同慢混池"),
    "混凝膠凝池":       ("SS", "0~5 (絮體形成但未沉澱)", "只形成絮體, 未分離"),
    "氧化池":           ("COD;BOD;真色色度", "30~70;30~70;30~60", "化學氧化分解有機物"),
    "氰系氧化槽":       ("氰化物;COD", "90~99;20~50", "NaClO 破氰: CN⁻ → CNO⁻ → N₂↑"),
    "鉻系還原槽":       ("六價鉻", "95~99", "NaHSO₃ 還原 Cr⁶⁺ → Cr³⁺ (總鉻不去除, 只變價)"),
    "批次反應槽":       ("", "", "視批次配方而定, 無固定去除率"),

    # ═══ 分離類 - 沉澱/浮除 (重金屬去除主力) ═══
    "沉澱池":     ("SS;銅;鎳;鋅;鉛;總鉻;COD;BOD",
                    "80~95;60~90;50~90;70~95;60~90;60~90;20~50;20~40",
                    "重力沉降 + 絮體氫氧化物沉澱, COD/BOD 隨 SS 略降"),
    "沉降池":     ("SS;銅;鎳;鋅;鉛;總鉻;COD;BOD",
                    "80~95;60~90;50~90;70~95;60~90;60~90;20~50;20~40",
                    "同沉澱池"),
    "浮除槽":     ("油脂;SS;COD", "80~98;60~90;20~50",
                    "加壓浮除, 油脂主力, SS 伴隨"),
    "油脂分離槽": ("油脂", "70~95", "重力分離, 只降油脂"),

    # ═══ 分離類 - 過濾/吸附 ═══
    "砂濾塔": ("SS;濁度", "70~95;70~95",
                 "顆粒物理過濾, 溶解性物質不去除, 反洗水 SS↑↑"),
    "砂濾器": ("SS;濁度", "70~95;70~95", "同砂濾塔"),
    "活性碳吸附塔": ("COD;BOD;色度;異味;游離氯",
                     "30~80;30~70;70~95;80~98;90~99",
                     "吸附有機物/色度/餘氯, SS 不應去除 (會堵塞)"),
    "活性碳吸附裝置": ("COD;BOD;色度;異味;游離氯",
                       "30~80;30~70;70~95;80~98;90~99",
                       "同活性碳吸附塔"),
    "離子交換樹脂塔": ("銅;鎳;鋅;鉻;鎘;硬度離子",
                       "90~99;90~99;90~99;90~99;90~99;50~90",
                       "離子交換, 對特定金屬去除極高"),

    # ═══ 分離類 - 其他 ═══
    "預處理池": ("SS;油脂", "20~60;30~70",
                   "粗篩/沉砂, 重金屬未沉澱不去除"),

    # ═══ 生物類 ═══
    "曝氣槽": ("BOD;COD;氨氮", "85~98;60~85;70~95",
                 "活性污泥好氧分解, 氨氮硝化, MLSS 微生物自增"),
    "接觸氧化池": ("BOD;COD;氨氮", "60~90;50~80;50~90",
                      "生物膜好氧分解"),
    "MBR": ("BOD;COD;氨氮;SS", "85~99;70~95;80~99;95~99",
              "生物 + 膜過濾, SS 極低"),
    "厭氧池": ("BOD;COD", "70~90;60~85", "厭氧分解產沼氣, pH 可能酸化"),

    # ═══ 污泥類 ═══
    "脫水機":            ("含水率", "8~15 (從 97% → 82~89%)", "機械脫水"),
    "污泥離心式脫水機": ("含水率", "12~22 (從 97% → 75~85%)", "離心高速脫水"),
    "污泥帶濾式脫水機": ("含水率", "8~17 (從 97% → 80~89%)", "帶濾中速脫水"),
    "濃縮槽":            ("含水率", "1~3 (從 97% → 94~96%)", "重力濃縮"),
    "污泥濃縮設施":     ("含水率", "1~3 (從 97% → 94~96%)", "同濃縮槽"),

    # ═══ 儲存/中性類 (不做去除) ═══
    "污泥儲槽":       ("", "", "儲存, 質量守恆; 略沉降可能微降含水率"),
    "暫存槽":         ("", "", "應守恆"),
    "貯留槽":         ("", "", "應守恆"),
    "調勻池":         ("", "", "均勻化水質波動, 質量守恆"),
    "廢水調整池":     ("", "", "同調勻池"),
    "廢水收集池":     ("", "", "僅匯流"),
    "放流池":         ("", "", "放流前緩衝, 應守恆"),
    "濾液池":         ("", "", "回收濾液, 質量守恆但含水率高"),
}


def main():
    if not XLSX.exists():
        print(f"[X] 找不到 {XLSX}")
        return

    wb = openpyxl.load_workbook(XLSX)
    if SHEET not in wb.sheetnames:
        print(f"[X] {XLSX} 沒 {SHEET}")
        return

    ws = wb[SHEET]
    print(f"[1/3] {SHEET} 現況: {ws.max_row} 列 x {ws.max_column} 欄")

    # 寫表頭
    for col, header in NEW_HEADERS.items():
        existing = ws.cell(1, col).value
        if existing and existing != header:
            print(f"[!] R1 C{col} 已有 '{existing}', 不覆蓋")
            continue
        ws.cell(1, col, header)
        print(f"  寫入 R1 C{col} = '{header}'")

    print()
    filled = 0
    missing = []
    for r in range(2, ws.max_row + 1):
        tank = ws.cell(r, 1).value
        status = ws.cell(r, 8).value or ""
        if not tank:
            continue
        if tank in RATES:
            existing = ws.cell(r, 25).value
            if existing not in (None, ""):
                print(f"[skip] R{r} {tank} 已有削減率")
                continue
            items, rates, note = RATES[tank]
            ws.cell(r, 25, items if items else None)
            ws.cell(r, 26, rates if rates else None)
            ws.cell(r, 27, note if note else None)
            filled += 1
        elif status == "V":
            missing.append((r, tank))

    wb.save(XLSX)
    print(f"[2/3] 已寫入 {filled} 條削減率資料")
    if missing:
        print(f"\n[!] 未在削減率表中的 active 槽體:")
        for r, tank in missing:
            print(f"    R{r}: {tank}")

    print(f"\n[3/3] 完成. {XLSX}")


if __name__ == "__main__":
    main()
