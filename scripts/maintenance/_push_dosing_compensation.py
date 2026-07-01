# -*- coding: utf-8 -*-
"""把 _加藥規則 的 R/S/T 三欄 (典型劑量/引入項目/轉換係數) 推到 Sheets。

只動這 3 欄, 不動其他欄/分頁/統計。
"""
import os
import openpyxl
import sheets_sync as ss

XLSX = os.path.join(os.path.dirname(__file__), "規則庫.xlsx")
SHEET = "_加藥規則"
COLS = [("R", 18), ("S", 19), ("T", 20)]


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    # 對齊 sheets API
    client, source = ss._get_gspread_client()
    print(f"認證來源: {source}")
    sh = client.open_by_key(ss.DEFAULT_SHEET_ID)
    ws_sheet = sh.worksheet(SHEET)

    # 確保 Sheet 欄數 >= 20
    if ws_sheet.col_count < 20:
        print(f"Sheet 目前 {ws_sheet.col_count} 欄, 擴充到 20 欄")
        ws_sheet.add_cols(20 - ws_sheet.col_count)

    # 一次推 3 欄: R 1..max_row, S 1..max_row, T 1..max_row
    for letter, idx in COLS:
        rows_to_push = []
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, idx).value
            rows_to_push.append([val if val is not None else ""])
        cell_range = f"{letter}1:{letter}{len(rows_to_push)}"
        ws_sheet.update(values=rows_to_push, range_name=cell_range)
        print(f"[OK] 推 {letter} 欄 ({len(rows_to_push)} 列)")

    print(f"\nURL: https://docs.google.com/spreadsheets/d/{ss.DEFAULT_SHEET_ID}")


if __name__ == "__main__":
    main()
