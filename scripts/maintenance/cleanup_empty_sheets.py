# -*- coding: utf-8 -*-
"""清掉規則庫.xlsx 裡的空殼分頁。

背景: 歷史上某次建表流程留下了「文件類」「現場設備類」兩個空分頁,
       但真實規則都在帶括號的「(文件類)」「(現場設備類)」裡。
       為避免混淆, 把空殼刪掉。

執行:
    python cleanup_empty_sheets.py        # dry run (只列出, 不動)
    python cleanup_empty_sheets.py --apply # 真的刪
"""
import sys
import openpyxl

TARGETS = ["文件類", "現場設備類"]   # 不帶括號的空殼


def main():
    apply = "--apply" in sys.argv
    wb = openpyxl.load_workbook("規則庫.xlsx")

    to_delete = []
    for name in TARGETS:
        if name not in wb.sheetnames:
            print(f"  [SKIP] '{name}' 不存在")
            continue
        ws = wb[name]
        n_data = sum(
            1 for r in ws.iter_rows(min_row=2, values_only=True)
            if any(c not in (None, "") for c in r)
        )
        if n_data == 0:
            to_delete.append(name)
            print(f"  [DEL]  '{name}' (0 筆資料, 將刪除)")
        else:
            print(f"  [KEEP] '{name}' ({n_data} 筆資料, 不刪)")

    if not to_delete:
        print("\n沒有需要刪除的分頁。")
        return

    if not apply:
        print(f"\n(dry-run) 預計刪除 {len(to_delete)} 個分頁: {to_delete}")
        print("加 --apply 才會真的執行。")
        return

    for name in to_delete:
        del wb[name]
        print(f"  已刪除: {name}")
    wb.save("規則庫.xlsx")
    print(f"\n完成! 已寫回 規則庫.xlsx (剩餘 {len(wb.sheetnames)} 個分頁)")
    print("\n下一步:")
    print("  1. 開 Streamlit App 的「🔄 規則庫管理」")
    print("  2. 按「⬆️ 上傳 xlsx → 協作表」, 讓雲端也同步刪掉空殼")


if __name__ == "__main__":
    main()
