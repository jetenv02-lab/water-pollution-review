# -*- coding: utf-8 -*-
"""一次性腳本: 把 step3f_synonyms.DEFAULT_SYNONYMS 寫進 規則庫.xlsx 的 _同義字 分頁。

執行前: 請先關閉 Excel
用法: python setup_synonyms_sheet.py
"""
import os
import shutil
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE, "規則庫.xlsx")
BACKUP_DIR = os.path.join(BASE, "backup")


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"找不到 {XLSX_PATH}")
        return 1

    try:
        with open(XLSX_PATH, "r+b"):
            pass
    except PermissionError:
        print("❌ 規則庫.xlsx 被 Excel 鎖住, 請先關閉 Excel")
        return 1

    # 備份
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = os.path.join(BACKUP_DIR, f"規則庫_{ts}_pre_synonyms.xlsx")
    shutil.copy2(XLSX_PATH, backup)
    print(f"✓ 備份 → {backup}")

    import step3f_synonyms
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(XLSX_PATH)

    # 移除舊的 _同義字 (若有)
    if "_同義字" in wb.sheetnames:
        wb.remove(wb["_同義字"])

    # 找 _說明 的位置, 把 _同義字 放在 _說明 之後 (跟其他 meta 分頁相鄰)
    insert_pos = 0
    for i, sn in enumerate(wb.sheetnames):
        if sn.startswith("_"):
            insert_pos = i + 1

    ws = wb.create_sheet("_同義字", insert_pos)

    # 表頭
    ws.append(["標準詞", "別名 (用 / 分隔)", "說明"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFE4B5")
        cell.alignment = Alignment(horizontal="center")

    # 寫資料
    syn_data = step3f_synonyms.DEFAULT_SYNONYMS
    # 分類提示 (照原本 dict 的順序)
    sections = [
        (["出水高度", "有效容量", "有效水深", "槽體尺寸"], "── 槽體尺寸 ──"),
        (["滯留時間", "曝氣量", "迴流比", "MLSS", "DO", "SVI", "F/M"], "── 操作參數 ──"),
        (["液位計", "pH計", "DO計", "流量計", "攪拌機", "曝氣機",
          "加藥泵", "刮泥機", "污泥泵"], "── 機具設施 ──"),
        (["反洗水", "加藥", "中和", "混凝", "膠凝"], "── 化學處理 ──"),
        (["BOD", "COD", "SS", "TKN", "TP", "TN"], "── 水質參數 ──"),
        (["進流水", "出流水", "原廢水", "放流口"], "── 流向 ──"),
        (["污泥含水率", "脫水", "濃縮"], "── 污泥處理 ──"),
    ]

    section_fill = PatternFill("solid", fgColor="E0E0E0")
    written_terms = set()

    for terms, label in sections:
        # 區塊標題列
        ws.append([label, "", ""])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, italic=True)
            cell.fill = section_fill

        for std in terms:
            if std not in syn_data:
                continue
            aliases = syn_data[std]
            ws.append([std, " / ".join(aliases), ""])
            written_terms.add(std)

    # 其他沒列在分類裡的, 補在最後
    misc = [k for k in syn_data if k not in written_terms]
    if misc:
        ws.append(["── 其他 ──", "", ""])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, italic=True)
            cell.fill = section_fill
        for std in misc:
            ws.append([std, " / ".join(syn_data[std]), ""])

    # 加 _說明 段
    ws.append(["", "", ""])
    ws.append(["── 使用說明 ──", "", ""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    instructions = [
        ["標準詞", "系統內部使用的標準詞彙", ""],
        ["別名", "技師可能用的不同寫法 (用 / 分隔, 大小寫不分)", ""],
        ["", "系統會自動把別名比對到標準詞", ""],
        ["範例", "「液面到槽頂距離」會自動對應到「出水高度」", ""],
        ["新增", "在最下方加新列就行, 不用刻意分類", ""],
        ["", "系統下次同步時自動更新", ""],
    ]
    for row in instructions:
        ws.append(row)

    # 欄寬
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30

    # 自動換行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"\n✅ 完成! _同義字 分頁已加到 規則庫.xlsx")
    print(f"   - {len(syn_data)} 個標準詞")
    print(f"   - 約 {sum(len(v) for v in syn_data.values())} 個別名")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
