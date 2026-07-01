# -*- coding: utf-8 -*-
"""把 V/W/X 三欄 (RPM_min/RPM_max/RPM 備註) 推到 Sheets _槽體學理。"""
import os
import openpyxl
import sheets_sync as ss

XLSX = os.path.join(os.path.dirname(__file__), "規則庫.xlsx")
SHEET = "_槽體學理"
COLS = [("V", 22), ("W", 23), ("X", 24)]


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    client, source = ss._get_gspread_client()
    print(f"認證來源: {source}")
    sh = client.open_by_key(ss.DEFAULT_SHEET_ID)
    ws_sheet = sh.worksheet(SHEET)

    if ws_sheet.col_count < 24:
        print(f"Sheet 目前 {ws_sheet.col_count} 欄, 擴充到 24 欄")
        ws_sheet.add_cols(24 - ws_sheet.col_count)

    for letter, idx in COLS:
        rows_to_push = []
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, idx).value
            rows_to_push.append([val if val is not None else ""])
        cell_range = f"{letter}1:{letter}{len(rows_to_push)}"
        ws_sheet.update(values=rows_to_push, range_name=cell_range)
        print(f"[OK] 推 {letter} 欄 ({len(rows_to_push)} 列)")

    print(f"\nURL: https://docs.google.com/spreadsheets/d/{ss.DEFAULT_SHEET_ID}")


if __name__ == "__main__":
    main()
