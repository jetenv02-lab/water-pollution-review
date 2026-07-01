# -*- coding: utf-8 -*-
"""把「典型 mg/L 劑量 + 引入水質項目 + 轉換係數」3 個新欄加到 _加藥規則。

新增欄位:
  R 典型劑量 (mg/L)  — 對進流水的典型劑量
  S 引入水質項目     — 加藥會引入哪些水質項目 (例: SS, 氯離子)
  T 轉換係數         — 每 kg 藥劑變多少 kg 該項目 (0~1)

不修改既有 17 欄, 不改其他分頁。
規則: 一條 _加藥規則 R 列 → 一個 (藥劑, 引入項目) pair 的補償資訊
若一個藥引入多個項目 (例 PAC → SS + Cl), 用「;」分隔多個項目, 對應分隔多個係數。

實際運作: chemical_calc 之後讀這 3 欄, 算 in_mass + 加藥 vs out_mass
"""
import openpyxl
from pathlib import Path

XLSX = Path(__file__).parent / "規則庫.xlsx"
SHEET = "_加藥規則"

# 新欄位定義
NEW_HEADERS = {
    18: "典型劑量 (mg/L)",
    19: "引入水質項目",
    20: "轉換係數",
}

# 對既有 20 條規則, 各填的補償資訊
# Key: (分流系統, 槽體類別, 藥劑) — 用 3 欄 tuple 對齊
# Value: (典型 mg/L, 引入項目串, 轉換係數串)
# - 引入項目串: 多個用 "; " 分隔 (例: "懸浮固體（mg/L）; 氯離子")
# - 轉換係數串: 多個用 "; " 分隔, 對齊引入項目順序 (例: "0.30; 0.05")
COMPENSATION = {
    # 鎳系 pH調整池
    ("鎳系", "pH調整池", "NaOH"):       (80, "", ""),   # 純鹼不引入 SS
    ("鎳系", "pH調整池", "Ca(OH)2"):    (60, "懸浮固體（mg/L）", "0.50"),  # 形成 CaCO3/CaF2
    ("鎳系", "pH調整池", "PAC"):        ("", "", ""),    # 禁加, 不填
    ("鎳系", "pH調整池", "Polymer"):    ("", "", ""),

    # 輕系 中和池
    ("輕系", "中和池", "NaOH"):         (50, "", ""),
    ("輕系", "中和池", "H2SO4"):        (30, "硫酸根", "0.98"),
    ("輕系", "中和池", "HCl"):          (30, "氯離子", "0.97"),
    ("輕系", "中和池", "PAC"):          ("", "", ""),    # 禁加

    # 各系 快混池
    ("各系", "快混池", "PAC"):          (100, "懸浮固體（mg/L）; 氯離子", "0.30; 0.05"),

    # 各系 慢混池
    ("各系", "慢混池", "Polymer (PAM)"): (2, "懸浮固體（mg/L）", "1.00"),

    # 各系 混凝池
    ("各系", "混凝池", "FeCl3"):        (80, "懸浮固體（mg/L）; 氯離子", "0.35; 0.65"),

    # 通用 氰系氧化池
    ("通用", "氰系氧化池", "NaOCl"):     (300, "氯離子", "0.48"),
    ("通用", "氰系氧化池", "H2O2"):      (200, "", ""),  # H2O2 反應後變 H2O + O2, 不留殘留

    # 通用 鉻系還原池
    ("通用", "鉻系還原池", "NaHSO3"):    (200, "硫酸根", "0.92"),
    ("通用", "鉻系還原池", "Na2S"):      (50, "", ""),   # 形成金屬硫化物沉澱, 少量殘 S²⁻

    # 通用 硫化物沉澱池
    ("通用", "硫化物沉澱池", "Na2S"):    (50, "", ""),

    # 通用 消毒池
    ("通用", "消毒池", "NaOCl"):         (50, "氯離子", "0.48"),

    # 通用 pH調整池
    ("通用", "pH調整池", "NaOH"):        (50, "", ""),

    # 通用 濾液池
    ("通用", "濾液池", "NaOH"):          (30, "", ""),
    ("通用", "濾液池", "FeCl3"):         (50, "懸浮固體（mg/L）; 氯離子", "0.35; 0.65"),
}


def main():
    wb = openpyxl.load_workbook(XLSX)
    if SHEET not in wb.sheetnames:
        print(f"[X] 找不到分頁 {SHEET}")
        return
    ws = wb[SHEET]
    print(f"[1/3] {SHEET} 現況: {ws.max_row} 列 × {ws.max_column} 欄")

    # 寫表頭
    for col, header in NEW_HEADERS.items():
        existing = ws.cell(1, col).value
        if existing and existing != header:
            print(f"[!] R{1} C{col} 已有 '{existing}', 不覆蓋")
            continue
        ws.cell(1, col, header)
        print(f"  寫入 R1 C{col} = '{header}'")
    print()

    # 對齊每條規則, 寫進 R/S/T
    filled = 0
    missing = []
    for r in range(2, ws.max_row + 1):
        sys_ = ws.cell(r, 2).value
        tank = ws.cell(r, 3).value
        drug = ws.cell(r, 4).value
        if not (sys_ and tank and drug):
            continue
        key = (str(sys_).strip(), str(tank).strip(), str(drug).strip())
        if key in COMPENSATION:
            mg_per_L, items, coefs = COMPENSATION[key]
            ws.cell(r, 18, mg_per_L if mg_per_L != "" else None)
            ws.cell(r, 19, items if items != "" else None)
            ws.cell(r, 20, coefs if coefs != "" else None)
            filled += 1
        else:
            missing.append((r, key))

    wb.save(XLSX)
    print(f"[2/3] 已寫入 {filled} 條補償資訊")

    if missing:
        print(f"\n[!] 以下規則未在補償表中, 留空:")
        for r, key in missing:
            print(f"    R{r}: {key}")

    print(f"\n[3/3] 完成. 規則庫已保存至 {XLSX}")


if __name__ == "__main__":
    main()
