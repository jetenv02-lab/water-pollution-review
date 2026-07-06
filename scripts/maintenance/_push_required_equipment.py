# -*- coding: utf-8 -*-
"""把 AB 欄 (28: 必備機具) 推到 Sheets _槽體學理."""
import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
import openpyxl
import sheets_sync as ss

XLSX = os.path.join(os.path.dirname(__file__), "..", "..", "規則庫.xlsx")
XLSX = os.path.abspath(XLSX)
SHEET = "_槽體學理"
COLS = [("AB", 28)]


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    client, source = ss._get_gspread_client()
    print("認證來源: {}".format(source))
    sh = client.open_by_key(ss.DEFAULT_SHEET_ID)
    ws_sheet = sh.worksheet(SHEET)

    if ws_sheet.col_count < 28:
        print("Sheet 目前 {} 欄, 擴充到 28 欄".format(ws_sheet.col_count))
        ws_sheet.add_cols(28 - ws_sheet.col_count)

    for letter, idx in COLS:
        rows_to_push = []
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, idx).value
            rows_to_push.append([val if val is not None else ""])
        cell_range = "{}1:{}{}".format(letter, letter, len(rows_to_push))
        ws_sheet.update(values=rows_to_push, range_name=cell_range)
        print("[OK] 推 {} 欄 ({} 列)".format(letter, len(rows_to_push)))

    print("\nURL: https://docs.google.com/spreadsheets/d/{}".format(ss.DEFAULT_SHEET_ID))


if __name__ == "__main__":
    main()
