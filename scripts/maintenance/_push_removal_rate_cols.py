# -*- coding: utf-8 -*-
"""把 Y/Z/AA 三欄推到 Sheets _槽體學理."""
import os
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(REPO_ROOT))

import openpyxl
import sheets_sync as ss

XLSX = REPO_ROOT / "規則庫.xlsx"
SHEET = "_槽體學理"
COLS = [("Y", 25), ("Z", 26), ("AA", 27)]


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    client, source = ss._get_gspread_client()
    print(f"認證來源: {source}")
    sh = client.open_by_key(ss.DEFAULT_SHEET_ID)
    ws_sheet = sh.worksheet(SHEET)

    if ws_sheet.col_count < 27:
        print(f"Sheet 目前 {ws_sheet.col_count} 欄, 擴到 27 欄")
        ws_sheet.add_cols(27 - ws_sheet.col_count)

    for letter, idx in COLS:
        rows = []
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, idx).value
            rows.append([val if val is not None else ""])
        cell_range = f"{letter}1:{letter}{len(rows)}"
        ws_sheet.update(values=rows, range_name=cell_range)
        print(f"[OK] 推 {letter} 欄 ({len(rows)} 列)")

    print(f"\nURL: https://docs.google.com/spreadsheets/d/{ss.DEFAULT_SHEET_ID}")


if __name__ == "__main__":
    main()
