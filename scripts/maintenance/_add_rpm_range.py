# -*- coding: utf-8 -*-
"""把「攪拌轉速 RPM 合理範圍」加到 _槽體學理 分頁。

實務優先判斷 — 廠商 PDF 通常寫 RPM (轉/分), 技師先用 RPM 判斷是否合理,
超出範圍才追查 G 值 / 槳葉直徑。

新增欄位:
  V RPM_min
  W RPM_max
  X RPM 備註
"""
import openpyxl
from pathlib import Path

XLSX = Path(__file__).parent / "規則庫.xlsx"
SHEET = "_槽體學理"

NEW_HEADERS = {
    22: "RPM_min (轉/分)",
    23: "RPM_max (轉/分)",
    24: "RPM 備註",
}

# 依實務經驗建的 RPM 合理範圍
# 空字串代表「該槽不常用轉速判斷」
RPM_RANGES = {
    # 反應類 (混合為主)
    "pH調整槽":         (60, 200, "中低速混合加藥點"),
    "pH調整池":         (60, 200, "中低速混合加藥點"),
    "中和池":           (60, 200, "中低速中和反應"),
    "pH調整暨快混池":   (120, 250, "兼具 pH 調整+快混"),
    "pH調整池暨快混池": (120, 250, "兼具 pH 調整+快混"),
    "快混槽":           (150, 300, "高速分散凝集劑"),
    "快混池":           (150, 300, "高速分散凝集劑"),
    "慢混池":           (20, 60, "低速不打散絮體"),
    "慢混槽":           (20, 60, "低速不打散絮體"),
    "混凝膠凝池":       (30, 120, "多段: 進口高出口低"),
    "氧化池":           (60, 150, "中速讓氧化劑接觸"),
    "氰系氧化槽":       (60, 150, "同氧化池, 需 ORP 儀"),
    "鉻系還原槽":       (60, 150, "同氧化池, 需 ORP 儀"),
    "批次反應槽":       (30, 200, "變頻: 反應期高沉降期停"),

    # 分離類 (通常不攪或極低速)
    "沉澱池":           (0, 15, "0=不攪, 5~15=刮泥機"),
    "沉降池":           (0, 15, "0=不攪, 5~15=刮泥機"),
    "浮除槽":           (0, 30, "靠氣泡, 一般不攪"),
    "砂濾塔":           ("", "", "無攪拌 (過濾)"),
    "砂濾器":           ("", "", "無攪拌 (過濾)"),
    "活性碳吸附塔":     ("", "", "無攪拌 (吸附)"),
    "活性碳吸附裝置":   ("", "", "無攪拌 (吸附)"),
    "離子交換樹脂塔":   ("", "", "無攪拌 (離子交換)"),
    "油脂分離槽":       (0, 15, "靠浮力"),
    "預處理池":         (30, 100, "均勻化 + 部分沉降"),

    # 生物類
    "曝氣槽":           (30, 150, "機械式; 鼓風式無此參數"),
    "厭氧池":           (20, 60, "保持微生物懸浮"),
    "接觸氧化池":       (30, 100, "低速循環水流"),
    "MBR":              (30, 100, "配合膜組件"),

    # 污泥/儲存類
    "脫水機":           ("", "", "視機型: 帶濾 5~30, 離心 1000~3000"),
    "污泥離心式脫水機": (1000, 3500, "離心式高速"),
    "污泥帶濾式脫水機": (5, 30, "帶濾式低速"),
    "濃縮槽":           (0, 20, "0=重力, 20=有刮泥"),
    "污泥濃縮設施":     (0, 20, "同上"),
    "污泥儲槽":         (10, 30, "低速防結塊"),
    "濾液池":           (0, 60, "視是否加藥"),

    # 通用/中性類
    "暫存槽":           (0, 60, "通常不攪或低速"),
    "貯留槽":           (0, 60, "同上"),
    "廢水調整池":       (30, 80, "均勻化用"),
    "廢水收集池":       (0, 60, "通常不攪"),
    "調勻池":           (30, 80, "均勻化 + 水質波動吸收"),
    "放流池":           ("", "", "無攪拌 (放流前緩衝)"),
}


def main():
    wb = openpyxl.load_workbook(XLSX)
    if SHEET not in wb.sheetnames:
        print(f"[X] 找不到分頁 {SHEET}")
        return
    ws = wb[SHEET]
    print(f"[1/3] {SHEET} 現況: {ws.max_row} 列 x {ws.max_column} 欄")

    # 寫表頭
    for col, header in NEW_HEADERS.items():
        existing = ws.cell(1, col).value
        if existing and existing != header:
            print(f"[!] R1 C{col} 已有 '{existing}', 不覆蓋")
            continue
        ws.cell(1, col, header)
        print(f"  寫入 R1 C{col} = '{header}'")

    print()
    filled = 0
    missing = []
    for r in range(2, ws.max_row + 1):
        tank = ws.cell(r, 1).value
        status = ws.cell(r, 8).value or ""
        if not tank:
            continue
        if tank in RPM_RANGES:
            existing = ws.cell(r, 22).value
            if existing not in (None, ""):
                print(f"[skip] R{r} {tank} 已有 RPM")
                continue
            rmin, rmax, note = RPM_RANGES[tank]
            ws.cell(r, 22, rmin if rmin != "" else None)
            ws.cell(r, 23, rmax if rmax != "" else None)
            ws.cell(r, 24, note if note else None)
            filled += 1
        elif status == "V":
            missing.append((r, tank))

    wb.save(XLSX)
    print(f"[2/3] 已寫入 {filled} 條 RPM 範圍")
    if missing:
        print(f"\n[!] 以下 active 槽體未在 RPM 範圍表中:")
        for r, tank in missing:
            print(f"    R{r}: {tank}")

    print(f"\n[3/3] 完成. {XLSX}")


if __name__ == "__main__":
    main()
