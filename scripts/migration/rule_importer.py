# -*- coding: utf-8 -*-
"""新審查意見半自動匯入模組。

工作流:
    1. 使用者貼上「NotebookLM 或自己整理好的規則表」文字
       支援格式: TSV / CSV / Markdown 表格
    2. parse_input_text() 解析成 list[dict]
    3. preview_import() 對照現有規則庫, 算「新增筆數 / 新槽體 / 衝突 ID / 缺漏欄」
    4. (使用者確認) commit_import() 寫進 規則庫.xlsx + _來源清單
"""
import csv
import io
import os
import re
from datetime import datetime
try:
    from tz_util import fmt_tpe as _fmt_tpe
except Exception:
    def _fmt_tpe(fmt="%Y-%m-%d %H:%M:%S"):
        return datetime.now().strftime(fmt)

BASE = os.path.dirname(os.path.abspath(__file__))
RULES_XLSX = os.path.join(BASE, "規則庫.xlsx")

# 必填欄位 (任何一個缺漏就標 warn, 該筆仍然匯入但會顯示)
REQUIRED_FIELDS = ["原文缺失", "檢查類型", "對照項目", "規則", "標準槽體名稱"]

# 預期欄位 (來源欄位若有差異, 用同義字映射)
FIELD_ALIASES = {
    "缺失ID": ["缺失id", "id", "缺失代號"],
    "原文缺失": ["原文", "原文意見", "缺失內容", "意見原文"],
    "檢查類型": ["類型", "缺失類型", "分類"],
    "對照項目": ["項目", "對照"],
    "規則": ["規則(萃取/可比對判斷式)", "判斷式", "規則內容", "規則描述"],
    "比對位置": ["位置", "比對位置(依標題,非頁碼)", "查找位置"],
    "判定邏輯": ["判定邏輯(條件→結論)", "判斷邏輯", "邏輯"],
    "技師姓名": ["技師", "姓名"],
    "序號": ["原文序號", "技師序號"],
    "標準槽體名稱": ["槽體", "標準槽體", "槽體名稱"],
    "原始槽體代號": ["槽體代號", "代號", "T 代號", "原槽體代號"],
    "狀態": [],
}


# ──────────────────────────────────────────────────
# 解析輸入文字
# ──────────────────────────────────────────────────

def _normalize_header(raw):
    """去空白 / 小寫 / 去除 markdown 的 ** *。"""
    if raw is None:
        return ""
    s = str(raw).strip()
    s = s.replace("**", "").replace("*", "").strip()
    return s


def _map_field(header):
    """用同義字表把任意輸入欄名 → 標準欄名。"""
    h = _normalize_header(header)
    if not h:
        return None
    for std, aliases in FIELD_ALIASES.items():
        if h == std:
            return std
        for a in aliases:
            if h == a:
                return std
    return None  # 無法對應 — 但保留


def _parse_tsv_or_csv(text):
    """嘗試 TSV → CSV。回傳 list[dict]。"""
    text = text.strip()
    if not text:
        return None

    # 偵測分隔符
    first_line = text.split("\n", 1)[0]
    if "\t" in first_line and first_line.count("\t") >= 2:
        delim = "\t"
    elif "," in first_line and first_line.count(",") >= 2:
        delim = ","
    else:
        return None

    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    rows = []
    for row in reader:
        # 表頭規範化 + 對應到標準欄
        clean = {}
        for raw_k, v in row.items():
            std = _map_field(raw_k) or _normalize_header(raw_k)
            if std:
                clean[std] = ("" if v is None else str(v).strip())
        if any(clean.values()):  # 至少有一個非空
            rows.append(clean)
    return rows


def _parse_markdown_table(text):
    """解析 markdown 表格: | col1 | col2 | / |---|---| / | v1 | v2 |。"""
    lines = [l.strip() for l in text.strip().split("\n") if l.strip()]
    if not lines:
        return None

    # 找第一條表格列 (有 | 開頭)
    table_lines = [l for l in lines if l.startswith("|") and l.endswith("|")]
    if len(table_lines) < 3:
        return None

    # 第一條 = 表頭
    header_cells = [c.strip() for c in table_lines[0].strip("|").split("|")]
    # 第二條 = 分隔線 (跳過)
    # 第三條起 = 資料
    rows = []
    for line in table_lines[2:]:
        cells = [c.strip() for c in line.strip("|").split("|")]
        if len(cells) != len(header_cells):
            continue
        clean = {}
        for k, v in zip(header_cells, cells):
            std = _map_field(k) or _normalize_header(k)
            if std:
                clean[std] = v
        if any(clean.values()):
            rows.append(clean)
    return rows or None


def parse_input_text(text):
    """解析貼上的文字 → list[dict] (標準欄位)。

    自動偵測 TSV / CSV / Markdown 表格。

    Returns:
        {"ok": True, "rows": [...], "format": "tsv/csv/markdown", "row_count": N}
        {"ok": False, "error": str}
    """
    if not text or not text.strip():
        return {"ok": False, "error": "輸入為空"}

    # 先試 markdown (因為有 | 比較好認)
    rows = _parse_markdown_table(text)
    if rows:
        return {"ok": True, "rows": rows, "format": "markdown", "row_count": len(rows)}

    # 再試 TSV / CSV
    rows = _parse_tsv_or_csv(text)
    if rows:
        # 判斷哪種分隔符
        delim = "tsv" if "\t" in text.split("\n", 1)[0] else "csv"
        return {"ok": True, "rows": rows, "format": delim, "row_count": len(rows)}

    return {"ok": False, "error": "無法解析 — 請確認格式為 TSV / CSV / Markdown 表格"}


# ──────────────────────────────────────────────────
# 解析上傳檔案 (CSV / XLSX)
# ──────────────────────────────────────────────────

def parse_csv_file(file_bytes, filename="(uploaded.csv)"):
    """解析上傳的 CSV 檔。自動偵測編碼 (UTF-8 / UTF-8-BOM / Big5)。

    Returns:
        {"ok": True, "rows": [...], "format": "csv-file", "row_count": N}
        {"ok": False, "error": str}
    """
    # 試多種編碼
    text = None
    for enc in ("utf-8-sig", "utf-8", "big5", "cp950"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return {"ok": False, "error": f"無法讀取 {filename} — 編碼不支援"}

    result = parse_input_text(text)
    if result.get("ok"):
        result["format"] = "csv-file"
        result["filename"] = filename
    return result


def parse_xlsx_file(file_bytes, filename="(uploaded.xlsx)"):
    """解析上傳的 xlsx 檔。

    支援兩種結構:
    1. 單一分頁 (扁平表): 第一個分頁當資料, 自動讀
    2. 多分頁 (跟主檔同結構): 每分頁名稱 = 標準槽體名稱, 合併所有分頁

    Returns:
        {"ok": True, "rows": [...], "format": "xlsx-file", "row_count": N, "sheets": [...]}
        {"ok": False, "error": str}
    """
    try:
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        return {"ok": False, "error": f"無法打開 xlsx: {e}"}

    all_rows = []
    used_sheets = []

    for sn in wb.sheetnames:
        if sn.startswith("_"):
            continue  # 跳過 meta 分頁
        ws = wb[sn]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers_raw = next(rows_iter)
        except StopIteration:
            continue
        if not headers_raw:
            continue
        # 對應欄名
        std_headers = []
        for h in headers_raw:
            std = _map_field(h) if h else None
            std_headers.append(std)

        sheet_row_count = 0
        for raw_row in rows_iter:
            if not raw_row or not any(c is not None and str(c).strip() for c in raw_row):
                continue
            row_dict = {}
            for std, v in zip(std_headers, raw_row):
                if std:
                    row_dict[std] = "" if v is None else str(v).strip()
            # 若分頁名是合法槽體 → 補進「標準槽體名稱」
            if "標準槽體名稱" not in row_dict or not row_dict.get("標準槽體名稱"):
                if not sn.startswith("_"):
                    row_dict["標準槽體名稱"] = sn
            if any(row_dict.values()):
                all_rows.append(row_dict)
                sheet_row_count += 1
        if sheet_row_count > 0:
            used_sheets.append({"name": sn, "rows": sheet_row_count})

    if not all_rows:
        return {"ok": False, "error": "xlsx 沒有任何有效資料 (檢查表頭跟欄位名稱)"}

    return {
        "ok": True,
        "rows": all_rows,
        "format": "xlsx-file",
        "row_count": len(all_rows),
        "sheets": used_sheets,
        "filename": filename,
    }


def parse_uploaded_file(uploaded_file):
    """通用上傳檔案解析 — 看副檔名自動決定走 CSV / XLSX。

    Args:
        uploaded_file: Streamlit 的 UploadedFile 物件 (有 .name / .getvalue())
    """
    if uploaded_file is None:
        return {"ok": False, "error": "沒有檔案"}
    name = uploaded_file.name.lower()
    data = uploaded_file.getvalue()

    if name.endswith(".csv"):
        return parse_csv_file(data, uploaded_file.name)
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        return parse_xlsx_file(data, uploaded_file.name)
    elif name.endswith(".pdf"):
        return {"ok": False,
                "error": "PDF 自動抽取尚未實作 — 請先用 NotebookLM/Gemini 抽出 CSV 後再上傳"}
    else:
        return {"ok": False, "error": f"不支援的副檔名: {name}"}


# ──────────────────────────────────────────────────
# 預覽 (對照現有 規則庫.xlsx)
# ──────────────────────────────────────────────────

def _get_existing_state():
    """讀現有 規則庫.xlsx, 取得:
        - 既有槽體分頁名清單
        - 既有所有缺失ID
        - 最大 D 流水號
        - 最大 S 流水號
    """
    state = {
        "tanks": set(),
        "ids": set(),
        "max_d_num": 0,
        "max_s_num": 0,
    }
    if not os.path.exists(RULES_XLSX):
        return state

    from openpyxl import load_workbook
    wb = load_workbook(RULES_XLSX, data_only=True)
    state["tanks"] = {sn for sn in wb.sheetnames if not sn.startswith("_")}

    # 掃所有缺失ID
    for sn in state["tanks"]:
        ws = wb[sn]
        if ws.max_row < 2:
            continue
        headers = [c.value for c in ws[1]]
        if "缺失ID" not in headers:
            continue
        id_col = headers.index("缺失ID") + 1
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=id_col).value
            if v:
                state["ids"].add(str(v).strip())

    # 算最大 D 流水號
    for rid in state["ids"]:
        m = re.match(r"^D(\d+)", str(rid))
        if m:
            state["max_d_num"] = max(state["max_d_num"], int(m.group(1)))

    # 算最大 S 流水號 (從 _來源清單)
    if "_來源清單" in wb.sheetnames:
        ws_src = wb["_來源清單"]
        if ws_src.max_row >= 2:
            headers = [c.value for c in ws_src[1]]
            if "來源代號" in headers:
                code_idx = headers.index("來源代號") + 1
                for row in range(2, ws_src.max_row + 1):
                    v = ws_src.cell(row=row, column=code_idx).value
                    if v:
                        m = re.match(r"^S(\d+)", str(v).strip())
                        if m:
                            state["max_s_num"] = max(state["max_s_num"], int(m.group(1)))

    return state


def preview_import(parsed_rows):
    """對解析後的列, 預覽匯入結果 (不寫檔)。

    Returns:
        {
            "ok": True,
            "total": N,
            "next_source_code": "S02",
            "next_d_number": 300,
            "tanks_in_import": [...],
            "new_tanks": [...],     # 不在現有 25 槽體裡
            "id_conflicts": [...],  # 跟現有 D ID 撞 — 會自動改新
            "missing_required": [{row_idx, missing_fields, sample_row}],
            "ok_to_import": N,      # 通過必填檢查的筆數
        }
    """
    state = _get_existing_state()

    next_source_code = f"S{state['max_s_num'] + 1:02d}"
    next_d_number = state["max_d_num"] + 1

    tanks_in_import = set()
    new_tanks = set()
    id_conflicts = []
    missing_required = []
    ok_count = 0

    for i, row in enumerate(parsed_rows):
        tank = (row.get("標準槽體名稱") or "").strip()
        if tank:
            tanks_in_import.add(tank)
            if tank not in state["tanks"]:
                new_tanks.add(tank)

        # 必填檢查
        missing = [f for f in REQUIRED_FIELDS if not (row.get(f) or "").strip()]
        if missing:
            missing_required.append({
                "row_idx": i + 1,
                "missing": missing,
                "row_preview": (row.get("原文缺失") or row.get("規則") or "")[:50],
            })
        else:
            ok_count += 1

        # ID 衝突
        rid = (row.get("缺失ID") or "").strip()
        if rid and rid in state["ids"]:
            id_conflicts.append(rid)

    return {
        "ok": True,
        "total": len(parsed_rows),
        "ok_to_import": ok_count,
        "next_source_code": next_source_code,
        "next_d_number": next_d_number,
        "tanks_in_import": sorted(tanks_in_import),
        "new_tanks": sorted(new_tanks),
        "id_conflicts": id_conflicts,
        "missing_required": missing_required,
    }


# ──────────────────────────────────────────────────
# 匯入 (寫進 xlsx)
# ──────────────────────────────────────────────────

# xlsx 各槽體分頁的標準表頭順序
XLSX_TANK_HEADERS = [
    "缺失ID", "來源", "原文缺失", "檢查類型", "對照項目",
    "規則(萃取/可比對判斷式)", "比對位置(依標題,非頁碼)",
    "判定邏輯(條件→結論)", "技師姓名", "序號", "原始槽體代號", "狀態",
]

# 標準欄 → xlsx 表頭 的反向映射
STD_TO_XLSX_HEADER = {
    "缺失ID": "缺失ID",
    "來源": "來源",
    "原文缺失": "原文缺失",
    "檢查類型": "檢查類型",
    "對照項目": "對照項目",
    "規則": "規則(萃取/可比對判斷式)",
    "比對位置": "比對位置(依標題,非頁碼)",
    "判定邏輯": "判定邏輯(條件→結論)",
    "技師姓名": "技師姓名",
    "序號": "序號",
    "原始槽體代號": "原始槽體代號",
    "狀態": "狀態",
}


def _backup_xlsx_before_import():
    """匯入前備份 xlsx。"""
    if not os.path.exists(RULES_XLSX):
        return None
    backup_dir = os.path.join(BASE, "backup")
    os.makedirs(backup_dir, exist_ok=True)
    ts = _fmt_tpe("%Y%m%d_%H%M%S")
    dst = os.path.join(backup_dir, f"規則庫_{ts}_pre_import.xlsx")
    with open(RULES_XLSX, "rb") as src, open(dst, "wb") as out:
        out.write(src.read())
    return dst


def commit_import(parsed_rows, source_metadata, skip_missing=True):
    """執行匯入。

    Args:
        parsed_rows: parse_input_text() 回傳的 rows
        source_metadata: dict 含這次匯入的 來源資訊 — keys:
            檔名, 技師姓名, 技師證書字號, 查核日期, 簽證事業名稱, 備註
        skip_missing: True = 必填欄缺漏的筆數會跳過, False = 強制匯入 (不建議)

    Returns:
        {ok, imported_count, skipped_count, source_code, new_tanks_created, backup, error}
    """
    if not parsed_rows:
        return {"ok": False, "error": "沒有資料可匯入"}
    if not os.path.exists(RULES_XLSX):
        return {"ok": False, "error": "找不到 規則庫.xlsx"}

    # 檢查 xlsx 是否被開啟
    try:
        with open(RULES_XLSX, "r+b"):
            pass
    except PermissionError:
        return {"ok": False, "error": "規則庫.xlsx 被開啟中 (請關閉 Excel)"}

    # 備份
    backup = _backup_xlsx_before_import()

    state = _get_existing_state()
    source_code = f"S{state['max_s_num'] + 1:02d}"
    source_filename = source_metadata.get("檔名", "")
    source_full = f"{source_code} {source_filename}".strip()

    # ID 分配器
    next_d = state["max_d_num"] + 1
    used_ids = set(state["ids"])

    def alloc_id():
        nonlocal next_d
        while f"D{next_d:03d}" in used_ids:
            next_d += 1
        rid = f"D{next_d:03d}"
        used_ids.add(rid)
        next_d += 1
        return rid

    from openpyxl import load_workbook
    from openpyxl.styles import Font

    wb = load_workbook(RULES_XLSX)

    # 把列按槽體分組
    rows_by_tank = {}
    skipped_count = 0
    for row in parsed_rows:
        tank = (row.get("標準槽體名稱") or "").strip()
        if not tank:
            skipped_count += 1
            continue
        # 必填檢查
        missing = [f for f in REQUIRED_FIELDS if not (row.get(f) or "").strip()]
        if missing and skip_missing:
            skipped_count += 1
            continue
        rows_by_tank.setdefault(tank, []).append(row)

    imported_count = 0
    new_tanks_created = []

    for tank, rows in rows_by_tank.items():
        # 若是新槽體 → 建立分頁
        if tank not in wb.sheetnames:
            ws = wb.create_sheet(tank)
            # 寫表頭
            ws.append(XLSX_TANK_HEADERS)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            new_tanks_created.append(tank)
        else:
            ws = wb[tank]

        # 找 xlsx 現有表頭順序 (各分頁可能不同)
        existing_headers = [c.value for c in ws[1]] if ws.max_row >= 1 else XLSX_TANK_HEADERS

        for row in rows:
            # 分配缺失ID (若是空 / 衝突 → 自動分配)
            rid = (row.get("缺失ID") or "").strip()
            if not rid or rid in state["ids"]:
                rid = alloc_id()

            # 來源 (若使用者沒填, 用 source_full)
            source_val = (row.get("來源") or "").strip() or source_full

            # 組 row data (按 existing_headers 順序)
            row_data = []
            for h in existing_headers:
                if h == "缺失ID":
                    row_data.append(rid)
                elif h == "來源":
                    row_data.append(source_val)
                else:
                    # 找對應的標準欄
                    std = None
                    for s, x in STD_TO_XLSX_HEADER.items():
                        if x == h:
                            std = s
                            break
                    if std:
                        row_data.append(row.get(std, ""))
                    else:
                        row_data.append("")
            ws.append(row_data)
            imported_count += 1

    # 寫進 _來源清單
    if "_來源清單" in wb.sheetnames:
        ws_src = wb["_來源清單"]
        headers = [c.value for c in ws_src[1]]
        row_data = []
        for h in headers:
            if h == "來源代號":
                row_data.append(source_code)
            elif h == "審查意見檔名":
                row_data.append(source_metadata.get("檔名", ""))
            elif h == "技師姓名":
                row_data.append(source_metadata.get("技師姓名", ""))
            elif h == "技師證書字號":
                row_data.append(source_metadata.get("技師證書字號", ""))
            elif h == "查核日期":
                row_data.append(source_metadata.get("查核日期", ""))
            elif h == "簽證事業名稱":
                row_data.append(source_metadata.get("簽證事業名稱", ""))
            elif h == "匯入日期":
                row_data.append(_fmt_tpe("%Y-%m-%d"))
            elif h == "備註":
                row_data.append(source_metadata.get("備註", "Streamlit 半自動匯入"))
            elif h == "涵蓋槽體數":
                row_data.append(len(rows_by_tank))
            elif h == "貢獻規則數":
                row_data.append(imported_count)
            elif h == "最後同步時間":
                row_data.append(_fmt_tpe())
            else:
                row_data.append("")
        ws_src.append(row_data)

    wb.save(RULES_XLSX)

    return {
        "ok": True,
        "imported_count": imported_count,
        "skipped_count": skipped_count,
        "source_code": source_code,
        "source_full": source_full,
        "new_tanks_created": new_tanks_created,
        "backup": backup,
        "timestamp": _fmt_tpe(),
    }


# ──────────────────────────────────────────────────
# CLI 測試
# ──────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    sample = """缺失ID\t原文缺失\t檢查類型\t對照項目\t規則\t標準槽體名稱\t原始槽體代號
\tT99-01 測試槽體出水口缺裝液位計\t機具設施\t液位計\t測試規則: 出水池應有液位計\t放流池\tT99-01
"""
    print("=== 解析測試 ===")
    r = parse_input_text(sample)
    print(r)
    if r["ok"]:
        print("\n=== 預覽測試 ===")
        p = preview_import(r["rows"])
        import json
        print(json.dumps(p, ensure_ascii=False, indent=2))
