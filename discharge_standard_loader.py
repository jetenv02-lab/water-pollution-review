# -*- coding: utf-8 -*-
"""放流水標準查詢 + 超標檢查 — 純新增模組。

三個資料來源分頁 (規則庫.xlsx):
    _放流標準_附表五   — 電鍍/PCB/金屬基本/金屬表面處理業
    _放流標準_附表十六 — 保護農地水體加嚴
    _放流標準_附表四   — 化工業

對外 API:
    get_standard(industry, item, conditions=None) → 限值 (float 或 None)
    check_discharge_water(放流水質 dict, industry, conditions=None) → list[finding]
    check_all_discharge_units(units, industry, conditions=None) → list[finding]

規則庫欄位 (Schema 7 欄):
    A 業別 / B 水質項目 / C 限值 / D 單位
    E 適用條件 / F 施行日期 / G 備註來源
"""
import os
import re
from functools import lru_cache

import openpyxl

DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "規則庫.xlsx")

# 業別 → 適用附表
INDUSTRY_TO_TABLE = {
    "電鍍": "_放流標準_附表五",
    "金屬電鍍": "_放流標準_附表五",
    "PCB": "_放流標準_附表五",
    "印刷電路板": "_放流標準_附表五",
    "金屬基本工業": "_放流標準_附表五",
    "金屬表面處理": "_放流標準_附表五",
    "化工業": "_放流標準_附表四",
    "化學工業": "_放流標準_附表四",
    "食品製造": "_放流標準_附表四",  # 食品業無專屬表, 走通用化工
    "紙板製造": "_放流標準_附表四",  # 造紙業有專屬附表七, 目前 fallback
    "造紙": "_放流標準_附表四",
}

# 水質項目同義字 → 標準名 (規則庫用的名)
ITEM_SYNONYM = {
    "pH": "pH", "pH值": "pH", "PH": "pH", "ph": "pH", "氫離子濃度指數": "pH",
    "SS": "懸浮固體", "懸浮固體(mg/L)": "懸浮固體", "懸浮固體（mg/L）": "懸浮固體",
    "COD": "化學需氧量", "化學需氧量(mg/L)": "化學需氧量", "化學需氧量（mg/L）": "化學需氧量",
    "BOD": "生化需氧量", "生化需氧量(mg/L)": "生化需氧量", "生化需氧量（mg/L）": "生化需氧量",
    "氯": "氯離子", "氯鹽": "氯離子", "氯化物": "氯離子",
    "硫酸": "硫酸根", "硫酸鹽": "硫酸根",
    "F": "氟鹽", "氟化物": "氟鹽", "氟": "氟鹽",
    "CN": "氰化物",
    "TP": "總磷", "總磷": "正磷酸鹽",  # 總磷有時對應正磷酸鹽 (以三價磷酸根計算)
    "NH3-N": "氨氮", "氨氮（mg/L）": "氨氮",
    "水溫(攝氏)": "水溫", "水溫（攝氏）": "水溫", "溫度": "水溫",
    "油脂（mg/L）": "油脂", "油及脂": "油脂", "礦物性油脂": "油脂",
    "真色色度（mg/L）": "真色色度", "色度": "真色色度",
    "Cu": "銅", "Ni": "鎳", "Zn": "鋅", "Pb": "鉛", "Cd": "鎘",
    "Cr": "總鉻", "Cr6+": "六價鉻", "Cr⁶⁺": "六價鉻",
    "As": "砷", "Hg": "總汞", "Sn": "錫", "Mo": "鉬", "Ag": "銀",
    "Se": "硒", "B": "硼",
    "Fe": "溶解性鐵", "Mn": "溶解性錳",
    "陰離子界面活性劑（mg/L）": "陰離子界面活性劑",
}


def _to_float(v):
    """把限值字串轉 float。支援 '<0.005', '3.0', '6.0~9.0' 等"""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # <0.005 → 0.005 (視為上限)
    s = s.lstrip("<≤")
    # 6.0~9.0 → 這種是 pH 範圍, 不好直接轉 float, 回傳 None (需另外處理)
    if "~" in s or "-" in s and not s.startswith("-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_range(v):
    """回傳 (min, max) 或 (None, val)。for pH 6.0~9.0 這種"""
    if v is None:
        return (None, None)
    s = str(v).strip()
    m = re.match(r"^(\d+(?:\.\d+)?)\s*[~-]\s*(\d+(?:\.\d+)?)$", s)
    if m:
        return (float(m.group(1)), float(m.group(2)))
    # 純數字視為 (None, val), 即單一上限
    f = _to_float(s)
    return (None, f)


def _normalize_item(item):
    """把 PDF 上的水質項目名, 標準化成規則庫用的名。"""
    if not item:
        return None
    item_str = str(item).strip()
    # 直接查
    if item_str in ITEM_SYNONYM:
        return ITEM_SYNONYM[item_str]
    # 掃描找子字串
    for syn, std in ITEM_SYNONYM.items():
        if syn == item_str or syn in item_str:
            return std
    return item_str  # fallback: 原字串


@lru_cache(maxsize=8)
def load_table(sheet_name, xlsx_path=None):
    """讀某個放流標準分頁, 回傳 list of dict。"""
    path = xlsx_path or DEFAULT_XLSX
    if not os.path.exists(path):
        return []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:  # 水質項目為空跳過
            continue
        out.append({
            "業別": str(row[0] or "").strip(),
            "水質項目": str(row[1] or "").strip(),
            "限值": row[2],
            "單位": str(row[3] or "").strip(),
            "適用條件": str(row[4] or "").strip(),
            "施行日期": str(row[5] or "").strip(),
            "備註": str(row[6] or "").strip(),
        })
    return out


def get_table_for_industry(industry):
    """業別 → 對應附表 sheet_name。找不到回 None。"""
    if not industry:
        return None
    industry_str = str(industry).strip()
    for key, table in INDUSTRY_TO_TABLE.items():
        if key in industry_str or industry_str in key:
            return table
    return None


def get_standard(industry, item, conditions=None):
    """查某業別某項目的放流標準 (最嚴的一筆)。

    Args:
        industry: 業別 (例: '電鍍', 'PCB')
        item: 水質項目 (例: '鋅', '銅', 'Cu')
        conditions: dict 可選過濾條件, 例:
            {"排放水量": 200, "設立日期": "106/12/25 前完成", "保護區": False}

    Returns:
        dict {"限值": float, "單位": str, "適用條件": str, "備註": str} 或 None
    """
    table_name = get_table_for_industry(industry)
    if not table_name:
        return None
    data = load_table(table_name)
    if not data:
        return None

    std_item = _normalize_item(item)
    # 過濾符合的
    matches = [d for d in data if _normalize_item(d["水質項目"]) == std_item]
    if not matches:
        return None

    # 目前簡化: 取「限值最嚴 (最小)」的那筆
    # 未來可根據 conditions 智慧選擇
    best = None
    for m in matches:
        val = _to_float(m["限值"])
        if val is None:
            # pH 之類的範圍
            continue
        if best is None or val < _to_float(best["限值"]):
            best = m
    if best is None:
        # 全是範圍值 (例如 pH), 直接回第一筆
        best = matches[0]
    return best


def check_discharge_water(water_quality, industry, conditions=None, tank_code=""):
    """檢查放流水質是否超標。

    Args:
        water_quality: dict {水質項目名: 濃度值 (float)}
        industry: 業別
        conditions: 可選過濾條件
        tank_code: 該放流單元代碼 (finding 顯示用)

    Returns:
        list[finding]
    """
    findings = []
    for item, val in water_quality.items():
        val_num = _to_float(val)
        if val_num is None:
            continue

        std = get_standard(industry, item, conditions)
        if not std:
            continue

        limit_str = std["限值"]
        # pH 特殊 (範圍)
        std_item = _normalize_item(item)
        if std_item == "pH":
            lo, hi = _parse_range(limit_str)
            if lo is not None and hi is not None:
                if val_num < lo or val_num > hi:
                    findings.append({
                        "嚴重度": "不合理",
                        "類型": "放流水標準",
                        "單元": tank_code or "放流口",
                        "標準槽體": "",
                        "對照項目": item,
                        "描述": (
                            f"pH 值 {val_num} 超出標準 {lo}~{hi}。"
                            f" 業別: {industry}, 依據: {std['備註']}"
                        ),
                        "依據": std["備註"],
                    })
            continue

        limit_num = _to_float(limit_str)
        if limit_num is None:
            continue

        # 一般數值: 超過 → finding
        if val_num > limit_num:
            over_pct = (val_num - limit_num) / limit_num * 100
            severity = "不合理" if over_pct > 50 else "待確認"
            findings.append({
                "嚴重度": severity,
                "類型": "放流水標準",
                "單元": tank_code or "放流口",
                "標準槽體": "",
                "對照項目": item,
                "描述": (
                    f"{std_item} {val_num} {std['單位']} 超過標準 {limit_num} {std['單位']}"
                    f" (+{over_pct:.0f}%)。業別: {industry}, 條件: {std['適用條件']}, "
                    f"依據: {std['備註']}"
                ),
                "依據": std["備註"],
            })
    return findings


def _find_real_discharge_units(units):
    """找出真正放流口單元 (廠內流程最末端).

    Nick 定調 (2026-07-06): 不是每個 std_tank = 放流池 都算放流口.
    例馥廷 T01-15 = 放流暫存槽, 但後面還有 T01-16 砂濾, 才是真正末端.

    判定順序 (由強到弱):
        1. 名稱含「放流」但不含「暫存/貯存/儲存/中間/緩衝」
           且 effluent 中至少一條 stream 沒接到其他單元 (真沒下游)
        2. 若沒有 (1) 命中, 找所有「不是污泥系統」且 effluent stream 沒下游的單元
        3. 都沒有 → 回舊邏輯的 std_tank='放流池' 兜底
    """
    if isinstance(units, dict):
        unit_dict = units
    else:
        unit_dict = {(u.get("raw_code") or u.get("code") or ""): u for u in units}

    # 收集所有 influent 用到的 stream codes (WTB xxxxx)
    used_by_downstream = set()
    for u in unit_dict.values():
        if not isinstance(u, dict):
            continue
        inf = u.get("influent") or {}
        for sc in inf.keys():
            # WTB01-13-1 對應上游 WTA01-13-1 (只是 B/A 對調)
            # 但實際上下游是靠 stream_q 或連結表, 這裡簡化用 code 對應
            used_by_downstream.add(str(sc))

    def _has_no_downstream(unit):
        """判該單元的 effluent 是否至少一條沒接到下游."""
        eff = unit.get("effluent") or {}
        if not eff:
            return False
        for sc in eff.keys():
            sc_str = str(sc)
            # WTA01-13-1 → 下游會用 WTB01-14-1 (換 A/B), 這裡簡單看有沒有任何單元
            # 的 influent 引用了這個 stream code 或其對應的 B 版本
            b_ver = sc_str.replace("WTA", "WTB", 1) if "WTA" in sc_str else sc_str
            if sc_str not in used_by_downstream and b_ver not in used_by_downstream:
                return True
        return False

    # 排除污泥/濾液/脫水類 (不是水流末端)
    def _is_sludge_or_side(unit):
        name = str(unit.get("name_in_doc") or "")
        std = str(unit.get("std_tank") or "")
        for kw in ["污泥", "脫水", "濾液", "濃縮", "貯泥"]:
            if kw in name or kw in std:
                return True
        return False

    # 排除「暫存/中間」名稱
    def _is_buffer_only(unit):
        name = str(unit.get("name_in_doc") or "")
        # 「放流暫存槽」「中間池」「緩衝槽」等 — 名稱有暫存字眼
        for kw in ["暫存", "貯存", "儲存", "中間", "緩衝"]:
            if kw in name:
                return True
        return False

    # 條件 1: 名稱含「放流」且不含 buffer 字眼 且 沒下游
    for code, u in unit_dict.items():
        if not isinstance(u, dict):
            continue
        name = str(u.get("name_in_doc") or "")
        if "放流" in name and not _is_buffer_only(u) and _has_no_downstream(u):
            return [(code, u)]

    # 條件 2: 非污泥系統 且 沒下游 (通常是砂濾/活性碳等末端)
    candidates = []
    for code, u in unit_dict.items():
        if not isinstance(u, dict):
            continue
        if _is_sludge_or_side(u):
            continue
        if _has_no_downstream(u):
            candidates.append((code, u))

    if candidates:
        # 選序號最大的 (T01-16 > T01-15)
        candidates.sort(key=lambda x: str(x[0]))
        return [candidates[-1]]

    # 兜底: std_tank='放流池' 且不是 buffer
    for code, u in unit_dict.items():
        if not isinstance(u, dict):
            continue
        std = str(u.get("std_tank") or "")
        if "放流" in std and not _is_buffer_only(u):
            return [(code, u)]

    return []


def check_all_discharge_units(units, industry, conditions=None):
    """對真正放流口單元檢查放流水質標準.

    Args:
        units: dict[code → unit] 或 list[unit]
        industry: 業別 (從 extract_production_scale 拿)
        conditions: 可選

    Returns:
        list[finding]
    """
    findings = []
    real_discharge = _find_real_discharge_units(units)

    for code, unit in real_discharge:
        if not isinstance(unit, dict):
            continue
        # 從 effluent 拿放流水質 — 若有多條 stream, 只挑「主流」(Q 最大),
        # 反洗/濃縮支流等小 Q 支流不算放流水 (它們會回流再處理)
        eff = unit.get("effluent") or {}
        stream_q = unit.get("stream_q") or {}

        def _stream_q_cmd(sc):
            info = stream_q.get(sc) or {}
            if isinstance(info, dict):
                q = info.get("q_cmd")
                if isinstance(q, (int, float)):
                    return float(q)
            return 0.0

        # 挑 Q 最大的一條當放流主流
        main_stream = None
        max_q = -1
        for sc in eff.keys():
            q = _stream_q_cmd(sc)
            if q > max_q:
                max_q = q
                main_stream = sc

        streams_to_check = [main_stream] if main_stream else list(eff.keys())

        for stream_code in streams_to_check:
            stream = eff.get(stream_code)
            if not isinstance(stream, dict):
                continue
            wq = {}
            for item, v in stream.items():
                if isinstance(v, dict):
                    c = v.get("濃度")
                    if c is not None:
                        wq[item] = c
            if wq:
                fs = check_discharge_water(wq, industry, conditions, tank_code=code)
                findings.extend(fs)
    return findings


def _self_test():
    """自我測試。"""
    print(f"[1] 業別 → 附表對照:")
    for k, v in INDUSTRY_TO_TABLE.items():
        print(f"  {k}: {v}")
    print()

    # Test 1: 電鍍業 鋅
    std = get_standard("電鍍", "鋅")
    print(f"Test 1 (電鍍 鋅): 限值 {std['限值']} {std['單位']} - {std['適用條件']}")

    # Test 2: 電鍍業 銅
    std2 = get_standard("電鍍", "銅")
    print(f"Test 2 (電鍍 銅): 限值 {std2['限值']} {std2['單位']} - {std2['適用條件']}")

    # Test 3: PCB 銀
    std3 = get_standard("PCB", "銀")
    print(f"Test 3 (PCB 銀): 限值 {std3['限值']} {std3['單位']}")

    # Test 4: 化工 鎘
    std4 = get_standard("化工業", "鎘")
    print(f"Test 4 (化工 鎘): 限值 {std4['限值']} {std4['單位']}")

    # Test 5: 假放流水 超標檢查
    wq = {
        "鋅": 6.5,      # 電鍍 標準 3.5, 超標
        "銅": 1.2,      # 電鍍 標準 1.0 or 1.5 or 2.0, 依規模
        "鎳": 0.5,      # 電鍍 標準 0.7, 未超
        "pH": 10,       # 標準 6~9, 超標
        "COD": 80,      # 電鍍 標準 100, 未超
    }
    fs = check_discharge_water(wq, "電鍍", tank_code="D01")
    print(f"\nTest 5 (電鍍 5 項目對照): {len(fs)} findings")
    for f in fs:
        print(f"  [{f['嚴重度']}] {f['對照項目']}: {f['描述'][:100]}")

    print("\n[OK] 自我測試完成")


if __name__ == "__main__":
    _self_test()
