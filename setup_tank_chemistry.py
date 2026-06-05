# -*- coding: utf-8 -*-
"""建立 / 重建 規則庫.xlsx 的 _槽體學理 分頁。

此分頁定義各標準槽體的「學理上應變動 / 不應變動」水質項目, 供
step3b_balance_check.check_tank_chemistry() 使用。

執行方式:
    python setup_tank_chemistry.py          # dry run (印出表)
    python setup_tank_chemistry.py --apply  # 真寫進 規則庫.xlsx
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "規則庫.xlsx"
SHEET_NAME = "_槽體學理"

HEADERS = [
    "標準槽體",        # key, 對應槽體分頁名 / std_tank
    "加藥類型",        # 文字描述, 給人看
    "應變動項目",      # 半形逗號分隔; 用類別詞 (SS, 重金屬, 離子) 或具體項目
    "不應變動項目",    # 半形逗號分隔; "*" = 除應變動外都不該變
    "容忍度(%)",       # 質量平衡比對的容差; 預設 5
    "違反嚴重度",      # 不合理 / 待人工
    "學理說明",        # 給使用者看的解釋
    "狀態",            # V / ? / 空白 (V=已核可, ?=待討論)
]

# ──────────────────────────────────────────────────
# 18 條學理規則 (依使用者整理的清單)
# 規則表達語意:
#   - "*" 在「不應變動項目」代表「除應變動外都不該變」
#   - 在「應變動項目」代表「都允許變」, 用於均化槽
#   - 類別詞: SS / 重金屬 / 離子 / 有機物 / 含水率 (程式會展開到具體項目)
# ──────────────────────────────────────────────────
RULES = [
    # === pH 調整 / 中和類 ===
    ("pH調整槽",          "酸/鹼",               "pH, 水溫",              "*",
     5, "不合理", "純 pH 調整槽只加酸/鹼, 無分離機制, 其他水質應守恆", "V"),
    ("pH調整池",          "酸/鹼",               "pH, 水溫",              "*",
     5, "不合理", "同 pH 調整槽 (申請文件命名變體)", "V"),
    ("中和池",            "酸/鹼",               "pH, 水溫",              "*",
     5, "不合理", "加酸鹼中和廢水, 學理上同純 pH 調整槽", "V"),
    ("pH調整暨快混池",    "酸鹼 + 混凝劑",       "pH, 水溫, SS",          "*",
     5, "不合理", "加酸鹼 + 混凝劑, 混凝劑本身為固體故 SS 增加合理, "
                  "但無分離機制, 重金屬/COD/BOD 須後端沉澱才能去除", "V"),
    ("pH調整池暨快混池",  "酸鹼 + 混凝劑",       "pH, 水溫, SS",          "*",
     5, "不合理", "同 pH 調整暨快混池 (申請文件命名變體)", "V"),

    # === 混合類 ===
    ("快混槽",            "混凝劑",              "SS, 水溫",              "*",
     5, "不合理", "加 PAC/明礬/FeCl₃ 混凝, SS 增加合理, 但無沉降, "
                  "重金屬/COD/BOD 不該減少", "V"),
    ("慢混池",            "助凝劑",              "SS, 水溫",              "*",
     5, "不合理", "加助凝劑 (PAM) 形成大絮凝體, SS 可微增, 仍無沉降, "
                  "重金屬/COD/BOD 不該減少", "V"),

    # === 沉降 / 浮除 ===
    ("沉澱池",            "無 (重力沉降)",       "SS, 重金屬, 濁度, COD, BOD", "pH, 氨氮, 硼, 氯, 硝酸鹽",
     10, "不合理", "重力沉降, 固體+附著重金屬下沉, 但溶解性離子 (氨氮/硼/Cl⁻/NO₃⁻) "
                   "不應改變", "V"),
    ("沉降池",            "無 (重力沉降)",       "SS, 重金屬, 濁度, COD, BOD", "pH, 氨氮, 硼, 氯, 硝酸鹽",
     10, "不合理", "同沉澱池", "V"),
    ("浮除槽",            "氣浮 (DAF)",          "SS, 油脂",              "pH, 重金屬, 離子",
     10, "不合理", "加壓溶氣浮除油脂與輕質懸浮物, 重金屬與溶解性物質不應改變", "V"),

    # === 過濾 / 吸附 ===
    ("砂濾塔",            "無 (物理過濾)",       "SS",                    "pH, 重金屬, 離子, 氨氮",
     10, "不合理", "物理截留懸浮固體, 對溶解物無作用 (COD/重金屬僅微減)", "V"),
    ("活性碳吸附塔",      "活性碳",              "COD, BOD, 有機物",      "pH, 離子, 氨氮, 硝酸鹽",
     15, "不合理", "活性碳吸附有機物與部分重金屬, 對溶解性離子無效", "V"),
    ("活性碳吸附裝置",    "活性碳",              "COD, BOD, 有機物",      "pH, 離子, 氨氮, 硝酸鹽",
     15, "不合理", "同活性碳吸附塔", "V"),
    ("離子交換樹脂塔",    "離子交換樹脂",        "指定離子 (Ca, Mg, Na, NO₃, 重金屬)", "pH, COD, BOD, SS",
     20, "待人工", "依樹脂類型 (陽/陰/螯合) 只去除特定離子, 對有機物/SS 無作用; "
                   "需人工判斷樹脂類型與目標離子是否相符", "V"),

    # === 生物處理 ===
    ("曝氣槽",            "供氧 + 微生物",       "BOD, COD, 氨氮, DO",    "pH, 重金屬, 氯, 硼",
     15, "不合理", "好氧微生物代謝有機物 + 硝化, 重金屬不應自行減少, "
                   "Cl⁻/硼為保守物質不該變", "V"),
    ("氧化池",            "氧化 (供氧 / 加氧化劑)", "COD, BOD, 真色色度",    "pH, 重金屬, 離子",
     15, "不合理", "氧化有機物與色度, 但對重金屬 (除非還原型 Cr⁶⁺) 不應顯著去除", "V"),
    ("厭氧池",            "厭氧菌",              "BOD, COD",              "pH, 重金屬, 離子",
     15, "不合理", "厭氧菌代謝有機物, 重金屬不應減少", "V"),

    # === 脫水 / 濃縮 (特殊: 質量守恆優先) ===
    ("脫水機",            "脫水助劑",            "含水率",                "*",
     10, "待人工", "物理脫水, 濃度暴升合理但污染物質量應守恆; 系統會用質量基準檢查", "V"),
    ("濃縮槽",            "無 (重力濃縮)",       "含水率",                "*",
     10, "待人工", "重力濃縮污泥, 同脫水機原則", "V"),
    ("污泥儲槽",          "無 (儲存)",           "含水率",                "*",
     15, "待人工", "暫存污泥, 質量應守恆", "V"),

    # === 暫存 / 貯留 / 調勻 (均化槽 → 進出可變動但需人工確認) ===
    ("暫存槽",            "無 (均化)",           "*",                     "重金屬, 離子",
     30, "待人工", "均化多股廢水, 進出水質可變動但污染物質量應大致守恆; "
                   "若大幅減少需確認是否誤填", "V"),
    ("貯留槽",            "無 (均化)",           "*",                     "重金屬, 離子",
     30, "待人工", "同暫存槽", "V"),
    ("調勻池",            "無 (均化)",           "*",                     "重金屬, 離子",
     30, "待人工", "同暫存槽", "V"),
    ("廢水調整池",        "無 (均化)",           "*",                     "重金屬, 離子",
     30, "待人工", "均化多股廢水穩定後端進流, 重金屬等不應自行減少", "V"),
    ("廢水收集池",        "無 (均化)",           "*",                     "重金屬, 離子",
     30, "待人工", "同廢水調整池", "V"),

    # === 其他 ===
    ("批次反應槽",        "視製程而定",          "視批次而定",            "視批次而定",
     20, "待人工", "依批次操作條件不同, 規則難以一概而論, 須人工判讀", "V"),
    ("放流池",            "無 (儲存)",           "(不應變動)",            "*",
     5, "不合理", "放流前的暫存, 水質應與前端處理出流完全一致", "V"),
]


HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def main():
    apply = "--apply" in sys.argv
    print(f"\n=== 預計寫入 _槽體學理 分頁 ({len(RULES)} 筆) ===\n")
    print(f"{'標準槽體':<22} {'加藥':<14} {'應變動':<28} {'不應變動':<20} {'容忍':<5} {'嚴重度'}")
    print("-" * 110)
    for r in RULES:
        std, dose, allow, deny, tol, sev, *_ = r
        print(f"{std:<20} {dose:<12} {allow[:26]:<26} {deny[:18]:<18} {tol:<4} {sev}")

    if not apply:
        print("\n(dry-run) 加 --apply 才會真的寫入 規則庫.xlsx")
        return

    wb = openpyxl.load_workbook(XLSX)
    # 已有就刪掉重建
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
        print(f"\n刪掉舊的 '{SHEET_NAME}' 分頁")

    ws = wb.create_sheet(SHEET_NAME)
    # 表頭
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # 內容
    for ri, rule in enumerate(RULES, 2):
        for ci, val in enumerate(rule, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            cell.alignment = WRAP

    # 欄寬
    widths = [22, 18, 32, 30, 10, 12, 50, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(XLSX)
    print(f"\n✅ 已寫入 {XLSX} → {SHEET_NAME} 分頁 ({len(RULES)} 筆規則)")
    print("\n下一步:")
    print("  1. 確認分頁內容正確")
    print("  2. 跑 python -c \"import sheets_sync; print(sheets_sync.upload_xlsx_to_sheets())\" 同步雲端")
    print("     (或在 Streamlit 按「⬆️ 上傳 xlsx → 協作表」)")


if __name__ == "__main__":
    main()
