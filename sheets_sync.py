# -*- coding: utf-8 -*-
"""規則庫.xlsx ←→ Google Sheet 同步模組。

設計理念 (方案 C):
    規則庫.xlsx 是 source of truth (跟同事在編的那份完全一樣)
    Google Sheet 結構也跟 xlsx 一模一樣 (每個槽體一個分頁 + meta 分頁)
    rules_extracted.csv 是「衍生物」, 自動由 xlsx 產生, 供 step3e 跑檢查

工作流:
    1. 上傳 xlsx → Sheet (push):
       - 把本機 規則庫.xlsx 推到 Sheet
       - Sheet 上每個分頁對應 xlsx 一個分頁 (槽體名/meta 名)
       - 完全覆蓋, 結構同 xlsx
    2. 下載 Sheet → xlsx (pull):
       - 把 Sheet 拉回, 覆寫 規則庫.xlsx
       - 自動備份舊 xlsx 到 backup/
       - 自動 export 一份扁平的 rules_extracted.csv (給 step3e 用)
    3. 直接重新 export csv (不動 Sheet):
       - 純粹從 xlsx 重產 csv

認證:
    Service Account JSON 從以下兩處讀 (依序):
    1. Streamlit secrets[gcp_service_account]
    2. 本機 service_account.json
"""
import csv
import json
import os
from datetime import datetime
try:
    from tz_util import fmt_tpe as _fmt_tpe
except Exception:
    def _fmt_tpe(fmt="%Y-%m-%d %H:%M:%S"):
        return datetime.now().strftime(fmt)

# 預設 Sheet 設定
DEFAULT_SHEET_ID = "1FOx4Wu1PVidbaC-89HBzyQSK7cBOOvxfu4RxLoBqwBQ"

BASE = os.path.dirname(os.path.abspath(__file__))
RULES_XLSX = os.path.join(BASE, "規則庫.xlsx")
RULES_CSV = os.path.join(BASE, "rules_extracted.csv")
BACKUP_DIR = os.path.join(BASE, "backup")
LOCAL_SA_PATH = os.path.join(BASE, "service_account.json")

# 扁平 csv 的欄位 (給 step3e 用)
FLAT_CSV_FIELDS = [
    "缺失ID", "來源", "技師姓名", "序號", "原文缺失",
    "檢查類型", "對照項目", "規則", "比對位置", "判定邏輯",
    "標準槽體名稱", "原始槽體代號", "狀態",
]

# xlsx 欄位名 → csv 欄位名 的映射 (xlsx 的「規則(萃取/可比對判斷式)」要映回 csv 的「規則」)
XLSX_TO_CSV_FIELD = {
    "缺失ID": "缺失ID",
    "來源": "來源",
    "原文缺失": "原文缺失",
    "檢查類型": "檢查類型",
    "對照項目": "對照項目",
    "規則(萃取/可比對判斷式)": "規則",
    "比對位置(依標題,非頁碼)": "比對位置",
    "判定邏輯(條件→結論)": "判定邏輯",
    "技師姓名": "技師姓名",
    "序號": "序號",
    "原始槽體代號": "原始槽體代號",
    "狀態": "狀態",
}


# ──────────────────────────────────────────────────
# 認證
# ──────────────────────────────────────────────────

def _get_service_account_info():
    """從 Streamlit Secrets 或本機檔案取得 service account 資訊。"""
    try:
        import streamlit as st
        if "gcp_service_account" in st.secrets:
            return dict(st.secrets["gcp_service_account"]), "streamlit"
    except Exception:
        pass

    if os.path.exists(LOCAL_SA_PATH):
        try:
            with open(LOCAL_SA_PATH, "r", encoding="utf-8") as f:
                return json.load(f), "local"
        except Exception:
            pass

    return None, None


def _get_gspread_client():
    sa_info, source = _get_service_account_info()
    if not sa_info:
        raise RuntimeError(
            "找不到 Google Service Account 認證。\n"
            "請完成以下其中一種設定:\n"
            "1. 本機: 把 service_account.json 放在專案根目錄\n"
            "2. Streamlit Cloud: Secrets 設 [gcp_service_account]\n"
            "詳見 SHEETS_SETUP.md"
        )

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError as e:
        raise RuntimeError(f"缺少套件: {e}. 請 pip install gspread google-auth")

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
    return gspread.authorize(creds), source


def get_service_account_email():
    sa_info, _ = _get_service_account_info()
    return sa_info.get("client_email") if sa_info else None


def check_auth_status():
    sa_info, source = _get_service_account_info()
    if not sa_info:
        return {"ok": False, "source": None, "email": None,
                "message": "未設定 Service Account"}
    return {
        "ok": True,
        "source": source,
        "email": sa_info.get("client_email", "(未知)"),
        "message": f"已認證 (來源: {source})",
    }


# ──────────────────────────────────────────────────
# _來源清單 自動統計
# ──────────────────────────────────────────────────

def refresh_source_list_stats():
    """重新計算 _來源清單 的「涵蓋槽體數 / 貢獻規則數 / 最後同步時間」欄。

    從各槽體分頁掃出每個「來源代號」的統計, 寫回 _來源清單。
    若 _來源清單 不存在或結構不符 (沒有自動欄), 就跳過。

    Returns:
        dict: {ok, updated_count, error}
    """
    from collections import defaultdict
    from openpyxl import load_workbook

    if not os.path.exists(RULES_XLSX):
        return {"ok": False, "error": "xlsx 不存在"}

    try:
        wb = load_workbook(RULES_XLSX)
    except Exception as e:
        return {"ok": False, "error": f"無法打開 xlsx: {e}"}

    if "_來源清單" not in wb.sheetnames:
        return {"ok": False, "error": "_來源清單 分頁不存在"}

    # 從各槽體分頁統計來源
    stats = defaultdict(lambda: {"tanks": set(), "rule_count": 0})
    for sn in wb.sheetnames:
        if sn.startswith("_"):
            continue
        ws = wb[sn]
        if ws.max_row < 2:
            continue
        headers = [c.value for c in ws[1]]
        try:
            source_idx = headers.index("來源") + 1
        except ValueError:
            continue
        for row in range(2, ws.max_row + 1):
            source = ws.cell(row=row, column=source_idx).value
            if not source:
                continue
            source_str = str(source).strip()
            code = source_str.split()[0] if source_str else ""
            if not code.startswith("S"):
                code = "S01"
            stats[code]["tanks"].add(sn)
            stats[code]["rule_count"] += 1

    # 更新 _來源清單
    ws_src = wb["_來源清單"]
    headers_src = [c.value for c in ws_src[1]]
    if "涵蓋槽體數" not in headers_src or "貢獻規則數" not in headers_src:
        return {"ok": False, "error": "_來源清單 沒有自動欄 (請先跑 setup_source_list_v3.py)"}

    code_idx = headers_src.index("來源代號") + 1
    tank_count_idx = headers_src.index("涵蓋槽體數") + 1
    rule_count_idx = headers_src.index("貢獻規則數") + 1
    sync_time_idx = (headers_src.index("最後同步時間") + 1
                     if "最後同步時間" in headers_src else None)

    sync_time = _fmt_tpe()
    updated = 0
    for row in range(2, ws_src.max_row + 1):
        code = ws_src.cell(row=row, column=code_idx).value
        if not code:
            continue
        code = str(code).strip()
        if code in stats:
            ws_src.cell(row=row, column=tank_count_idx, value=len(stats[code]["tanks"]))
            ws_src.cell(row=row, column=rule_count_idx, value=stats[code]["rule_count"])
            if sync_time_idx:
                ws_src.cell(row=row, column=sync_time_idx, value=sync_time)
            updated += 1
        else:
            # 該來源在槽體分頁找不到任何規則
            ws_src.cell(row=row, column=tank_count_idx, value=0)
            ws_src.cell(row=row, column=rule_count_idx, value=0)
            if sync_time_idx:
                ws_src.cell(row=row, column=sync_time_idx, value=sync_time)

    wb.save(RULES_XLSX)
    return {"ok": True, "updated_count": updated, "stats": {
        code: {"tank_count": len(s["tanks"]), "rule_count": s["rule_count"]}
        for code, s in stats.items()
    }}


# ──────────────────────────────────────────────────
# xlsx 讀寫工具
# ──────────────────────────────────────────────────

def _read_xlsx_to_dict():
    """讀 規則庫.xlsx → {sheet_name: [[row1], [row2], ...]} (含表頭)。"""
    from openpyxl import load_workbook
    if not os.path.exists(RULES_XLSX):
        raise RuntimeError(f"找不到 {RULES_XLSX}")
    wb = load_workbook(RULES_XLSX, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # 把 None 轉空字串
            rows.append([("" if v is None else str(v)) for v in row])
        # 移除尾端全空白列
        while rows and not any(c.strip() for c in rows[-1]):
            rows.pop()
        out[sn] = rows
    return out, wb.sheetnames  # 保留分頁順序


def _write_xlsx_from_dict(sheets_data, sheet_order):
    """把 {sheet_name: [[row1], ...]} 寫回 規則庫.xlsx, 用 sheet_order 控制分頁順序。"""
    from openpyxl import Workbook
    wb = Workbook()
    # 移除預設分頁
    default_ws = wb.active
    wb.remove(default_ws)

    for sn in sheet_order:
        if sn not in sheets_data:
            continue
        ws = wb.create_sheet(title=sn)
        for row in sheets_data[sn]:
            ws.append(row)
    wb.save(RULES_XLSX)


def export_xlsx_to_csv():
    """從 規則庫.xlsx 各槽體分頁產出扁平 rules_extracted.csv。

    Returns:
        dict: {ok, rows_written, error}
    """
    try:
        from openpyxl import load_workbook
        if not os.path.exists(RULES_XLSX):
            return {"ok": False, "error": f"找不到 {RULES_XLSX}"}

        wb = load_workbook(RULES_XLSX, data_only=True)
        flat_rows = []
        for sn in wb.sheetnames:
            if sn.startswith("_"):
                continue  # 跳過 meta 分頁
            ws = wb[sn]
            if ws.max_row < 2:
                continue
            xlsx_headers = [c.value for c in ws[1]]
            for row_idx in range(2, ws.max_row + 1):
                row_dict = {}
                for col_idx, h in enumerate(xlsx_headers, start=1):
                    if h is None:
                        continue
                    csv_field = XLSX_TO_CSV_FIELD.get(h)
                    if not csv_field:
                        continue
                    v = ws.cell(row=row_idx, column=col_idx).value
                    row_dict[csv_field] = "" if v is None else str(v)
                # 標準槽體名稱 = 分頁名
                row_dict["標準槽體名稱"] = sn
                # 跳過全空白列
                if not any(row_dict.get(f, "").strip() for f in FLAT_CSV_FIELDS):
                    continue
                flat_rows.append(row_dict)

        # 寫 csv
        with open(RULES_CSV, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=FLAT_CSV_FIELDS, extrasaction="ignore")
            writer.writeheader()
            for r in flat_rows:
                writer.writerow({k: r.get(k, "") for k in FLAT_CSV_FIELDS})

        return {
            "ok": True,
            "rows_written": len(flat_rows),
            "csv_path": RULES_CSV,
            "timestamp": _fmt_tpe(),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ──────────────────────────────────────────────────
# 備份
# ──────────────────────────────────────────────────

def _backup_current_xlsx():
    """把現有 規則庫.xlsx 跟 rules_extracted.csv 一起備份。"""
    if not os.path.exists(RULES_XLSX):
        return None
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = _fmt_tpe("%Y%m%d_%H%M%S")

    backup_info = {}
    # 備份 xlsx
    xlsx_bak = os.path.join(BACKUP_DIR, f"規則庫_{ts}.xlsx")
    with open(RULES_XLSX, "rb") as src, open(xlsx_bak, "wb") as dst:
        dst.write(src.read())
    backup_info["xlsx"] = xlsx_bak

    # 順手備份 csv
    if os.path.exists(RULES_CSV):
        csv_bak = os.path.join(BACKUP_DIR, f"rules_{ts}.csv")
        with open(RULES_CSV, "rb") as src, open(csv_bak, "wb") as dst:
            dst.write(src.read())
        backup_info["csv"] = csv_bak

    return backup_info


def list_backups():
    if not os.path.exists(BACKUP_DIR):
        return []
    items = []
    for f in os.listdir(BACKUP_DIR):
        if not (f.endswith(".csv") or f.endswith(".xlsx")):
            continue
        path = os.path.join(BACKUP_DIR, f)
        items.append({
            "name": f,
            "path": path,
            "size_kb": round(os.path.getsize(path) / 1024, 1),
            "mtime": datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S"),
        })
    items.sort(key=lambda x: x["mtime"], reverse=True)
    return items


# ──────────────────────────────────────────────────
# 上傳: xlsx → Sheet (每個分頁對應)
# ──────────────────────────────────────────────────

def upload_xlsx_to_sheets(sheet_id=None):
    """把 規則庫.xlsx 整份推到 Google Sheet, 每個分頁對應。

    上傳前會自動更新 _來源清單 的統計欄位。

    Returns:
        dict: {ok, sheets_written: [(name, rows, cols)], sheet_url, source, error}
    """
    sheet_id = sheet_id or DEFAULT_SHEET_ID

    # 先重算來源清單統計 (讓 Sheet 上看到的數字也是新的)
    refresh_source_list_stats()

    try:
        sheets_data, sheet_order = _read_xlsx_to_dict()
    except Exception as e:
        return {"ok": False, "error": f"讀 xlsx 失敗: {e}"}

    if not sheets_data:
        return {"ok": False, "error": "xlsx 是空的"}

    try:
        client, source = _get_gspread_client()
        sh = client.open_by_key(sheet_id)

        # 取得目前 Sheet 上所有分頁
        existing_ws = {ws.title: ws for ws in sh.worksheets()}

        sheets_written = []
        for sn in sheet_order:
            rows = sheets_data[sn]
            if not rows:
                rows = [[""]]  # 至少要 1 個 cell

            if sn in existing_ws:
                ws = existing_ws[sn]
                ws.clear()
            else:
                ws = sh.add_worksheet(title=sn, rows=max(len(rows) + 50, 100),
                                      cols=max(len(rows[0]) if rows else 1, 12))

            # 寫入 (range_name 用 A1 起算)
            ws.update(values=rows, range_name="A1")
            sheets_written.append({
                "name": sn,
                "rows": len(rows),
                "cols": len(rows[0]) if rows else 0,
            })

        # 刪除 Sheet 上多餘的分頁 (xlsx 沒有的)
        removed = []
        for ws_name, ws in existing_ws.items():
            if ws_name not in sheets_data:
                # Sheet 至少要保留 1 個分頁, 不能刪到 0
                if len(sh.worksheets()) > 1:
                    sh.del_worksheet(ws)
                    removed.append(ws_name)

        # 為「狀態」欄加 Data Validation (下拉選單: V / ?)
        validations_added = _apply_status_validation(sh, sheets_data)

        return {
            "ok": True,
            "sheets_written": sheets_written,
            "sheets_removed": removed,
            "validations_added": validations_added,
            "total_data_rows": sum(s["rows"] - 1 for s in sheets_written if s["rows"] > 1),
            "sheet_url": f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit",
            "source": source,
            "timestamp": _fmt_tpe(),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _apply_status_validation(sh, sheets_data):
    """為每個槽體分頁的「狀態」欄加 Data Validation 下拉選單 (V / ?)。

    用 Google Sheets API 的 batchUpdate 一次處理所有分頁。
    """
    requests = []
    for sn, rows in sheets_data.items():
        if sn.startswith("_"):
            continue  # meta 分頁不加
        if not rows or len(rows) < 1:
            continue
        headers = rows[0]
        if "狀態" not in headers:
            continue
        status_col_idx = headers.index("狀態")  # 0-based

        # 找到該分頁的 sheetId
        try:
            ws = sh.worksheet(sn)
        except Exception:
            continue
        sheet_gid = ws.id  # gspread Worksheet.id 就是 sheetId

        # 對狀態欄 (從第 2 列到第 1000 列) 加 Data Validation
        requests.append({
            "setDataValidation": {
                "range": {
                    "sheetId": sheet_gid,
                    "startRowIndex": 1,  # 跳過表頭
                    "endRowIndex": 1000,
                    "startColumnIndex": status_col_idx,
                    "endColumnIndex": status_col_idx + 1,
                },
                "rule": {
                    "condition": {
                        "type": "ONE_OF_LIST",
                        "values": [
                            {"userEnteredValue": "V"},
                            {"userEnteredValue": "?"},
                        ],
                    },
                    "showCustomUi": True,
                    "strict": False,  # 允許空白
                    "inputMessage": "V = 已核可 (空白也視為 V) / ? = 待討論 (系統會跳過)",
                },
            }
        })

    if requests:
        try:
            sh.batch_update({"requests": requests})
        except Exception as e:
            # Data Validation 失敗不致命, 只記錯誤
            return f"FAILED: {e}"
    return len(requests)


# ──────────────────────────────────────────────────
# 下載: Sheet → xlsx (自動備份 + 自動產 csv)
# ──────────────────────────────────────────────────

def download_sheets_to_xlsx(sheet_id=None):
    """從 Sheet 拉資料覆寫 規則庫.xlsx (覆寫前自動備份, 完成後自動產 csv)。"""
    sheet_id = sheet_id or DEFAULT_SHEET_ID

    try:
        client, source = _get_gspread_client()
        sh = client.open_by_key(sheet_id)

        sheets_data = {}
        sheet_order = []
        for ws in sh.worksheets():
            sheet_order.append(ws.title)
            sheets_data[ws.title] = ws.get_all_values()

        if not sheets_data:
            return {"ok": False, "error": "Sheet 是空的"}

        # 先備份舊版
        backup = _backup_current_xlsx()

        # 寫回 xlsx
        _write_xlsx_from_dict(sheets_data, sheet_order)

        # 重算 _來源清單 統計 (同事可能新增了規則)
        refresh_source_list_stats()

        # 順便重新 export csv
        export_result = export_xlsx_to_csv()

        # 統計
        total_data_rows = sum(
            len(rows) - 1 for sn, rows in sheets_data.items()
            if not sn.startswith("_") and len(rows) > 1
        )

        return {
            "ok": True,
            "sheets_read": len(sheets_data),
            "total_data_rows": total_data_rows,
            "csv_export": export_result,
            "backup": backup,
            "sheet_url": f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit",
            "source": source,
            "timestamp": _fmt_tpe(),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ──────────────────────────────────────────────────
# 預覽差異 (依槽體分頁)
# ──────────────────────────────────────────────────

def preview_diff(sheet_id=None):
    """比對 Sheet 跟 xlsx 在「規則層級」(以 缺失ID 為主鍵) 有什麼差異。

    Returns:
        dict: {
            ok, by_tank: {槽體: {added: [], removed: [], changed: []}},
            sheets_only: [], xlsx_only: [], total_changed,
            sheet_total, xlsx_total, error
        }
    """
    sheet_id = sheet_id or DEFAULT_SHEET_ID

    try:
        client, source = _get_gspread_client()
        sh = client.open_by_key(sheet_id)
        sheet_sheets = {ws.title: ws.get_all_values() for ws in sh.worksheets()}

        xlsx_sheets, _ = _read_xlsx_to_dict()
    except Exception as e:
        return {"ok": False, "error": str(e)}

    # 取所有 (非 meta) 分頁名
    all_tanks = sorted({n for n in list(sheet_sheets) + list(xlsx_sheets)
                        if not n.startswith("_")})

    by_tank = {}
    sheet_total = 0
    xlsx_total = 0
    total_changed = 0
    total_added = 0
    total_removed = 0

    for tank in all_tanks:
        sheet_rows = sheet_sheets.get(tank, [])
        xlsx_rows = xlsx_sheets.get(tank, [])

        # 第一列當表頭, 用 缺失ID 當 key
        def rows_to_dict(rows):
            if not rows:
                return {}, []
            headers = rows[0]
            try:
                id_idx = headers.index("缺失ID")
            except ValueError:
                return {}, headers
            d = {}
            for r in rows[1:]:
                if id_idx < len(r):
                    rid = r[id_idx].strip()
                    if rid:
                        d[rid] = r
            return d, headers

        sheet_dict, sheet_headers = rows_to_dict(sheet_rows)
        xlsx_dict, xlsx_headers = rows_to_dict(xlsx_rows)

        added = [rid for rid in sheet_dict if rid not in xlsx_dict]
        removed = [rid for rid in xlsx_dict if rid not in sheet_dict]
        changed = []
        for rid in sheet_dict:
            if rid in xlsx_dict:
                if [str(x).strip() for x in sheet_dict[rid]] != \
                   [str(x).strip() for x in xlsx_dict[rid]]:
                    changed.append(rid)

        if added or removed or changed:
            by_tank[tank] = {
                "added": added, "removed": removed, "changed": changed,
            }

        sheet_total += len(sheet_dict)
        xlsx_total += len(xlsx_dict)
        total_added += len(added)
        total_removed += len(removed)
        total_changed += len(changed)

    return {
        "ok": True,
        "by_tank": by_tank,
        "sheet_total": sheet_total,
        "xlsx_total": xlsx_total,
        "total_added": total_added,
        "total_removed": total_removed,
        "total_changed": total_changed,
        "source": source,
    }


# ──────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    import io
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    except Exception:
        pass

    print("=== 認證狀態 ===")
    status = check_auth_status()
    print(status)
    if not status["ok"]:
        sys.exit(0)

    cmd = sys.argv[1] if len(sys.argv) > 1 else "preview"

    if cmd == "upload":
        print("\n=== 上傳 xlsx → Sheet ===")
        r = upload_xlsx_to_sheets()
    elif cmd == "download":
        print("\n=== 下載 Sheet → xlsx ===")
        r = download_sheets_to_xlsx()
    elif cmd == "preview":
        print("\n=== 預覽差異 ===")
        r = preview_diff()
    elif cmd == "export":
        print("\n=== 從 xlsx 重產 csv ===")
        r = export_xlsx_to_csv()
    elif cmd == "backups":
        print("\n=== 備份清單 ===")
        for b in list_backups():
            print(f"  {b['mtime']}  {b['name']}  ({b['size_kb']} KB)")
        sys.exit(0)
    else:
        print(f"用法: python sheets_sync.py [upload|download|preview|export|backups]")
        sys.exit(0)

    print(json.dumps(r, ensure_ascii=False, indent=2))
