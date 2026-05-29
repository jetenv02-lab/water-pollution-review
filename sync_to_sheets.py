# -*- coding: utf-8 -*-
"""Step 5: 同步 規則庫.xlsx 到 Google Sheets。

兩種模式:
  A) 手動匯入(推薦,免設定): 開啟 Google Sheets → 檔案 → 匯入 → 上傳 規則庫.xlsx
  B) 自動同步: 需先設定 Google Cloud 專案 + Service Account 金鑰

本程式預設用「模式 B 自動同步」。
若尚未設定金鑰,執行時會印出設定步驟。

依賴: gspread, google-auth (pip install gspread google-auth)
設定:
  1) Google Cloud Console 開啟 Google Sheets API + Google Drive API
  2) 建立 Service Account → 下載 JSON 金鑰 → 存成 service_account.json (與本檔同目錄)
  3) 把目標 Google Sheet 分享給 service account 的 email (在 JSON 裡的 client_email)
  4) 把 Sheet 的 URL 填到下面 SHEET_URL

用法: python sync_to_sheets.py
"""
import os
import openpyxl

BASE = r"C:\Users\jeten\Desktop\AI\水措審查"
XLSX_IN = os.path.join(BASE, "規則庫.xlsx")
SA_JSON = os.path.join(BASE, "service_account.json")

# 在這裡填你 Google Sheet 的 URL(三人協作的那份)
SHEET_URL = ""  # 例如: "https://docs.google.com/spreadsheets/d/1AbCdEfGhIjKlMnOpQrStUvWxYz/edit"


def print_setup_guide():
    print("=" * 60)
    print("Google Sheets 自動同步設定步驟")
    print("=" * 60)
    print("1) 前往 https://console.cloud.google.com/")
    print("2) 建立或選擇一個專案")
    print("3) 「API 和服務 → 程式庫」啟用:")
    print("   - Google Sheets API")
    print("   - Google Drive API")
    print("4) 「IAM 和管理 → 服務帳戶」建立服務帳戶")
    print("5) 在該服務帳戶 → 金鑰 → 新增金鑰 → JSON → 下載")
    print(f"6) 把下載的 JSON 重新命名為 service_account.json,放到 {BASE}")
    print("7) 用 Google Sheets 建立一份試算表,把 service account 的 email")
    print("   (在 JSON 裡的 client_email 欄位)加為編輯者")
    print(f"8) 編輯 {os.path.abspath(__file__)} 的 SHEET_URL 變數")
    print("9) pip install gspread google-auth")
    print("10) 重新執行: python sync_to_sheets.py")
    print()
    print("替代方案: 直接開 Google Sheets → 檔案 → 匯入 → 上傳 規則庫.xlsx (手動)")


def sync():
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("未安裝 gspread / google-auth")
        print("請執行: pip install gspread google-auth")
        return

    if not os.path.exists(SA_JSON):
        print(f"找不到 service account 金鑰: {SA_JSON}")
        print_setup_guide()
        return
    if not SHEET_URL:
        print("尚未設定 SHEET_URL")
        print_setup_guide()
        return
    if not os.path.exists(XLSX_IN):
        print(f"找不到規則庫: {XLSX_IN}")
        return

    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(SA_JSON, scopes=SCOPES)
    gc = gspread.authorize(creds)
    sh = gc.open_by_url(SHEET_URL)
    print(f"已連線到: {sh.title}")

    wb = openpyxl.load_workbook(XLSX_IN, read_only=True)
    for sname in wb.sheetnames:
        ws_src = wb[sname]
        data = []
        for row in ws_src.iter_rows(values_only=True):
            data.append(["" if v is None else str(v) for v in row])
        if not data:
            continue
        # 取得或建立對應分頁
        try:
            ws_dst = sh.worksheet(sname)
            ws_dst.clear()
        except gspread.WorksheetNotFound:
            ws_dst = sh.add_worksheet(title=sname, rows=max(len(data) + 10, 100), cols=max(len(data[0]) + 5, 20))
        ws_dst.update("A1", data, value_input_option="RAW")
        print(f"  同步分頁: {sname} ({len(data)} 列)")

    print(f"\n同步完成 → {sh.url}")


if __name__ == "__main__":
    if not SHEET_URL or not os.path.exists(SA_JSON):
        print_setup_guide()
    else:
        sync()
