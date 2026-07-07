# -*- coding: utf-8 -*-
"""加藥規則 → 質量補償計算 (從規則庫 _加藥規則 讀, 非 hardcode)。

設計理念:
    - chemical_calc.py 用 hardcode 典型值 (v1 第一版)
    - 此模組改從 規則庫.xlsx 的 _加藥規則 讀 (v2)
    - chemical_calc.py 不動, 此模組純新增, 對外提供同名 API
    - tank_chemistry.py 之後 patch 時, 可選擇用哪個版本

讀的欄位:
    B 分流系統 / C 槽體類別 / D 藥劑 / E 加藥角色
    R(18) 典型劑量 (mg/L) / S(19) 引入水質項目 / T(20) 轉換係數

對外 API (跟 chemical_calc 一致, 可直接 swap):
    compute_chemical_mass(unit, item, q_cmd=None) → kg/d
    get_typical_dosing(std_tank) → dict[chem → mg/L]
    describe_chemical_contribution(unit, item, q_cmd=None) → str
"""
import os
import re
from functools import lru_cache

import openpyxl

DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "規則庫.xlsx")
SHEET_NAME = "_加藥規則"


def _split_multi(s):
    """處理多項目格式: 'a; b; c' → ['a', 'b', 'c']"""
    if not s:
        return []
    return [x.strip() for x in re.split(r"[;,，；]", str(s)) if x.strip()]


def _to_float(v):
    try:
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


@lru_cache(maxsize=1)
def load_dosing_compensation(xlsx_path=None):
    """從 _加藥規則 讀「應加」藥劑 + 補償欄, 整理成可查表結構。

    回傳:
        {
            標準槽體名: [
                {
                    "drug": "PAC",
                    "system": "各系",
                    "mg_per_L": 100,
                    "items": ["懸浮固體（mg/L）", "氯離子"],
                    "coefs": [0.30, 0.05],
                },
                ...
            ],
            ...
        }
    """
    path = xlsx_path or DEFAULT_XLSX
    if not os.path.exists(path):
        return {}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[dosing_rules_loader] 讀 {path} 失敗: {e}")
        return {}

    if SHEET_NAME not in wb.sheetnames:
        return {}
    ws = wb[SHEET_NAME]

    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 20:
            continue
        system = row[1]   # B
        tank = row[2]     # C
        drug = row[3]     # D
        role = row[4]     # E
        mg_per_L = _to_float(row[17])  # R (idx 17 = col 18)
        items_str = row[18]              # S
        coefs_str = row[19]              # T

        if not tank or not drug:
            continue
        # 只處理「應加」(禁加/可加不算進補償)
        if str(role or "").strip() != "應加":
            continue
        if mg_per_L is None or mg_per_L <= 0:
            continue

        items = _split_multi(items_str)
        coefs = [_to_float(c) or 0.0 for c in _split_multi(coefs_str)]

        # 對齊 items 與 coefs 長度 (短的補 0)
        n = max(len(items), len(coefs))
        items = (items + [""] * n)[:n]
        coefs = (coefs + [0.0] * n)[:n]

        entry = {
            "drug": str(drug).strip(),
            "system": str(system or "").strip(),
            "mg_per_L": mg_per_L,
            "items": items,
            "coefs": coefs,
        }

        tank_key = str(tank).strip()
        out.setdefault(tank_key, []).append(entry)

    return out


# 槽體別名: PDF 上可能的槽體名 → _加藥規則 的標準名
TANK_ALIAS = {
    "pH調整槽": "pH調整池",
    "pH 調整槽": "pH調整池",
    "中和槽": "中和池",
    "快混槽": "快混池",
    "慢混槽": "慢混池",
    "氰系氧化槽": "氰系氧化池",
    "鉻系還原槽": "鉻系還原池",
    "硫化物沉澱槽": "硫化物沉澱池",
    "pH調整暨快混池": "快混池",       # 用 PAC 同
    "pH調整池暨快混池": "快混池",
    "混凝膠凝池": "混凝池",
    "混凝池": "混凝池",
}


def get_typical_dosing(std_tank):
    """回傳該標準槽體的藥劑 → mg/L (跟 chemical_calc API 相同)。"""
    db = load_dosing_compensation()
    key = TANK_ALIAS.get(std_tank, std_tank)
    entries = db.get(key, [])
    return {e["drug"]: e["mg_per_L"] for e in entries}


# ──────────────────────────────────────────────────
# 從廠商申報 (measure_params) 讀實際 kg/day
# 2026-07-01: Nick 反映廠商實際劑量在處理設施表中, step2 已抽出
# 例: measure_params["加藥量(NaOH（45％）)"] = {"min": 0.87, "max": 8.672}
# ──────────────────────────────────────────────────

# 藥劑名同義字對照 (廠商 PDF 上的名稱 → 規則庫的標準名)
DRUG_NAME_ALIAS = {
    "NaOH": ["NaOH", "氫氧化鈉", "苛性鈉", "燒鹼", "片鹼"],
    "H2SO4": ["H2SO4", "硫酸", "H₂SO₄"],
    "HCl": ["HCl", "鹽酸", "氫氯酸"],
    "Ca(OH)2": ["Ca(OH)2", "氫氧化鈣", "石灰", "石灰乳", "熟石灰"],
    "CaCl2": ["CaCl2", "氯化鈣", "CaCl₂"],
    "PAC": ["PAC", "多氯化鋁", "聚合氯化鋁", "AL13"],
    "PAM": ["PAM", "Polymer", "polymer", "高分子", "聚丙烯醯胺",
            "助凝劑", "陽離子高分子", "陰離子高分子"],
    "FeCl3": ["FeCl3", "三氯化鐵", "氯化鐵"],
    "Al2(SO4)3": ["Al2(SO4)3", "硫酸鋁", "鋁明礬", "Al₂(SO₄)₃"],
    "FeSO4": ["FeSO4", "硫酸亞鐵", "FeSO₄"],
    "NaClO": ["NaClO", "NaOCl", "次氯酸鈉", "漂白水"],
    "NaHSO3": ["NaHSO3", "亞硫酸氫鈉", "NaHSO₃"],
    "H2O2": ["H2O2", "過氧化氫", "雙氧水", "H₂O₂"],
    "Na2S": ["Na2S", "硫化鈉", "Na₂S"],
    "尿素": ["尿素", "urea", "CO(NH2)2"],
    "磷酸": ["磷酸", "H3PO4", "H₃PO₄"],
    "活性碳": ["活性碳", "activated carbon", "粉末活性碳"],  # 移除 "AC" 避免跟 PAC 混淆
}

# 藥劑「片段對照」— 用於 step2 排版壞的 case
# 例 "氫氧化依加藥桶液位鈉" 應該是「氫氧化鈉」被插入了「依加藥桶液位」
# 判斷: 每組 tuple 內全部片段都在 text 中 (順序無所謂)
# 單字元 tuple 也支援 (例 ("化鈉",) — 用於 PDF 抽取只留末段的 case)
DRUG_FRAGMENT_ALIAS = {
    "NaOH": [("氫氧化", "鈉"), ("化鈉",)],           # 「化鈉（45」也算 NaOH
    "H2SO4": [("硫", "酸")],
    "Ca(OH)2": [("氫氧化", "鈣"), ("化鈣",)],         # 「化鈣」也算 Ca(OH)2 (但可能跟 CaCl2 衝突, 見下)
    "CaCl2": [("氯化", "鈣")],
    "FeSO4": [("硫酸", "亞鐵"), ("硫酸", "鐵")],
    "FeCl3": [("氯化", "鐵"), ("三氯化", "鐵")],
    "NaClO": [("次氯酸", "鈉")],
    "NaHSO3": [("亞硫酸", "鈉"), ("亞硫", "鈉")],
    "Na2S": [("硫化", "鈉")],
}


def _match_drug_name(text, drug_key):
    """比對 measure_params 的參數名是否包含指定藥劑.
    支援兩種比對:
    1. 完整字串比對 (DRUG_NAME_ALIAS)
    2. 片段對照 (DRUG_FRAGMENT_ALIAS) — 兩片段都在 text 中 (order-free), 支援排版壞的 case
    """
    if not text:
        return False
    text_str = str(text)
    # (1) 完整字串比對
    aliases = DRUG_NAME_ALIAS.get(drug_key, [drug_key])
    for alias in aliases:
        if alias in text_str:
            return True
    # (2) 片段對照
    for fragments in DRUG_FRAGMENT_ALIAS.get(drug_key, []):
        if all(frag in text_str for frag in fragments):
            return True
    return False


def _parse_concentration_pct(param_name):
    """從參數名解析商品濃度 (%).
    支援多種變形:
    - 標準: '加藥量(NaOH（45％）)' → 45.0
    - step2 排版壞的: '加藥量(氫氧化依加藥桶液位鈉45％)' → 45.0
    - 純數字後接%: '硫酸 10%' → 10.0
    找不到回 None."""
    if not param_name:
        return None
    text = str(param_name)
    # 找任何位置的 "數字 %" 或 "數字 ％" (兩種百分符號)
    # step2 可能把濃度插入雜訊: "硫酸10依加藥桶液位％" 這種數字跟%被拆開
    # 用兩階段:
    # 1) 連續 數字+% (標準)
    # 2) 數字後 <= 12 字元有 %  (排版壞)
    m = re.search(r"(\d+(?:\.\d+)?)\s*[％%]", text)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    m = re.search(r"(\d+(?:\.\d+)?)[\u4e00-\u9fff]{1,12}[％%]", text)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    return None


def get_declared_dosing_kg_per_day(unit, drug_key):
    """從 measure_params 讀廠商申報的藥劑 kg/day (純物質量, 已扣商品濃度).

    Args:
        unit: 處理單元 dict
        drug_key: 藥劑 key (例 "NaOH", "PAC")

    Returns:
        (avg_kg_per_day_純, 商品濃度%) 或 (None, None) 若未申報

    例:
        param "加藥量(NaOH（45％）)" min=0.87 max=8.672
        → 商品平均 4.77 kg/d, NaOH 純量 = 4.77 × 45% = 2.15 kg/d
        → 回傳 (2.15, 45.0)
    """
    if not isinstance(unit, dict):
        return (None, None)
    measure = unit.get("measure_params") or {}
    for pname, pval in measure.items():
        if not _match_drug_name(pname, drug_key):
            continue
        if "加藥" not in str(pname):
            continue
        if not isinstance(pval, dict):
            continue
        vmin = pval.get("min")
        vmax = pval.get("max")
        try:
            vmin = float(vmin) if vmin is not None else None
            vmax = float(vmax) if vmax is not None else None
        except (TypeError, ValueError):
            continue
        # 取平均當估算 (min/max 為操作範圍)
        if vmin is not None and vmax is not None:
            avg_commercial = (vmin + vmax) / 2
        elif vmax is not None:
            avg_commercial = vmax
        elif vmin is not None:
            avg_commercial = vmin
        else:
            continue

        # 扣商品濃度得純物質
        pct = _parse_concentration_pct(pname)
        if pct is not None and pct > 0:
            pure_kg = avg_commercial * pct / 100.0
        else:
            pure_kg = avg_commercial  # 假設是純物質 (若沒標濃度)
            pct = 100.0

        return (pure_kg, pct)

    return (None, None)


def _has_any_declared_dosing(unit):
    """判斷該單元是否有廠商申報「任何加藥量」.

    比對條件 (符合任一即可):
        (1) 欄位名含 "加藥" 二字 + 值有數字
        (2) 欄位名含任何**藥劑名** (見 DRUG_NAME_ALIAS / DRUG_FRAGMENT_ALIAS) + 值有數字
            — 用於 PDF 抽取壞版的 case, 例如「化鈉（45」被截斷但仍含「化鈉」片段

    這樣: 廠商申報氫氧化鈉一種就算「有申報」, 我們不再幫他補其他藥的典型值.
    """
    if not isinstance(unit, dict):
        return False
    measure = unit.get("measure_params") or {}
    for pname, pval in measure.items():
        text = str(pname)
        if not isinstance(pval, dict):
            continue
        try:
            vmin = float(pval.get("min")) if pval.get("min") is not None else None
            vmax = float(pval.get("max")) if pval.get("max") is not None else None
        except (TypeError, ValueError):
            continue
        has_number = (vmin is not None and vmin > 0) or (vmax is not None and vmax > 0)
        if not has_number:
            continue
        # (1) 有 "加藥" 二字
        if "加藥" in text:
            return True
        # (2) 有任何已知藥劑名 (完整別名或片段)
        for drug_key in DRUG_NAME_ALIAS.keys():
            if _match_drug_name(text, drug_key):
                return True
    return False


def compute_chemical_mass(unit, item, q_cmd=None):
    """計算該單元因加藥引入指定水質項目的質量 (kg/d).

    優先級:
    (1) 若廠商申報**任何**加藥 → 只用廠商申報 (不 fallback 典型值)
        避免「幫廠商想像沒申報的藥」
    (2) 若廠商完全沒申報 → 用規則庫典型值 mg/L × Q 全套估算

    2026-07-01 v3 (Nick 反映): 舊版對每個藥獨立判斷,
    導致廠商只申報 A 藥時, 系統會幫 B/C 藥用典型值補. 這是不對的.

    Args:
        unit: 處理單元 dict (有 std_tank + measure_params)
        item: 水質項目名 (例: "懸浮固體（mg/L）")
        q_cmd: 進流總 Q (m³/d). None 時自動加總 stream_q 的 WTB.

    Returns:
        kg/d (float)
    """
    std_tank = unit.get("std_tank") if isinstance(unit, dict) else None
    if not std_tank:
        return 0.0

    db = load_dosing_compensation()
    key = TANK_ALIAS.get(std_tank, std_tank)
    entries = db.get(key, [])
    if not entries:
        return 0.0

    if q_cmd is None:
        sq = unit.get("stream_q") or {}
        q_cmd = 0.0
        for stream_code, info in sq.items():
            if stream_code.startswith("WTB") and isinstance(info, dict):
                q = info.get("q_cmd")
                if isinstance(q, (int, float)):
                    q_cmd += float(q)

    # v4 (Nick 2026-07-07): 廠商沒申報 kg/d 時, 不再自動用典型值 fallback
    # 原因: 典型值多藥累加會爆掉 (T04-08 pH調整槽 沒申報 → 3 藥典型值 →
    #      Ca(OH)2 60 + 活性碳 100 + CaCl2 50 = 232 kg/d SS → 誤判 SS 減 80%)
    # 廠商沒申報加藥的問題交給 step3d/e (加藥機制) 檢查, 不靠質量平衡撈
    total = 0.0
    for entry in entries:
        # 找該藥劑對該水質項目的轉換係數
        item_coef = 0.0
        for it, coef in zip(entry["items"], entry["coefs"]):
            if it == item and coef > 0:
                item_coef = coef
                break
        if item_coef <= 0:
            continue
        # 只用廠商實際申報的 kg/d 值 (v4: 拿掉典型值 fallback)
        declared_kg, _pct = get_declared_dosing_kg_per_day(unit, entry["drug"])
        if declared_kg is not None and declared_kg > 0:
            total += declared_kg * item_coef
    return total


def describe_chemical_contribution(unit, item, q_cmd=None):
    """描述加藥引入細節, 給 finding 用。

    回傳: str (可空字串), 例:
        "[加藥引入估算: PAC 100mg/L×0.3=710 → +710 kg/d]"
    """
    std_tank = unit.get("std_tank") if isinstance(unit, dict) else None
    if not std_tank:
        return ""

    db = load_dosing_compensation()
    key = TANK_ALIAS.get(std_tank, std_tank)
    entries = db.get(key, [])
    if not entries:
        return ""

    if q_cmd is None:
        sq = unit.get("stream_q") or {}
        q_cmd = 0.0
        for stream_code, info in sq.items():
            if stream_code.startswith("WTB") and isinstance(info, dict):
                q = info.get("q_cmd")
                if isinstance(q, (int, float)):
                    q_cmd += float(q)

    # v4 (Nick 2026-07-07): 只描述廠商真正申報的加藥, 不再幫他猜典型值
    parts = []
    total = 0.0
    for entry in entries:
        for it, coef in zip(entry["items"], entry["coefs"]):
            if it != item or coef <= 0:
                continue
            declared_kg, pct = get_declared_dosing_kg_per_day(unit, entry["drug"])
            if declared_kg is not None and declared_kg > 0:
                added = declared_kg * coef
                parts.append(
                    f"{entry['drug']} 廠商申報 {declared_kg:.2f}kg/d純物質"
                    f"(@{pct:.0f}%)×{coef}={added:.1f}"
                )
                total += added

    if not parts:
        return ""
    return f"[加藥引入估算 {item}: {'; '.join(parts)} → +{total:.1f} kg/d]"


def _self_test():
    """跑跟 chemical_calc 一樣的 5 個測試, 確認從 Sheets 讀的結果一致。"""
    db = load_dosing_compensation()
    print(f"從 _加藥規則 讀到 {len(db)} 個標準槽體有加藥配方:")
    for k, v in db.items():
        chems = [(e["drug"], e["mg_per_L"]) for e in v]
        print(f"  {k}: {chems}")
    print()

    # Test 1: 快混池加 PAC 100 mg/L, Q = 23673 m³/d
    fake_unit = {
        "std_tank": "快混池",
        "stream_q": {"WTB01-01-1": {"q_cmd": 23673}},
    }
    ss = compute_chemical_mass(fake_unit, "懸浮固體（mg/L）")
    expected = 100 * 23673 / 1000 * 0.30
    print(f"Test 1 (快混池 PAC → SS): {ss:.2f} kg/d (預期 {expected:.2f})")
    assert abs(ss - expected) < 0.1

    # Test 2: 中和池 NaOH 不引入 SS
    n_unit = {"std_tank": "中和池", "stream_q": {"WTB": {"q_cmd": 1000}}}
    ss2 = compute_chemical_mass(n_unit, "懸浮固體（mg/L）")
    print(f"Test 2 (中和池 NaOH → SS): {ss2:.2f} kg/d (預期 0)")
    assert ss2 == 0.0

    # Test 3: 中和池 H2SO4 引入 SO4
    so4 = compute_chemical_mass(n_unit, "硫酸根")
    expected = 30 * 1000 / 1000 * 0.98
    print(f"Test 3 (中和池 H2SO4 → SO4): {so4:.2f} kg/d (預期 {expected:.2f})")
    assert abs(so4 - expected) < 0.1

    # Test 4: 描述
    desc = describe_chemical_contribution(fake_unit, "懸浮固體（mg/L）")
    print(f"Test 4 描述: {desc}")
    assert "PAC" in desc and "+710" in desc

    # Test 5: pH 槽 別名測試 - pH 調整池 含 Ca(OH)2 會引入 SS
    # (跟 hardcode v1 不同: v1 pH 槽只有 NaOH/H2SO4 不引 SS, v2 Sheets 多了鎳系 Ca(OH)2 → 會引)
    ph_unit = {"std_tank": "pH調整槽", "stream_q": {"WTB": {"q_cmd": 1000}}}
    ph_ss = compute_chemical_mass(ph_unit, "懸浮固體（mg/L）")
    # 預期: Ca(OH)2 60 mg/L × 0.50 = 30 kg/d
    print(f"Test 5 (pH調整槽 別名→pH調整池, Ca(OH)2 → SS): {ph_ss:.2f} kg/d (預期 30, 來自鎳系 Ca(OH)2)")
    assert ph_ss == 30.0

    print("\n[OK] 全部測試通過 — 從 Sheets 讀的補償計算等同 hardcode 版")


if __name__ == "__main__":
    _self_test()
