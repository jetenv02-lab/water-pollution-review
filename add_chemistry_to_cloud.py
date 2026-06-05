# -*- coding: utf-8 -*-
"""把本地 _槽體學理 分頁新增到雲端協作表。

對比 sheets_sync.upload_xlsx_to_sheets():
    那個是「整本覆蓋」, 會 ws.clear() + ws.update() 所有分頁。
    雖然 Drive comments 理論上會保留, 但 anchor 可能 invalidate, 風險較高。

這支只做一件事: 在雲端新增 _槽體學理 分頁 (如果已存在則先刪再建)。
其他任何分頁的內容、格式、註解, 完全不動。

用法:
    python add_chemistry_to_cloud.py            # dry run
    python add_chemistry_to_cloud.py --apply    # 真執行
"""
import sys
import sheets_sync
import openpyxl
import gspread
from google.oauth2.service_account import Credentials

SHEET_NAME = "_槽體學理"


def main():
    apply = "--apply" in sys.argv

    # 1. 從本地 xlsx 讀 _槽體學理 內容
    wb = openpyxl.load_workbook(sheets_sync.RULES_XLSX, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ 本地 xlsx 沒有 '{SHEET_NAME}' 分頁, 請先跑 python setup_tank_chemistry.py --apply")
        wb.close()
        return
    ws_local = wb[SHEET_NAME]
    rows = []
    for row in ws_local.iter_rows(values_only=True):
        rows.append(["" if v is None else v for v in row])
    wb.close()
    print(f"本地 '{SHEET_NAME}' 讀到 {len(rows)} 列 (含表頭)")

    # 2. 連雲端
    creds = Credentials.from_service_account_file(
        "service_account.json",
        scopes=["https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"],
    )
    client = gspread.authorize(creds)
    sh = client.open_by_key(sheets_sync.DEFAULT_SHEET_ID)
    existing = {ws.title: ws for ws in sh.worksheets()}
    print(f"雲端目前 {len(existing)} 個分頁")

    if SHEET_NAME in existing:
        print(f"  雲端已有 '{SHEET_NAME}', 將先刪除再重建")

    if not apply:
        print("\n(dry-run) 加 --apply 才會真的執行")
        print(f"預計動作: 在雲端{'重建' if SHEET_NAME in existing else '新增'} '{SHEET_NAME}' 分頁 ({len(rows)} 列)")
        return

    # 3. 真執行
    if SHEET_NAME in existing:
        sh.del_worksheet(existing[SHEET_NAME])
        print(f"  已刪除舊的 '{SHEET_NAME}'")

    ws_new = sh.add_worksheet(
        title=SHEET_NAME,
        rows=max(len(rows) + 20, 100),
        cols=max(len(rows[0]) if rows else 1, 10),
    )
    # 寫入
    ws_new.update(values=rows, range_name="A1")

    # 凍結首列
    sh.batch_update({
        "requests": [{
            "updateSheetProperties": {
                "properties": {
                    "sheetId": ws_new.id,
                    "gridProperties": {"frozenRowCount": 1},
                },
                "fields": "gridProperties.frozenRowCount",
            }
        }]
    })

    print(f"\n✅ 已在雲端新增 '{SHEET_NAME}' ({len(rows)} 列)")
    print(f"   雲端現有 {len(sh.worksheets())} 個分頁")
    print("   其他分頁的內容、格式、Faye 的 Drive 註解全部沒動到")


if __name__ == "__main__":
    main()
