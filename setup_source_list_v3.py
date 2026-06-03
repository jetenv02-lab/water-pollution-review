# -*- coding: utf-8 -*-
"""一次性腳本: 把 規則庫.xlsx 的 _來源清單 分頁升級為 v3 結構。

新增 3 欄 (自動算):
  - 涵蓋槽體數
  - 貢獻規則數
  - 最後同步時間

執行前: 請先關閉 Excel
用法: python setup_source_list_v3.py
"""
import os
import shutil
from collections import defaultdict
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE, "規則庫.xlsx")
BACKUP_DIR = os.path.join(BASE, "backup")

# v3 完整欄位 (排原本 8 欄 + 新增 3 欄)
V3_HEADERS = [
    "來源代號", "審查意見檔名", "技師姓名", "技師證書字號",
    "查核日期", "簽證事業名稱", "匯入日期", "備註",
    "涵蓋槽體數", "貢獻規則數", "最後同步時間",
]


def calc_source_stats(wb):
    """從各槽體分頁掃出每個「來源代號」的統計。

    Returns:
        {來源代號: {"tank_count": N, "rule_count": M, "tanks": [...]}}
    """
    stats = defaultdict(lambda: {"tanks": set(), "rule_count": 0})

    for sn in wb.sheetnames:
        if sn.startswith("_"):
            continue
        ws = wb[sn]
        if ws.max_row < 2:
            continue
        headers = [c.value for c in ws[1]]
        try:
            source_idx = headers.index("來源") + 1
        except ValueError:
            continue
        for row in range(2, ws.max_row + 1):
            source = ws.cell(row=row, column=source_idx).value
            if not source:
                continue
            # 「來源」欄通常是 "S01 水污染業務查核缺失.pdf", 取前綴 S01
            source_str = str(source).strip()
            # 從 S01 開頭抓代號
            source_code = source_str.split()[0] if source_str else ""
            if not source_code.startswith("S"):
                source_code = "S01"  # 預設
            stats[source_code]["tanks"].add(sn)
            stats[source_code]["rule_count"] += 1

    # 轉成可序列化
    return {
        code: {
            "tank_count": len(info["tanks"]),
            "rule_count": info["rule_count"],
            "tanks": sorted(info["tanks"]),
        }
        for code, info in stats.items()
    }


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"找不到 {XLSX_PATH}")
        return 1

    try:
        with open(XLSX_PATH, "r+b"):
            pass
    except PermissionError:
        print("❌ 規則庫.xlsx 被 Excel 鎖住, 請先關閉 Excel")
        return 1

    # 備份
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = os.path.join(BACKUP_DIR, f"規則庫_{ts}_pre_source_v3.xlsx")
    shutil.copy2(XLSX_PATH, backup)
    print(f"✓ 備份 → {backup}")

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(XLSX_PATH)

    # 算統計
    stats = calc_source_stats(wb)
    print(f"\n統計結果:")
    for code, s in stats.items():
        print(f"  {code}: 涵蓋 {s['tank_count']} 槽體, 貢獻 {s['rule_count']} 筆規則")

    # 重建 _來源清單
    if "_來源清單" in wb.sheetnames:
        # 保留原本的人工填值 (來源代號 → row dict)
        old_ws = wb["_來源清單"]
        old_headers = [c.value for c in old_ws[1]]
        old_data = {}
        for row in range(2, old_ws.max_row + 1):
            row_vals = {old_headers[i]: old_ws.cell(row=row, column=i + 1).value
                        for i in range(len(old_headers))}
            code = row_vals.get("來源代號", "")
            if code:
                old_data[code] = row_vals
        wb.remove(old_ws)
    else:
        old_data = {}

    # 建新 _來源清單, 放在第 2 個位置 (第 1 個是 _說明)
    insert_pos = 1 if "_說明" in wb.sheetnames else 0
    ws = wb.create_sheet("_來源清單", insert_pos)

    # 表頭
    ws.append(V3_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFE4B5")
        cell.alignment = Alignment(horizontal="center")

    # 寫資料 — 每個來源都寫一列
    # 找出所有來源 (從統計 + 既有人工資料 取聯集)
    all_codes = sorted(set(list(stats.keys()) + list(old_data.keys())))
    sync_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for code in all_codes:
        old = old_data.get(code, {})
        stat = stats.get(code, {"tank_count": 0, "rule_count": 0})
        row = [
            code,
            old.get("審查意見檔名", ""),
            old.get("技師姓名", ""),
            old.get("技師證書字號", ""),
            old.get("查核日期", ""),
            old.get("簽證事業名稱", ""),
            old.get("匯入日期", ""),
            old.get("備註", ""),
            stat["tank_count"],   # 涵蓋槽體數 (自動)
            stat["rule_count"],   # 貢獻規則數 (自動)
            sync_time,            # 最後同步時間 (自動)
        ]
        ws.append(row)

    # 標記「自動」欄背景色
    auto_fill = PatternFill("solid", fgColor="E0F0E0")
    for col_idx in (9, 10, 11):
        ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor="C8E6C9")
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).fill = auto_fill

    # 欄寬
    widths = [10, 30, 12, 14, 12, 16, 12, 30, 12, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 凍結首列
    ws.freeze_panes = "A2"

    wb.save(XLSX_PATH)
    print(f"\n✅ 完成! _來源清單 已升級為 v3 ({len(V3_HEADERS)} 欄)")
    print(f"   - 新欄 (淺綠底色): 涵蓋槽體數 / 貢獻規則數 / 最後同步時間 (自動)")
    print(f"   - 來源筆數: {len(all_codes)}")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
