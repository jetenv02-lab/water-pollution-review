# -*- coding: utf-8 -*-
"""水措審查系統 — Web UI (Flask)。

啟動: python web_app.py
瀏覽器開啟: http://localhost:5000

功能:
  - 顯示規則庫狀態 (各槽體分頁筆數)
  - 顯示參考/需審查之文件/ 內的 PDF 清單
  - 一鍵跑 Step 1c (CSV → xlsx)
  - 一鍵跑完整流程 (Step 2-4) 並即時看輸出
  - 下載審查報告
  - 上傳新的審查意見 PDF / 申請 PDF
"""
import os
import subprocess
import sys
import json
import csv
from datetime import datetime
from flask import Flask, render_template, jsonify, request, send_file, Response, abort
from werkzeug.utils import secure_filename
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
RULES_XLSX = os.path.join(BASE, "規則庫.xlsx")
RULES_CSV = os.path.join(BASE, "rules_extracted.csv")
APP_DIR = os.path.join(BASE, "參考", "需審查之文件")
REVIEW_DIR = os.path.join(BASE, "參考", "審查意見")
REPORT_DIR = os.path.join(BASE, "審查報告")

os.makedirs(APP_DIR, exist_ok=True)
os.makedirs(REVIEW_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200 MB


# ───────────────────────────── 工具函式 ─────────────────────────────

def get_rules_summary():
    """讀取 規則庫.xlsx 各分頁筆數與規則摘要。"""
    if not os.path.exists(RULES_XLSX):
        return None
    wb = openpyxl.load_workbook(RULES_XLSX, read_only=True)
    sheets = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        is_system = sname.startswith("_")
        row_count = ws.max_row - 1 if not is_system else ws.max_row - 1
        sheets.append({
            "name": sname,
            "is_system": is_system,
            "rows": max(row_count, 0),
        })
    return sheets


def get_csv_count():
    if not os.path.exists(RULES_CSV):
        return 0
    with open(RULES_CSV, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        next(reader, None)
        return sum(1 for _ in reader)


def list_application_pdfs():
    if not os.path.exists(APP_DIR):
        return []
    return [{
        "name": f,
        "size_kb": os.path.getsize(os.path.join(APP_DIR, f)) // 1024
    } for f in sorted(os.listdir(APP_DIR)) if f.lower().endswith(".pdf")]


def list_review_pdfs():
    if not os.path.exists(REVIEW_DIR):
        return []
    return [{
        "name": f,
        "size_kb": os.path.getsize(os.path.join(REVIEW_DIR, f)) // 1024
    } for f in sorted(os.listdir(REVIEW_DIR)) if f.lower().endswith(".pdf")]


def list_reports():
    if not os.path.exists(REPORT_DIR):
        return []
    files = []
    for f in sorted(os.listdir(REPORT_DIR), reverse=True):
        fp = os.path.join(REPORT_DIR, f)
        if not os.path.isfile(fp):
            continue
        files.append({
            "name": f,
            "size_kb": os.path.getsize(fp) // 1024,
            "modified": datetime.fromtimestamp(os.path.getmtime(fp)).strftime("%Y-%m-%d %H:%M"),
        })
    return files


def get_review_log():
    """讀規則庫.xlsx 的 _審查紀錄 分頁。"""
    if not os.path.exists(RULES_XLSX):
        return []
    wb = openpyxl.load_workbook(RULES_XLSX, read_only=True)
    if "_審查紀錄" not in wb.sheetnames:
        return []
    ws = wb["_審查紀錄"]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        d = {h: ("" if v is None else str(v)) for h, v in zip(headers, row)}
        rows.append(d)
    return rows


# ───────────────────────────── 路由 ─────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/status")
def api_status():
    """系統狀態總覽。"""
    return jsonify({
        "csv_exists": os.path.exists(RULES_CSV),
        "csv_count": get_csv_count(),
        "xlsx_exists": os.path.exists(RULES_XLSX),
        "rules_summary": get_rules_summary(),
        "application_pdfs": list_application_pdfs(),
        "review_pdfs": list_review_pdfs(),
        "reports": list_reports(),
        "review_log": get_review_log(),
    })


@app.route("/api/run", methods=["POST"])
def api_run():
    """執行某個 step 並回傳輸出。"""
    data = request.json or {}
    step = data.get("step", "")
    target = data.get("target", "")  # 可指定 PDF 路徑

    cmd_map = {
        "1c": [sys.executable, os.path.join(BASE, "step1c_csv_to_xlsx.py")],
        "2": [sys.executable, os.path.join(BASE, "step2_extract_application.py")],
        "3": [sys.executable, os.path.join(BASE, "step3_compare.py")],
        "4": [sys.executable, os.path.join(BASE, "step4_output.py")],
        "run": [sys.executable, os.path.join(BASE, "run.py")],
    }
    if step not in cmd_map:
        return jsonify({"ok": False, "error": f"未知 step: {step}"}), 400

    cmd = cmd_map[step]
    if target:
        target_path = os.path.join(APP_DIR, target)
        if os.path.exists(target_path):
            cmd.append(target_path)

    try:
        result = subprocess.run(
            cmd, cwd=BASE, capture_output=True,
            encoding="utf-8", errors="replace", timeout=600
        )
        return jsonify({
            "ok": result.returncode == 0,
            "returncode": result.returncode,
            "stdout": result.stdout,
            "stderr": result.stderr,
        })
    except subprocess.TimeoutExpired:
        return jsonify({"ok": False, "error": "執行逾時 (10 分鐘)"}), 504
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/upload", methods=["POST"])
def api_upload():
    """上傳 PDF。kind=application 或 review。"""
    kind = request.form.get("kind", "application")
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "沒有檔案"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "error": "檔名為空"}), 400
    if not f.filename.lower().endswith(".pdf"):
        return jsonify({"ok": False, "error": "只接受 PDF"}), 400

    fname = secure_filename(f.filename)
    # 為了支援中文檔名,改用簡單清理
    fname = f.filename.replace("\\", "").replace("/", "").replace("..", "")
    target_dir = APP_DIR if kind == "application" else REVIEW_DIR
    save_path = os.path.join(target_dir, fname)
    f.save(save_path)
    return jsonify({"ok": True, "saved_to": save_path, "filename": fname})


@app.route("/download/<path:subpath>")
def download(subpath):
    """下載審查報告或規則庫。"""
    # 支援: report/檔名、xlsx (規則庫)、csv (rules_extracted)
    if subpath == "xlsx":
        if not os.path.exists(RULES_XLSX):
            abort(404)
        return send_file(RULES_XLSX, as_attachment=True)
    if subpath == "csv":
        if not os.path.exists(RULES_CSV):
            abort(404)
        return send_file(RULES_CSV, as_attachment=True)
    if subpath.startswith("report/"):
        fname = subpath[len("report/"):]
        fp = os.path.join(REPORT_DIR, fname)
        if not os.path.exists(fp):
            abort(404)
        return send_file(fp, as_attachment=True)
    abort(404)


@app.route("/api/rules/<sheet_name>")
def api_rules_sheet(sheet_name):
    """讀規則庫某一分頁的內容。"""
    if not os.path.exists(RULES_XLSX):
        return jsonify({"ok": False, "error": "規則庫不存在"}), 404
    wb = openpyxl.load_workbook(RULES_XLSX, read_only=True)
    if sheet_name not in wb.sheetnames:
        return jsonify({"ok": False, "error": "分頁不存在"}), 404
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append(["" if v is None else str(v) for v in row])
    return jsonify({"ok": True, "headers": headers, "rows": rows})


if __name__ == "__main__":
    print("=" * 60)
    print("水措審查系統 Web UI")
    print("=" * 60)
    print(f"工作目錄: {BASE}")
    print(f"瀏覽器開啟: http://localhost:5000")
    print("按 Ctrl+C 停止")
    print("=" * 60)
    app.run(host="127.0.0.1", port=5000, debug=False)
