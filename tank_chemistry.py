# -*- coding: utf-8 -*-
"""讀取 規則庫.xlsx 的 _槽體學理 分頁, 提供「某槽體某水質項目該不該變動」查詢。

對外 API:
    load_rules() → dict[標準槽體 → 規則]
    classify_item(tank, item_name, rules) → "應變動" | "不應變動" | "未列入"
    check_unit(unit, rules) → list[finding] (跑學理檢查)

設計:
    - 規則表的「類別詞」(SS / 重金屬 / 離子 / 有機物 / 含水率) 自動展開到具體項目
    - 同義詞統一 (例: SS == 懸浮固體（mg/L）== 懸浮固體(mg/L))
    - 比對基準用「質量 (Σ進 vs Σ出)」, 進出 Q 不同也算得準
    - 表讀不到時 fallback 到 hardcoded 預設 (純 pH 槽 / pH 暨快混池)
"""
import os
from functools import lru_cache

# 規則庫 xlsx 預設位置
DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), "規則庫.xlsx")
SHEET_NAME = "_槽體學理"

# ──────────────────────────────────────────────────
# 類別詞展開 — 把「重金屬」等類別詞展開到具體水質項目
# 同個項目可能有多種寫法 (「銅」「銅（mg/L）」「Cu」), 全部列入
# ──────────────────────────────────────────────────
CATEGORY_EXPANSION = {
    # 懸浮固體類
    "SS": ["SS", "ss", "懸浮固體", "懸浮固體（mg/L）", "懸浮固體(mg/L)", "TSS"],

    # 重金屬類
    "重金屬": ["銅", "鎳", "鋅", "鉛", "鎘", "鉻", "總鉻", "六價鉻",
              "錫", "鐵", "錳", "汞", "總汞", "砷", "鉬", "銀", "鈷",
              "Cu", "Ni", "Zn", "Pb", "Cd", "Cr", "Cr⁶⁺", "Cr6+",
              "Sn", "Fe", "Mn", "Hg", "As", "Mo"],

    # 離子類 (溶解性, 不該被沉澱去除)
    "離子": ["氯", "Cl-", "Cl⁻", "氯鹽", "氯化物", "硼", "B",
            "硝酸鹽", "硝酸鹽氮", "NO3-", "NO₃⁻",
            "硫酸鹽", "SO4", "SO₄",
            "氟", "氟鹽", "F-", "F⁻",
            "鈣", "Ca", "鎂", "Mg", "鈉", "Na", "鉀", "K"],

    # 有機物類
    "有機物": ["COD", "BOD", "TOC", "DOC", "有機物",
              "化學需氧量", "化學需氧量（mg/L）",
              "生化需氧量", "生化需氧量（mg/L）",
              "酚", "甲醛", "界面活性劑", "陰離子界面活性劑"],

    # 油脂類
    "油脂": ["油脂", "油脂（mg/L）", "礦物性油脂", "動植物性油脂", "油及脂"],

    # 含水率類
    "含水率": ["含水率", "含水率(%)", "含水率（%）"],

    # 真色色度
    "真色色度": ["真色色度", "色度", "真色色度（mg/L）"],

    # 氨氮類
    "氨氮": ["氨氮", "氨氮（mg/L）", "NH3-N", "NH₃-N", "NH4-N", "NH₄-N"],

    # pH (這個不該被展開, 已經是具體項目)
    "pH": ["pH", "pH值", "PH"],

    # 水溫
    "水溫": ["水溫", "水溫(攝氏)", "水溫（攝氏）", "溫度"],

    # DO
    "DO": ["DO", "溶氧", "溶氧量"],

    # 濁度
    "濁度": ["濁度", "濁度(NTU)", "濁度（NTU）", "NTU"],
}


def _normalize(name):
    """把水質項目名稱「正規化」, 用於同義詞比對。

    去除單位後綴 (mg/L)、空白、全形括號等, 但保留中文項目名。
    """
    if name is None:
        return ""
    s = str(name).strip()
    # 去掉常見單位後綴
    for suffix in ["（mg/L）", "(mg/L)", "（%）", "(%)", "（攝氏）", "(攝氏)",
                   "（NTU）", "(NTU)", "（mg-N/L）", "(mg-N/L)"]:
        if s.endswith(suffix):
            s = s[: -len(suffix)].strip()
    return s


def expand_category(category_or_item):
    """類別詞展開: '重金屬' → 所有重金屬項目列表;
    若已是具體項目 (不在類別表), 直接回 [自己]。
    """
    name = str(category_or_item).strip()
    # 完全比對類別表
    if name in CATEGORY_EXPANSION:
        return list(CATEGORY_EXPANSION[name])
    # 正規化後比對
    norm = _normalize(name)
    if norm in CATEGORY_EXPANSION:
        return list(CATEGORY_EXPANSION[norm])
    # 否則當作具體項目
    return [name]


def expand_list(comma_str):
    """把 "SS, 重金屬, 油脂" 展開成具體項目集合。

    "*" 代表「萬用」, 留給呼叫端特殊處理。
    """
    if not comma_str:
        return set()
    s = str(comma_str).strip()
    if s == "*":
        return {"*"}
    # 半形/全形逗號都拆
    parts = [p.strip() for p in s.replace("，", ",").split(",") if p.strip()]
    result = set()
    for p in parts:
        result.update(expand_category(p))
    return result


# ──────────────────────────────────────────────────
# 從 xlsx 讀規則表
# ──────────────────────────────────────────────────

@lru_cache(maxsize=1)
def load_rules(xlsx_path=None):
    """讀 _槽體學理 分頁, 回傳 dict[標準槽體 → 規則]。

    每筆規則:
        {
            "標準槽體": "pH調整槽",
            "加藥類型": "酸/鹼",
            "應變動原始": "pH, 水溫",         # 原始字串 (給 UI 顯示)
            "應變動集合": {"pH", "pH值", ...}, # 展開後
            "不應變動原始": "*",
            "不應變動集合": {"*"} or {具體項目},
            "容忍度": 5.0,                     # %
            "嚴重度": "不合理",
            "學理說明": "...",
            "狀態": "V",
        }
    """
    if xlsx_path is None:
        xlsx_path = DEFAULT_XLSX
    if not os.path.exists(xlsx_path):
        return {}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb[SHEET_NAME]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()
    except Exception:
        return {}

    def _to_float_or_none(v):
        if v in (None, "", "-"):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    rules = {}
    for r in rows:
        if not r or not r[0]:
            continue
        # 容錯: 欄數不夠就補 None (現在 13 欄)
        r = list(r) + [None] * (14 - len(r))
        (std_tank, dose, allow_raw, deny_raw, tol, sev, desc, status,
         hrt_min, hrt_max, sor_max, g_min, g_max, tank_type) = r[:14]

        # 狀態為 "?" 視為「待討論, 暫不啟用」
        if status and str(status).strip() == "?":
            continue

        try:
            tol_pct = float(tol) if tol not in (None, "") else 10.0
        except (TypeError, ValueError):
            tol_pct = 10.0

        rules[str(std_tank).strip()] = {
            "標準槽體": str(std_tank).strip(),
            "加藥類型": str(dose or "").strip(),
            "應變動原始": str(allow_raw or "").strip(),
            "應變動集合": expand_list(allow_raw),
            "不應變動原始": str(deny_raw or "").strip(),
            "不應變動集合": expand_list(deny_raw),
            "容忍度": tol_pct,
            "嚴重度": str(sev or "待人工").strip(),
            "學理說明": str(desc or "").strip(),
            "狀態": str(status or "").strip(),
            # 設計參數學理範圍 (新增)
            "HRT_min": _to_float_or_none(hrt_min),
            "HRT_max": _to_float_or_none(hrt_max),
            "SOR_max": _to_float_or_none(sor_max),
            "G_min": _to_float_or_none(g_min),
            "G_max": _to_float_or_none(g_max),
            "類型": (str(tank_type).strip() if tank_type else ""),
        }
    return rules


def get_rule_for_unit(unit, rules=None):
    """給一個 unit dict, 找對應的學理規則。

    優先 name_in_doc, 後 std_tank (因為 std_tank 偶爾被誤分類)。
    """
    if rules is None:
        rules = load_rules()
    if not rules:
        return None

    for key in ("name_in_doc", "std_tank"):
        v = unit.get(key)
        if v and v in rules:
            return rules[v]
    return None


def classify_item(item_name, rule):
    """這個水質項目在這個槽體裡應該屬於哪一類?

    Returns: "應變動" / "不應變動" / "未明確"
    """
    if not rule:
        return "未明確"
    norm = _normalize(item_name)

    allow = rule["應變動集合"]
    deny = rule["不應變動集合"]

    # 1. 直接命中
    if item_name in allow or norm in allow:
        return "應變動"
    if item_name in deny or norm in deny:
        return "不應變動"

    # 2. 萬用 "*"
    # 「應變動」是 * → 全部允許 (例: 暫存槽); 但「不應變動」明列的優先
    # 「不應變動」是 * → 除應變動外都不該變
    if "*" in deny:
        return "不應變動"  # 隱含: 不在應變動裡的都歸這
    if "*" in allow:
        return "應變動"    # 隱含: 不在不應變動裡的都允許

    return "未明確"


# ──────────────────────────────────────────────────
# 主檢查函式
# ──────────────────────────────────────────────────

def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def check_unit(unit, rules=None):
    """跑「槽體學理」檢查。

    對每個進+出兩邊都有資料的水質項目:
        - 算 Σ進質量 vs Σ出質量, 變動 %
        - 查該項目在此槽體應該「應變動 / 不應變動 / 未明確」
        - 若標 "不應變動" 但實際變動 > 容忍度 → 產生 finding

    Returns: list of finding dicts
    """
    if rules is None:
        rules = load_rules()
    rule = get_rule_for_unit(unit, rules)
    if not rule:
        return []

    code = unit.get("raw_code") or unit.get("code") or "?"
    std_tank = rule["標準槽體"]
    tol = rule["容忍度"]
    severity = rule["嚴重度"]
    desc_text = rule["學理說明"]

    influent = unit.get("influent", {}) or {}
    effluent = unit.get("effluent", {}) or {}
    if not influent or not effluent:
        return []

    # 收集所有水質項目
    all_items = set()
    for stream in list(influent.values()) + list(effluent.values()):
        if isinstance(stream, dict):
            all_items.update(stream.keys())

    # 是否為「分流結構」(本單元 effluent ≥ 2 條) — 用於濃度閘判斷
    self_split = len(effluent) >= 2

    findings = []
    for item in sorted(all_items):
        cls = classify_item(item, rule)
        if cls != "不應變動":
            continue

        # 算質量平衡, 同時累加 質量 + 濃度×Q (用於加權平均濃度)
        in_mass = 0.0
        out_mass = 0.0
        in_mass_for_conc = 0.0   # 進: Σ (濃度 × 質量) — 但這裡其實質量已含 Q 因素
        in_q_sum = 0.0
        out_q_sum = 0.0
        in_conc_x_q = 0.0  # 進: Σ (濃度 × Q) — 用於算加權平均濃度
        out_conc_x_q = 0.0
        in_has = False
        out_has = False
        for stream in influent.values():
            if not isinstance(stream, dict):
                continue
            v = stream.get(item)
            if isinstance(v, dict):
                m = _to_float(v.get("質量"))
                c = _to_float(v.get("濃度"))
                q = _to_float(v.get("Q") or v.get("q") or v.get("q_cmd"))
                if m is not None:
                    in_mass += m
                    in_has = True
                if c is not None and q is not None and q > 0:
                    in_conc_x_q += c * q
                    in_q_sum += q
        for stream in effluent.values():
            if not isinstance(stream, dict):
                continue
            v = stream.get(item)
            if isinstance(v, dict):
                m = _to_float(v.get("質量"))
                c = _to_float(v.get("濃度"))
                q = _to_float(v.get("Q") or v.get("q") or v.get("q_cmd"))
                if m is not None:
                    out_mass += m
                    out_has = True
                if c is not None and q is not None and q > 0:
                    out_conc_x_q += c * q
                    out_q_sum += q

        if not (in_has and out_has and in_mass > 0):
            continue

        diff_pct = abs(out_mass - in_mass) / in_mass * 100
        # 絕對最小容忍 0.5% (避免 PDF 抽取的舍入誤差被當成違規)
        effective_tol = max(tol, 0.5)
        if diff_pct <= effective_tol:
            continue  # 在容忍度內

        # ── 加權平均濃度閘 ──
        # 若本單元為「水量分流」, 質量看起來變化大但濃度其實沒變 → 跳過
        # (技師備註: T01-21 慢混分流, 鋅濃度未改變, 但質量按 Σ 出/Σ 進看起來減 50%)
        if self_split and in_q_sum > 0 and out_q_sum > 0:
            in_avg_c = in_conc_x_q / in_q_sum
            out_avg_c = out_conc_x_q / out_q_sum
            if in_avg_c > 0:
                conc_diff_pct = abs(out_avg_c - in_avg_c) / in_avg_c * 100
                if conc_diff_pct <= max(tol, 5.0):
                    # 濃度沒明顯變 → 水量分流造成的質量變化, 不算去除
                    # 寫進 unit 的 topology_notes 當備註, 不產 finding
                    note_lines = unit.setdefault("topology_notes", [])
                    note_lines.append(
                        f"ℹ️ 拓樸提示: {item} 進出加權平均濃度 "
                        f"{in_avg_c:.2f} → {out_avg_c:.2f} (Δ {conc_diff_pct:.1f}%), "
                        f"質量差異 {diff_pct:.1f}% 來自水量分流而非實際去除。"
                    )
                    continue

        direction = "減少" if out_mass < in_mass else "增加"
        findings.append({
            "嚴重度": severity,
            "類型": "質量平衡",
            "單元": code,
            "標準槽體": std_tank,
            "對照項目": item,
            "描述": (
                f"{item} 質量 進 {in_mass:.3f} → 出 {out_mass:.3f} kg/d "
                f"({direction} {diff_pct:.1f}%, 容忍 {tol}%)。{desc_text}"
            ),
            "依據": f"_槽體學理 規則: {std_tank} ({rule['加藥類型']})",
        })

    return findings


def clear_cache():
    """規則庫被改後可呼叫此清快取。"""
    load_rules.cache_clear()
