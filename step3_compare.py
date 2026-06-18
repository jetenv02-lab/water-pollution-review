# -*- coding: utf-8 -*-
"""Step 3: 規則 × 申請資料 → 觸發比對結果。

輸入:
  - 規則庫.xlsx (各槽體分頁的規則)
  - application_*.json (Step 2 萃取的申請資料)

輸出:
  - comparison_{案名}.json — 結構化比對結果
  - comparison_{案名}.csv — 同份扁平 CSV (易於人工檢視)

比對策略:
  對申請資料的每個處理單元(如 T01-03 批次反應槽):
    1) 找到該單元對應的『標準槽體類型』分頁(批次反應槽)
    2) 把該分頁的所有規則套用到該單元
    3) 嘗試從『判定邏輯』欄位機械化解析 IF-THEN
       - 解析失敗 → 標記為「待確認判定」+ 規則原文
    4) 另外把『(文件類)』『(現場設備類)』的規則也納入比對(全域規則)

依賴: openpyxl
用法: python step3_compare.py "application_xxx.json"
"""
import csv
import json
import os
import re
import sys
from datetime import datetime
import openpyxl

BASE = r"C:\Users\jeten\Desktop\AI\水措審查"
RULES_XLSX = os.path.join(BASE, "規則庫.xlsx")

# 機械化解析的判定邏輯 pattern (簡易版,後續可擴充)
# 例如: "若 用途含『去除重金屬』且 pH下限<=6 → 標記:..."
PH_LOWER_PATTERN = re.compile(r"pH\s*下限\s*[<≤]=?\s*(\d+(?:\.\d+)?)")
PH_UPPER_PATTERN = re.compile(r"pH\s*上限\s*[>≥]=?\s*(\d+(?:\.\d+)?)")


def load_rules(xlsx_path):
    """讀規則庫所有槽體分頁,回傳 { 標準槽體名稱: [規則 dict, ...] }。"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    rules_by_tank = {}
    for sname in wb.sheetnames:
        if sname.startswith("_"):
            continue
        ws = wb[sname]
        headers = [c.value for c in ws[1]]
        rules = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            d = {h: v for h, v in zip(headers, row) if h is not None}
            rules.append(d)
        rules_by_tank[sname] = rules
    return rules_by_tank


def parse_ph_range(s):
    """從『6~9』『8.5-10』之類字串解析 (min, max)。"""
    if not s:
        return None, None
    s = str(s).replace("～", "~").replace("－", "-").replace("–", "-")
    m = re.search(r"(\d+(?:\.\d+)?)\s*[~\-]\s*(\d+(?:\.\d+)?)", s)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def evaluate_rule(rule, unit_info):
    """套用單一規則到單元資料,回傳 (是否觸發, 觸發描述)。
    目前只實作 pH 範圍檢查,其他規則先回 (None, 規則原文+待確認判定)。
    """
    logic = (rule.get("判定邏輯(條件→結論)") or rule.get("判定邏輯") or "").strip()
    rule_text = (rule.get("規則(萃取/可比對判斷式)") or rule.get("規則") or "").strip()
    check_type = (rule.get("檢查類型") or "").strip()

    ph_value = unit_info.get("design_params", {}).get("pH", "")
    if "pH" in (rule.get("對照項目") or "") or "pH" in logic:
        lo, hi = parse_ph_range(ph_value)
        if lo is None:
            return None, f"無法解析 pH(申請文件: {ph_value!r}) — 待確認"
        # 規則中 pH 下限門檻
        m = PH_LOWER_PATTERN.search(logic) or re.search(r"pH\s*下限\s*<\s*(\d+(?:\.\d+)?)", logic)
        if m:
            threshold = float(m.group(1))
            if lo < threshold:
                return True, f"申請 pH 下限 {lo} < 規則門檻 {threshold}"
            else:
                return False, f"申請 pH 下限 {lo} ≥ 規則門檻 {threshold}"

    return None, "規則需人工判定: " + rule_text[:80]


def compare(app_json_path):
    print(f"=== 載入申請資料: {app_json_path} ===")
    with open(app_json_path, "r", encoding="utf-8") as f:
        app = json.load(f)

    print(f"=== 載入規則庫: {RULES_XLSX} ===")
    rules_by_tank = load_rules(RULES_XLSX)
    print(f"規則庫分頁: {list(rules_by_tank.keys())}")

    findings = []  # 每筆: {unit_code, std_tank, rule_id, 判定, 描述, 規則原文}

    # 全域規則(不綁特定槽體)
    global_rule_tanks = [k for k in rules_by_tank if k.startswith("(") or k in ("文件類", "現場設備類")]

    for code, info in app.get("units", {}).items():
        std_tank = info.get("std_tank", "")
        # 對該槽體的所有規則
        candidate_rules = list(rules_by_tank.get(std_tank, []))
        for rule in candidate_rules:
            triggered, desc = evaluate_rule(rule, info)
            findings.append({
                "申請單元": code,
                "標準槽體": std_tank,
                "缺失ID": rule.get("缺失ID", ""),
                "規則來源": rule.get("來源", ""),
                "檢查類型": rule.get("檢查類型", ""),
                "對照項目": rule.get("對照項目", ""),
                "判定": ("不合理" if triggered else "合理" if triggered is False else "待確認"),
                "描述": desc,
                "規則原文": rule.get("規則(萃取/可比對判斷式)", rule.get("規則", "")),
                "原文缺失": (rule.get("原文缺失") or "")[:200],
            })

    # 輸出
    base = os.path.splitext(os.path.basename(app_json_path))[0]
    out_json = os.path.join(BASE, f"comparison_{base}.json")
    out_csv = os.path.join(BASE, f"comparison_{base}.csv")

    result = {
        "application": app.get("source_pdf"),
        "compared_at": datetime.now().isoformat(),
        "rules_xlsx": RULES_XLSX,
        "total_findings": len(findings),
        "findings": findings,
    }
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    with open(out_csv, "w", encoding="utf-8-sig", newline="") as f:
        if findings:
            w = csv.DictWriter(f, fieldnames=list(findings[0].keys()))
            w.writeheader()
            w.writerows(findings)

    # 統計
    stats = {"不合理": 0, "合理": 0, "待確認": 0}
    for fi in findings:
        stats[fi["判定"]] += 1

    print(f"\n=== 比對結果 ===")
    print(f"總比對數: {len(findings)}")
    print(f"  不合理: {stats['不合理']}")
    print(f"  合理: {stats['合理']}")
    print(f"  待確認: {stats['待確認']}")
    print(f"輸出: {out_json}")
    print(f"輸出: {out_csv}")
    return result


def main():
    if len(sys.argv) >= 2:
        app_json = sys.argv[1]
    else:
        # 預設掃工作目錄裡的 application_*.json (取最新一份)
        jsons = sorted([f for f in os.listdir(BASE) if f.startswith("application_") and f.endswith(".json")])
        if not jsons:
            print("找不到 application_*.json,請先跑 step2")
            return
        app_json = os.path.join(BASE, jsons[-1])
    if not os.path.exists(app_json):
        print(f"找不到: {app_json}")
        return
    compare(app_json)


if __name__ == "__main__":
    main()
