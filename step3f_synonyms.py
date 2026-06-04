# -*- coding: utf-8 -*-
"""水措審查 — 同義字儲。

設計:
    - 維護一個「標準詞 → 別名清單」的 dict
    - 提供 normalize() 把任何別名 → 標準詞
    - 提供 are_synonyms() 判斷兩個詞是否為同義字
    - 提供 expand() 給一個詞, 回所有變體 (含自己) 給比對用
    - 從 規則庫.xlsx 的 _同義字 分頁載入 (若存在, 覆蓋預設)

使用情境:
    step3e 比對「對照項目」時, 用 normalize() 統一詞彙
    gemini_extractor 的 prompt 加上「這幾個詞算同一件事」
"""
import os

BASE = os.path.dirname(os.path.abspath(__file__))
RULES_XLSX = os.path.join(BASE, "規則庫.xlsx")

# ──────────────────────────────────────────────────
# 預設同義字 (萬一 xlsx 沒有 _同義字 分頁就用這個)
# 每組: 第一個 = 標準詞, 後面是別名
# ──────────────────────────────────────────────────
DEFAULT_SYNONYMS = {
    # ── 槽體尺寸 ──
    "出水高度": ["液面到槽頂距離", "自由水面高度", "餘裕高度", "乾舷", "freeboard"],
    "有效容量": ["有效體積", "有效水量", "operating volume", "工作容積"],
    "有效水深": ["有效深度", "操作水深", "operating depth"],
    "槽體尺寸": ["槽體規格", "tank dimension"],

    # ── 操作參數 ──
    "滯留時間": ["停留時間", "HRT", "水力停留時間", "接觸時間", "反應時間"],
    "曝氣量": ["風量", "曝氣風量", "供氣量", "air flow"],
    "迴流比": ["污泥迴流率", "迴流污泥比", "RAS"],
    "MLSS": ["混合液懸浮固體", "活性污泥濃度"],
    "DO": ["溶氧", "溶氧濃度", "dissolved oxygen"],
    "SVI": ["污泥容積指數", "沉降比"],
    "F/M": ["食微比", "BOD 負荷率"],

    # ── 機具設施 ──
    "液位計": ["水位計", "液面計", "level transmitter", "LT", "level sensor"],
    "pH計": ["pH meter", "酸鹼度計", "pH 計", "pH 感測器"],
    "DO計": ["溶氧計", "dissolved oxygen meter", "DO 計"],
    "流量計": ["FT", "flow meter", "流速計", "flowmeter"],
    "攪拌機": ["攪拌器", "攪拌槳", "agitator", "mixer", "攪拌設備"],
    "曝氣機": ["鼓風機", "blower", "氣泵", "air blower"],
    "加藥泵": ["藥劑泵", "投藥泵", "dosing pump", "計量泵"],
    "刮泥機": ["刮砂機", "scraper", "sludge scraper"],
    "污泥泵": ["污泥輸送泵", "sludge pump"],

    # ── 化學處理 ──
    "反洗水": ["反沖洗水", "回沖洗", "backwash", "逆洗水"],
    "加藥": ["投藥", "添加藥品", "dosing", "藥品投加"],
    "中和": ["pH 調整", "酸鹼中和", "neutralization"],
    "混凝": ["膠凝", "coagulation"],
    "膠凝": ["絮凝", "flocculation"],

    # ── 水質參數 ──
    "BOD": ["生化需氧量", "BOD5", "biochemical oxygen demand"],
    "COD": ["化學需氧量", "chemical oxygen demand"],
    "SS": ["懸浮固體", "懸浮物", "suspended solids"],
    "TKN": ["總凱氏氮", "total Kjeldahl nitrogen"],
    "TP": ["總磷", "total phosphorus"],
    "TN": ["總氮", "total nitrogen"],

    # ── 流向 ──
    "進流水": ["入流水", "進水", "influent", "原水"],
    "出流水": ["放流水", "排水", "effluent", "出水"],
    "原廢水": ["原始廢水", "raw wastewater", "原廢"],
    "放流口": ["排放口", "outlet", "discharge point"],

    # ── 污泥 ──
    "污泥含水率": ["含水率", "moisture content", "水含量"],
    "脫水": ["污泥脫水", "dewatering"],
    "濃縮": ["污泥濃縮", "thickening"],
}


# ──────────────────────────────────────────────────
# 載入 (從 xlsx 的 _同義字 分頁覆蓋預設, 若有)
# ──────────────────────────────────────────────────

_SYNONYMS_CACHE = None
_REVERSE_CACHE = None  # 別名 → 標準詞


def _load_from_xlsx():
    """從 規則庫.xlsx 的 _同義字 分頁載入。

    分頁格式:
        | 標準詞 | 別名 (用 / 或 , 或 ; 分隔) |
        | 出水高度 | 液面到槽頂距離 / 自由水面高度 / 餘裕高度 |
    """
    if not os.path.exists(RULES_XLSX):
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(RULES_XLSX, data_only=True, read_only=True)
        if "_同義字" not in wb.sheetnames:
            return None
        ws = wb["_同義字"]
        if ws.max_row < 2:
            return None

        # 找欄位 (容錯: 接受 "標準詞"/"別名" 或 "標準"/"同義字" 等)
        headers = [c.value for c in ws[1]]
        std_col = None
        alias_col = None
        for i, h in enumerate(headers):
            if not h:
                continue
            h_str = str(h).strip()
            if h_str in ("標準詞", "標準", "primary"):
                std_col = i
            elif h_str in ("別名", "同義字", "alias", "synonyms"):
                alias_col = i
        if std_col is None or alias_col is None:
            return None

        out = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            std = row[std_col] if std_col < len(row) else None
            aliases_raw = row[alias_col] if alias_col < len(row) else None
            if not std:
                continue
            std = str(std).strip()
            if not std:
                continue
            if aliases_raw:
                aliases = [
                    a.strip()
                    for a in str(aliases_raw).replace(";", "/").replace(",", "/").split("/")
                    if a.strip()
                ]
            else:
                aliases = []
            out[std] = aliases
        return out if out else None
    except Exception:
        return None


def _build_reverse(synonyms):
    """從 {標準詞: [別名...]} 建反向 {任一詞 (lower): 標準詞}。"""
    reverse = {}
    for std, aliases in synonyms.items():
        reverse[std.lower()] = std  # 標準詞也對應到自己
        for a in aliases:
            reverse[a.lower()] = std
    return reverse


def _ensure_loaded():
    """lazy load: 第一次使用時才載入 xlsx 並建 cache。"""
    global _SYNONYMS_CACHE, _REVERSE_CACHE
    if _SYNONYMS_CACHE is None:
        # 先 deepcopy 預設
        merged = {k: list(v) for k, v in DEFAULT_SYNONYMS.items()}
        # 用 xlsx 覆蓋 (若有同樣標準詞, xlsx 版本贏)
        xlsx_data = _load_from_xlsx()
        if xlsx_data:
            for std, aliases in xlsx_data.items():
                merged[std] = aliases
        _SYNONYMS_CACHE = merged
        _REVERSE_CACHE = _build_reverse(merged)
    return _SYNONYMS_CACHE, _REVERSE_CACHE


def reload():
    """強制重載 (例如使用者剛在 Sheet 改了 _同義字, 下載回 xlsx 後)。"""
    global _SYNONYMS_CACHE, _REVERSE_CACHE
    _SYNONYMS_CACHE = None
    _REVERSE_CACHE = None
    return _ensure_loaded()


# ──────────────────────────────────────────────────
# 公開 API
# ──────────────────────────────────────────────────

def normalize(term):
    """把任何詞 (標準或別名) 轉成標準詞。

    >>> normalize("液面到槽頂距離") -> "出水高度"
    >>> normalize("出水高度") -> "出水高度"
    >>> normalize("不認識的詞") -> "不認識的詞" (回傳原值)
    """
    if not term:
        return term
    _, reverse = _ensure_loaded()
    term_str = str(term).strip()
    return reverse.get(term_str.lower(), term_str)


def are_synonyms(a, b):
    """判斷兩個詞是否為同義字 (含「相同」也算)。"""
    if not a or not b:
        return False
    return normalize(a) == normalize(b)


def expand(term):
    """給一個詞, 回傳 [標準詞, 所有別名] (含自己)。

    用途: 比對「對照項目」時, 把規則上寫的詞展開成所有變體, 再去文件裡找。
    """
    if not term:
        return [term] if term else []
    synonyms, _ = _ensure_loaded()
    std = normalize(term)
    if std in synonyms:
        return [std] + synonyms[std]
    return [term]


def get_all_synonyms():
    """回傳完整 {標準詞: [別名...]} dict (給 UI 顯示用)。"""
    synonyms, _ = _ensure_loaded()
    return dict(synonyms)


def stats():
    """統計資訊。"""
    synonyms, reverse = _ensure_loaded()
    return {
        "標準詞數": len(synonyms),
        "總詞彙數 (含別名)": len(reverse),
        "別名總數": len(reverse) - len(synonyms),
    }


# ──────────────────────────────────────────────────
# CLI 測試
# ──────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    import io
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    except Exception:
        pass

    print("=== 同義字儲統計 ===")
    print(stats())
    print()

    print("=== 測試 normalize ===")
    for t in ["液面到槽頂距離", "出水高度", "HRT", "停留時間",
              "水位計", "液位計", "不認識的詞", "lt"]:
        print(f"  {t!r} -> {normalize(t)!r}")
    print()

    print("=== 測試 expand ===")
    for t in ["出水高度", "液位計", "DO"]:
        print(f"  {t!r} -> {expand(t)}")
    print()

    print("=== 測試 are_synonyms ===")
    for a, b in [("液面到槽頂距離", "出水高度"),
                 ("HRT", "停留時間"),
                 ("BOD", "COD")]:
        print(f"  {a!r} ~ {b!r}: {are_synonyms(a, b)}")
