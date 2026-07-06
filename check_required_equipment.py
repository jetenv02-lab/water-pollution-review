# -*- coding: utf-8 -*-
"""必備機具檢查 — 用 _槽體學理.必備機具 欄過濾, 只對該槽體真的需要的機具做檢查.

改進點 (Nick 定調, 2026-07-06):
    - 舊版 step3e 對每條 rules_extracted.csv 都跳「機具未列 X」, 誤觸發率高
      (例: 砂濾塔被套「攪拌機未列」, T01-11 暫存槽被套「排泥未列」...)
    - 新版: 讀 _槽體學理.必備機具 欄, 只有該槽體必備的才檢查
    - equipment_list 空時: 改成「疑抽取失敗, 請人工確認」而非斷定「未列」
"""
import os
from functools import lru_cache

BASE = os.path.dirname(os.path.abspath(__file__))
RULES_XLSX = os.path.join(BASE, "規則庫.xlsx")


# 機具名關鍵字 → 描述
EQ_HINT = {
    "攪拌機": "反應槽需攪拌以確保混合均勻",
    "加藥機": "加藥槽需加藥機以控制劑量",
    "pH計": "pH 反應槽應有 pH 監控",
    "ORP計": "氧化還原槽應有 ORP 監控",
    "排泥": "沉澱/濃縮槽應有排泥機具或方式",
    "反洗": "過濾/吸附裝置應有反洗系統",
    "鼓風機": "曝氣槽/生物處理需鼓風機供氧",
    "液位計": "儲存/緩衝槽應有液位計避免溢流",
    "流量計": "放流/計量點應有流量計",
    "再生系統": "離子交換樹脂需再生系統",
}


@lru_cache(maxsize=1)
def load_required_equipment():
    """讀 _槽體學理.必備機具 欄 → {標準槽體: [機具名, ...]}."""
    try:
        import openpyxl
    except ImportError:
        return {}
    try:
        wb = openpyxl.load_workbook(RULES_XLSX, data_only=True, read_only=True)
    except Exception:
        return {}
    if "_槽體學理" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["_槽體學理"]

    # 找「必備機具」欄
    header_row = 1
    col_idx = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v and str(v).strip() == "必備機具":
            col_idx = c
            break
    if not col_idx:
        wb.close()
        return {}

    result = {}
    for r in range(2, ws.max_row + 1):
        tank = ws.cell(r, 1).value
        eq_str = ws.cell(r, col_idx).value
        if not tank:
            continue
        tank = str(tank).strip()
        if not eq_str:
            result[tank] = []  # 該槽體沒必備機具需求
            continue
        eq_list = [x.strip() for x in str(eq_str).split(",") if x.strip()]
        result[tank] = eq_list

    wb.close()
    return result


EQ_ALIAS = {
    # 排泥 → 泵浦類/氣動泵/刮泥機都算
    "排泥": ["排泥", "污泥泵", "氣動泵", "污泥抽送", "污泥輸送", "污泥排出",
             "刮泥機", "污泥迴流泵", "螺旋輸送", "螺旋泵"],
    "攪拌機": ["攪拌機", "攪拌器", "攪拌", "混合機", "膠凝機"],
    "加藥機": ["加藥機", "加藥泵", "計量泵", "定量泵", "藥液泵", "藥劑泵"],
    # "pH計" 也認「pH 計」(廠商常打中間空格) 及氟離子計 (中和槽同類監控)
    "pH計": ["pH計", "pH 計", "pH", "酸鹼度", "酸鹼度計", "氟離子計", "離子計"],
    "ORP計": ["ORP", "氧化還原", "還原電位"],
    # 反洗 → 差壓計常用於反洗觸發 (含差壓計代表有反洗系統的 signal)
    "反洗": ["反洗", "逆洗", "backwash", "自動反洗", "差壓計"],
    "鼓風機": ["鼓風機", "空壓機", "曝氣機", "送風機"],
    "液位計": ["液位計", "液位", "液面計", "浮球", "浮球開關"],
    "流量計": ["流量計", "電磁流量", "累計流量", "積算器"],
    "再生系統": ["再生", "再生塔", "再生槽", "再生泵"],
}


def has_equipment(equipment_list, name_keyword):
    """檢查機具清單是否含某關鍵字 (支援同義字)."""
    if not equipment_list:
        return False
    aliases = EQ_ALIAS.get(name_keyword, [name_keyword])
    for eq in equipment_list:
        if not isinstance(eq, dict):
            continue
        eq_name = str(eq.get("name") or "")
        for a in aliases:
            if a in eq_name:
                return True
    return False


def check_unit_required_equipment(unit, required_map=None):
    """檢查單元的必備機具.

    Args:
        unit: 處理單元 dict (要有 std_tank + equipment)
        required_map: 必備機具對照表 (None 自動載入)

    Returns:
        list[finding]
    """
    if required_map is None:
        required_map = load_required_equipment()

    std_tank = str(unit.get("std_tank") or "").strip()
    if not std_tank:
        return []
    required = required_map.get(std_tank)
    if required is None or not required:
        return []  # 該槽體無必備機具需求

    code = unit.get("raw_code") or unit.get("code_id") or "?"
    # step2 存的是 "equipment", 有些舊路徑用 "equipment_list"
    eq_list = unit.get("equipment")
    if eq_list is None:
        eq_list = unit.get("equipment_list") or []

    findings = []
    for req_eq in required:
        if has_equipment(eq_list, req_eq):
            continue
        hint = EQ_HINT.get(req_eq, "")
        if not eq_list:
            # equipment_list 完全空 → 疑抽取失敗
            findings.append({
                "嚴重度": "待確認",
                "類型": "機具設施",
                "單元": code,
                "標準槽體": std_tank,
                "對照項目": f"必備機具: {req_eq}",
                "描述": (
                    f"{std_tank} 應具備 {req_eq}, 但單元機具清單完全空白 "
                    f"(可能 PDF 表格抽取失敗, 或廠商未登載)。"
                    f"{('學理: ' + hint) if hint else ''} 請人工核對 PDF 該單元「相關機具設施」欄."
                ),
                "依據": f"_槽體學理.必備機具 ({std_tank})",
            })
        else:
            # equipment_list 有東西但沒這一項 → 斷定「未列」
            findings.append({
                "嚴重度": "不合理",
                "類型": "機具設施",
                "單元": code,
                "標準槽體": std_tank,
                "對照項目": f"必備機具: {req_eq}",
                "描述": (
                    f"{std_tank} 應具備 {req_eq}, 但單元機具清單未列。"
                    f"{('學理: ' + hint) if hint else ''}"
                ),
                "依據": f"_槽體學理.必備機具 ({std_tank})",
            })

    return findings


def _collect_series_equipment(units):
    """收集每個系列 (T01/T02/T03..) 的所有 equipment name.

    廠商常用共用單元序號登加藥機/攪拌機, step2 沒展開,
    → 用系列內共享, 若同系列其他單元登了, 該單元 pass.
    """
    series_eq = {}  # {series_prefix: set of eq_name}
    if isinstance(units, dict):
        unit_iter = units.values()
    else:
        unit_iter = units
    for u in unit_iter:
        if not isinstance(u, dict):
            continue
        code = str(u.get("raw_code") or u.get("code_id") or "")
        if "-" not in code:
            continue
        prefix = code.split("-")[0]  # T01, T02, ...
        eq_list = u.get("equipment") or u.get("equipment_list") or []
        if prefix not in series_eq:
            series_eq[prefix] = set()
        for e in eq_list:
            if isinstance(e, dict):
                name = str(e.get("name") or "")
                if name:
                    series_eq[prefix].add(name)
    return series_eq


def check_all_units_required_equipment(units, required_map=None):
    """批次檢查全廠必備機具.

    含系列共用邏輯: 廠商常用「共用單元序號」欄登共用機具,
    step2 目前沒展開, 所以在此以「系列 (T01, T02..) 共享」近似補償.
    """
    if required_map is None:
        required_map = load_required_equipment()

    # 收集每系列的所有 equipment name (共享)
    series_eq = _collect_series_equipment(units)

    findings = []
    if isinstance(units, dict):
        unit_iter = list(units.values())
    else:
        unit_iter = list(units)

    for unit in unit_iter:
        if not isinstance(unit, dict):
            continue
        code = str(unit.get("raw_code") or unit.get("code_id") or "")
        prefix = code.split("-")[0] if "-" in code else ""
        series_all_eq = series_eq.get(prefix, set())

        # 拿本單元 + 系列共享 的 equipment 一起檢查
        eq_self = unit.get("equipment") or unit.get("equipment_list") or []
        eq_combined = list(eq_self)
        for name in series_all_eq:
            # 用 fake dict 加入以便 has_equipment 掃描
            if not any(isinstance(e, dict) and str(e.get("name") or "") == name
                       for e in eq_combined):
                eq_combined.append({"name": name, "_from_series": True})

        # 手動跑一次 check_unit_required_equipment 邏輯 (用 eq_combined 取代 eq_list)
        std_tank = str(unit.get("std_tank") or "").strip()
        if not std_tank:
            continue
        required = required_map.get(std_tank)
        if required is None or not required:
            continue
        eq_code = unit.get("raw_code") or unit.get("code_id") or "?"

        for req_eq in required:
            if has_equipment(eq_combined, req_eq):
                continue
            hint = EQ_HINT.get(req_eq, "")
            if not eq_self:
                findings.append({
                    "嚴重度": "待確認",
                    "類型": "機具設施",
                    "單元": eq_code,
                    "標準槽體": std_tank,
                    "對照項目": f"必備機具: {req_eq}",
                    "描述": (
                        f"{std_tank} 應具備 {req_eq}, 但單元機具清單完全空白 "
                        f"(可能 PDF 表格抽取失敗, 或廠商未登載)。"
                        f"{('學理: ' + hint) if hint else ''} 請人工核對 PDF."
                    ),
                    "依據": f"_槽體學理.必備機具 ({std_tank})",
                })
            else:
                findings.append({
                    "嚴重度": "不合理",
                    "類型": "機具設施",
                    "單元": eq_code,
                    "標準槽體": std_tank,
                    "對照項目": f"必備機具: {req_eq}",
                    "描述": (
                        f"{std_tank} 應具備 {req_eq}, 但單元機具清單未列 "
                        f"(同系列 {prefix} 也未見)。"
                        f"{('學理: ' + hint) if hint else ''}"
                    ),
                    "依據": f"_槽體學理.必備機具 ({std_tank})",
                })

    return findings


def clear_cache():
    load_required_equipment.cache_clear()


if __name__ == "__main__":
    # 自我測試
    m = load_required_equipment()
    print("讀到 {} 個槽體必備機具規則".format(len(m)))
    for tank in ["快混槽", "沉澱池", "砂濾塔", "暫存槽", "曝氣槽"]:
        print("  {}: {}".format(tank, m.get(tank)))
